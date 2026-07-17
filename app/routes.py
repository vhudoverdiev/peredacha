from __future__ import annotations

from collections import Counter
from copy import copy
from datetime import date, datetime, timedelta
import json
import random
import re
import traceback
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import urlparse
from sqlalchemy import or_, and_, exists
from werkzeug.exceptions import HTTPException

from flask import Blueprint, abort, current_app, flash, redirect, render_template, request, send_file, session, url_for, jsonify
from flask_login import current_user, login_required
from sqlalchemy import Integer, cast, distinct, func
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side
from openpyxl.utils import get_column_letter

EXCEL_HEADER_FILL_COLOR = "FFE2F0D9"
EXCEL_DOWNLOAD_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
MEASUREMENT_REQUEST_COMMENT = "Заявка создана из раздела «Замеры»"
MEASUREMENT_REQUEST_TITLE_PATTERN = re.compile(r"^Заявка из замеров №(\d+)$")
MEASUREMENT_WRITEOFF_COMMENT_PREFIX = "Автоматическое списание из замеров"
from app import csrf, db
from app.forms import CommentForm, ProjectForm, TaskEditForm, UploadExcelForm, UserForm, UserPasswordForm
from app.models import (
    Apartment,
    Contractor,
    MaterialRequest,
    MaterialRequestItem,
    MaterialWriteOff,
    MaterialWriteOffItem,
    material_writeoff_tasks,
    GlassMeasurement,
    GlassMeasurementItem,
    Project,
    SyncConflict,
    SyncLog,
    SiteErrorReport,
    SiteVisit,
    DeletionActionLog,
    ChangeLog,
    Task,
    TaskComment,
    TASK_STATUSES,
    User,
    WorkCategory,
    WorkPoint,
    ROLE_ADMIN,
    ROLE_MANAGER,
    ROLE_EXECUTOR,
    ROLE_PAINTER,
    ROLE_HANDYMAN,
    ROLE_GLAZIER,
    ROLE_VERIFIER,
    ROLE_VIEWER,
    ROLE_LABELS,
    WORKER_ROLES,
    STATUS_DONE,
    STATUS_NOT_STARTED,
    STATUS_FINISHERS,
    STATUS_CONTRACTOR,
    STATUS_GUARANTEE,
    STATUS_CONCESSION,
    DONE_STATUSES,
)
from app.permissions import can_change_task, can_export, can_manage_mapping, can_manage_sync, role_required
from app.security import client_ip, hit_rate_limit, resolve_site_visit_project_id
from app.services.changelog_service import log_change
from app.services.document_flow import (
    addendum_field_keys,
    addendum_fields_for_template,
    addendum_options_for_template,
    build_addendum_docx,
    create_builtin_addendum_template,
    prepare_uploaded_word_file,
    safe_docx_filename,
    validate_addendum_template,
)
from app.services.avr_document import (
    build_avr_docx,
    default_avr_phrase,
    format_doc_date,
    format_input_date,
    safe_avr_filename,
)
from app.services.excel_export import _excel_premise_label, _safe_filename_part, build_export_path, export_glass_measurements_excel, export_remark_tasks_excel, export_report_tasks_excel, export_simple_tasks_excel, export_source_excel_reconstructed, export_source_excel_with_strikes, export_tasks_to_excel, resolve_source_excel_with_strikes_path
from app.services.pdf_export import export_assignment_worker_pdf, export_table_pdf, export_tasks_pdf
from app.services.excel_import import inspect_remarks_workbook, preview_excel, save_upload, sync_excel_file
from app.services.google_sheets_sync import sync_google_sheets, update_task_strike_in_google_sheet
from app.services.mapping_service import ensure_default_categories, update_category_points
from app.services.pdf_recognition import is_no_remark_text, recognize_pdf_act
from app.services.transfer_import import _is_app_mode, _parse_app_date, inspect_transfer_workbook, sync_transfer_statistics
from app.services.task_service import (
    DOP_AGREEMENT_POINT_NUMBERS,
    MAIN_WORK_POINT_NUMBERS,
    VISIBLE_WORK_POINT_NUMBERS,
    build_task_query,
    category_stats,
    change_task_status,
    dashboard_stats,
    detect_search_mode,
    get_setting,
    is_apartment_unsold,
    looks_like_apartment_identifier,
    parse_multi_premise_search,
    parse_date,
    premise_matches_search,
    premise_matches_selector,
    set_setting,
    AVR_STATUS_NEEDED,
    AVR_STATUS_SIGNED,
    APP_DEADLINE_EXPIRING,
    APP_DEADLINE_EXPIRED,
    APP_DEADLINE_NO_REMARKS,
)
from app.services.status_rules import is_problem_details_required
from app.services.sync_rollback import apply_sync_rollback, build_project_rollback_data
from app.time_utils import to_moscow_datetime
from app.services.uid_service import build_task_uid, cell_hash, normalize_text, stable_hash
from app.services.remark_format import remark_text_html
from app.security import hit_rate_limit, security_event, validate_upload
from app.two_factor import generate_totp_secret, provisioning_uri, qr_svg_data_uri, verify_totp

bp = Blueprint("main", __name__)


@bp.route("/service-worker.js")
def service_worker():
    response = current_app.send_static_file("service-worker.js")
    response.headers["Content-Type"] = "application/javascript; charset=utf-8"
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Service-Worker-Allowed"] = "/"
    return response


@bp.route("/service-worker-reset")
def service_worker_reset():
    html = """<!doctype html><html lang=\"ru\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\"><title>Обновление CRM</title></head><body><p>Обновляем приложение...</p><script>
    (() => {
      let finished = false;
      const openCrm = () => {
        if (finished) return;
        finished = true;
        window.location.replace('/?worker=v17');
      };
      navigator.serviceWorker?.addEventListener('controllerchange', openCrm, { once: true });
      navigator.serviceWorker?.register('/service-worker.js?v=v17-mobile-po-gap-half', { scope: '/', updateViaCache: 'none' })
        .then(registration => registration.update())
        .catch(() => {})
        .finally(() => window.setTimeout(openCrm, 1200));
    })();
    </script></body></html>"""
    response = current_app.response_class(html, content_type="text/html; charset=utf-8")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


MANUAL_WRITEOFF_COMMENT_PREFIX = "__manual_writeoff__:"
TRUE_SETTING_VALUES = {"1", "true", "yes", "on", "да", "checked"}


def _setting_bool(key: str, default: bool = False) -> bool:
    value = get_setting(key, "1" if default else "0")
    return str(value or "").strip().lower() in TRUE_SETTING_VALUES


def _set_setting_bool(key: str, enabled: bool) -> None:
    set_setting(key, "1" if enabled else "0")


def _setting_csv(key: str) -> set[str]:
    value = get_setting(key, "") or ""
    return {item.strip() for item in str(value).split(",") if item.strip()}


def _set_setting_csv(key: str, values) -> None:
    cleaned = sorted({str(value).strip() for value in values if str(value).strip()})
    set_setting(key, ",".join(cleaned))


SECTION_LOCK_CHOICES = [
    {
        "key": "objects",
        "label": "Объекты",
        "icon": "bi-buildings",
        "endpoints": {
            "main.objects", "main.object_new", "main.object_edit", "main.object_delete",
            "main.object_delete_confirm", "main.object_open",
        },
    },
    {"key": "dashboard", "label": "Главная", "icon": "bi-grid-1x2", "endpoints": {"main.dashboard", "main.dashboard_legacy"}},
    {
        "key": "remarks",
        "label": "Замечания",
        "icon": "bi-card-checklist",
        "endpoints": {
            "main.task_list", "main.task_detail", "main.task_new", "main.task_delete",
            "main.task_recognition", "main.update_task", "main.add_task_comment", "main.quick_status",
            "main.inline_update_text", "main.split_task_remark",
        },
    },
    {
        "key": "contractors",
        "label": "Подрядчики",
        "icon": "bi-person-gear",
        "endpoints": {
            "main.contractors_list", "main.contractors_export", "main.contractors_excel_selection",
            "main.contractor_directory", "main.contractor_new", "main.contractor_edit", "main.contractor_delete",
        },
    },
    {
        "key": "apartments",
        "label": "Квартиры",
        "icon": "bi-building",
        "endpoints": {
            "main.apartments", "main.apartments_export", "main.apartment_detail", "main.update_apartment_po_status",
            "main.update_apartment_inspection_status", "main.update_apartment_inspection_date", "main.update_apartment_inspection_note",
            "main.update_apartment_comment", "main.update_apartment_avr_status",
        },
    },
    {"key": "avr", "label": "АВР", "icon": "bi-file-earmark-check", "endpoints": {"main.avr"}},
    {
        "key": "assignments",
        "label": "Выдача задач",
        "icon": "bi-person-check",
        "endpoints": {
            "main.assignments", "main.assignment_unassign", "main.assignment_delete_from_employee",
            "main.assignment_issued_employee_export", "main.assignment_manual_task_new",
        },
    },
    {
        "key": "report",
        "label": "Отчет",
        "icon": "bi-file-earmark-bar-graph",
        "endpoints": {
            "main.work_report", "main.work_report_export", "main.assignments_report",
            "main.assignments_report_worker_pdf",
        },
    },
    {
        "key": "materials",
        "label": "Расход материалов",
        "icon": "bi-box-seam",
        "endpoints": {
            "main.materials", "main.material_balance_delete", "main.material_request_detail",
            "main.material_request_rename", "main.material_request_update", "main.material_request_delete",
            "main.material_requests_bulk_delete", "main.material_request_export", "main.material_writeoff_edit",
            "main.material_writeoff_delete", "main.material_writeoffs_bulk_delete", "main.material_expense_export",
            "main.material_request_new", "main.material_writeoff_new", "main.material_manual_task_new",
        },
    },
    {
        "key": "glass",
        "label": "Замеры",
        "icon": "bi-window",
        "endpoints": {
            "main.glass_measurements", "main.glass_need_measure", "main.glass_measurement_save",
            "main.glass_measurement_return_to_all", "main.glass_status_update", "main.glass_order_export", "main.glass_create_material_request",
            "main.glass_measurements_delete", "main.glass_order",
        },
    },
    {
        "key": "documents",
        "label": "Документы",
        "icon": "bi-files",
        "endpoints": {"main.documents", "main.documents_addendum", "main.documents_download"},
    },
    {
        "key": "notifications",
        "label": "Уведомления",
        "icon": "bi-bell",
        "endpoints": {"main.notifications", "main.archive_apartment_avr"},
    },
    {
        "key": "users",
        "label": "Пользователи",
        "icon": "bi-people",
        "endpoints": {
            "main.users", "main.user_set_password", "main.user_set_password_page",
            "main.user_delete_confirm", "main.user_delete",
        },
    },
    {
        "key": "service",
        "label": "Импорт, синхронизация и распределение",
        "icon": "bi-sliders",
        "endpoints": {
            "main.upload_excel", "main.sync_google", "main.sync_logs", "main.sync_log_details", "main.delete_sync_log",
            "main.rollback_sync_log", "main.sync_conflicts", "main.resolve_conflict",
            "main.resolve_conflicts_bulk", "main.mapping_settings",
        },
    },
    {
        "key": "site_errors",
        "label": "Для разработчика",
        "icon": "bi-bug",
        "endpoints": {
            "main.site_errors",
            "main.site_error_close",
            "main.site_error_delete",
            "main.developer_delete_logs",
            "main.developer_delete_log_undo",
            "main.developer_statistics",
        },
    },
    {
        "key": "worker_tasks",
        "label": "Мои задачи сотрудников",
        "icon": "bi-check2-square",
        "endpoints": {"main.my_tasks", "main.my_task_done", "main.my_task_return"},
    },
]


def _section_lock_choice_for_endpoint(endpoint: str | None):
    if not endpoint or endpoint == "main.site_settings":
        return None
    for choice in SECTION_LOCK_CHOICES:
        if endpoint in choice["endpoints"]:
            return choice
    return None


def _future_features():
    return [
        {
            "icon": "bi-file-earmark-text",
            "title": "Быстрое редактирование и генерация документов",
            "text": "АПП, дополнительные соглашения и другие формы можно будет быстро подготовить, отредактировать и отправить на подписание.",
        },
        {
            "icon": "bi-house-add",
            "title": "Добавление объектов с кладовками",
            "text": "Структура объектов станет шире: можно будет добавлять и вести кладовые помещения вместе с квартирами и коммерцией.",
        },
        {
            "icon": "bi-box-seam",
            "title": "Учёт затраченного материала",
            "text": "По каждой задаче появится понятная отчётность: какие материалы были использованы, в каком объёме и кем.",
        },
    ]


def _documents_under_development_response():
    return render_template(
        "under_development.html",
        layout_mode="documents",
        title="Документы",
        message="Технические работы",
        subtitle="Мы временно скрыли раздел, чтобы не показывать незавершённый функционал.",
        icon="bi-tools",
    )


def _maintenance_response(
    *,
    title: str = "Технические работы",
    message: str = "На сайте ведутся технические работы",
    subtitle: str = "Мы уже обновляем систему и скоро вернём доступ. Пожалуйста, попробуйте зайти немного позже.",
    hint: str = "Доступ для инженера остаётся открытым, чтобы завершить настройку сайта.",
):
    return render_template(
        "maintenance.html",
        title=title,
        message=message,
        subtitle=subtitle,
        hint=hint,
        future_features=_future_features(),
    ), 503


def _blocked_section_response(section_label: str):
    return _maintenance_response(
        title=f"{section_label} временно закрыт",
        message=f"Раздел «{section_label}» временно закрыт",
        subtitle="Мы готовим обновление этого раздела. Как только работы будут завершены, доступ снова появится автоматически.",
    )


def _refresh_sync_dashboard_settings(project_id: int) -> None:
    latest_log = (
        SyncLog.query.filter(
            SyncLog.project_id == project_id,
            SyncLog.status == "success",
            SyncLog.rolled_back_at.is_(None),
        )
        .order_by(SyncLog.started_at.desc())
        .first()
    )
    if latest_log:
        set_setting("last_sync_at", latest_log.started_at.isoformat())
        set_setting("last_sync_source", latest_log.source_name or latest_log.source_type)
    else:
        set_setting("last_sync_at", "—")
        set_setting("last_sync_source", "—")


RU_WEEKDAYS = {
    0: "понедельник",
    1: "вторник",
    2: "среда",
    3: "четверг",
    4: "пятница",
    5: "суббота",
    6: "воскресенье",
}
RU_MONTHS_GENITIVE = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}


def format_ru_date(value: date | datetime | None = None) -> str:
    value = value or date.today()
    if isinstance(value, datetime):
        value = value.date()
    return f"{value.day} {RU_MONTHS_GENITIVE.get(value.month, '')} {value.year}".strip()


def format_ru_day_month(value: date | datetime | None = None) -> str:
    value = value or date.today()
    if isinstance(value, datetime):
        value = value.date()
    return f"{value.day} {RU_MONTHS_GENITIVE.get(value.month, '')}".strip()


def format_ru_datetime(value: datetime | None = None) -> str:
    value = to_moscow_datetime(value or datetime.utcnow())
    return f"{format_ru_date(value)} {value.strftime('%H:%M')}"


def format_ru_weekday(value: date | datetime | None = None) -> str:
    value = value or date.today()
    if isinstance(value, datetime):
        value = value.date()
    return RU_WEEKDAYS.get(value.weekday(), "")


def _parse_history_date_value(value: object) -> date | datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, (date, datetime)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        if len(text) <= 10:
            return date.fromisoformat(text[:10])
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def _history_actor_label(change: ChangeLog) -> str:
    if change.user:
        label = (change.user.full_name or change.user.username or "").strip()
        normalized_label = re.sub(r"[\s_-]+", " ", label).strip().casefold()
        if normalized_label == "codex mobile":
            return "Моб.Устройство"
        return label
    return "Синхронизация"


def _history_responsible_label(value: object, users_cache: dict[int, str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "не назначен"
    if not text.isdigit():
        return text
    user_id = int(text)
    cached = users_cache.get(user_id)
    if cached is not None:
        return cached
    user = db.session.get(User, user_id)
    label = (user.full_name or user.username) if user else f"пользователь #{user_id}"
    users_cache[user_id] = label
    return label


def _history_field_value(field_name: str | None, value: object, users_cache: dict[int, str]) -> str:
    text = str(value or "").strip()
    if field_name == "status":
        return TASK_STATUSES.get(text, {}).get("label", text or "не задан")
    if field_name == "avr_status":
        return "Подписан" if text == AVR_STATUS_SIGNED else ("Нужен" if text == AVR_STATUS_NEEDED else (text or "не задан"))
    if field_name == "avr_signed_date":
        parsed = _parse_history_date_value(value)
        return format_ru_date(parsed) if parsed else "не задана"
    if field_name == "responsible_id":
        return _history_responsible_label(value, users_cache)
    if field_name in {"planned_date", "completed_date"}:
        parsed = _parse_history_date_value(value)
        return format_ru_date(parsed) if parsed else "не задана"
    if field_name == "comment":
        return text or "пусто"
    if field_name in {"apartment_comment", "apartment_inspection_note"}:
        return text or "пусто"
    if field_name == "po_status":
        return PO_STATUS_LABELS.get(text, text or "не задан")
    if field_name == "description":
        return text or "пусто"
    if field_name == "is_missing_in_latest_sync":
        return "да" if text.lower() in {"1", "true", "yes"} else "нет"
    return text or "не задано"


def _split_glass_size_comment(value: str | None) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "", ""
    if " — " in text:
        size_text, comment = text.split(" — ", 1)
        return size_text.strip(), comment.strip()
    return text, ""


def _compact_history_text(value: object) -> str:
    return " ".join(str(value or "").split()).strip()


def _manual_split_history_text(change: ChangeLog, task_text: str) -> str:
    first_part = ""
    second_part = ""
    try:
        payload = json.loads(str(change.new_value or ""))
        if isinstance(payload, dict):
            first_part = _compact_history_text(payload.get("first"))
            second_part = _compact_history_text(payload.get("second"))
    except (TypeError, ValueError, json.JSONDecodeError):
        pass

    if not first_part and change.action == "manual_split":
        first_part = _compact_history_text(change.old_value)
    if not second_part:
        second_part = _compact_history_text(change.new_value if change.action == "manual_split_created" else task_text)
    if not second_part and task_text:
        second_part = _compact_history_text(task_text)

    if first_part and second_part:
        return f"Замечание разделено на две части. 1 часть — «{first_part}». 2 часть — «{second_part}»."
    if second_part:
        return f"Замечание разделено на две части. 2 часть — «{second_part}»."
    return "Замечание разделено на две части."


def _build_change_history_entry(change: ChangeLog, task: Task | None = None, users_cache: dict[int, str] | None = None) -> dict[str, str]:
    users_cache = users_cache or {}
    field_name = change.field_name or ""
    old_value = _history_field_value(field_name, change.old_value, users_cache)
    new_value = _history_field_value(field_name, change.new_value, users_cache)
    actor = _history_actor_label(change)
    point_label = ""
    if task and task.work_point:
        point_label = task.work_point.display_name
    task_text = str(task.description or task.source_cell_value or "").strip() if task else ""
    summary_class = ""

    if change.action == "status_change":
        if task_text:
            point_label = task_text
        if (change.new_value or "") == "problem":
            problem_text = str(task.comment or "").strip() if task else ""
            summary = f"Проблема: «{problem_text}»." if problem_text else "Проблема."
            summary_class = "timeline-summary-danger"
        elif not change.user and (change.old_value or "") not in DONE_STATUSES and (change.new_value or "") in DONE_STATUSES:
            summary = f"Синхронизация отметила замечание выполненным и зачеркнула его: было «{old_value}», стало «{new_value}»."
        elif not change.user:
            summary = f"Синхронизация изменила статус замечания: было «{old_value}», стало «{new_value}»."
        elif field_name == "avr_status":
            summary = (
                f"Статус АВР помещения изменён: было «{old_value}», стало «{new_value}»."
                if change.user
                else f"Синхронизация изменила статус АВР помещения: было «{old_value}», стало «{new_value}»."
            )
        elif field_name == "avr_signed_date":
            summary = (
                f"Дата подписания АВР изменена: была «{old_value}», стала «{new_value}»."
                if change.user
                else f"Синхронизация изменила дату подписания АВР: была «{old_value}», стала «{new_value}»."
            )
        else:
            summary = f"Статус замечания изменён: было «{old_value}», стало «{new_value}»."
    elif change.action == "field_update":
        if field_name == "status":
            if task_text:
                point_label = task_text
            summary = (
                f"Синхронизация изменила статус замечания: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"Статус замечания изменён: было «{old_value}», стало «{new_value}»."
            )
        elif field_name == "responsible_id":
            summary = (
                f"Синхронизация изменила исполнителя: был «{old_value}», стал «{new_value}»."
                if not change.user
                else f"Исполнитель изменён: был «{old_value}», стал «{new_value}»."
            )
        elif field_name == "planned_date":
            summary = (
                f"Синхронизация изменила плановую дату: была «{old_value}», стала «{new_value}»."
                if not change.user
                else f"Плановая дата изменена: была «{old_value}», стала «{new_value}»."
            )
        elif field_name == "comment":
            summary = (
                f"Синхронизация изменила комментарий к замечанию: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"Комментарий к замечанию изменён: было «{old_value}», стало «{new_value}»."
            )
        elif field_name == "description":
            summary = (
                f"Синхронизация изменила текст замечания: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"Замечание отредактировано. Было — «{old_value}». Стало — «{new_value}»."
            )
        else:
            field_label = {
                "status": "Статус",
                "responsible_id": "Исполнитель",
                "planned_date": "Плановая дата",
                "comment": "Комментарий",
                "description": "Текст замечания",
            }.get(field_name, field_name or "Поле")
            summary = (
                f"Синхронизация изменила поле «{field_label}»: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"{field_label} изменён: было «{old_value}», стало «{new_value}»."
            )
    elif change.action == "apartment_field_update":
        if field_name == "apartment_inspection_note":
            summary = (
                f"Синхронизация изменила комментарий осмотра помещения: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"Комментарий осмотра помещения изменён: было «{old_value}», стало «{new_value}»."
            )
        elif field_name == "po_status":
            summary = (
                f"Синхронизация изменила внутренний статус помещения: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"Внутренний статус помещения изменён: было «{old_value}», стало «{new_value}»."
            )
        else:
            summary = (
                f"Синхронизация изменила комментарий помещения: было «{old_value}», стало «{new_value}»."
                if not change.user
                else f"Комментарий помещения изменён: было «{old_value}», стало «{new_value}»."
            )
        if field_name == "avr_status":
            summary = (
                f"Статус АВР помещения изменён: было «{old_value}», стало «{new_value}»."
                if change.user
                else f"Синхронизация изменила статус АВР помещения: было «{old_value}», стало «{new_value}»."
            )
        elif field_name == "avr_signed_date":
            summary = (
                f"Дата подписания АВР изменена: была «{old_value}», стала «{new_value}»."
                if change.user
                else f"Синхронизация изменила дату подписания АВР: была «{old_value}», стала «{new_value}»."
            )
        point_label = "Помещение"
    elif change.action == "manual_created":
        summary = f"Замечание добавлено вручную: «{str(change.new_value or '').strip() or 'без текста'}»."
    elif change.action in {"manual_split", "manual_split_created"}:
        summary = _manual_split_history_text(change, task_text)
    elif change.action == "manual_assignment_created":
        summary = f"Задача выдана вручную: «{str(change.new_value or '').strip() or 'без текста'}»."
    elif change.action == "manual_act_created":
        summary = f"Замечание добавлено вручную из акта: «{str(change.new_value or '').strip() or 'без текста'}»."
    elif change.action == "pdf_recognition_created":
        summary = f"Замечание добавлено после распознавания PDF: «{str(change.new_value or '').strip() or 'без текста'}»."
    elif change.action == "comment_added":
        summary = f"Добавлен новый комментарий: «{str(change.new_value or '').strip() or 'без текста'}»."
    elif change.action == "created_from_sync":
        summary = f"Замечание добавлено из синхронизации: «{str(change.new_value or '').strip() or 'без названия источника'}»."
    elif change.action == "missing_in_latest_sync":
        summary = "После синхронизации замечание пропало из последней загруженной таблицы."
    else:
        summary = "В задаче выполнено системное изменение."

    return {
        "timestamp": format_ru_datetime(change.created_at),
        "actor": actor,
        "point_label": point_label,
        "summary": summary,
        "summary_class": summary_class,
    }


def _is_legacy_problem_comment_change(change: ChangeLog, task: Task | None) -> bool:
    """Hide the old extra comment event that was written with Problem status."""
    if not task or change.action != "field_update" or change.field_name != "comment":
        return False
    if not change.created_at:
        return False
    for related in task.changes:
        if related.action != "status_change" or related.field_name != "status" or related.new_value != "problem":
            continue
        if related.user_id != change.user_id or not related.created_at:
            continue
        if abs((related.created_at - change.created_at).total_seconds()) <= 5:
            return True
    return False


def _save_site_error(message: str, kind: str = "user", traceback_text: str | None = None, page_url: str | None = None) -> SiteErrorReport | None:
    message = (message or "").strip()
    if not message:
        return None
    try:
        project = selected_project()
    except Exception:
        project = None
    try:
        user_id = current_user.id if current_user.is_authenticated else None
    except Exception:
        user_id = None
    report = SiteErrorReport(
        project_id=project.id if project else None,
        user_id=user_id,
        kind=kind,
        message=message[:5000],
        page_url=(page_url or request.form.get("page_url") or request.referrer or request.url or "")[:500],
        user_agent=(request.headers.get("User-Agent") or "")[:500],
        traceback_text=(traceback_text or None),
        status="new",
    )
    db.session.add(report)
    db.session.commit()
    return report


def _record_deletion_action(
    action_key: str,
    entity_type: str,
    entity_id: int | None,
    entity_title: str,
    description: str,
    snapshot: dict | None,
    project_id: int | None = None,
) -> DeletionActionLog | None:
    """Write a restorable deletion/unassignment action into developer logs."""
    try:
        log = DeletionActionLog(
            project_id=project_id,
            user_id=current_user.id if current_user.is_authenticated else None,
            action_key=(action_key or "")[:80],
            entity_type=(entity_type or "")[:80],
            entity_id=entity_id,
            entity_title=(entity_title or "")[:255],
            description=(description or "")[:2000],
            snapshot_json=json.dumps(snapshot or {}, ensure_ascii=False, default=str),
            is_undone=False,
        )
        db.session.add(log)
        return log
    except Exception:
        current_app.logger.exception("Failed to record deletion action")
        return None


def _snapshot_model(obj, extra: dict | None = None) -> dict:
    snapshot = {}
    for column in getattr(getattr(obj, "__table__", None), "columns", []):
        value = getattr(obj, column.name, None)
        if isinstance(value, (datetime, date)):
            value = value.isoformat()
        snapshot[column.name] = value
    if extra:
        snapshot.update(extra)
    return snapshot


def _record_simple_deletion(
    action_key: str,
    entity_type: str,
    obj,
    entity_title: str,
    description: str,
    project_id: int | None = None,
    extra: dict | None = None,
) -> DeletionActionLog | None:
    return _record_deletion_action(
        action_key,
        entity_type,
        getattr(obj, "id", None),
        entity_title,
        description,
        _snapshot_model(obj, extra),
        project_id=project_id,
    )


def _snapshot_children(items) -> list[dict]:
    return [_snapshot_model(item) for item in list(items or [])]


def _coerce_snapshot_value(model_cls, field_name: str, value):
    if value is None:
        return None
    column = getattr(getattr(model_cls, "__table__", None), "columns", {}).get(field_name)
    if column is None:
        return value
    try:
        python_type = column.type.python_type
    except NotImplementedError:
        return value
    except Exception:
        return value
    if python_type is datetime:
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            return None
    if python_type is date:
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            return None
    if python_type is bool:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "on", "да"}
    if python_type is int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    if python_type is float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if python_type is str:
        return str(value)
    return value


def _restore_model_from_snapshot(model_cls, snapshot: dict, *, force: dict | None = None, skip: set[str] | None = None):
    if not isinstance(snapshot, dict) or not snapshot:
        return None
    force = force or {}
    skip = set(skip or set())
    columns = getattr(getattr(model_cls, "__table__", None), "columns", [])
    entity_id = snapshot.get("id")
    if entity_id:
        try:
            existing_id = int(entity_id)
        except (TypeError, ValueError):
            existing_id = None
        if existing_id and db.session.get(model_cls, existing_id):
            raise ValueError("Запись уже существует")
    obj = model_cls()
    for column in columns:
        name = column.name
        if name in skip or name in force or name not in snapshot:
            continue
        setattr(obj, name, _coerce_snapshot_value(model_cls, name, snapshot.get(name)))
    for name, value in force.items():
        setattr(obj, name, _coerce_snapshot_value(model_cls, name, value))
    db.session.add(obj)
    db.session.flush()
    return obj


def _restore_children_from_snapshots(model_cls, rows: list[dict], *, force: dict) -> int:
    restored = 0
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        try:
            _restore_model_from_snapshot(model_cls, row, force=force)
            restored += 1
        except ValueError:
            continue
    return restored


def _mark_deletion_log_undone(log: DeletionActionLog) -> None:
    log.is_undone = True
    log.undone_at = datetime.utcnow()
    log.undone_by_user_id = current_user.id


PO_STATUS_NOT_READY = "not_ready"
PO_STATUS_DO_NOT_INVITE = "do_not_invite"
PO_STATUS_TO_THROW = "to_throw"
PO_STATUS_THROWN = "thrown"
PO_STATUS_PO = "po"
PO_STATUS_LABELS = {
    PO_STATUS_NOT_READY: "Не готова",
    PO_STATUS_DO_NOT_INVITE: "Не приглашать",
    PO_STATUS_TO_THROW: "Кинуть",
    PO_STATUS_THROWN: "Кинуто",
    PO_STATUS_PO: "По",
}
PO_STATUS_CLASSES = {
    PO_STATUS_NOT_READY: "status-pill-danger",
    PO_STATUS_DO_NOT_INVITE: "status-pill-muted",
    PO_STATUS_TO_THROW: "status-pill-warning",
    PO_STATUS_THROWN: "status-pill-info",
    PO_STATUS_PO: "status-pill-success",
}
WALL_POINT_NUMBERS = {"11"}

CONTRACTOR_POINT_LABELS = {
    "10": "Вентиляция",
    "11": "Стены. Потолки (штукатурка/шпаклёвка)",
    "12": "Возведение коробки здания",
    "13": "Работы по устройству подстилающего слоя",
    "14": "Работы по возведению блочных перегородок",
    "15": "Разнорабочие",
    "16": "Работы по монтажу ПВХ блоков",
    "17": "Работы по монтажу откосов и подоконников",
    "18": "Работы по монтажу холодного витражного остекления балконов",
    "19": "Работы по монтажу системы отопления, в/с, канализации",
    "20": "Работы по монтажу входных дверей",
    "21": "Электрика",
    "22": "Прочее",
}


@bp.app_context_processor
def inject_globals():
    has_all_projects = current_user.is_authenticated and current_user.can_access_all_projects
    project_id = session.get("current_project_id") if has_all_projects else (
        current_user.project_id if current_user.is_authenticated and current_user.project_id else session.get("current_project_id")
    )
    current_project = db.session.get(Project, project_id) if project_id else None
    if current_project and current_user.is_authenticated and not current_user.can_access_project(current_project):
        current_project = None
    mobile_switch_projects = []
    if current_user.is_authenticated:
        projects_query = Project.query.order_by(Project.name.asc(), Project.id.asc())
        mobile_switch_projects = [project for project in projects_query.all() if current_user.can_access_project(project)]
    new_site_errors_count = 0
    if current_user.is_authenticated and current_user.role == ROLE_ADMIN:
        error_query = SiteErrorReport.query.filter(SiteErrorReport.status == "new")
        if current_project:
            error_query = error_query.filter(or_(SiteErrorReport.project_id == current_project.id, SiteErrorReport.project_id.is_(None)))
        new_site_errors_count = error_query.count()
    return {
        "TASK_STATUSES": TASK_STATUSES,
        "STATUS_DONE": STATUS_DONE,
        "DONE_STATUSES": DONE_STATUSES,
        "ROLE_LABELS": ROLE_LABELS,
        "current_project": current_project,
        "mobile_switch_projects": mobile_switch_projects,
        "new_site_errors_count": new_site_errors_count,
        "hide_documents_section": _setting_bool("hide_documents_section"),
        "mobile_version_under_development": _setting_bool("mobile_version_under_development"),
        "site_maintenance_mode": _setting_bool("site_maintenance_mode"),
        "blocked_site_sections": _setting_csv("blocked_site_sections"),
        "section_lock_choices": SECTION_LOCK_CHOICES,
        "is_mobile_phone_request": _is_mobile_phone_request(),
        "fmt_quantity": fmt_quantity,
        "display_material_name": display_material_name,
        "ru_plural": ru_plural,
        "task_word": task_word,
        "task_count_label": task_count_label,
        "format_ru_date": format_ru_date,
        "format_ru_weekday": format_ru_weekday,
        "remark_text": remark_text_html,
        "glass_measurement_action_label": glass_measurement_action_label,
    }


def selected_project() -> Project | None:
    # Если пользователь привязан к конкретному объекту, нельзя подменить объект
    # через session/current_project_id или прямую ссылку.
    project_id = session.get("current_project_id")
    if project_id:
        project = db.session.get(Project, project_id)
        if project and (not current_user.is_authenticated or current_user.can_access_project(project)):
            return project
        session.pop("current_project_id", None)
    if current_user.is_authenticated and not current_user.can_access_all_projects:
        project_ids = current_user.project_access_ids
        if len(project_ids) == 1:
            project = db.session.get(Project, next(iter(project_ids)))
            if project:
                session["current_project_id"] = project.id
                return project
    return None


def _is_local_redirect(target: str | None) -> bool:
    if not target:
        return False
    parsed = urlparse(target)
    return parsed.scheme == "" and parsed.netloc == "" and target.startswith("/")


def _safe_redirect(target: str | None, fallback_endpoint: str = "main.dashboard"):
    return redirect(target if _is_local_redirect(target) else url_for(fallback_endpoint))


def _project_access_allowed(project: Project | None) -> bool:
    if project is None:
        return False
    return current_user.can_access_project(project)


def _abort_if_project_forbidden(project: Project | None) -> Project:
    if not _project_access_allowed(project):
        abort(404)
    return project


def _user_can_work_in_project(user: User | None, project: Project | None) -> bool:
    if not user or not project or not user.is_active:
        return False
    return user.role in WORKER_ROLES and user.can_access_project(project)


def _abort_if_user_outside_current_project(user: User | None, project: Project | None = None) -> User:
    project = project or selected_project()
    if user is None:
        abort(404)
    if current_user.role == ROLE_ADMIN:
        return user
    if project and not user.can_access_project(project):
        abort(404)
    return user


def _task_for_current_project(task_id: int, project: Project | None = None) -> Task:
    project = project or selected_project()
    if project is None:
        abort(404)
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    return task


def _role_home_endpoint() -> str:
    if current_user.role in WORKER_ROLES:
        return "main.my_tasks"
    if current_user.role == ROLE_VERIFIER:
        return "main.work_report"
    if current_user.role == ROLE_VIEWER:
        return "main.dashboard"
    return "main.dashboard"


def _deny_or_redirect():
    if request.method in {"GET", "HEAD", "OPTIONS"}:
        return redirect(url_for(_role_home_endpoint()))
    abort(403)


WORKER_ALLOWED_ENDPOINTS = {
    "main.dashboard_legacy",
    "main.my_tasks",
    "main.my_task_done",
    "main.my_task_return",
    "main.account",
    "main.report_error",
}

VERIFIER_ALLOWED_ENDPOINTS = {
    "main.dashboard_legacy",
    "main.objects",
    "main.object_open",
    "main.work_report",
    "main.work_report_export",
    "main.documents",
    "main.documents_addendum",
    "main.documents_download",
    "main.account",
    "main.report_error",
}

VIEWER_ALLOWED_GET_ENDPOINTS = {
    "main.objects",
    "main.object_open",
    "main.dashboard",
    "main.dashboard_legacy",
    "main.task_list",
    "main.task_detail",
    "main.contractors_list",
    "main.apartments",
    "main.apartment_detail",
    "main.avr",
    "main.glass_measurements",
    "main.materials",
    "main.material_request_detail",
    "main.work_report",
    "main.documents",
    "main.documents_download",
    "main.account",
}


@bp.before_request
def enforce_role_access():
    # Единая защита всех main-маршрутов: сначала аутентификация, затем доступ по роли.
    endpoint = request.endpoint or ""
    # Репорт ошибки доступен и со стартового экрана. Он всё равно защищён CSRF и rate limit.
    if endpoint in {"main.report_error", "main.service_worker", "main.service_worker_reset"}:
        return None

    if not current_user.is_authenticated:
        if request.method in {"GET", "HEAD"} and endpoint in {"main.dashboard", "main.dashboard_legacy"}:
            from app.auth import login as login_view

            return login_view()
        next_url = request.full_path if request.query_string else request.path
        return redirect(url_for("auth.login", next=next_url))

    locked_section = _section_lock_choice_for_endpoint(endpoint)
    if locked_section and locked_section["key"] in _setting_csv("blocked_site_sections"):
        return _blocked_section_response(locked_section["label"])

    if _setting_bool("site_maintenance_mode") and current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        return _maintenance_response()

    if _is_mobile_phone_request():
        if endpoint not in _mobile_phone_allowed_endpoints():
            if request.method in {"GET", "HEAD", "OPTIONS"}:
                return redirect(url_for(_mobile_phone_home_endpoint()))
            abort(403)

    if current_user.role in {ROLE_ADMIN, ROLE_MANAGER}:
        return None

    if current_user.role in WORKER_ROLES:
        if endpoint in WORKER_ALLOWED_ENDPOINTS:
            return None
        return _deny_or_redirect()

    if current_user.role == ROLE_VERIFIER:
        if endpoint in VERIFIER_ALLOWED_ENDPOINTS:
            return None
        return _deny_or_redirect()

    if current_user.role == ROLE_VIEWER:
        if endpoint == "main.account":
            return None
        if request.method in {"GET", "HEAD", "OPTIONS"} and endpoint in VIEWER_ALLOWED_GET_ENDPOINTS:
            return None
        if endpoint == "main.report_error":
            return None
        return _deny_or_redirect()

    abort(403)


def fmt_quantity(value) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return "0"
    if number.is_integer():
        return str(int(number))
    return (f"{number:.3f}".rstrip("0").rstrip(".")) or "0"


def display_material_name(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "стеклопакет" not in text.lower():
        return text
    return re.sub(r"\s+кв\.?\s*\d+[^\s,;—-]*(?=\s*$)", "", text, flags=re.IGNORECASE).strip()


def ru_plural(value, one: str, few: str, many: str) -> str:
    try:
        number = abs(int(value or 0))
    except (TypeError, ValueError):
        number = 0
    mod10 = number % 10
    mod100 = number % 100
    if mod10 == 1 and mod100 != 11:
        return one
    if 2 <= mod10 <= 4 and not (12 <= mod100 <= 14):
        return few
    return many


def task_word(value) -> str:
    return ru_plural(value, "задача", "задачи", "задач")


def task_count_label(value) -> str:
    try:
        number = int(value or 0)
    except (TypeError, ValueError):
        number = 0
    return f"{number} {task_word(number)}"


def _can_edit_materials() -> bool:
    return current_user.is_authenticated and current_user.role != "viewer"


def _parse_quantity(value: str | None) -> float | None:
    text_value = str(value or "").strip().replace(",", ".")
    if not text_value:
        return None
    try:
        quantity = float(text_value)
    except ValueError:
        return None
    if quantity <= 0:
        return None
    return quantity


MATERIAL_KEY_SEP = "|||"


def _read_material_rows_from_form(limit: int = 10) -> list[dict[str, object]]:
    names = request.form.getlist("name[]")
    quantities = request.form.getlist("quantity[]")
    units = request.form.getlist("unit[]")
    source_item_ids = request.form.getlist("item_id[]")
    rows = []
    for idx in range(min(limit, max(len(names), len(quantities), len(units), len(source_item_ids)))):
        name = (names[idx] if idx < len(names) else "").strip()
        unit = (units[idx] if idx < len(units) else "").strip()
        quantity = _parse_quantity(quantities[idx] if idx < len(quantities) else "")
        if not name and quantity is None and not unit:
            continue
        if not name or quantity is None or not unit:
            raise ValueError("Заполните наименование, количество и единицу измерения у каждой позиции")
        source_item_id = None
        if idx < len(source_item_ids):
            try:
                source_item_id = int(source_item_ids[idx])
            except (TypeError, ValueError):
                pass
        rows.append({"name": name, "quantity": quantity, "unit": unit, "source_item_id": source_item_id})
    return rows


def _material_key(name: str, unit: str) -> str:
    return f"{name}{MATERIAL_KEY_SEP}{unit}"


def _split_material_key(value: str | None) -> tuple[str, str] | None:
    text = str(value or "")
    if MATERIAL_KEY_SEP not in text:
        return None
    name, unit = text.split(MATERIAL_KEY_SEP, 1)
    name = name.strip()
    unit = unit.strip()
    if not name or not unit:
        return None
    return name, unit


def _normalize_material_identity(name: str, unit: str) -> tuple[str, str]:
    return (" ".join(str(name or "").strip().lower().split()), " ".join(str(unit or "").strip().lower().split()))


def _balance_options(project_id: int) -> list[dict[str, object]]:
    options = []
    for row in _material_balance_rows(project_id):
        balance = float(row.get("balance") or 0)
        if balance <= 0:
            continue
        name = str(row.get("name") or "").strip()
        unit = str(row.get("unit") or "").strip()
        options.append({**row, "key": _material_key(name, unit), "label": f"{name} — остаток {fmt_quantity(balance)} {unit}"})
    return options


def _read_balance_writeoff_row(project_id: int) -> dict[str, object]:
    parsed = _split_material_key(request.form.get("material_key"))
    quantity = _parse_quantity(request.form.get("quantity"))
    if parsed is None:
        raise ValueError("Выберите материал с баланса")
    if quantity is None:
        raise ValueError("Введите корректное количество")
    name, unit = parsed
    balance_rows = _material_balance_rows(project_id)
    row = next((item for item in balance_rows if str(item.get("name")) == name and str(item.get("unit")) == unit), None)
    if row is None or float(row.get("balance") or 0) <= 0:
        raise ValueError("Этого материала нет на балансе")
    balance = float(row.get("balance") or 0)
    if quantity > balance + 0.000001:
        raise ValueError(f"Нельзя списать больше остатка. Доступно: {fmt_quantity(balance)} {unit}")
    return {"name": name, "unit": unit, "quantity": quantity, "balance": balance}


def _read_balance_writeoff_rows(project_id: int) -> list[dict[str, object]]:
    material_keys = request.form.getlist("material_key")
    quantities = request.form.getlist("quantity")
    balance_rows = _material_balance_rows(project_id)
    balance_by_identity = {
        _normalize_material_identity(str(item.get("name") or ""), str(item.get("unit") or "")): item
        for item in balance_rows
        if float(item.get("balance") or 0) > 0
    }

    requested_rows: list[tuple[int, str, str, float]] = []
    max_len = max(len(material_keys), len(quantities))
    for index in range(max_len):
        material_key = material_keys[index].strip() if index < len(material_keys) and isinstance(material_keys[index], str) else ""
        quantity_raw = quantities[index].strip() if index < len(quantities) and isinstance(quantities[index], str) else ""
        if not material_key and not quantity_raw:
            continue
        parsed = _split_material_key(material_key)
        if parsed is None:
            raise ValueError(f"Выберите материал в строке {index + 1}")
        quantity = _parse_quantity(quantity_raw)
        if quantity is None:
            raise ValueError(f"Введите корректное количество в строке {index + 1}")
        name, unit = parsed
        requested_rows.append((index, name, unit, float(quantity)))

    if not requested_rows:
        raise ValueError("Выберите хотя бы один материал с баланса")

    consolidated: dict[tuple[str, str], dict[str, object]] = {}
    for index, name, unit, quantity in requested_rows:
        identity = _normalize_material_identity(name, unit)
        row = balance_by_identity.get(identity)
        if row is None:
            raise ValueError(f"Материал из строки {index + 1} уже отсутствует на балансе")
        existing = consolidated.get(identity)
        if existing is None:
            consolidated[identity] = {
                "name": str(row.get("name") or "").strip(),
                "unit": str(row.get("unit") or "").strip(),
                "quantity": float(quantity),
                "balance": float(row.get("balance") or 0),
            }
        else:
            existing["quantity"] = float(existing["quantity"]) + float(quantity)

    result = list(consolidated.values())
    for row in result:
        if float(row["quantity"]) > float(row["balance"]) + 0.000001:
            raise ValueError(
                f"Нельзя списать больше остатка для {row['name']}. "
                f"Доступно: {fmt_quantity(float(row['balance']))} {row['unit']}"
            )
    return result


def _task_material_weight(task: Task) -> float:
    text = f"{task.description or ''} {task.source_cell_value or ''}".strip()
    length = max(len(text), 20)
    return float(length) * random.uniform(0.75, 1.35)


def _distribute_material_quantity(tasks: list[Task], quantity: float) -> dict[int, float]:
    if not tasks:
        return {}
    total_int = int(round(float(quantity or 0)))
    if total_int <= 0:
        return {}

    weights = [_task_material_weight(task) for task in tasks]
    if total_int < len(tasks):
        winners = sorted(range(len(tasks)), key=lambda idx: weights[idx], reverse=True)[:total_int]
        return {tasks[idx].id: 1.0 for idx in winners}

    weight_sum = sum(weights) or float(len(tasks))
    raw = [max(1, int(round(total_int * weight / weight_sum))) for weight in weights]
    diff = total_int - sum(raw)
    order = sorted(range(len(tasks)), key=lambda idx: weights[idx], reverse=(diff > 0))
    while diff != 0 and order:
        for idx in order:
            if diff == 0:
                break
            if diff > 0:
                raw[idx] += 1
                diff -= 1
            elif raw[idx] > 1:
                raw[idx] -= 1
                diff += 1
    return {task.id: float(value) for task, value in zip(tasks, raw)}


def _material_balance_rows(project_id: int) -> list[dict[str, object]]:
    balances: dict[tuple[str, str], dict[str, object]] = {}

    def row_for(name: str, unit: str) -> dict[str, object]:
        key = _normalize_material_identity(name, unit)
        if key not in balances:
            balances[key] = {"name": name.strip(), "unit": unit.strip(), "key": _material_key(name.strip(), unit.strip()), "received": 0.0, "spent": 0.0, "balance": 0.0}
        return balances[key]

    request_items = (
        MaterialRequestItem.query.join(MaterialRequest)
        .filter(MaterialRequest.project_id == project_id)
        .all()
    )
    for item in request_items:
        row = row_for(item.name, item.unit)
        row["received"] = float(row["received"]) + float(item.quantity or 0)

    writeoff_items = (
        MaterialWriteOffItem.query.join(MaterialWriteOff)
        .filter(MaterialWriteOff.project_id == project_id)
        .all()
    )
    for item in writeoff_items:
        row = row_for(item.name, item.unit)
        row["spent"] = float(row["spent"]) + float(item.quantity or 0)

    for row in balances.values():
        row["balance"] = float(row["received"]) - float(row["spent"])

    active_balances = [row for row in balances.values() if abs(float(row["balance"])) > 0.000001]
    return sorted(active_balances, key=lambda row: (str(row["name"]).lower(), str(row["unit"]).lower()))


def _next_measurement_request_number(project_id: int) -> int:
    requests = (
        MaterialRequest.query
        .filter(
            MaterialRequest.project_id == project_id,
            MaterialRequest.comment == MEASUREMENT_REQUEST_COMMENT,
        )
        .order_by(MaterialRequest.id)
        .all()
    )
    used_numbers: set[int] = set()
    legacy_requests = []
    for material_request in requests:
        match = MEASUREMENT_REQUEST_TITLE_PATTERN.fullmatch((material_request.title or "").strip())
        if match:
            number = int(match.group(1))
            if number not in used_numbers:
                used_numbers.add(number)
            else:
                legacy_requests.append(material_request)
        else:
            legacy_requests.append(material_request)
    next_available = 1
    for material_request in legacy_requests:
        while next_available in used_numbers:
            next_available += 1
        material_request.title = f"Заявка из замеров №{next_available}"
        used_numbers.add(next_available)
        next_available += 1
    return max(used_numbers, default=0) + 1


def _measurement_ids_for_writeoff(writeoff_id: int) -> list[int]:
    return [
        row.id
        for row in (
            GlassMeasurement.query
            .with_entities(GlassMeasurement.id)
            .filter(GlassMeasurement.material_writeoff_id == writeoff_id)
            .all()
        )
    ]


def _unlink_measurements_from_writeoff(writeoff_id: int) -> list[int]:
    measurement_ids = _measurement_ids_for_writeoff(writeoff_id)
    if measurement_ids:
        GlassMeasurement.query.filter(GlassMeasurement.id.in_(measurement_ids)).update(
            {"material_writeoff_id": None},
            synchronize_session=False,
        )
    return measurement_ids


def _is_measurement_material_request(material_request: MaterialRequest | None) -> bool:
    if material_request is None:
        return False
    title = (material_request.title or "").strip()
    return material_request.comment == MEASUREMENT_REQUEST_COMMENT or bool(MEASUREMENT_REQUEST_TITLE_PATTERN.fullmatch(title))


def _measurement_writeoffs_for_request(material_request: MaterialRequest | None) -> list[MaterialWriteOff]:
    if material_request is None:
        return []
    item_ids = [item.id for item in material_request.items if item.id]
    if not item_ids:
        return []
    measurements = (
        GlassMeasurement.query.options(selectinload(GlassMeasurement.material_writeoff))
        .filter(GlassMeasurement.material_request_item_id.in_(item_ids))
        .all()
    )
    writeoffs = []
    seen_ids: set[int] = set()
    for measurement in measurements:
        writeoff = measurement.material_writeoff
        if writeoff is None or writeoff.id in seen_ids:
            continue
        seen_ids.add(writeoff.id)
        writeoffs.append(writeoff)
    return writeoffs


def _delete_measurement_request_writeoffs(material_request: MaterialRequest, *, record: bool = True) -> list[int]:
    writeoffs = _measurement_writeoffs_for_request(material_request)
    deleted_ids: list[int] = []
    for writeoff in writeoffs:
        measurement_ids = _measurement_ids_for_writeoff(writeoff.id)
        if record:
            _record_simple_deletion(
                "material_writeoff_delete",
                "material_writeoff",
                writeoff,
                f"Списание #{writeoff.id}",
                f"Удалено автосписание по заявке из замеров: {material_request.title or material_request.id}.",
                project_id=material_request.project_id,
                extra={
                    "items": [_snapshot_model(item) for item in writeoff.items],
                    "task_ids": [task.id for task in writeoff.tasks],
                    "measurement_ids": measurement_ids,
                },
            )
        _unlink_measurements_from_writeoff(writeoff.id)
        writeoff.tasks.clear()
        deleted_ids.append(writeoff.id)
        db.session.delete(writeoff)
    return deleted_ids


def _measurement_request_groups(old_items: list[MaterialRequestItem]) -> list[dict[str, object]]:
    old_item_ids = [item.id for item in old_items if item.id]
    if not old_item_ids:
        return []
    linked_measurements = (
        GlassMeasurement.query.options(selectinload(GlassMeasurement.material_writeoff))
        .filter(GlassMeasurement.material_request_item_id.in_(old_item_ids))
        .all()
    )
    linked_by_old_item_id = {measurement.material_request_item_id: measurement for measurement in linked_measurements}
    groups: list[dict[str, object]] = []
    current_group: dict[str, object] | None = None
    for index, old_item in enumerate(old_items):
        measurement = linked_by_old_item_id.get(old_item.id)
        if measurement is not None:
            current_group = {"measurement": measurement, "start": index, "end": index + 1}
            groups.append(current_group)
        elif current_group is not None:
            current_group["end"] = index + 1
    for group in groups:
        start = int(group.get("start") or 0)
        end = int(group.get("end") or start + 1)
        group["item_ids"] = [item.id for item in old_items[start:end] if item.id]
    return groups


def _sync_measurement_writeoffs_from_request_groups(
    material_request: MaterialRequest,
    groups: list[dict[str, object]],
    new_items_by_old_id: dict[int, MaterialRequestItem],
) -> None:
    if not groups:
        return
    for group in groups:
        measurement = group.get("measurement")
        if measurement is None:
            continue
        group_item_ids = [int(item_id) for item_id in group.get("item_ids") or []]
        group_items = [new_items_by_old_id[item_id] for item_id in group_item_ids if item_id in new_items_by_old_id]
        if group_items:
            measurement.material_request_item = group_items[0]
        else:
            measurement.material_request_item = None
        writeoff = getattr(measurement, "material_writeoff", None)
        if writeoff is None:
            continue
        if not group_items:
            _unlink_measurements_from_writeoff(writeoff.id)
            measurement.material_writeoff = None
            writeoff.tasks.clear()
            db.session.delete(writeoff)
            continue
        writeoff.writeoff_date = material_request.request_date
        writeoff.comment = f"{MEASUREMENT_WRITEOFF_COMMENT_PREFIX}: {material_request.title or material_request.id}"
        writeoff.items.clear()
        for item in group_items:
            writeoff.items.append(
                MaterialWriteOffItem(
                    name=item.name,
                    quantity=item.quantity,
                    unit=item.unit,
                )
            )


def _material_task_options(project_id: int, params=None) -> list[Task]:
    # Отдельно нормализуем параметры: раньше фильтр/сортировка в списании материала
    # мог работать нестабильно из-за ImmutableMultiDict и пустых значений.
    raw_params = dict(params or {})
    normalized = {key: (value.strip() if isinstance(value, str) else value) for key, value in raw_params.items()}
    normalized = {key: value for key, value in normalized.items() if value not in (None, "")}
    # В списании материалов сортировку/фильтр по статусу убрали: поиск должен работать
    # как в замечаниях и не прятать строки из-за старого выбранного статуса.
    dop_only = str(normalized.get("dop_only") or "").strip() == "1"
    normalized.pop("status", None)
    normalized.setdefault("sort", "apartment")

    category_id = None
    if dop_only:
        dop_category = next(
            (
                category
                for category in WorkCategory.query.filter_by(is_active=True).all()
                if re.sub(r"[\s.]+", "", category.name or "").casefold() == "допсоглашение"
            ),
            None,
        )
        category_id = dop_category.id if dop_category else None

    query = build_task_query(normalized, category_id=category_id, project_id=project_id)
    acceptance_status = normalized.get("acceptance_status")
    if acceptance_status == "accepted":
        query = query.filter(Apartment.is_app_mode.is_(True))
    elif acceptance_status == "waiting":
        query = query.filter(Apartment.is_app_mode.is_(False))
    if dop_only:
        query = query.filter(Task.work_point.has(WorkPoint.point_number.in_(DOP_AGREEMENT_POINT_NUMBERS)))
    query = query.filter(
        ~exists().where(material_writeoff_tasks.c.task_id == Task.id)
    )
    if dop_only:
        query = query.order_by(None).order_by(
            Task.is_done.desc(),
            cast(Apartment.apartment_number, Integer).asc(),
            Apartment.apartment_number.asc(),
            WorkPoint.point_number.asc(),
            Task.id.asc(),
        )
    return (
        query.options(selectinload(Task.apartment), selectinload(Task.work_point))
        .limit(500)
        .all()
    )

def project_stats(project: Project) -> dict[str, int]:
    stats = dashboard_stats(project.id)
    return {
        "tasks": int(stats.get("tasks") or 0),
        "done": int(stats.get("done") or 0),
        "apartments": int(stats.get("apartment_count") or 0),
        "commercial_count": int(stats.get("commercial_count") or 0),
        "premises": int(stats.get("apartments") or 0),
        "transferred": int(stats.get("accepted") or 0),
        "not_transferred": int(stats.get("not_accepted") or 0),
        "unsold": int(stats.get("unsold") or 0),
    }


@bp.route("/objects")
@login_required
def objects():
    projects_query = Project.query.order_by(Project.created_at.desc())
    projects = [project for project in projects_query.all() if current_user.can_access_project(project)]
    changed = False
    for project in projects:
        if project.name == "100 Квартал 7 очередь":
            if not project.address:
                project.address = "Архангельск, ул. Поморская, 34"
                changed = True
    if changed:
        db.session.commit()
    project_cards = [{"project": project, "stats": project_stats(project)} for project in projects]
    return render_template("objects.html", project_cards=project_cards)


@bp.route("/objects/new", methods=["GET", "POST"])
@login_required
def object_new():
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    form = ProjectForm()
    if form.validate_on_submit():
        if form.has_storerooms.data:
            flash("Кладовки пока нельзя добавить: раздел находится в разработке.", "warning")
            return render_template("object_form.html", form=form, form_title="Добавить объект", submit_label="Создать объект")
        if not (form.has_apartments.data or form.has_commercial.data):
            flash("Выберите хотя бы один тип помещений: квартиры или коммерции.", "warning")
            return render_template("object_form.html", form=form, form_title="Добавить объект", submit_label="Создать объект")
        project = Project(
            name=form.name.data.strip(),
            address=form.address.data.strip() if form.address.data else None,
            has_apartments=bool(form.has_apartments.data),
            has_commercial=bool(form.has_commercial.data),
            has_storerooms=False,
        )
        db.session.add(project)
        db.session.commit()
        session["current_project_id"] = project.id
        flash("Объект создан", "success")
        return redirect(url_for("main.dashboard"))
    return render_template("object_form.html", form=form, form_title="Добавить объект", submit_label="Создать объект")


@bp.route("/objects/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
def object_edit(project_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = _abort_if_project_forbidden(db.session.get(Project, project_id) or abort(404))
    form = ProjectForm(obj=project)
    if form.validate_on_submit():
        if form.has_storerooms.data:
            flash("Кладовки пока нельзя добавить: раздел находится в разработке.", "warning")
            return render_template("object_form.html", form=form, form_title="Редактировать объект", submit_label="Сохранить")
        if not (form.has_apartments.data or form.has_commercial.data):
            flash("Выберите хотя бы один тип помещений: квартиры или коммерции.", "warning")
            return render_template("object_form.html", form=form, form_title="Редактировать объект", submit_label="Сохранить")
        project.name = form.name.data.strip()
        project.address = form.address.data.strip() if form.address.data else None
        project.has_apartments = bool(form.has_apartments.data)
        project.has_commercial = bool(form.has_commercial.data)
        project.has_storerooms = False
        db.session.commit()
        flash("Объект обновлён", "success")
        return redirect(url_for("main.objects"))
    return render_template("object_form.html", form=form, form_title="Редактировать объект", submit_label="Сохранить")


@bp.route("/objects/<int:project_id>/delete", methods=["POST"])
@login_required
def object_delete(project_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = _abort_if_project_forbidden(db.session.get(Project, project_id) or abort(404))
    _record_simple_deletion(
        "object_delete",
        "project",
        project,
        project.name,
        f"Удалён объект «{project.name}».",
        project_id=None,
        extra={
            "apartments_count": Apartment.query.filter(Apartment.project_id == project.id).count(),
            "tasks_count": Task.query.filter(Task.project_id == project.id).count(),
        },
    )
    task_ids = [task_id for (task_id,) in db.session.query(Task.id).filter(Task.project_id == project.id).all()]
    if task_ids:
        SyncConflict.query.filter(SyncConflict.task_id.in_(task_ids)).delete(synchronize_session=False)
    db.session.delete(project)
    if session.get("current_project_id") == project_id:
        session.pop("current_project_id", None)
    db.session.commit()
    flash("Объект удалён", "success")
    return redirect(url_for("main.objects"))


@bp.route("/objects/<int:project_id>/delete/confirm", methods=["GET"])
@login_required
def object_delete_confirm(project_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = _abort_if_project_forbidden(db.session.get(Project, project_id) or abort(404))
    return render_template("object_delete_confirm.html", project=project)


@bp.route("/objects/<int:project_id>/open")
@login_required
def object_open(project_id: int):
    project = _abort_if_project_forbidden(db.session.get(Project, project_id) or abort(404))
    session["current_project_id"] = project.id
    if current_user.role == ROLE_VERIFIER:
        response = redirect(url_for("main.work_report"))
    else:
        response = redirect(url_for("main.dashboard"))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response



@bp.route("/")
@login_required
def dashboard():
    if current_user.role in WORKER_ROLES:
        return redirect(url_for("main.my_tasks"))
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    ensure_default_categories()
    db.session.commit()
    stats = dashboard_stats(project.id)
    categories = category_stats(project.id)
    return render_template("dashboard.html", stats=stats, categories=categories, project=project)


@bp.route("/object")
@login_required
def dashboard_legacy():
    """Keep the legacy CRM entry URL usable for bookmarks and offline copies."""
    if current_user.role == ROLE_VERIFIER:
        return redirect(url_for("main.work_report"))
    return dashboard()


@bp.route("/sync/google", methods=["POST"])
@login_required
def sync_google():
    if not can_manage_sync(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    try:
        result = sync_google_sheets(project_name=project.name, spreadsheet_id_override=project.google_sheet_id)
        pending_conflicts = _project_pending_conflicts_query(project.id).count()
        flash(
            f"Добавлено новых - {result.get('created_count', 0)} , несостыковок - {pending_conflicts}",
            "success",
        )
    except Exception as exc:
        current_app.logger.exception("Google Sheets sync failed")
        flash(f"Ошибка синхронизации Google Sheets: {exc}", "danger")
    return redirect(url_for("main.dashboard"))


@bp.route("/report-error", methods=["POST"])
def report_error():
    if hit_rate_limit("report-error", 6, 300):
        security_event("report_error_rate_limited", "Слишком много сообщений об ошибке", severity="warning")
        abort(429)
    message = (request.form.get("message") or "").strip()
    if not message:
        flash("Опишите ошибку текстом", "warning")
    else:
        _save_site_error(message[:5000], kind="user")
        flash("Ошибка отправлена. Спасибо, разберём.", "success")
    return _safe_redirect(request.form.get("page_url") or request.referrer, "auth.login")


def _visit_browser_label(user_agent: str | None) -> str:
    ua = (user_agent or "").lower()
    if "yabrowser" in ua:
        return "Yandex Browser"
    if "edg/" in ua:
        return "Microsoft Edge"
    if "opr/" in ua or "opera" in ua:
        return "Opera"
    if "chrome/" in ua and "chromium" not in ua:
        return "Google Chrome"
    if "firefox/" in ua:
        return "Mozilla Firefox"
    if "safari/" in ua and "chrome/" not in ua:
        return "Safari"
    if "postmanruntime" in ua:
        return "Postman"
    if "python-requests" in ua:
        return "Python Requests"
    return "Неизвестный браузер"


def _visit_os_label(user_agent: str | None) -> str:
    ua = (user_agent or "").lower()
    if "windows" in ua:
        return "Windows"
    if "iphone" in ua or "ios" in ua:
        return "iPhone (iOS)"
    if "ipad" in ua:
        return "iPadOS"
    if "android" in ua:
        return "Android"
    if "mac os x" in ua or "macintosh" in ua:
        return "macOS"
    if "linux" in ua:
        return "Linux"
    return "Неизвестная ОС"


def _visit_device_label(user_agent: str | None) -> str:
    ua = (user_agent or "").lower()
    if "ipad" in ua or "tablet" in ua:
        return "Планшет"
    if "iphone" in ua or "android" in ua and "mobile" in ua:
        return "Телефон"
    if "mobile" in ua:
        return "Телефон"
    return "Компьютер"


def _is_mobile_phone_request(user_agent: str | None = None) -> bool:
    ua = (user_agent or request.headers.get("User-Agent") or "").lower()
    return (
        "iphone" in ua
        or "ipod" in ua
        or "windows phone" in ua
        or "webos" in ua
        or "blackberry" in ua
        or "opera mini" in ua
        or "iemobile" in ua
        or ("android" in ua and "mobile" in ua)
        or ("mobile" in ua and "ipad" not in ua and "tablet" not in ua)
    )


def _mobile_phone_home_endpoint() -> str:
    if current_user.role in WORKER_ROLES:
        return "main.my_tasks"
    if current_user.role == ROLE_VERIFIER:
        return "main.work_report"
    project_id = session.get("current_project_id") if current_user.can_access_all_projects else (
        current_user.project_id if current_user.is_authenticated and current_user.project_id else session.get("current_project_id")
    )
    return "main.dashboard" if project_id else "main.objects"


def _mobile_phone_allowed_endpoints() -> set[str]:
    if current_user.role in WORKER_ROLES:
        return set(WORKER_ALLOWED_ENDPOINTS)
    if current_user.role == ROLE_VERIFIER:
        return set(VERIFIER_ALLOWED_ENDPOINTS)

    allowed = {
        "main.report_error",
        "main.objects",
        "main.object_open",
        "main.dashboard",
        "main.dashboard_legacy",
        "main.account",
        "main.task_list",
        "main.task_detail",
        "main.task_new",
        "main.task_delete",
        "main.task_recognition",
        "main.quick_status",
        "main.update_task",
        "main.add_task_comment",
        "main.inline_update_text",
        "main.split_task_remark",
        "main.remarks_excel_selection",
        "main.export_category_tasks",
        "main.export_category_tasks_pdf",
        "main.glass_measurements",
        "main.glass_need_measure",
        "main.glass_measurement_save",
        "main.glass_measurement_return_to_all",
        "main.glass_status_update",
        "main.glass_order_export",
        "main.glass_create_material_request",
        "main.glass_measurements_delete",
        "main.glass_order",
        "main.glass_manual_task_new",
        "main.material_request_detail",
        "main.apartments",
        "main.apartments_export",
        "main.apartment_detail",
        "main.update_apartment_po_status",
        "main.update_apartment_inspection_status",
        "main.update_apartment_inspection_date",
        "main.update_apartment_inspection_note",
        "main.update_apartment_comment",
        "main.update_apartment_avr_status",
    }
    if current_user.role in {ROLE_ADMIN, ROLE_MANAGER}:
        allowed.update({
            "main.assignments",
            "main.assignment_manual_task_new",
            "main.assignment_issued_employee_export",
        })
    return allowed


def _visit_status_tone(status_code: int | None) -> str:
    if not status_code:
        return "muted"
    if status_code >= 500:
        return "danger"
    if status_code >= 400:
        return "warning"
    if status_code >= 300:
        return "info"
    return "success"


def _visit_status_bucket_label(status_code: int | None) -> str:
    if not status_code:
        return "Без ответа"
    if status_code >= 500:
        return "5xx"
    if status_code >= 400:
        return "4xx"
    if status_code >= 300:
        return "3xx"
    if status_code >= 200:
        return "2xx"
    return "Прочее"


def _visit_user_label(user: User | None) -> str:
    if not user:
        return "Гость"
    return (user.full_name or user.username or f"ID {user.id}").strip()


def _visit_short_path(path: str | None, endpoint: str | None = None) -> str:
    parsed = urlparse(path or "")
    short = parsed.path or path or endpoint or "—"
    if len(short) > 44:
        return f"{short[:41]}..."
    return short


def _build_site_visit_daily_series(base_query, start_date: date, end_date: date) -> list[dict]:
    span = max(1, min((end_date - start_date).days + 1, 30))
    chart_start_date = end_date - timedelta(days=span - 1)
    use_compact_labels = span >= 21
    rows = (
        base_query
        .with_entities(func.date(SiteVisit.created_at).label("day"), func.count(SiteVisit.id).label("hits"))
        .filter(SiteVisit.created_at >= datetime.combine(chart_start_date, datetime.min.time()))
        .group_by(func.date(SiteVisit.created_at))
        .all()
    )
    day_map = {str(day): int(hits or 0) for day, hits in rows}
    peak = max(day_map.values(), default=0)
    series = []
    for offset in range(span):
        day = chart_start_date + timedelta(days=offset)
        hits = day_map.get(day.isoformat(), 0)
        if peak <= 0 or hits <= 0:
            bar = 0
            if False and duplicate_count:
                flash("Такие замечания уже есть в базе", "warning")
                return render_template(
                    "task_form.html",
                    project=project,
                    apartments=apartments,
                    points=points,
                    add_mode=add_mode,
                    manual_kind=manual_kind,
                )
        else:
            # Scale strictly against the current peak for the visible period.
            bar = max(1, min(100, int(round((hits / peak) * 100))))
        series.append({
            "date": day,
            "label": day.strftime("%d"),
            "full_label": day.strftime("%d.%m.%Y"),
            "compact_label": use_compact_labels,
            "hits": hits,
            "bar": bar,
        })
    return series


def _build_site_visit_chart_month_label(start_date: date, end_date: date) -> str:
    month_names = {
        1: "Январь",
        2: "Февраль",
        3: "Март",
        4: "Апрель",
        5: "Май",
        6: "Июнь",
        7: "Июль",
        8: "Август",
        9: "Сентябрь",
        10: "Октябрь",
        11: "Ноябрь",
        12: "Декабрь",
    }
    start_month = month_names.get(start_date.month, "")
    end_month = month_names.get(end_date.month, "")
    if start_date.year == end_date.year:
        if start_date.month == end_date.month:
            return f"{start_month} {start_date.year}"
        return f"{start_month} - {end_month} {start_date.year}"
    return f"{start_month} {start_date.year} - {end_month} {end_date.year}"


def _build_site_visit_future_notice(start_date: date, end_date: date, today: date, total_visits: int) -> str:
    if total_visits > 0 or start_date <= today:
        return ""
    if start_date == end_date:
        return "Информации за эту дату еще нет."
    return "Информации за этот период еще нет."


def _build_site_visit_period_label(start_date: date, end_date: date) -> str:
    if start_date == end_date:
        return format_ru_day_month(start_date)
    return f"с {format_ru_date(start_date)} по {format_ru_date(end_date)}"


def _build_site_visit_ip_summary(base_query, ip_address: str | None) -> dict | None:
    if not ip_address:
        return None

    path_counter: Counter[str] = Counter()
    users_seen: dict[int, User] = {}
    projects_seen: Counter[str] = Counter()
    browser_counter: Counter[str] = Counter()
    first_seen = None
    last_seen = None

    visits = (
        base_query
        .filter(SiteVisit.ip_address == ip_address)
        .options(selectinload(SiteVisit.user), selectinload(SiteVisit.project))
        .order_by(SiteVisit.created_at.asc(), SiteVisit.id.asc())
        .limit(500)
        .all()
    )
    if not visits:
        return None

    for visit in visits:
        path_counter[_visit_short_path(visit.path, visit.endpoint)] += 1
        browser_counter[_visit_browser_label(visit.user_agent)] += 1
        if visit.user:
            users_seen[visit.user.id] = visit.user
        if visit.project:
            projects_seen[visit.project.name] += 1
        elif not visit.project_id:
            projects_seen["Р‘РµР· РѕР±СЉРµРєС‚Р°"] += 1
        if first_seen is None:
            first_seen = visit.created_at
        last_seen = visit.created_at

    return {
        "ip": ip_address,
        "first_seen": first_seen,
        "last_seen": last_seen,
        "top_paths": path_counter.most_common(5),
        "users": list(users_seen.values())[:6],
        "projects": projects_seen.most_common(4),
        "browsers": browser_counter.most_common(4),
    }


def _build_site_visit_visitor_groups(visits: list[SiteVisit]) -> list[dict]:
    groups: dict[tuple[str, str], dict] = {}
    ordered_groups: list[dict] = []

    for visit in visits:
        if visit.user_id:
            key = ("user", str(visit.user_id))
            label = _visit_user_label(visit.user)
            kind_label = "Авторизован"
            kind_class = "is-auth"
        else:
            guest_ip = (visit.ip_address or "").strip()
            key = ("guest", guest_ip or f"guest-{visit.id}")
            label = f"Гость • {guest_ip}" if guest_ip else "Гость"
            kind_label = "Гость"
            kind_class = "is-guest"

        group = groups.get(key)
        if group is None:
            group = {
                "key": key,
                "label": label,
                "kind_label": kind_label,
                "kind_class": kind_class,
                "hits": 0,
                "ip_set": set(),
                "last_seen": visit.created_at,
                "visits": [],
            }
            groups[key] = group
            ordered_groups.append(group)

        group["hits"] += 1
        if visit.ip_address:
            group["ip_set"].add(visit.ip_address)
        group["visits"].append(visit)

    for group in ordered_groups:
        group["ip_count"] = len(group["ip_set"])
        group.pop("ip_set", None)

    return ordered_groups


def _build_site_visit_agent_stats(base_query) -> tuple[list[dict], list[dict], list[dict]]:
    browser_counter: Counter[str] = Counter()
    os_counter: Counter[str] = Counter()
    device_counter: Counter[str] = Counter()
    rows = base_query.with_entities(SiteVisit.user_agent).order_by(SiteVisit.created_at.desc()).limit(1500).all()
    for (user_agent,) in rows:
        browser_counter[_visit_browser_label(user_agent)] += 1
        os_counter[_visit_os_label(user_agent)] += 1
        device_counter[_visit_device_label(user_agent)] += 1

    def build_items(counter: Counter[str]) -> list[dict]:
        peak = max(counter.values(), default=0)
        return [
            {
                "label": label,
                "count": count,
                "bar": 0 if peak <= 0 else max(12, int((count / peak) * 100)),
            }
            for label, count in counter.most_common(5)
        ]

    return build_items(browser_counter), build_items(os_counter), build_items(device_counter)


def _ru_plural(number: int, forms: tuple[str, str, str]) -> str:
    value = abs(int(number))
    if 11 <= value % 100 <= 14:
        return forms[2]
    if value % 10 == 1:
        return forms[0]
    if 2 <= value % 10 <= 4:
        return forms[1]
    return forms[2]


def _format_ru_milliseconds(value: float | int | None) -> str:
    milliseconds = max(int(round(value or 0)), 0)
    return f"{milliseconds} м/сек"


def _developer_statistics_filters() -> dict:
    allowed_days = [7, 14, 30]
    max_period_days = max(allowed_days)
    try:
        days = int(request.args.get("days") or 7)
    except (TypeError, ValueError):
        days = 7
    if days not in allowed_days:
        days = 7
    today = date.today()
    period_end_date = parse_date(request.args.get("end_date")) or parse_date(request.args.get("date")) or today
    period_start_date = parse_date(request.args.get("start_date"))
    if period_start_date and period_start_date > period_end_date:
        period_start_date, period_end_date = period_end_date, period_start_date
    if period_start_date:
        current_period_days = max(1, (period_end_date - period_start_date).days + 1)
        if current_period_days > max_period_days:
            period_start_date = period_end_date - timedelta(days=max_period_days - 1)
            current_period_days = max_period_days
        days = current_period_days
    else:
        period_start_date = period_end_date - timedelta(days=days - 1)
    scope = (request.args.get("scope") or "all").strip().lower()
    if scope not in {"all", "auth", "guest"}:
        scope = "all"
    ip_filter = (request.args.get("ip") or "").strip()
    path_filter = (request.args.get("path") or "").strip()
    user_filter_raw = (request.args.get("user_id") or "").strip()
    try:
        user_filter_id = int(user_filter_raw) if user_filter_raw else None
    except (TypeError, ValueError):
        user_filter_id = None
    return {
        "allowed_days": allowed_days,
        "days": days,
        "selected_date": period_end_date,
        "selected_date_iso": period_end_date.isoformat(),
        "period_start_date": period_start_date,
        "period_start_date_iso": period_start_date.isoformat(),
        "period_end_date": period_end_date,
        "period_end_date_iso": period_end_date.isoformat(),
        "max_period_days": max_period_days,
        "today": today,
        "scope": scope,
        "ip_filter": ip_filter,
        "path_filter": path_filter,
        "user_filter_id": user_filter_id,
    }


def _apply_site_visit_filters(query, *, project, range_start, range_end, scope, ip_filter, path_filter, user_filter_id, visit_kind: str | None = None):
    query = query.filter(SiteVisit.created_at >= range_start, SiteVisit.created_at < range_end)
    if project:
        query = query.filter(or_(SiteVisit.project_id == project.id, SiteVisit.project_id.is_(None)))
    if visit_kind == "request":
        query = query.filter(or_(SiteVisit.visit_kind == "request", SiteVisit.visit_kind.is_(None)))
    elif visit_kind:
        query = query.filter(SiteVisit.visit_kind == visit_kind)
    if scope == "auth":
        query = query.filter(SiteVisit.is_authenticated.is_(True))
    elif scope == "guest":
        query = query.filter(SiteVisit.is_authenticated.is_(False))
    if ip_filter:
        like_value = f"%{ip_filter[:80]}%"
        query = query.filter(or_(SiteVisit.ip_address.ilike(like_value), SiteVisit.forwarded_for.ilike(like_value)))
    if path_filter:
        like_value = f"%{path_filter[:200]}%"
        query = query.filter(or_(SiteVisit.path.ilike(like_value), SiteVisit.endpoint.ilike(like_value)))
    if user_filter_id:
        query = query.filter(SiteVisit.user_id == user_filter_id)
    return query


def _build_developer_statistics_context() -> dict:
    project = selected_project()
    filters = _developer_statistics_filters()
    period_start_date = filters["period_start_date"]
    period_end_date = filters["period_end_date"]
    scope = filters["scope"]
    ip_filter = filters["ip_filter"]
    path_filter = filters["path_filter"]
    user_filter_id = filters["user_filter_id"]
    range_start = datetime.combine(period_start_date, datetime.min.time())
    range_end = datetime.combine(period_end_date + timedelta(days=1), datetime.min.time())

    tab_query = _apply_site_visit_filters(
        SiteVisit.query,
        project=project,
        range_start=range_start,
        range_end=range_end,
        scope=scope,
        ip_filter=ip_filter,
        path_filter=path_filter,
        user_filter_id=user_filter_id,
        visit_kind="tab_open",
    )
    request_query = _apply_site_visit_filters(
        SiteVisit.query,
        project=project,
        range_start=range_start,
        range_end=range_end,
        scope=scope,
        ip_filter=ip_filter,
        path_filter=path_filter,
        user_filter_id=user_filter_id,
        visit_kind="request",
    )

    total_visits = tab_query.order_by(None).count()
    avg_duration = request_query.with_entities(func.avg(SiteVisit.duration_ms)).scalar() or 0
    unique_ips = (
        tab_query
        .with_entities(func.count(distinct(SiteVisit.ip_address)))
        .filter(SiteVisit.ip_address.isnot(None))
        .scalar()
        or 0
    )
    known_users = (
        tab_query
        .with_entities(SiteVisit.user_id)
        .filter(SiteVisit.user_id.isnot(None))
        .distinct()
        .all()
    )
    guest_ips = (
        tab_query
        .with_entities(SiteVisit.ip_address)
        .filter(SiteVisit.user_id.is_(None), SiteVisit.ip_address.isnot(None))
        .distinct()
        .all()
    )
    latest_visit_at = tab_query.with_entities(func.max(SiteVisit.created_at)).scalar()
    unique_visitors = len(known_users) + len(guest_ips)

    top_ip_rows = (
        tab_query
        .with_entities(
            SiteVisit.ip_address,
            func.count(SiteVisit.id).label("hits"),
            func.count(distinct(SiteVisit.user_id)).label("users_count"),
            func.max(SiteVisit.created_at).label("last_seen"),
        )
        .filter(SiteVisit.ip_address.isnot(None))
        .group_by(SiteVisit.ip_address)
        .order_by(func.count(SiteVisit.id).desc(), func.max(SiteVisit.created_at).desc())
        .limit(10)
        .all()
    )
    top_ips = [
        {
            "ip_address": ip_address or "—",
            "hits": int(hits or 0),
            "users_count": int(users_count or 0),
            "last_seen": last_seen,
        }
        for ip_address, hits, users_count, last_seen in top_ip_rows
    ]

    top_user_rows = (
        tab_query
        .with_entities(
            SiteVisit.user_id,
            func.count(SiteVisit.id).label("hits"),
            func.count(distinct(SiteVisit.ip_address)).label("ip_count"),
            func.max(SiteVisit.created_at).label("last_seen"),
        )
        .filter(SiteVisit.user_id.isnot(None))
        .group_by(SiteVisit.user_id)
        .order_by(func.count(SiteVisit.id).desc(), func.max(SiteVisit.created_at).desc())
        .limit(10)
        .all()
    )
    top_user_ids = [user_id for user_id, _, _, _ in top_user_rows if user_id]
    top_user_map = {user.id: user for user in User.query.filter(User.id.in_(top_user_ids)).all()} if top_user_ids else {}
    top_users = [
        {
            "user": top_user_map.get(user_id),
            "hits": int(hits or 0),
            "ip_count": int(ip_count or 0),
            "last_seen": last_seen,
        }
        for user_id, hits, ip_count, last_seen in top_user_rows
    ]

    recent_visits = (
        tab_query
        .options(selectinload(SiteVisit.user), selectinload(SiteVisit.project))
        .order_by(SiteVisit.created_at.desc(), SiteVisit.id.desc())
        .limit(240)
        .all()
    )
    visitor_groups = _build_site_visit_visitor_groups(recent_visits)

    focused_ip_summary = None
    if ip_filter and total_visits:
        path_counter: Counter[str] = Counter()
        users_seen: dict[int, User] = {}
        projects_seen: Counter[str] = Counter()
        browser_counter: Counter[str] = Counter()
        first_seen = None
        for visit in (
            tab_query
            .options(selectinload(SiteVisit.user), selectinload(SiteVisit.project))
            .order_by(SiteVisit.created_at.asc(), SiteVisit.id.asc())
            .limit(500)
            .all()
        ):
            path_counter[_visit_short_path(visit.path, visit.endpoint)] += 1
            browser_counter[_visit_browser_label(visit.user_agent)] += 1
            if visit.user:
                users_seen[visit.user.id] = visit.user
            if visit.project:
                projects_seen[visit.project.name] += 1
            elif not visit.project_id:
                projects_seen["Без объекта"] += 1
            if first_seen is None:
                first_seen = visit.created_at
        focused_ip_summary = {
            "ip": ip_filter,
            "first_seen": first_seen,
            "last_seen": latest_visit_at,
            "top_paths": path_counter.most_common(5),
            "users": list(users_seen.values())[:6],
            "projects": projects_seen.most_common(4),
            "browsers": browser_counter.most_common(4),
        }

    ip_summaries = {
        item["ip_address"]: _build_site_visit_ip_summary(tab_query, item["ip_address"])
        for item in top_ips
    }
    if ip_filter and not focused_ip_summary:
        focused_ip_summary = ip_summaries.get(ip_filter)

    browser_stats, os_stats, device_stats = _build_site_visit_agent_stats(tab_query)
    environment_summary = {
        "browser": browser_stats[0] if browser_stats else None,
        "os": os_stats[0] if os_stats else None,
        "device": device_stats[0] if device_stats else None,
    }
    daily_series = _build_site_visit_daily_series(tab_query, period_start_date, period_end_date)
    chart_month_label = _build_site_visit_chart_month_label(period_start_date, period_end_date)
    period_display_label = _build_site_visit_period_label(period_start_date, period_end_date)
    future_notice = _build_site_visit_future_notice(period_start_date, period_end_date, filters["today"], total_visits)
    return {
        "project": project,
        **filters,
        "total_visits": total_visits,
        "unique_visitors": unique_visitors,
        "unique_ips": unique_ips,
        "avg_duration": int(avg_duration or 0),
        "avg_duration_label": _format_ru_milliseconds(avg_duration),
        "latest_visit_at": latest_visit_at,
        "top_ips": top_ips,
        "top_users": top_users,
        "ip_summaries": ip_summaries,
        "recent_visits": recent_visits,
        "visitor_groups": visitor_groups,
        "browser_stats": browser_stats,
        "os_stats": os_stats,
        "device_stats": device_stats,
        "environment_summary": environment_summary,
        "daily_series": daily_series,
        "chart_month_label": chart_month_label,
        "period_display_label": period_display_label,
        "future_notice": future_notice,
        "focused_ip_summary": focused_ip_summary,
        "visit_browser_label": _visit_browser_label,
        "visit_os_label": _visit_os_label,
        "visit_device_label": _visit_device_label,
        "visit_user_label": _visit_user_label,
        "visit_short_path": _visit_short_path,
    }


def _render_developer_statistics_page(template_name: str, active_page: str, subtitle: str):
    if current_user.role != ROLE_ADMIN:
        abort(403)
    context = _build_developer_statistics_context()
    context.update(
        active_statistics_page=active_page,
        statistics_subtitle=subtitle,
    )
    return render_template(template_name, **context)


@bp.route("/developer/statistics")
@login_required
def developer_statistics():
    return _render_developer_statistics_page(
        "developer_statistics.html",
        active_page="overview",
        subtitle="",
    )


@bp.route("/developer/statistics/visits")
@login_required
def developer_statistics_visits():
    return _render_developer_statistics_page(
        "developer_statistics_visits.html",
        active_page="visits",
        subtitle="",
    )


@bp.route("/developer/statistics/sources")
@login_required
def developer_statistics_sources():
    return _render_developer_statistics_page(
        "developer_statistics_sources.html",
        active_page="sources",
        subtitle="",
    )


@bp.route("/analytics/tab-open", methods=["POST"])
@csrf.exempt
def analytics_tab_open():
    if hit_rate_limit("site-tab-open", 180, 60):
        return ("", 204)
    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}
    tab_id = str(payload.get("tab_id") or "").strip()[:80]
    if not tab_id:
        return ("", 204)

    opened_path = str(payload.get("path") or "").strip()
    if not opened_path and request.referrer:
        parsed_referrer = urlparse(request.referrer)
        opened_path = parsed_referrer.path or "/"
        if parsed_referrer.query:
            opened_path = f"{opened_path}?{parsed_referrer.query}"
    referrer = str(payload.get("referrer") or "").strip()[:500] or None
    forwarded_for = (request.headers.get("X-Forwarded-For") or "")[:255] or None

    try:
        existing_visit = (
            SiteVisit.query
            .with_entities(SiteVisit.id)
            .filter(SiteVisit.visit_kind == "tab_open", SiteVisit.tab_id == tab_id)
            .first()
        )
        if existing_visit:
            return ("", 204)

        visit = SiteVisit(
            project_id=resolve_site_visit_project_id(),
            user_id=current_user.id if getattr(current_user, "is_authenticated", False) else None,
            ip_address=client_ip()[:80],
            forwarded_for=forwarded_for,
            endpoint="analytics_tab_open",
            method="TAB",
            path=(opened_path or "/")[:500],
            referrer=referrer,
            user_agent=(request.headers.get("User-Agent") or "")[:500] or None,
            status_code=204,
            duration_ms=None,
            is_authenticated=bool(getattr(current_user, "is_authenticated", False)),
            visit_kind="tab_open",
            tab_id=tab_id,
        )
        db.session.add(visit)
        db.session.commit()
    except Exception:
        db.session.rollback()
    return ("", 204)


@bp.route("/site-errors")
@login_required
def site_errors():
    if current_user.role != ROLE_ADMIN:
        abort(403)
    project = selected_project()
    query = SiteErrorReport.query.options(selectinload(SiteErrorReport.user), selectinload(SiteErrorReport.project))
    if project:
        query = query.filter(or_(SiteErrorReport.project_id == project.id, SiteErrorReport.project_id.is_(None)))
    status = request.args.get("status") or ""
    kind = request.args.get("kind") or ""
    if status:
        query = query.filter(SiteErrorReport.status == status)
    if kind:
        query = query.filter(SiteErrorReport.kind == kind)
    reports = query.order_by(SiteErrorReport.created_at.desc()).limit(300).all()
    return render_template("site_errors.html", reports=reports, project=project, status=status, kind=kind)


@bp.route("/site-errors/<int:report_id>/close", methods=["POST"])
@login_required
def site_error_close(report_id: int):
    if current_user.role != ROLE_ADMIN:
        abort(403)
    report = db.session.get(SiteErrorReport, report_id) or abort(404)
    project = selected_project()
    if project and report.project_id not in {None, project.id}:
        abort(404)
    report.status = "closed" if report.status != "closed" else "new"
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify(ok=True, status=report.status, message="Статус ошибки обновлен")
    flash("Статус ошибки обновлен", "success")
    return redirect(request.referrer or url_for("main.site_errors"))


@bp.route("/site-errors/<int:report_id>/delete", methods=["POST"])
@login_required
def site_error_delete(report_id: int):
    if current_user.role != ROLE_ADMIN:
        abort(403)
    report = db.session.get(SiteErrorReport, report_id) or abort(404)
    project = selected_project()
    if project and report.project_id not in {None, project.id}:
        abort(404)
    snapshot = {
        "kind": report.kind,
        "message": report.message,
        "page_url": report.page_url,
        "user_agent": report.user_agent,
        "traceback_text": report.traceback_text,
        "status": report.status,
        "project_id": report.project_id,
        "user_id": report.user_id,
    }
    _record_deletion_action(
        "site_error_delete",
        "site_error_report",
        report.id,
        "Заявка на регистрацию" if report.kind == "registration" else "Запись для разработчика",
        "Удалена запись из раздела «Для разработчика».",
        snapshot,
        project_id=report.project_id,
    )
    db.session.delete(report)
    db.session.commit()
    flash("Запись удалена. Действие можно отменить в логах удалений.", "success")
    return redirect(url_for("main.site_errors"))


@bp.route("/developer/delete-logs")
@login_required
def developer_delete_logs():
    if current_user.role != ROLE_ADMIN:
        abort(403)
    project = selected_project()
    query = DeletionActionLog.query.options(
        selectinload(DeletionActionLog.user),
        selectinload(DeletionActionLog.project),
        selectinload(DeletionActionLog.undone_by),
    )
    if project:
        query = query.filter(or_(DeletionActionLog.project_id == project.id, DeletionActionLog.project_id.is_(None)))
    logs = query.order_by(DeletionActionLog.created_at.desc(), DeletionActionLog.id.desc()).limit(300).all()
    return render_template("developer_delete_logs.html", logs=logs, project=project)


def _restore_glass_measurement_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    task_id = snapshot.get("task_id")
    if not task_id or not db.session.get(Task, int(task_id)):
        return False, "Не удалось восстановить замер: связанное замечание уже отсутствует."
    if GlassMeasurement.query.filter(GlassMeasurement.task_id == int(task_id)).first():
        return False, "Не удалось восстановить замер: у замечания уже есть замер."
    if snapshot.get("material_request_item_id") and not db.session.get(MaterialRequestItem, int(snapshot.get("material_request_item_id"))):
        snapshot = dict(snapshot)
        snapshot["material_request_item_id"] = None
    if snapshot.get("material_writeoff_id") and not db.session.get(MaterialWriteOff, int(snapshot.get("material_writeoff_id"))):
        snapshot = dict(snapshot)
        snapshot["material_writeoff_id"] = None
    measurement = _restore_model_from_snapshot(
        GlassMeasurement,
        snapshot,
        skip={"items"},
    )
    _restore_children_from_snapshots(
        GlassMeasurementItem,
        snapshot.get("items") or [],
        force={"measurement_id": measurement.id},
    )
    return True, "Замер восстановлен."


def _restore_task_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    project_id = snapshot.get("project_id")
    apartment_id = snapshot.get("apartment_id")
    work_point_id = snapshot.get("work_point_id")
    if not project_id or not db.session.get(Project, int(project_id)):
        return False, "Не удалось восстановить замечание: объект уже отсутствует."
    if not apartment_id or not db.session.get(Apartment, int(apartment_id)):
        return False, "Не удалось восстановить замечание: помещение уже отсутствует."
    if not work_point_id or not db.session.get(WorkPoint, int(work_point_id)):
        return False, "Не удалось восстановить замечание: пункт работ уже отсутствует."
    task_snapshot = dict(snapshot)
    source_uid = str(task_snapshot.get("source_uid") or "").strip()
    if source_uid and Task.query.filter(Task.source_uid == source_uid).first():
        task_snapshot["source_uid"] = stable_hash([source_uid, "restored", datetime.utcnow().isoformat()])
    task = _restore_model_from_snapshot(
        Task,
        task_snapshot,
        skip={"comments", "changes", "glass_measurement", "material_writeoff_ids", "comment_count", "change_count", "apartment_label", "work_point"},
    )
    comments = [
        row for row in (snapshot.get("comments") or [])
        if not row.get("user_id") or db.session.get(User, int(row.get("user_id")))
    ]
    changes = [
        row for row in (snapshot.get("changes") or [])
        if not row.get("user_id") or db.session.get(User, int(row.get("user_id")))
    ]
    _restore_children_from_snapshots(TaskComment, comments, force={"task_id": task.id})
    _restore_children_from_snapshots(ChangeLog, changes, force={"task_id": task.id})
    glass_snapshot = snapshot.get("glass_measurement")
    if isinstance(glass_snapshot, dict) and glass_snapshot:
        glass_snapshot = dict(glass_snapshot)
        glass_snapshot["task_id"] = task.id
        _restore_glass_measurement_from_snapshot(glass_snapshot)
    for writeoff_id in snapshot.get("material_writeoff_ids") or []:
        writeoff = db.session.get(MaterialWriteOff, int(writeoff_id)) if writeoff_id else None
        if writeoff and writeoff.project_id == task.project_id and task not in writeoff.tasks:
            writeoff.tasks.append(task)
    return True, "Замечание восстановлено."


def _restore_material_request_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    project_id = snapshot.get("project_id")
    if not project_id or not db.session.get(Project, int(project_id)):
        return False, "Не удалось восстановить заявку: объект уже отсутствует."
    if snapshot.get("author_id") and not db.session.get(User, int(snapshot.get("author_id"))):
        snapshot = dict(snapshot)
        snapshot["author_id"] = None
    material_request = _restore_model_from_snapshot(MaterialRequest, snapshot, skip={"items"})
    _restore_children_from_snapshots(
        MaterialRequestItem,
        snapshot.get("items") or [],
        force={"request_id": material_request.id},
    )
    return True, "Заявка на материалы восстановлена."


def _restore_material_writeoff_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    project_id = snapshot.get("project_id")
    if not project_id or not db.session.get(Project, int(project_id)):
        return False, "Не удалось восстановить списание: объект уже отсутствует."
    if snapshot.get("author_id") and not db.session.get(User, int(snapshot.get("author_id"))):
        snapshot = dict(snapshot)
        snapshot["author_id"] = None
    writeoff = _restore_model_from_snapshot(MaterialWriteOff, snapshot, skip={"items", "task_ids", "measurement_ids"})
    _restore_children_from_snapshots(
        MaterialWriteOffItem,
        snapshot.get("items") or [],
        force={"writeoff_id": writeoff.id},
    )
    for task_id in snapshot.get("task_ids") or []:
        task = db.session.get(Task, int(task_id)) if task_id else None
        if task and task.project_id == writeoff.project_id and task not in writeoff.tasks:
            writeoff.tasks.append(task)
    for measurement_id in snapshot.get("measurement_ids") or []:
        measurement = db.session.get(GlassMeasurement, int(measurement_id)) if measurement_id else None
        if measurement and measurement.project_id == writeoff.project_id:
            measurement.material_writeoff = writeoff
    return True, "Списание материалов восстановлено."


def _restore_material_balance_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    restored = 0
    skipped = 0
    for item_snapshot in snapshot.get("request_items") or []:
        request_id = item_snapshot.get("request_id")
        if request_id and db.session.get(MaterialRequest, int(request_id)):
            try:
                _restore_model_from_snapshot(MaterialRequestItem, item_snapshot)
                restored += 1
            except ValueError:
                skipped += 1
        else:
            skipped += 1
    for item_snapshot in snapshot.get("writeoff_items") or []:
        writeoff_id = item_snapshot.get("writeoff_id")
        if writeoff_id and db.session.get(MaterialWriteOff, int(writeoff_id)):
            try:
                _restore_model_from_snapshot(MaterialWriteOffItem, item_snapshot)
                restored += 1
            except ValueError:
                skipped += 1
        else:
            skipped += 1
    if restored:
        message = f"Строки баланса восстановлены: {restored}."
        if skipped:
            message += f" Не восстановлено: {skipped}."
        return True, message
    return False, "Не удалось восстановить баланс: в логе нет полного снимка строк или родительские заявки уже отсутствуют."


def _restore_user_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    username = str(snapshot.get("username") or "").strip()
    if username and User.query.filter(User.username == username).first():
        return False, "Не удалось восстановить пользователя: такой логин уже существует."
    if snapshot.get("project_id") and not db.session.get(Project, int(snapshot.get("project_id"))):
        snapshot = dict(snapshot)
        snapshot["project_id"] = None
    _restore_model_from_snapshot(User, snapshot, skip={"project_ids"})
    return True, "Пользователь восстановлен."


def _restore_site_error_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    if snapshot.get("user_id") and not db.session.get(User, int(snapshot.get("user_id"))):
        snapshot = dict(snapshot)
        snapshot["user_id"] = None
    if snapshot.get("project_id") and not db.session.get(Project, int(snapshot.get("project_id"))):
        snapshot = dict(snapshot)
        snapshot["project_id"] = None
    _restore_model_from_snapshot(SiteErrorReport, snapshot)
    return True, "Запись для разработчика восстановлена."


def _restore_sync_log_from_snapshot(snapshot: dict) -> tuple[bool, str]:
    if snapshot.get("project_id") and not db.session.get(Project, int(snapshot.get("project_id"))):
        snapshot = dict(snapshot)
        snapshot["project_id"] = None
    _restore_model_from_snapshot(SyncLog, snapshot)
    return True, "Запись синхронизации восстановлена."


@bp.route("/developer/delete-logs/<int:log_id>/undo", methods=["POST"])
@login_required
def developer_delete_log_undo(log_id: int):
    if current_user.role != ROLE_ADMIN:
        abort(403)
    log = db.session.get(DeletionActionLog, log_id) or abort(404)
    project = selected_project()
    if project and log.project_id not in {None, project.id}:
        abort(404)
    if log.is_undone:
        flash("Это действие уже отменено.", "info")
        return redirect(request.referrer or url_for("main.developer_delete_logs"))
    try:
        snapshot = json.loads(log.snapshot_json or "{}")
    except Exception:
        snapshot = {}

    try:
        ok = False
        message = "Для этого действия отмена пока недоступна."
        if log.action_key == "assignment_delete_from_employee":
            task_id = int(snapshot.get("task_id") or log.entity_id or 0)
            task = db.session.get(Task, task_id) or abort(404)
            if log.project_id and task.project_id != log.project_id:
                abort(404)
            responsible_id = snapshot.get("responsible_id")
            responsible = db.session.get(User, int(responsible_id)) if responsible_id else None
            if not responsible or not _user_can_work_in_project(responsible, task.project):
                flash("Не удалось вернуть исполнителя: пользователь больше недоступен для этого объекта.", "warning")
                return redirect(request.referrer or url_for("main.developer_delete_logs"))
            old_responsible = task.responsible_id
            old_date = task.planned_date
            restored_date = parse_date(snapshot.get("planned_date"))
            task.responsible_id = responsible.id
            task.planned_date = restored_date
            task.manually_edited = True
            log_change(task, "field_update", "responsible_id", old_responsible, responsible.id, user_id=current_user.id)
            if old_date != restored_date:
                log_change(task, "field_update", "planned_date", old_date, restored_date, user_id=current_user.id)
            ok, message = True, "Удаление задачи у исполнителя отменено. Исполнитель восстановлен."
        elif log.action_key == "site_error_delete":
            ok, message = _restore_site_error_from_snapshot(snapshot)
        elif log.action_key == "task_delete":
            ok, message = _restore_task_from_snapshot(snapshot)
        elif log.action_key == "glass_measurement_delete":
            ok, message = _restore_glass_measurement_from_snapshot(snapshot)
        elif log.action_key in {"material_request_delete", "material_request_delete_empty_after_balance"}:
            ok, message = _restore_material_request_from_snapshot(snapshot)
        elif log.action_key in {"material_writeoff_delete", "material_writeoff_delete_empty_after_balance"}:
            ok, message = _restore_material_writeoff_from_snapshot(snapshot)
        elif log.action_key == "material_balance_delete":
            ok, message = _restore_material_balance_from_snapshot(snapshot)
        elif log.action_key == "user_delete":
            ok, message = _restore_user_from_snapshot(snapshot)
        elif log.action_key == "sync_log_delete":
            ok, message = _restore_sync_log_from_snapshot(snapshot)
        elif log.action_key == "object_delete":
            ok, message = False, "Объект нельзя восстановить полностью: в старом логе нет полного снимка квартир и замечаний."

        if ok:
            _mark_deletion_log_undone(log)
            db.session.commit()
            flash(message, "success")
        else:
            db.session.rollback()
            if False:
                flash("Такие замечания уже есть в базе", "warning")
            flash(message, "warning")
    except ValueError as exc:
        db.session.rollback()
        flash(f"Не удалось отменить действие: {exc}", "warning")
    except HTTPException:
        db.session.rollback()
        raise
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to undo deletion action")
        flash("Не удалось отменить действие. Подробности записаны в ошибки сайта.", "danger")
    return redirect(request.referrer or url_for("main.developer_delete_logs"))


@bp.app_errorhandler(Exception)
def handle_unexpected_exception(exc):
    if isinstance(exc, HTTPException):
        return exc
    current_app.logger.exception("Unhandled site error")
    try:
        db.session.rollback()
        _save_site_error(str(exc), kind="system", traceback_text=traceback.format_exc(), page_url=request.url)
    except Exception:
        current_app.logger.exception("Failed to save site error report")
        db.session.rollback()
    return render_template("site_error_500.html"), 500


@bp.route("/tasks")
@login_required
def task_list():
    return _task_list_response(contractor_page=False)


@bp.route("/contractors")
@login_required
def contractors_list():
    return _task_list_response(contractor_page=True)


def _project_contractor(project_id: int, contractor_id: int | None) -> Contractor | None:
    if not contractor_id:
        return None
    return Contractor.query.filter_by(id=contractor_id, project_id=project_id).first()


def _filter_tasks_for_contractor(query, contractor: Contractor | None):
    if contractor is None:
        return query
    work_point_ids = [point.id for point in contractor.work_points]
    apartment_ids = [apartment.id for apartment in contractor.apartments]
    if not work_point_ids or not apartment_ids:
        return query.filter(False)
    return query.filter(
        Task.work_point_id.in_(work_point_ids),
        Task.apartment_id.in_(apartment_ids),
    )


@bp.route("/contractors/directory")
@login_required
def contractor_directory():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    contractors = (
        Contractor.query
        .options(selectinload(Contractor.work_points), selectinload(Contractor.apartments))
        .filter_by(project_id=project.id)
        .order_by(func.lower(Contractor.name).asc(), Contractor.id.asc())
        .all()
    )
    contractor_rows = []
    for contractor in contractors:
        point_labels = sorted({
            CONTRACTOR_POINT_LABELS.get(str(point.point_number).strip(), point.display_name)
            for point in contractor.work_points
        })
        apartment_groups = {
            _apartment_group_key(apartment)
            for apartment in contractor.apartments
            if _is_visible_apartment_row(apartment)
        }
        contractor_rows.append({
            "contractor": contractor,
            "point_labels": point_labels,
            "apartment_count": len(apartment_groups),
        })
    return render_template(
        "contractor_directory.html",
        project=project,
        contractor_rows=contractor_rows,
    )


@bp.route("/contractors/new", methods=["GET", "POST"])
@login_required
def contractor_new():
    return _contractor_form_response()


@bp.route("/contractors/<int:contractor_id>/edit", methods=["GET", "POST"])
@login_required
def contractor_edit(contractor_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    contractor = _project_contractor(project.id, contractor_id) or abort(404)
    return _contractor_form_response(contractor)


@bp.route("/contractors/<int:contractor_id>/delete", methods=["POST"])
@login_required
def contractor_delete(contractor_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    contractor = _project_contractor(project.id, contractor_id) or abort(404)
    contractor_name = contractor.name
    contractor.work_points = []
    contractor.apartments = []
    db.session.delete(contractor)
    db.session.commit()
    flash(f"Подрядчик «{contractor_name}» удалён.", "success")
    return redirect(url_for("main.contractor_directory"))


def _contractor_form_response(contractor: Contractor | None = None):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        flash("Сначала выберите объект.", "warning")
        return redirect(url_for("main.objects"))
    if contractor is not None and contractor.project_id != project.id:
        abort(404)

    point_options = _contractor_point_options()
    apartment_options = _contractor_apartment_options(project.id)
    if request.method == "POST":
        name = str(request.form.get("name") or "").strip()
        selected_point_numbers = {
            str(value).strip()
            for value in request.form.getlist("work_points")
            if str(value).strip()
        }
        selected_apartment_group_ids = {
            str(value).strip()
            for value in request.form.getlist("apartment_groups")
            if str(value).strip()
        }
    else:
        name = contractor.name if contractor else ""
        selected_point_numbers = {
            str(point.point_number).strip()
            for point in (contractor.work_points if contractor else [])
        }
        contractor_apartment_ids = {
            apartment.id for apartment in (contractor.apartments if contractor else [])
        }
        selected_apartment_group_ids = {
            str(option["id"])
            for option in apartment_options
            if contractor_apartment_ids.intersection(option["apartment_ids"])
        }

    if request.method == "POST":
        errors = []
        available_point_numbers = {option["number"] for option in point_options}
        selected_point_numbers &= available_point_numbers
        apartment_group_map = {str(option["id"]): option for option in apartment_options}
        selected_apartment_group_ids &= set(apartment_group_map)

        if len(name) < 2:
            errors.append("Укажите наименование подрядчика.")
        elif len(name) > 180:
            errors.append("Наименование подрядчика не должно превышать 180 символов.")
        duplicate_query = Contractor.query.filter(
            Contractor.project_id == project.id,
            func.lower(Contractor.name) == name.lower(),
        )
        if contractor is not None:
            duplicate_query = duplicate_query.filter(Contractor.id != contractor.id)
        if len(name) >= 2 and duplicate_query.first():
            errors.append("Подрядчик с таким наименованием уже создан на этом объекте.")
        if not selected_point_numbers:
            errors.append("Выберите хотя бы один пункт ответственности.")
        if not selected_apartment_group_ids:
            errors.append("Выберите хотя бы одну квартиру.")

        if not errors:
            work_points = (
                WorkPoint.query.filter(WorkPoint.point_number.in_(selected_point_numbers))
                .order_by(WorkPoint.point_number.asc(), WorkPoint.id.asc())
                .all()
            )
            apartment_ids = {
                apartment_id
                for group_id in selected_apartment_group_ids
                for apartment_id in apartment_group_map[group_id]["apartment_ids"]
            }
            apartments = (
                Apartment.query.filter(
                    Apartment.project_id == project.id,
                    Apartment.id.in_(apartment_ids),
                )
                .order_by(Apartment.id.asc())
                .all()
            )
            if contractor is None:
                contractor = Contractor(project=project)
                db.session.add(contractor)
            contractor.name = name
            contractor.work_points = work_points
            contractor.apartments = apartments
            db.session.commit()
            flash(
                f"Подрядчик «{contractor.name}» {'обновлён' if request.endpoint == 'main.contractor_edit' else 'создан'}.",
                "success",
            )
            return redirect(url_for("main.contractor_directory"))

        for error in errors:
            flash(error, "warning")

    return render_template(
        "contractor_form.html",
        project=project,
        point_options=point_options,
        apartment_options=apartment_options,
        contractor_name=name,
        selected_point_numbers=selected_point_numbers,
        selected_apartment_group_ids=selected_apartment_group_ids,
        contractor=contractor,
        contractor_form_title="Редактировать подрядчика" if contractor else "Создать подрядчика",
        contractor_submit_label="Сохранить изменения" if contractor else "Создать подрядчика",
    )


def _premise_options_from_apartments(apartments: list[Apartment]) -> list[dict]:
    premise_options = []
    seen_premise_ids = set()
    for apartment in apartments:
        if not apartment or apartment.id in seen_premise_ids:
            continue
        seen_premise_ids.add(apartment.id)
        premise_options.append({
            "id": apartment.id,
            "label": apartment.label(),
            "finish": apartment.finishing_type or "",
            "_sort_key": (
                0 if (apartment.premise_type or "apartment") == "commercial" else 1,
                _apartment_number_sort_value(apartment.apartment_number or apartment.construction_number),
                str(apartment.building or "").strip().lower(),
                str(apartment.label() or "").strip().lower(),
            ),
        })
    premise_options.sort(key=lambda item: item.get("_sort_key") or ())
    for item in premise_options:
        item.pop("_sort_key", None)
    return premise_options


def _premise_options_from_tasks(tasks: list[Task]) -> list[dict]:
    return _premise_options_from_apartments([task.apartment for task in tasks if task.apartment])


def _excel_premise_storage_key(base_key: str, project_id: int, export_args: dict[str, object]) -> str:
    normalized_args = {}
    for key, value in sorted((export_args or {}).items()):
        if key in {"page", "task_ids", "premise_ids"}:
            continue
        values = value if isinstance(value, (list, tuple, set)) else [value]
        cleaned_values = [str(item).strip() for item in values if item not in {None, ""} and str(item).strip()]
        if cleaned_values:
            normalized_args[str(key)] = cleaned_values
    signature = stable_hash([json.dumps(normalized_args, ensure_ascii=False, sort_keys=True, separators=(",", ":"))])[:16]
    return f"{base_key}:project:{project_id}:{signature}"


@bp.route("/contractors/excel-selection")
@login_required
def contractors_excel_selection():
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    ensure_default_categories()
    categories = WorkCategory.query.filter_by(is_active=True).order_by(WorkCategory.sort_order.asc()).all()
    all_cat = next((category for category in categories if (category.name or "").strip().lower() == "все"), None)
    query_args = request.args.copy()
    query_args.pop("page", None)
    query_args.pop("task_ids", None)
    query_args["sort"] = "apartment"
    query = build_task_query(query_args, category_id=all_cat.id if all_cat else None, project_id=project.id)
    selected_contractor = _project_contractor(project.id, request.args.get("contractor_id", type=int))
    query = _filter_tasks_for_contractor(query, selected_contractor)
    if current_user.role in WORKER_ROLES:
        query = query.filter(Task.responsible_id == current_user.id)
    tasks = (
        query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.glass_measurement))
        .all()
    )
    export_args = {key: values for key, values in request.args.lists() if key not in {"page", "task_ids", "premise_ids"}}
    back_url = url_for("main.contractors_list", **export_args)
    storage_key = _excel_premise_storage_key("contractors-premises", project.id, export_args)
    return render_template(
        "contractors_excel_selection.html",
        page_title="Выбор для Excel",
        export_form_id="contractorExportForm",
        export_endpoint="main.contractors_export",
        storage_key=storage_key,
        reset_button_class="js-contractor-premise-reset",
        export_button_class="js-contractor-export-submit",
        args=request.args,
        export_args=export_args,
        back_url=back_url,
        premise_options=_premise_options_from_tasks(tasks),
        today=date.today(),
    )


@bp.route("/tasks/<int:category_id>/excel-selection")
@login_required
def remarks_excel_selection(category_id: int):
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    ensure_default_categories()
    category = db.session.get(WorkCategory, category_id) or abort(404)
    query_args = request.args.copy()
    query_args.pop("page", None)
    query_args.pop("task_ids", None)
    query = build_task_query(query_args, category_id=category_id, project_id=project.id)
    if current_user.role in WORKER_ROLES:
        query = query.filter(Task.responsible_id == current_user.id)
    tasks = (
        query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.glass_measurement))
        .all()
    )
    export_args = {key: values for key, values in request.args.lists() if key not in {"page", "task_ids", "premise_ids"}}
    export_args["category_id"] = [str(category_id)]
    back_url = url_for("main.task_list", **export_args)
    storage_key = _excel_premise_storage_key(f"remarks-{category_id}-premises", project.id, export_args)
    return render_template(
        "contractors_excel_selection.html",
        page_title="Выбор для Excel",
        export_form_id="remarksExportForm",
        export_endpoint="main.export_category_tasks",
        export_endpoint_args={"category_id": category_id},
        storage_key=storage_key,
        reset_button_class="js-contractor-premise-reset",
        export_button_class="js-contractor-export-submit",
        args=request.args,
        export_args=export_args,
        back_url=back_url,
        premise_options=_premise_options_from_tasks(tasks),
        active_category=category,
        today=date.today(),
    )


@bp.route("/contractors/export")
@login_required
def contractors_export():
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    query_args = request.args.copy()
    query_args["sort"] = "apartment"
    selected_contractor = _project_contractor(project.id, request.args.get("contractor_id", type=int))
    tasks = _filter_tasks_for_contractor(_export_tasks_from_request(query_args, project.id), selected_contractor).all()
    contractor_label = selected_contractor.name if selected_contractor else "Все подрядчики"
    filename_prefix = f"{project.name}_Подрядчики"
    if selected_contractor:
        filename_prefix = f"{filename_prefix}_{selected_contractor.name}"
    title = f"Подрядчик: {contractor_label}" if selected_contractor else contractor_label
    path = export_remark_tasks_excel(tasks, filename_prefix, title=title)
    return send_file(path, as_attachment=True, download_name=Path(path).name)


def _unique_int_list(values) -> list[int]:
    result: list[int] = []
    seen: set[int] = set()
    for raw_value in values or []:
        try:
            item_id = int(raw_value)
        except (TypeError, ValueError):
            continue
        if item_id in seen:
            continue
        seen.add(item_id)
        result.append(item_id)
    return result


def _selected_task_ids_from_request() -> list[int]:
    return _unique_int_list(request.args.getlist("task_ids"))


def _selected_task_ids_from_form() -> list[int]:
    return _unique_int_list(request.form.getlist("task_ids"))


def _selected_premise_ids_from_request() -> list[int]:
    return _unique_int_list(request.args.getlist("premise_ids"))


def _export_tasks_from_request(query_args: dict, project_id: int, category_id: int | None = None):
    export_options = (
        selectinload(Task.apartment),
        selectinload(Task.work_point),
        selectinload(Task.responsible),
    )
    task_ids = _selected_task_ids_from_request()
    if task_ids:
        return (
            Task.query.options(*export_options)
            .join(Apartment)
            .join(WorkPoint)
            .filter(Task.project_id == project_id, Task.id.in_(task_ids))
            .order_by(Task.is_done.asc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc(), WorkPoint.point_number.asc(), Task.id.asc())
        )
    premise_ids = _selected_premise_ids_from_request()
    if premise_ids:
        return build_task_query(query_args, category_id=category_id, project_id=project_id).options(*export_options).filter(Apartment.id.in_(premise_ids))
    return build_task_query(query_args, category_id=category_id, project_id=project_id).options(*export_options)


def _active_users_for_project(project_id: int | None = None, roles=None) -> list[User]:
    query = User.query.filter(User.is_active.is_(True))
    if roles:
        query = query.filter(User.role.in_(roles))
    users = query.order_by(User.full_name.asc().nullslast(), User.username.asc()).all()
    if project_id is None:
        return users
    return [user for user in users if user.can_access_project(project_id)]


def _executor_users(project_id: int | None = None) -> list[User]:
    return _active_users_for_project(project_id, WORKER_ROLES)


ASSIGNMENT_ROLE_CATEGORY_NAMES = {
    ROLE_EXECUTOR: ("Маляры",),
    ROLE_PAINTER: ("Маляры",),
    ROLE_HANDYMAN: ("Разнорабочие",),
    ROLE_GLAZIER: ("Витражники",),
}


def _assignment_allowed_point_ids_for_user(user: User | None) -> set[int]:
    if not user or user.role not in WORKER_ROLES:
        return set()
    category_names = ASSIGNMENT_ROLE_CATEGORY_NAMES.get(user.role, ())
    if not category_names:
        return set()
    categories = (
        WorkCategory.query.options(selectinload(WorkCategory.work_points))
        .filter(WorkCategory.is_active.is_(True), WorkCategory.name.in_(category_names))
        .all()
    )
    return {
        point.id
        for category in categories
        for point in category.work_points
        if point.is_active
    }


def _assignment_filter_by_user_mapping(query, user: User | None):
    point_ids = _assignment_allowed_point_ids_for_user(user)
    if not point_ids:
        return query.filter(False)
    return query.filter(Task.work_point_id.in_(point_ids))


def _assignment_task_allowed_for_user(task: Task, user: User | None, allowed_point_ids: set[int] | None = None) -> bool:
    if not user or user.role not in WORKER_ROLES:
        return False
    point_ids = allowed_point_ids if allowed_point_ids is not None else _assignment_allowed_point_ids_for_user(user)
    return bool(task.work_point_id and task.work_point_id in point_ids)


def _parse_id_list(value: str | None) -> set[int]:
    result: set[int] = set()
    for raw_item in str(value or "").replace(",", " ").split():
        try:
            result.add(int(raw_item))
        except (TypeError, ValueError):
            continue
    return result


def _assignment_base_query(project_id: int, params=None):
    params = dict(params or {})
    params.setdefault("sort", "apartment")
    query = build_task_query(params, project_id=project_id).options(
        selectinload(Task.apartment),
        selectinload(Task.work_point),
        selectinload(Task.responsible),
    )
    query = query.filter(
        Task.status.notin_([STATUS_DONE, STATUS_FINISHERS, STATUS_CONTRACTOR]),
        Task.is_done.is_(False),
        Task.is_archived.is_(False),
        Task.is_missing_in_latest_sync.is_(False),
    )
    return query


ASSIGNMENT_SMART_BATCH_LIMIT = 10
ASSIGNMENT_DAILY_MAX = None  # дневной лимит выдачи задач отключён


def _display_user_name(user: User | None) -> str:
    if user is None:
        return "Сотрудник"
    return user.full_name or user.username or f"ID {user.id}"


def _assignment_issue_date_map(tasks: list[Task]) -> dict[int, datetime]:
    task_ids = [task.id for task in tasks if task.id]
    if not task_ids:
        return {}
    rows = (
        ChangeLog.query.filter(
            ChangeLog.task_id.in_(task_ids),
            ChangeLog.field_name == "responsible_id",
            ChangeLog.new_value.isnot(None),
            ChangeLog.new_value != "",
        )
        .order_by(ChangeLog.created_at.asc())
        .all()
    )
    result: dict[int, datetime] = {}
    current_responsible = {task.id: str(task.responsible_id or "") for task in tasks}
    for row in rows:
        if current_responsible.get(row.task_id) and str(row.new_value or "") == current_responsible[row.task_id]:
            result[row.task_id] = row.created_at
    for task in tasks:
        result.setdefault(task.id, task.updated_at or task.created_at or datetime.utcnow())
    return result


def _issued_status_payload(task: Task) -> dict[str, str]:
    if task.status == STATUS_DONE or task.is_done:
        return {"label": "Выполнено", "class": "success"}
    return {"label": "Не выполнено", "class": "secondary"}


def _assignment_issued_day_filter(value: str | None) -> tuple[str, date | None]:
    normalized = (value or "today").strip().lower()
    today = date.today()
    if normalized == "tomorrow":
        return "tomorrow", today + timedelta(days=1)
    if normalized == "overdue":
        return "overdue", None
    if normalized == "all":
        return "all", None
    return "today", today


def _assigned_tasks_count_for_day(project_id: int, user_id: int, planned_day: date, exclude_task_ids: list[int] | None = None) -> int:
    query = Task.query.filter(
        Task.project_id == project_id,
        Task.responsible_id == user_id,
        Task.planned_date == planned_day,
        Task.status.notin_([STATUS_DONE, STATUS_FINISHERS, STATUS_CONTRACTOR]),
        Task.is_done.is_(False),
        Task.is_archived.is_(False),
        Task.is_missing_in_latest_sync.is_(False),
    )
    if exclude_task_ids:
        query = query.filter(~Task.id.in_(exclude_task_ids))
    return query.count()


def _assignment_export_filename_part(value: str | None) -> str:
    bad_chars = '\\/:*?"<>|'
    cleaned = ''.join(' ' if char in bad_chars else char for char in str(value or '').strip())
    cleaned = ' '.join(cleaned.split())
    return cleaned or 'tasks'


def _assignment_issued_filter_label(issued_filter: str, issued_filter_date: date | None) -> str:
    if issued_filter == 'tomorrow':
        return f"Завтра, {issued_filter_date.strftime('%d.%m.%Y')}" if issued_filter_date else 'Завтра'
    if issued_filter == 'overdue':
        return 'Просроченные задачи'
    if issued_filter == 'all':
        return 'Все выданные задачи'
    return f"Сегодня, {issued_filter_date.strftime('%d.%m.%Y')}" if issued_filter_date else 'Сегодня'


@bp.route("/assignments", methods=["GET", "POST"])
@login_required
def assignments():
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    users = _executor_users(project.id)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    view_mode = (request.args.get("view") or "issue").strip()
    if view_mode not in {"issue", "issued"}:
        view_mode = "issue"

    def assignment_task_payload(task: Task):
        responsible = task.responsible
        is_overdue = bool(
            task.responsible_id
            and task.planned_date
            and task.planned_date < date.today()
            and task.status != STATUS_DONE
            and not task.is_done
        )
        return {
            "ok": True,
            "task_id": task.id,
            "responsible_id": task.responsible_id,
            "responsible_name": (responsible.full_name or responsible.username) if responsible else None,
            "planned_date": task.planned_date.strftime("%d.%m.%Y") if task.planned_date else "—",
            "planned_date_iso": task.planned_date.isoformat() if task.planned_date else "",
            "is_overdue": is_overdue,
        }

    def assignment_error(message: str, status_code: int = 400):
        if wants_json:
            return jsonify({"ok": False, "message": message}), status_code
        flash(message, "warning")
        return redirect(request.referrer or url_for("main.assignments"))

    if request.method == "POST":
        update_date_task_id = request.form.get("update_date_task_id", type=int)
        toggle_employee_status_task_id = request.form.get("toggle_employee_status_task_id", type=int)
        postpone_task_id = request.form.get("postpone_task_id", type=int)
        remove_assignee_task_id = request.form.get("remove_assignee_task_id", type=int)
        change_assignee_task_id = request.form.get("change_assignee_task_id", type=int)

        if change_assignee_task_id:
            task = db.session.get(Task, change_assignee_task_id) or abort(404)
            if task.project_id != project.id:
                abort(404)
            new_responsible_id = request.form.get("new_responsible_id", type=int)
            new_responsible = db.session.get(User, new_responsible_id) if new_responsible_id else None
            if not _user_can_work_in_project(new_responsible, project):
                return assignment_error("Выберите нового исполнителя")
            if not _assignment_task_allowed_for_user(task, new_responsible):
                return assignment_error("Эта задача не относится к пунктам, отмеченным для выбранного исполнителя в распределении")
            if task.responsible_id == new_responsible.id:
                return assignment_error("Этот исполнитель уже назначен на задачу")
            old_responsible = task.responsible_id
            task.responsible_id = new_responsible.id
            task.manually_edited = True
            log_change(task, "field_update", "responsible_id", old_responsible, new_responsible.id, user_id=current_user.id)
            db.session.commit()
            responsible_name = new_responsible.full_name or new_responsible.username
            if wants_json:
                return jsonify(assignment_task_payload(task) | {"changed_assignee": True, "message": f"Исполнитель изменён: {responsible_name}"})
            flash(f"Исполнитель изменён: {responsible_name}", "success")
            return redirect(request.referrer or url_for("main.assignments", view="issued"))

        if remove_assignee_task_id:
            task = db.session.get(Task, remove_assignee_task_id) or abort(404)
            if task.project_id != project.id:
                abort(404)
            if not task.responsible_id:
                return assignment_error("У задачи уже нет исполнителя")
            old_responsible = task.responsible_id
            task.responsible_id = None
            task.manually_edited = True
            log_change(task, "field_update", "responsible_id", old_responsible, None, user_id=current_user.id)
            db.session.commit()
            if wants_json:
                return jsonify(assignment_task_payload(task) | {"removed": True, "message": "Исполнитель снят. Задача снова без исполнителя."})
            flash("Исполнитель снят. Задача снова без исполнителя.", "success")
            return redirect(request.referrer or url_for("main.assignments", view="issued"))

        if toggle_employee_status_task_id:
            task = db.session.get(Task, toggle_employee_status_task_id) or abort(404)
            if task.project_id != project.id:
                abort(404)
            if not task.responsible_id:
                return assignment_error("Сначала назначьте исполнителя")
            new_status = STATUS_NOT_STARTED if (task.is_done or task.status == STATUS_DONE) else STATUS_DONE
            change_task_status(task, new_status, user_id=current_user.id)
            if wants_json:
                return jsonify(assignment_task_payload(task) | {"status_label": task.status_label(), "status_class": task.status_class(), "is_done": task.is_done, "message": f"Статус сотрудника изменён: {task.status_label()}"})
            flash(f"Статус сотрудника изменён: {task.status_label()}", "success")
            return redirect(request.referrer or url_for("main.assignments", view="issued"))

        if update_date_task_id:
            task = db.session.get(Task, update_date_task_id) or abort(404)
            if task.project_id != project.id:
                abort(404)
            if task.is_done or task.status == STATUS_DONE:
                return assignment_error("Выполненной задаче нельзя изменить дату")
            if not task.responsible_id:
                return assignment_error("Сначала назначьте исполнителя")
            planned_date = parse_date(request.form.get(f"planned_date_{update_date_task_id}") or request.form.get("planned_date"))
            if not planned_date:
                return assignment_error("Выберите дату выполнения задачи")
            old_date = task.planned_date
            task.planned_date = planned_date
            task.manually_edited = True
            if old_date != planned_date:
                log_change(task, "field_update", "planned_date", old_date, planned_date, user_id=current_user.id)
            db.session.commit()
            if wants_json:
                return jsonify(assignment_task_payload(task) | {"message": "Дата выполнения задачи изменена"})
            flash("Дата выполнения задачи изменена", "success")
            return redirect(request.referrer or url_for("main.assignments", view="issued"))

        if postpone_task_id:
            task = db.session.get(Task, postpone_task_id) or abort(404)
            if task.project_id != project.id:
                abort(404)
            if task.is_done or task.status == STATUS_DONE:
                return assignment_error("Выполненную задачу нельзя перенести")
            if not task.responsible_id:
                return assignment_error("Сначала назначьте исполнителя")
            old_date = task.planned_date
            base_date = task.planned_date or date.today()
            task.planned_date = base_date + timedelta(days=1)
            task.manually_edited = True
            log_change(task, "field_update", "planned_date", old_date, task.planned_date, user_id=current_user.id)
            db.session.commit()
            if wants_json:
                return jsonify(assignment_task_payload(task) | {"message": "Задача перенесена на следующий день"})
            flash("Задача перенесена на следующий день", "success")
            return redirect(request.referrer or url_for("main.assignments", view="issued"))

        task_ids = []
        seen_task_ids = set()
        for raw_id in request.form.getlist("task_ids"):
            try:
                task_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if task_id in seen_task_ids:
                continue
            seen_task_ids.add(task_id)
            task_ids.append(task_id)
        responsible_id = request.form.get("responsible_id", type=int)
        planned_date = parse_date(request.form.get("planned_date")) or date.today()
        responsible = db.session.get(User, responsible_id) if responsible_id else None
        if not task_ids:
            flash("Выберите хотя бы одну задачу", "warning")
        elif not _user_can_work_in_project(responsible, project):
            flash("Выберите исполнителя", "warning")
        else:
            existing_tasks = (
                Task.query.filter(Task.project_id == project.id, Task.id.in_(task_ids))
                .filter(Task.status.notin_([STATUS_DONE, STATUS_FINISHERS, STATUS_CONTRACTOR]), Task.is_done.is_(False))
                .all()
            )
            already_assigned = [task for task in existing_tasks if task.responsible_id]
            allowed_point_ids = _assignment_allowed_point_ids_for_user(responsible)
            candidate_tasks = [task for task in existing_tasks if not task.responsible_id]
            incompatible_tasks = [
                task for task in candidate_tasks
                if not _assignment_task_allowed_for_user(task, responsible, allowed_point_ids)
            ]
            tasks = [
                task for task in candidate_tasks
                if _assignment_task_allowed_for_user(task, responsible, allowed_point_ids)
            ]
            if incompatible_tasks:
                redirect_args = {
                    key: value
                    for key, value in request.args.to_dict(flat=True).items()
                    if key not in {"selected_task_ids", "invalid_task_ids", "responsible_id", "planned_date"}
                }
                redirect_args["selected_task_ids"] = ",".join(str(task_id) for task_id in task_ids)
                redirect_args["invalid_task_ids"] = ",".join(str(task.id) for task in incompatible_tasks)
                if responsible:
                    redirect_args["responsible_id"] = str(responsible.id)
                if planned_date:
                    redirect_args["planned_date"] = planned_date.isoformat()
                flash("Выбранные задачи не относятся к пунктам, отмеченным для этого исполнителя в распределении. Проблемные задачи подсвечены красным, снимите с них отметку и повторите выдачу.", "danger")
                return redirect(url_for("main.assignments", **redirect_args))
            if not tasks:
                skipped = len(already_assigned)
                if skipped:
                    flash(f"Задача уже выдана сотруднику. Повторная выдача не выполнена. Пропущено: {task_count_label(skipped)}", "warning")
                elif incompatible_tasks:
                    flash("Выбранные задачи не относятся к пунктам, отмеченным для этого исполнителя в распределении", "warning")
                else:
                    flash("Нет доступных задач для выдачи", "warning")
                return redirect(url_for("main.assignments"))
            for task in tasks:
                old_responsible = task.responsible_id
                old_date = task.planned_date
                task.responsible_id = responsible.id
                task.planned_date = planned_date
                task.manually_edited = True
                if old_responsible != responsible.id:
                    log_change(task, "field_update", "responsible_id", old_responsible, responsible.id, user_id=current_user.id)
                if old_date != planned_date:
                    log_change(task, "field_update", "planned_date", old_date, planned_date, user_id=current_user.id)
            db.session.commit()
            message = f"Выдано: {task_count_label(len(tasks))}"
            if already_assigned:
                message += f". Уже были выданы и пропущены: {task_count_label(len(already_assigned))}"
            if incompatible_tasks:
                message += f". Не подходят по распределению и пропущены: {task_count_label(len(incompatible_tasks))}"
            flash(message, "success" if not already_assigned and not incompatible_tasks else "warning")
            return redirect(url_for("main.assignments"))

    # В выдаче задач оставляем один умный поиск. Старые параметры сортировки/узких фильтров
    # намеренно не тащим дальше, чтобы экран не путал пользователя.
    query_args = {"q": (request.args.get("q") or "").strip()}
    smart_mode = request.args.get("smart") == "1"
    smart_user_id = request.args.get("smart_user_id", type=int)
    smart_date = parse_date(request.args.get("smart_date")) or date.today()
    smart_user = db.session.get(User, smart_user_id) if smart_user_id else None
    selected_task_ids = _parse_id_list(request.args.get("selected_task_ids"))
    invalid_task_ids = _parse_id_list(request.args.get("invalid_task_ids"))
    issue_responsible_id = request.args.get("responsible_id", type=int)
    issue_planned_date = parse_date(request.args.get("planned_date"))
    if smart_user and not _user_can_work_in_project(smart_user, project):
        smart_user = None
    smart_limit = ASSIGNMENT_SMART_BATCH_LIMIT
    assigned_today_count = 0

    if smart_mode and smart_user:
        assigned_today_count = (
            Task.query.filter(
                Task.project_id == project.id,
                Task.responsible_id == smart_user.id,
                Task.planned_date == smart_date,
                Task.status.notin_([STATUS_DONE, STATUS_FINISHERS, STATUS_CONTRACTOR]),
                Task.is_done.is_(False),
                Task.is_archived.is_(False),
                Task.is_missing_in_latest_sync.is_(False),
            ).count()
        )
        smart_limit = ASSIGNMENT_SMART_BATCH_LIMIT
        smart_args = {k: v for k, v in query_args.items() if k not in {"smart", "smart_user_id", "smart_date", "page"}}
        smart_args["status"] = "not_done"
        query = _assignment_base_query(project.id, smart_args).filter(Task.responsible_id.is_(None))
        query = _assignment_filter_by_user_mapping(query, smart_user)
        tasks = query.limit(smart_limit).all() if smart_limit > 0 else []
        pagination = None
        prev_args = {}
        next_args = {}
        assignment_total = len(tasks)
    else:
        query = _assignment_base_query(project.id, query_args).filter(Task.responsible_id.is_(None))
        page = request.args.get("page", 1, type=int)
        per_page = 10 if _is_mobile_phone_request() else 20
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        tasks = pagination.items
        prev_args = {k: v for k, v in query_args.items() if v}
        next_args = {k: v for k, v in query_args.items() if v}
        if pagination.has_prev:
            prev_args["page"] = pagination.prev_num
        if pagination.has_next:
            next_args["page"] = pagination.next_num
        assignment_total = pagination.total

    if selected_task_ids and view_mode != "issued":
        visible_task_ids = {task.id for task in tasks}
        missing_selected_ids = [task_id for task_id in selected_task_ids if task_id not in visible_task_ids]
        if missing_selected_ids:
            selected_order = {task_id: index for index, task_id in enumerate(missing_selected_ids)}
            missing_selected_tasks = (
                Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.responsible))
                .filter(
                    Task.project_id == project.id,
                    Task.id.in_(missing_selected_ids),
                    Task.status.notin_([STATUS_DONE, STATUS_FINISHERS, STATUS_CONTRACTOR]),
                    Task.is_done.is_(False),
                    Task.is_archived.is_(False),
                    Task.is_missing_in_latest_sync.is_(False),
                )
                .all()
            )
            missing_selected_tasks.sort(key=lambda task: selected_order.get(task.id, len(selected_order)))
            tasks = missing_selected_tasks + tasks

    today_value = date.today()
    issued_filter, issued_filter_date = _assignment_issued_day_filter(request.args.get("issued_day"))
    issued_query_builder = (
        Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.responsible))
        .filter(
            Task.project_id == project.id,
            Task.responsible_id.isnot(None),
            Task.status.notin_([STATUS_FINISHERS, STATUS_CONTRACTOR]),
            Task.is_archived.is_(False),
            Task.is_missing_in_latest_sync.is_(False),
        )
    )
    overdue_filter = (
        Task.planned_date.isnot(None),
        Task.planned_date < today_value,
        Task.status != STATUS_DONE,
        Task.is_done.is_(False),
    )
    overdue_total = issued_query_builder.filter(*overdue_filter).count()
    if issued_filter == "overdue":
        issued_query_builder = issued_query_builder.filter(*overdue_filter)
        issued_order = (Task.planned_date.asc().nullslast(), Task.responsible_id.asc(), Task.id.asc())
    else:
        if issued_filter_date is not None:
            issued_query_builder = issued_query_builder.filter(Task.planned_date == issued_filter_date)
        issued_order = (Task.responsible_id.asc(), Task.is_done.asc(), Task.planned_date.asc().nullslast(), Task.id.asc())
    issued_query = (
        issued_query_builder
        .order_by(*issued_order)
        .limit(1000 if issued_filter == "overdue" else 800)
        .all()
    )
    issue_dates = _assignment_issue_date_map(issued_query)
    issued_groups = []
    issued_overdue_day_groups = []
    issued_total = len(issued_query)

    def make_issued_row(task: Task) -> dict:
        overdue_days = (today_value - task.planned_date).days if task.planned_date else 0
        return {
            "task": task,
            "issued_at": issue_dates.get(task.id),
            "employee_status": _issued_status_payload(task),
            "overdue_days": max(overdue_days, 0),
        }

    if issued_filter == "overdue":
        rows_by_day: dict[date, list[dict]] = {}
        for task in issued_query:
            if task.planned_date is None:
                continue
            rows_by_day.setdefault(task.planned_date, []).append(make_issued_row(task))
        for planned_day in sorted(rows_by_day):
            day_rows = rows_by_day[planned_day]
            day_groups = []
            for user in users:
                user_rows = [row for row in day_rows if row["task"].responsible_id == user.id]
                if user_rows:
                    day_groups.append({"user": user, "rows": user_rows, "tasks": [row["task"] for row in user_rows]})
            if day_groups:
                issued_overdue_day_groups.append({
                    "date": planned_day,
                    "label": planned_day.strftime("%d.%m.%Y"),
                    "total": len(day_rows),
                    "groups": day_groups,
                })
    else:
        for user in users:
            user_rows = [
                make_issued_row(task)
                for task in issued_query
                if task.responsible_id == user.id
            ]
            if user_rows:
                issued_groups.append({"user": user, "rows": user_rows, "tasks": [row["task"] for row in user_rows]})

    finishing_types = [
        x[0]
        for x in db.session.query(distinct(Apartment.finishing_type))
        .filter(Apartment.project_id == project.id, Apartment.finishing_type.isnot(None))
        .all()
    ]
    return render_template(
        "assignments.html",
        project=project,
        tasks=tasks,
        pagination=pagination,
        users=users,
        points=_contractor_point_options(),
        finishing_types=finishing_types,
        args=request.args,
        prev_args=prev_args,
        next_args=next_args,
        smart_mode=smart_mode,
        smart_user=smart_user,
        smart_date=smart_date,
        issue_responsible_id=issue_responsible_id,
        issue_planned_date=issue_planned_date,
        selected_task_ids=selected_task_ids,
        invalid_task_ids=invalid_task_ids,
        smart_limit=smart_limit,
        assigned_today_count=assigned_today_count,
        assignment_total=assignment_total,
        view_mode=view_mode,
        issued_groups=issued_groups,
        issued_overdue_day_groups=issued_overdue_day_groups,
        issued_total=issued_total,
        overdue_total=overdue_total,
        issued_filter=issued_filter,
        issued_filter_date=issued_filter_date,
        today=today_value,
    )


@bp.route("/assignments/<int:task_id>/unassign", methods=["POST"])
@login_required
def assignment_unassign(task_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "message": "Выберите объект"}), 400
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    if not task.responsible_id:
        if wants_json:
            return jsonify({"ok": False, "message": "У задачи уже нет исполнителя"}), 400
        flash("У задачи уже нет исполнителя", "warning")
        return redirect(request.referrer or url_for("main.assignments", view="issued"))

    old_responsible = task.responsible_id
    task.responsible_id = None
    task.manually_edited = True
    log_change(task, "field_update", "responsible_id", old_responsible, None, user_id=current_user.id)
    db.session.commit()

    payload = {
        "ok": True,
        "task_id": task.id,
        "responsible_id": None,
        "responsible_name": None,
        "planned_date": task.planned_date.strftime("%d.%m.%Y") if task.planned_date else "—",
        "planned_date_iso": task.planned_date.isoformat() if task.planned_date else "",
        "removed": True,
        "message": "Исполнитель снят. Задача снова без исполнителя.",
    }
    if wants_json:
        return jsonify(payload)
    flash(payload["message"], "success")
    return redirect(request.referrer or url_for("main.assignments", view="issued"))


@bp.route("/assignments/<int:task_id>/delete-from-employee", methods=["POST"])
@login_required
def assignment_delete_from_employee(task_id: int):
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        if wants_json:
            return jsonify({"ok": False, "message": "Выберите объект"}), 400
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if not task.responsible_id:
        if wants_json:
            return jsonify({"ok": False, "message": "У задачи уже нет исполнителя"}), 400
        flash("У задачи уже нет исполнителя", "warning")
        return redirect(request.referrer or url_for("main.assignments", view="issued"))

    old_responsible = task.responsible_id
    old_responsible_user = task.responsible
    old_date = task.planned_date
    _record_deletion_action(
        "assignment_delete_from_employee",
        "task_assignment",
        task.id,
        f"Задача #{task.id}",
        f"Задача удалена у исполнителя: {((old_responsible_user.full_name or old_responsible_user.username) if old_responsible_user else '—')}.",
        {
            "task_id": task.id,
            "responsible_id": old_responsible,
            "planned_date": old_date.isoformat() if old_date else None,
            "status": task.status,
            "is_done": bool(task.is_done),
        },
        project_id=task.project_id,
    )
    task.responsible_id = None
    task.planned_date = None
    task.manually_edited = True
    log_change(task, "field_update", "responsible_id", old_responsible, None, user_id=current_user.id)
    if old_date is not None:
        log_change(task, "field_update", "planned_date", old_date, None, user_id=current_user.id)
    db.session.commit()

    message = "Задача удалена у сотрудника и снова доступна без исполнителя. Действие можно отменить в логах удалений."
    if wants_json:
        return jsonify({
            "ok": True,
            "task_id": task.id,
            "responsible_id": None,
            "responsible_name": None,
            "planned_date": "—",
            "planned_date_iso": "",
            "removed": True,
            "message": message,
        })
    flash(message, "success")
    return redirect(request.referrer or url_for("main.assignments", view="issued"))


@bp.route("/assignments/issued/<int:user_id>/export")
@login_required
def assignment_issued_employee_export(user_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))

    employee = db.session.get(User, user_id) or abort(404)
    if employee.role not in WORKER_ROLES or (employee.project_id is not None and employee.project_id != project.id):
        abort(404)

    today_value = date.today()
    issued_filter, issued_filter_date = _assignment_issued_day_filter(request.args.get("issued_day"))
    query = (
        Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.responsible))
        .filter(
            Task.project_id == project.id,
            Task.responsible_id == employee.id,
            Task.status.notin_([STATUS_FINISHERS, STATUS_CONTRACTOR]),
            Task.is_archived.is_(False),
            Task.is_missing_in_latest_sync.is_(False),
        )
    )
    if issued_filter == "overdue":
        query = query.filter(
            Task.planned_date.isnot(None),
            Task.planned_date < today_value,
            Task.status != STATUS_DONE,
            Task.is_done.is_(False),
        )
        order_by = (Task.planned_date.asc().nullslast(), Task.id.asc())
    else:
        if issued_filter_date is not None:
            query = query.filter(Task.planned_date == issued_filter_date)
        order_by = (Task.is_done.asc(), Task.planned_date.asc().nullslast(), Task.id.asc())

    tasks = (
        query.order_by(*order_by)
        .limit(1000)
        .all()
    )
    issue_dates = _assignment_issue_date_map(tasks)

    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    employee_name = _display_user_name(employee)
    role_label = ROLE_LABELS.get(employee.role, employee.role)
    filter_label = _assignment_issued_filter_label(issued_filter, issued_filter_date)

    ws.append(["Сотрудник", employee_name, "Роль", role_label])
    ws.append(["Объект", project.name, "Период", filter_label])
    ws.append(["Всего задач", len(tasks), "Дата выгрузки", datetime.now().strftime("%d.%m.%Y %H:%M")])
    ws.append([])
    ws.append(["Помещение", "Задача", "Дата выполнения", "Статус", "Отделка", "Дата выдачи"])

    for task in tasks:
        apartment = task.apartment
        issued_at = issue_dates.get(task.id)
        employee_status = _issued_status_payload(task)["label"]
        ws.append([
            _excel_premise_label(apartment) if apartment else "—",
            task.description or task.source_cell_value or "",
            task.planned_date.strftime("%d.%m.%Y") if task.planned_date else "—",
            employee_status,
            apartment.finishing_type if apartment else "",
            issued_at.strftime("%d.%m.%Y %H:%M") if issued_at else "",
        ])
        if task.is_done:
            for cell in ws[ws.max_row]:
                cell.font = Font(strike=True, color="667085")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
    for cell in ws[2]:
        cell.font = Font(bold=True, color="111827")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="111827")

    header_fill = PatternFill(fill_type="solid", start_color=EXCEL_HEADER_FILL_COLOR, end_color=EXCEL_HEADER_FILL_COLOR)
    for cell in ws[5]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [16, 90, 18, 18, 22, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(6, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42

    ws.freeze_panes = "A6"
    if ws.max_row >= 5:
        ws.auto_filter.ref = f"A5:F{ws.max_row}"

    safe_project = _assignment_export_filename_part(project.name)
    safe_employee = _assignment_export_filename_part(employee_name)
    safe_period = _assignment_export_filename_part(filter_label.replace(',', ''))
    filename = f"{safe_project}_{safe_employee}_задачи_{safe_period}_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return _make_excel_response(wb, filename)


def _manual_assignment_work_point() -> WorkPoint:
    point = WorkPoint.query.filter_by(point_number="ручная", source_sheet_name="assignment_manual").first()
    if point is None:
        point = WorkPoint(
            point_number="ручная",
            source_sheet_name="assignment_manual",
            original_column_name="Ручная задача",
            short_name="Ручная задача",
            is_active=True,
        )
        db.session.add(point)
        db.session.flush()
    return point


@bp.route("/assignments/manual/new", methods=["GET", "POST"])
@login_required
def assignment_manual_task_new():
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    users = _executor_users(project.id)
    apartments = _project_apartment_options(project.id)

    if request.method == "POST":
        apartment_id = request.form.get("apartment_id", type=int)
        responsible_id = request.form.get("responsible_id", type=int)
        planned_date = parse_date(request.form.get("planned_date"))
        text = (request.form.get("description") or "").strip()
        apartment = db.session.get(Apartment, apartment_id) if apartment_id else None
        work_point = _manual_assignment_work_point()
        responsible = db.session.get(User, responsible_id) if responsible_id else None

        if not apartment or apartment.project_id != project.id:
            flash("Выберите квартиру", "warning")
        elif not text:
            flash("Введите ручную задачу", "warning")
        elif responsible_id and not _user_can_work_in_project(responsible, project):
            flash("Выберите корректного исполнителя", "warning")
        else:
            source_uid = build_task_uid(
                project.name,
                apartment.construction_number or "",
                apartment.apartment_number or "",
                work_point.point_number,
                work_point.display_name,
                text,
            )
            if Task.query.filter_by(source_uid=source_uid).first():
                source_uid = stable_hash([source_uid, "assignment-manual", datetime.utcnow().isoformat()])
            task = Task(
                source_uid=source_uid,
                project_id=project.id,
                apartment_id=apartment.id,
                work_point_id=work_point.id,
                title=work_point.display_name,
                description=text,
                source_cell_value=text,
                source_sheet_name="assignment_manual",
                status=STATUS_NOT_STARTED,
                is_done=False,
                responsible_id=responsible.id if responsible else None,
                planned_date=planned_date,
                manually_edited=True,
                last_seen_at=datetime.utcnow(),
                source_hash=stable_hash([text]),
            )
            db.session.add(task)
            db.session.flush()
            log_change(task, "manual_assignment_created", None, None, text, user_id=current_user.id)
            db.session.commit()
            flash("Ручная задача добавлена", "success")
            return redirect(url_for("main.assignments"))

    return render_template(
        "assignment_task_form.html",
        project=project,
        apartments=apartments,
        users=users,
    )


@bp.route("/assignments/report")
@login_required
def assignments_report():
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    report_date = parse_date(request.args.get("date")) or date.today()
    users = _executor_users(project.id)
    user_id = request.args.get("user_id", type=int)
    if user_id and not any(user.id == user_id for user in users):
        abort(404)
    query = (
        Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.responsible))
        .join(User, Task.responsible_id == User.id)
        .join(Apartment)
        .filter(
            Task.project_id == project.id,
            Task.responsible_id.isnot(None),
            Task.status.notin_([STATUS_FINISHERS, STATUS_CONTRACTOR]),
        )
        .filter(
            or_(
                Task.planned_date == report_date,
                func.date(Task.completed_date) == report_date.isoformat(),
            )
        )
    )
    if user_id:
        query = query.filter(Task.responsible_id == user_id)
    tasks = query.order_by(User.full_name.asc(), Task.is_done.asc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc()).all()
    grouped = []
    for user in users:
        user_tasks = [task for task in tasks if task.responsible_id == user.id]
        if user_tasks or not user_id or user.id == user_id:
            grouped.append({
                "user": user,
                "tasks": user_tasks,
                "done": sum(1 for task in user_tasks if task.status == STATUS_DONE),
                "left": sum(1 for task in user_tasks if task.status != STATUS_DONE),
            })
    report_totals = {
        "workers": len([group for group in grouped if group["tasks"]]),
        "tasks": sum(len(group["tasks"]) for group in grouped),
        "done": sum(group["done"] for group in grouped),
        "left": sum(group["left"] for group in grouped),
    }
    return render_template(
        "assignment_report.html",
        project=project,
        grouped=grouped,
        users=users,
        report_date=report_date,
        selected_user_id=user_id,
        report_totals=report_totals,
    )


@bp.route("/assignments/report/<int:user_id>/pdf")
@bp.route("/assignments/report/<int:user_id>/excel")
@login_required
def assignments_report_worker_pdf(user_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    report_date = parse_date(request.args.get("date")) or date.today()
    user = db.session.get(User, user_id) or abort(404)
    if not _user_can_work_in_project(user, project):
        abort(404)
    tasks = (
        Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.responsible))
        .filter(
            Task.project_id == project.id,
            Task.responsible_id == user.id,
            Task.status.notin_([STATUS_FINISHERS, STATUS_CONTRACTOR]),
        )
        .filter(
            or_(
                Task.planned_date == report_date,
                func.date(Task.completed_date) == report_date.isoformat(),
            )
        )
        .join(Apartment)
        .order_by(Task.is_done.asc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc(), Task.id.asc())
        .all()
    )
    worker_name = user.full_name or user.username or f"worker_{user.id}"
    title = f"Ежедневный отчет - {worker_name} - {report_date.strftime('%d.%m.%Y')}"
    filename_prefix = f"{project.name}_{worker_name}_{report_date.strftime('%d.%m.%Y')}"
    path = export_simple_tasks_excel(tasks, filename_prefix=filename_prefix, title=title, report_header=True, include_point_in_remarks=False)
    return send_file(path, as_attachment=True, download_name=Path(path).name)


@bp.route("/my-tasks")
@login_required
def my_tasks():
    if current_user.role in {ROLE_ADMIN, ROLE_MANAGER}:
        return redirect(url_for("main.dashboard"))
    project = selected_project()
    today_value = date.today()
    query = Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point)).filter(Task.responsible_id == current_user.id)
    if project:
        query = query.filter(Task.project_id == project.id)
    query = query.filter(Task.is_archived.is_(False), Task.is_missing_in_latest_sync.is_(False))
    query = query.filter(
        or_(
            Task.planned_date == today_value,
            and_(Task.planned_date.is_(None), Task.status != STATUS_DONE),
            and_(Task.status == STATUS_DONE, func.date(Task.completed_date) == today_value.isoformat()),
        )
    )
    tasks = query.outerjoin(Apartment).order_by(Task.is_done.asc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc(), Task.updated_at.desc()).all()
    counters = {
        "done": sum(1 for task in tasks if task.status == STATUS_DONE),
        "left": sum(1 for task in tasks if task.status != STATUS_DONE),
        "all": len(tasks),
    }
    role_label = ROLE_LABELS.get(current_user.role, "Исполнитель")
    return render_template("my_tasks.html", project=project, tasks=tasks, counters=counters, today=today_value, role_label=role_label)


def _worker_status_payload(task: Task):
    return {
        "ok": True,
        "task_id": task.id,
        "status": task.status,
        "status_label": task.status_label(),
        "status_class": task.status_class(),
        "is_done": task.status == STATUS_DONE,
    }


@bp.route("/my-tasks/<int:task_id>/done", methods=["POST"])
@login_required
def my_task_done(task_id: int):
    task = db.session.get(Task, task_id) or abort(404)
    project = selected_project()
    if current_user.role in {ROLE_ADMIN, ROLE_MANAGER}:
        if project is None or task.project_id != project.id:
            abort(404)
    elif task.responsible_id != current_user.id:
        abort(403)
    change_task_status(task, STATUS_DONE, user_id=current_user.id)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify(_worker_status_payload(task))
    flash("Задача отмечена выполненной", "success")
    return redirect(url_for("main.my_tasks"))


@bp.route("/my-tasks/<int:task_id>/return", methods=["POST"])
@login_required
def my_task_return(task_id: int):
    task = db.session.get(Task, task_id) or abort(404)
    project = selected_project()
    if current_user.role in {ROLE_ADMIN, ROLE_MANAGER}:
        if project is None or task.project_id != project.id:
            abort(404)
    elif task.responsible_id != current_user.id:
        abort(403)
    change_task_status(task, STATUS_NOT_STARTED, user_id=current_user.id)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify(_worker_status_payload(task))
    flash("Задача возвращена в работу", "success")
    return redirect(url_for("main.my_tasks"))



GLASS_STATUS_NONE = "none"
GLASS_STATUS_MEASURE_NEEDED = "measure_needed"
GLASS_STATUS_MEASURED = "measured"
GLASS_STATUS_ORDERED = "ordered"
GLASS_STATUS_REPLACED = "replaced"

GLASS_STATUS_LABELS = {
    GLASS_STATUS_NONE: "Без замера",
    "not_ordered": "Сделать замер",
    GLASS_STATUS_MEASURE_NEEDED: "Сделать замер",
    GLASS_STATUS_MEASURED: "Замер внесён",
    GLASS_STATUS_ORDERED: "Заказано",
    GLASS_STATUS_REPLACED: "Поменяно",
}

GLASS_ITEM_TYPES = ["Стеклопакет", "Стекло", "Фр.стекло", "Рама/Створка", "Подоконник"]

GLASS_ITEM_WORDS = {
    "стеклопакет": {"singular": "стеклопакет", "plural": "стеклопакеты", "gender": "m"},
    "стекло": {"singular": "стекло", "plural": "стекла", "gender": "n"},
    "рама/створка": {"singular": "рама/створка", "plural": "рамы/створки", "gender": "f"},
    "подоконник": {"singular": "подоконник", "plural": "подоконники", "gender": "m"},
}


def _glass_item_key(value: str | None) -> str:
    return " ".join(str(value or "").strip().lower().replace("ё", "е").split())


def _glass_item_quantity(value) -> int:
    try:
        quantity = int(float(value or 1))
    except (TypeError, ValueError):
        quantity = 1
    return max(quantity, 1)


def _glass_status_verb(status: str, gender: str, quantity: int) -> str:
    if status == GLASS_STATUS_REPLACED:
        if quantity != 1:
            return "Поменяны"
        return {"m": "Поменян", "f": "Поменяна", "n": "Поменяно"}.get(gender, "Поменяно")
    if quantity != 1:
        return "Заказаны"
    return {"m": "Заказан", "f": "Заказана", "n": "Заказано"}.get(gender, "Заказано")


def _glass_item_noun(item_type: str | None, quantity: int) -> tuple[str, str]:
    key = _glass_item_key(item_type)
    meta = GLASS_ITEM_WORDS.get(key)
    if meta:
        return (meta["singular"] if quantity == 1 else meta["plural"], meta["gender"])
    fallback = str(item_type or "позиция").strip().lower() or "позиция"
    return (fallback, "f")


def glass_measurement_action_label(measurement: GlassMeasurement | None) -> str:
    status = _measurement_status(measurement)
    if status not in {GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED}:
        return GLASS_STATUS_LABELS.get(status, status)
    items = _glass_item_rows(measurement)
    type_quantities: dict[str, dict[str, object]] = {}
    for item in items:
        item_type = str(item.get("item_type") or "Стеклопакет").strip() or "Стеклопакет"
        key = _glass_item_key(item_type)
        entry = type_quantities.setdefault(key, {"type": item_type, "quantity": 0})
        entry["quantity"] = int(entry["quantity"]) + _glass_item_quantity(item.get("quantity"))
    if not type_quantities and measurement:
        item_type = measurement.glass_type or "Стеклопакет"
        type_quantities[_glass_item_key(item_type)] = {
            "type": item_type,
            "quantity": _glass_item_quantity(measurement.quantity),
        }
    if not type_quantities:
        return GLASS_STATUS_LABELS.get(status, status)
    if len(type_quantities) == 1:
        entry = next(iter(type_quantities.values()))
        quantity = _glass_item_quantity(entry.get("quantity"))
        noun, gender = _glass_item_noun(str(entry.get("type") or ""), quantity)
        return f"{_glass_status_verb(status, gender, quantity)} {noun}"
    prefix = "Поменяно" if status == GLASS_STATUS_REPLACED else "Заказано"
    nouns = []
    for entry in type_quantities.values():
        quantity = _glass_item_quantity(entry.get("quantity"))
        noun, _gender = _glass_item_noun(str(entry.get("type") or ""), quantity)
        nouns.append(noun)
    return f"{prefix}: {', '.join(nouns)}"


def _glass_item_rows(measurement: GlassMeasurement | None) -> list[dict[str, object]]:
    if measurement is None:
        return []
    rows = []
    for item in getattr(measurement, "items", []) or []:
        size_input, item_comment = _split_glass_size_comment(item.size_label())
        rows.append({
            "id": item.id,
            "item_type": item.item_type or "Стеклопакет",
            "width": item.width,
            "height": item.height,
            "quantity": item.quantity or 1,
            "size_label": item.size_label(),
            "size_input": size_input or item.size_label(),
            "item_comment": item_comment,
            "title_label": item.title_label(),
        })
    if rows:
        return rows
    if (measurement.width and measurement.height) or (measurement.size or "").strip():
        item_type = measurement.glass_type or "Стеклопакет"
        size_input, item_comment = _split_glass_size_comment(measurement.size_label())
        rows.append({
            "id": None,
            "item_type": item_type,
            "width": measurement.width,
            "height": measurement.height,
            "quantity": measurement.quantity or 1,
            "size_label": measurement.size_label(),
            "size_input": size_input or measurement.size_label(),
            "item_comment": item_comment,
            "title_label": f"{item_type} {measurement.size_label()}".strip(),
        })
    return rows


def _task_search_blob(task: Task) -> str:
    apartment = task.apartment
    categories = []
    if task.work_point:
        categories = [category.name for category in task.work_point.categories]
    parts = [
        apartment.label() if apartment else "",
        apartment.apartment_number if apartment else "",
        apartment.construction_number if apartment else "",
        apartment.building if apartment else "",
        "коммерция" if apartment and apartment.premise_type == "commercial" else "квартира",
        task.description or "",
        task.source_cell_value or "",
        task.title or "",
        task.comment or "",
        task.status_label(),
        task.work_point.point_number if task.work_point else "",
        task.work_point.display_name if task.work_point else "",
        " ".join(categories),
    ]
    return " ".join(str(part or "") for part in parts).lower().replace("ё", "е")


GLASS_WORK_POINT_NUMBERS = {"16", "17", "18", "19", "20"}


def _all_project_tasks(project_id: int, point_numbers: set[str] | None = None) -> list[Task]:
    query = (
        Task.query.options(
            selectinload(Task.apartment),
            selectinload(Task.work_point).selectinload(WorkPoint.categories),
            selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items),
        )
        .join(Apartment)
        .join(WorkPoint)
        .filter(Task.project_id == project_id)
    )
    if point_numbers:
        query = query.filter(WorkPoint.point_number.in_(sorted(point_numbers)))
    return (
        query.order_by(Apartment.premise_type.asc(), Apartment.building.asc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc(), WorkPoint.point_number.asc(), Task.id.asc())
        .limit(3000)
        .all()
    )


def _glass_tasks(project_id: int) -> list[Task]:
    # В «Замерах» работают только замечания по оконным пунктам 16–20.
    # Основной статус задачи не влияет на статус замера.
    return _all_project_tasks(project_id, point_numbers=GLASS_WORK_POINT_NUMBERS)


def _glass_tasks_without_ordered_apartments(tasks: list[Task]) -> list[Task]:
    ordered_apartment_keys = {
        _apartment_group_key(task.apartment)
        for task in tasks
        if task.apartment is not None
        and _measurement_status(task.glass_measurement) in {GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED}
    }
    if not ordered_apartment_keys:
        return tasks
    return [
        task for task in tasks
        if task.apartment is not None and _apartment_group_key(task.apartment) not in ordered_apartment_keys
    ]


def _glass_point_options(project_id: int) -> list[dict[str, str]]:
    points = (
        WorkPoint.query.join(Task)
        .filter(Task.project_id == project_id, WorkPoint.point_number.in_(sorted(GLASS_WORK_POINT_NUMBERS)))
        .distinct()
        .order_by(WorkPoint.point_number.asc(), WorkPoint.original_column_name.asc())
        .all()
    )
    return [
        {"number": str(point.point_number or ""), "label": f"{point.point_number} — {point.display_name}"}
        for point in points
        if str(point.point_number or "").strip()
    ]


def _get_or_create_glass_measurement(task: Task, status: str = GLASS_STATUS_MEASURE_NEEDED) -> GlassMeasurement:
    measurement = task.glass_measurement
    if measurement is None:
        measurement = GlassMeasurement(
            project_id=task.project_id,
            apartment_id=task.apartment_id,
            task_id=task.id,
            status=status,
            quantity=1,
        )
        db.session.add(measurement)
        db.session.flush()
    elif status and (measurement.status in (None, "", GLASS_STATUS_NONE, "not_ordered")):
        measurement.status = status
    if not measurement.apartment_id:
        measurement.apartment_id = task.apartment_id
    return measurement


def _measurement_status(measurement: GlassMeasurement | None) -> str:
    if measurement is None:
        return GLASS_STATUS_NONE
    return GLASS_STATUS_MEASURE_NEEDED if measurement.status == "not_ordered" else (measurement.status or GLASS_STATUS_NONE)


def _filter_glass_rows(tasks: list[Task], q: str = "", status: str = "", point: str = "") -> list[dict[str, object]]:
    premise_selectors, tail_query = parse_multi_premise_search(q)
    search_mode, search_value = detect_search_mode(q)
    needle_source = tail_query if premise_selectors else search_value
    needle = needle_source.strip().lower().replace("ё", "е")
    selector_count = len(premise_selectors)
    rows = []
    for task in tasks:
        measurement = task.glass_measurement
        current_status = _measurement_status(measurement)
        if status and current_status != status:
            continue
        if point and (not task.work_point or str(task.work_point.point_number).strip() != point):
            continue
        if premise_selectors and not any(premise_matches_selector(task.apartment, selector) for selector in premise_selectors):
            continue
        if needle:
            if search_mode in {"premise_number", "premise_number_or_building", "commercial_pair", "construction_number"}:
                if premise_selectors:
                    pass
                elif not premise_matches_search(task.apartment, search_mode, search_value):
                    continue
            elif needle not in _task_search_blob(task):
                continue
        rows.append({
            "task": task,
            "measurement": measurement,
            "items": _glass_item_rows(measurement),
            "status": current_status,
            "status_label": glass_measurement_action_label(measurement),
        })
    def _glass_row_sort_key(row: dict[str, object]):
        task = row["task"]
        apartment = task.apartment if isinstance(task, Task) else None
        requested_order = selector_count
        for index, selector in enumerate(premise_selectors):
            if premise_matches_selector(apartment, selector):
                requested_order = index
                break
        return (requested_order, *_task_apartment_sort_value_no_done(task))

    return sorted(rows, key=_glass_row_sort_key)


@bp.route("/glass")
@bp.route("/glass-measurements")
@login_required
def glass_measurements():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    tab = (request.args.get("tab") or "all").strip()
    if tab not in {"all", "order", "ordered"}:
        tab = "all"
    if _is_mobile_phone_request():
        tab = "ordered"
    q = (request.args.get("q") or "").strip()
    point = (request.args.get("point") or "").strip()
    ordered_status = (request.args.get("ordered_status") or "").strip()
    if ordered_status not in {"", GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED}:
        ordered_status = ""
    tasks = _glass_tasks(project.id)
    available_tasks = _glass_tasks_without_ordered_apartments(tasks)
    rows = []
    order_rows = []
    ordered_rows = []
    if tab == "all":
        rows = _filter_glass_rows(available_tasks, q=q, point=point)
    elif tab == "order":
        order_rows = _filter_glass_rows(available_tasks, q=q, status=GLASS_STATUS_MEASURE_NEEDED)
        order_rows.sort(key=lambda row: _task_apartment_sort_value_no_done(row["task"]))
    elif tab == "ordered":
        ordered_rows = [
            row for row in _filter_glass_rows(tasks, q=q)
            if row["status"] in {GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED}
        ]
        if ordered_status:
            ordered_rows = [row for row in ordered_rows if row["status"] == ordered_status]
        ordered_rows.sort(key=lambda row: _task_apartment_sort_value_no_done(row["task"]))

    active_rows = rows if tab == "all" else order_rows if tab == "order" else ordered_rows
    active_total = len(active_rows)
    glass_pagination = None
    glass_prev_args = {}
    glass_next_args = {}
    per_page = 10 if _is_mobile_phone_request() else 20
    requested_page = max(request.args.get("page", 1, type=int), 1)
    page_count = max((active_total + per_page - 1) // per_page, 1)
    page = min(requested_page, page_count)
    page_start = (page - 1) * per_page
    page_rows = active_rows[page_start:page_start + per_page]
    if tab == "all":
        rows = page_rows
    elif tab == "order":
        order_rows = page_rows
    else:
        ordered_rows = page_rows
    glass_pagination = {
        "page": page,
        "pages": page_count,
        "has_prev": page > 1,
        "has_next": page < page_count,
    }
    glass_prev_args = request.args.to_dict(flat=False)
    glass_next_args = request.args.to_dict(flat=False)
    glass_prev_args["tab"] = tab
    glass_next_args["tab"] = tab
    if page > 1:
        glass_prev_args["page"] = page - 1
    if page < page_count:
        glass_next_args["page"] = page + 1
    return render_template(
        "glass_measurements.html",
        project=project,
        glass_apartments=_project_apartment_options(project.id),
        rows=rows,
        order_rows=order_rows,
        ordered_rows=ordered_rows,
        tab=tab,
        q=q,
        status_labels=GLASS_STATUS_LABELS,
        glass_item_types=GLASS_ITEM_TYPES,
        glass_point_options=_glass_point_options(project.id),
        selected_point=point,
        ordered_status=ordered_status,
        active_total=active_total,
        glass_pagination=glass_pagination,
        glass_prev_args=glass_prev_args,
        glass_next_args=glass_next_args,
        today=date.today(),
    )


@bp.route("/glass/<int:task_id>/need-measure", methods=["POST"])
@login_required
def glass_need_measure(task_id: int):
    if current_user.role == "viewer":
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    measurement = _get_or_create_glass_measurement(task, status=GLASS_STATUS_MEASURE_NEEDED)
    measurement.status = GLASS_STATUS_MEASURE_NEEDED
    if not measurement.apartment_id:
        measurement.apartment_id = task.apartment_id
    db.session.commit()
    return_tab = (request.form.get("return_tab") or request.args.get("tab") or "order").strip().lower()
    if return_tab not in {"all", "order"}:
        return_tab = "order"
    success_message = "Позиция переведена в статус «В заказе»"
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify({
            "ok": True,
            "message": success_message,
            "status": GLASS_STATUS_MEASURE_NEEDED,
            "task_id": task.id,
            "measurement_id": measurement.id,
        })
    flash(success_message, "success")
    return redirect(url_for("main.glass_measurements", tab=return_tab))


@bp.route("/glass/<int:task_id>/save", methods=["POST"])
@login_required
def glass_measurement_save(task_id: int):
    if current_user.role == "viewer":
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    measurement = _get_or_create_glass_measurement(task, status=GLASS_STATUS_MEASURE_NEEDED)

    item_types = request.form.getlist("item_type[]") or request.form.getlist("item_type")
    sizes = request.form.getlist("size[]") or request.form.getlist("size")
    comments = request.form.getlist("item_comment[]") or request.form.getlist("item_comment")
    quantities = request.form.getlist("quantity[]") or request.form.getlist("quantity")
    max_len = max(len(item_types), len(sizes), len(comments), len(quantities), 1)

    def _size_dimensions(size_text: str) -> tuple[float, float]:
        import re
        numbers = re.findall(r"\d+(?:[,.]\d+)?", size_text or "")
        if len(numbers) >= 2:
            return float(numbers[0].replace(",", ".")), float(numbers[1].replace(",", "."))
        return 0.0, 0.0

    parsed_items: list[dict[str, object]] = []
    for index in range(max_len):
        item_type = (item_types[index] if index < len(item_types) else "Стеклопакет").strip() or "Стеклопакет"
        if item_type not in GLASS_ITEM_TYPES:
            item_type = "Стеклопакет"
        size_text = (sizes[index] if index < len(sizes) else "").strip()
        item_comment = (comments[index] if index < len(comments) else "").strip()
        quantity = _parse_quantity(quantities[index] if index < len(quantities) else None) or 1
        if not size_text and not item_comment:
            continue
        if not size_text:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
                return jsonify({"ok": False, "message": "В каждой добавленной позиции укажите размер"}), 400
            flash("В каждой добавленной позиции укажите размер", "warning")
            return redirect(url_for("main.glass_measurements", tab="order"))
        width, height = _size_dimensions(size_text)
        display_size = size_text
        if item_comment:
            display_size = f"{size_text} — {item_comment}"
        parsed_items.append({
            "item_type": item_type,
            "width": width,
            "height": height,
            "quantity": int(quantity) if float(quantity).is_integer() else int(round(quantity)),
            "size": display_size,
        })

    if not parsed_items:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": "Добавьте хотя бы один размер"}), 400
        flash("Добавьте хотя бы один размер", "warning")
        return redirect(url_for("main.glass_measurements", tab="order"))

    measurement.items.clear()
    for item_data in parsed_items:
        measurement.items.append(GlassMeasurementItem(**item_data))

    first_item = parsed_items[0]
    measurement.width = first_item["width"]
    measurement.height = first_item["height"]
    measurement.quantity = first_item["quantity"]
    measurement.glass_type = first_item["item_type"]
    measurement.comment = (request.form.get("comment") or "").strip() or None
    measurement.measured_at = parse_date(request.form.get("measured_at")) or date.today()
    measurement.size = first_item["size"]
    measurement.status = GLASS_STATUS_ORDERED
    measurement.ordered_at = parse_date(request.form.get("ordered_at")) or date.today()
    if not measurement.apartment_id:
        measurement.apartment_id = task.apartment_id
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify({
            "ok": True,
            "message": "Размеры перенесены в заказ",
            "measurement_id": measurement.id,
            "task_id": task.id,
            "status": measurement.status,
            "items": _glass_item_rows(measurement),
        })
    flash("Размеры внесены. Позиция перемещена во вкладку «Заказано»", "success")
    return redirect(url_for("main.glass_measurements", tab="ordered"))


@bp.route("/glass/<int:measurement_id>/return-to-all", methods=["POST"])
@login_required
def glass_measurement_return_to_all(measurement_id: int):
    if current_user.role == "viewer":
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    measurement = (
        GlassMeasurement.query.options(selectinload(GlassMeasurement.task))
        .filter(GlassMeasurement.id == measurement_id, GlassMeasurement.project_id == project.id)
        .first()
        or abort(404)
    )
    measurement.status = GLASS_STATUS_NONE
    measurement.width = None
    measurement.height = None
    measurement.quantity = 1
    measurement.glass_type = None
    measurement.size = None
    measurement.comment = None
    measurement.measured_at = None
    measurement.ordered_at = None
    measurement.replaced_at = None
    measurement.material_request_item = None
    measurement.items.clear()
    db.session.commit()
    return_tab = (request.form.get("return_tab") or request.args.get("tab") or "all").strip().lower()
    if return_tab not in {"all", "order"}:
        return_tab = "all"
    if wants_json:
        return jsonify({
            "ok": True,
            "message": "Позиция снова доступна как «Сделать замер»",
            "status": GLASS_STATUS_NONE,
            "task_id": measurement.task_id,
        })
    flash("Позиция снова доступна как «Сделать замер»", "success")
    return redirect(url_for("main.glass_measurements", tab=return_tab))


@bp.route("/glass/<int:measurement_id>/status", methods=["POST"])
@login_required
def glass_status_update(measurement_id: int):
    if current_user.role == "viewer":
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    measurement = (
        GlassMeasurement.query.options(selectinload(GlassMeasurement.task))
        .filter(GlassMeasurement.id == measurement_id, GlassMeasurement.project_id == project.id)
        .first()
        or abort(404)
    )
    next_status = (request.form.get("status") or "").strip()
    status_date = parse_date(request.form.get("status_date")) or date.today()
    if next_status not in {GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED}:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": "Некорректный статус стеклопакета"}), 400
        flash("Некорректный статус стеклопакета", "warning")
        return redirect(url_for("main.glass_measurements", tab="ordered"))
    measurement.status = next_status
    if next_status == GLASS_STATUS_ORDERED:
        if not measurement.ordered_at:
            measurement.ordered_at = status_date
        measurement.replaced_at = None
    else:
        measurement.replaced_at = status_date
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify({
            "ok": True,
            "message": "Статус стеклопакета обновлён",
            "status": measurement.status,
            "status_label": glass_measurement_action_label(measurement),
            "ordered_at": format_ru_date(measurement.ordered_at) if measurement.ordered_at else "—",
            "replaced_at": format_ru_date(measurement.replaced_at) if measurement.replaced_at else "",
        })
    flash("Статус стеклопакета обновлён", "success")
    return redirect(url_for("main.glass_measurements", tab="ordered"))


@bp.route("/glass/order/export")
@login_required
def glass_order_export():
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    scope = (request.args.get("scope") or "ordered").strip().lower()
    q = (request.args.get("q") or "").strip()
    tasks = _glass_tasks(project.id)
    if scope == "order":
        all_rows = _filter_glass_rows(_glass_tasks_without_ordered_apartments(tasks), q=q)
        rows = [row for row in all_rows if row["status"] == GLASS_STATUS_MEASURE_NEEDED]
    else:
        all_rows = _filter_glass_rows(tasks, q=q)
        rows = [row for row in all_rows if row["status"] in {GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED}]
    rows.sort(key=lambda row: _task_apartment_sort_value_no_done(row["task"]))

    wb = Workbook()
    ws = wb.active
    if scope == "order":
        ws.title = "Заказать"
        ws.append(["Номер квартиры", "Замечание", "Тип", "Размеры"])
        for row in rows:
            task = row["task"]
            ws.append([
                _excel_premise_label(task.apartment) if task.apartment else "",
                task.description or task.source_cell_value or "",
                "",
                "",
            ])
    else:
        ws.title = "Заказано"
        ws.append([
            "Помещение",
            "Замечание",
            "Тип",
            "Размер / комментарий",
            "Количество",
        ])

        for row in rows:
            task = row["task"]
            measurement = row["measurement"]
            items = row.get("items") or []
            if not items:
                items = [{
                    "item_type": getattr(measurement, "glass_type", "") or "",
                    "size_label": measurement.size_label() if measurement else "",
                    "quantity": getattr(measurement, "quantity", "") or "",
                }]
            for item in items:
                ws.append([
                    _excel_premise_label(task.apartment) if task.apartment else "",
                    task.description or task.source_cell_value or "",
                    item.get("item_type") or "",
                    item.get("size_label") or item.get("title_label") or "",
                    item.get("quantity") or "",
                ])

    _style_excel_header(ws)
    if scope == "order":
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 56
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 38
    else:
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 56
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 42
        ws.column_dimensions["E"].width = 14
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center" if cell.row == 1 else "left",
                vertical="center" if cell.row == 1 else "top",
                wrap_text=True,
            )
    last_column = "D" if scope == "order" else "E"
    ws.auto_filter.ref = f"A1:{last_column}{ws.max_row}"
    suffix = "замеры_заказать" if scope == "order" else "замеры_заказано"
    filename = f"{project.name}_{suffix}_{date.today().strftime('%Y-%m-%d')}.xlsx".replace("/", "-")
    return _make_excel_response(wb, filename)


@bp.route("/glass/ordered/create-material-request", methods=["POST"])
@login_required
def glass_create_material_request():
    if current_user.role == "viewer":
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    selected_ids = []
    for raw_id in request.form.getlist("measurement_ids"):
        try:
            selected_ids.append(int(raw_id))
        except (TypeError, ValueError):
            pass
    if not selected_ids:
        flash("Выберите хотя бы один стеклопакет", "warning")
        return redirect(url_for("main.glass_measurements", tab="ordered"))
    measurements = (
        GlassMeasurement.query.options(
            selectinload(GlassMeasurement.task).selectinload(Task.apartment),
            selectinload(GlassMeasurement.items),
            selectinload(GlassMeasurement.material_request_item).selectinload(MaterialRequestItem.request),
            selectinload(GlassMeasurement.material_writeoff),
        )
        .filter(GlassMeasurement.project_id == project.id, GlassMeasurement.id.in_(selected_ids))
        .all()
    )
    if len(measurements) != len(set(selected_ids)):
        flash("Часть выбранных стеклопакетов не найдена", "warning")
        return redirect(url_for("main.glass_measurements", tab="ordered"))
    already_requested = [m for m in measurements if m.material_request_item_id or m.material_writeoff_id]
    if already_requested:
        flash("Выбранные стеклопакеты уже внесены в заявку. Повторную заявку создать нельзя", "warning")
        return redirect(url_for("main.glass_measurements", tab="ordered"))
    without_size = [m for m in measurements if not _glass_item_rows(m)]
    if without_size:
        flash("У выбранных позиций должен быть указан хотя бы один размер", "warning")
        return redirect(url_for("main.glass_measurements", tab="ordered"))
    request_number = _next_measurement_request_number(project.id)
    material_request = MaterialRequest(
        project_id=project.id,
        author_id=current_user.id,
        request_date=date.today(),
        title=f"Заявка из замеров №{request_number}",
        comment=MEASUREMENT_REQUEST_COMMENT,
    )
    db.session.add(material_request)
    for measurement in measurements:
        task = measurement.task
        apt = task.apartment.label() if task and task.apartment else ""
        first_item_for_measurement = None
        writeoff = MaterialWriteOff(
            project_id=project.id,
            author_id=current_user.id,
            writeoff_date=date.today(),
            comment=f"{MEASUREMENT_WRITEOFF_COMMENT_PREFIX}: {material_request.title}",
        )
        if task is not None:
            writeoff.tasks.append(task)
        for item_row in _glass_item_rows(measurement):
            title = str(item_row.get("title_label") or "").strip()
            item_name = f"{title} {apt}".strip()
            quantity = item_row.get("quantity") or 1
            request_item = MaterialRequestItem(
                name=item_name,
                quantity=quantity,
                unit="шт",
            )
            material_request.items.append(request_item)
            writeoff.items.append(MaterialWriteOffItem(
                name=item_name,
                quantity=quantity,
                unit="шт",
            ))
            if first_item_for_measurement is None:
                first_item_for_measurement = request_item
        if first_item_for_measurement is not None:
            measurement.material_request_item = first_item_for_measurement
            measurement.material_writeoff = writeoff
            db.session.add(writeoff)
    db.session.commit()
    flash("Заявка создана, материалы автоматически списаны по связанным замечаниям", "success")
    return redirect(url_for("main.material_request_detail", request_id=material_request.id))


@bp.route("/glass/delete", methods=["POST"])
@login_required
def glass_measurements_delete():
    if current_user.role == "viewer":
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    scope = (request.form.get("scope") or "ordered").strip()
    if scope == "order":
        statuses = [GLASS_STATUS_MEASURE_NEEDED]
        redirect_tab = "order"
    else:
        statuses = [GLASS_STATUS_ORDERED, GLASS_STATUS_REPLACED]
        redirect_tab = "ordered"

    selected_ids = []
    for raw_id in request.form.getlist("measurement_ids"):
        try:
            selected_ids.append(int(raw_id))
        except (TypeError, ValueError):
            pass

    query = GlassMeasurement.query.filter(
        GlassMeasurement.project_id == project.id,
        GlassMeasurement.status.in_(statuses),
    )
    if request.form.get("delete_all") != "1":
        if not selected_ids:
            flash("Выберите хотя бы одну позицию для удаления", "warning")
            return redirect(url_for("main.glass_measurements", tab=redirect_tab))
        query = query.filter(GlassMeasurement.id.in_(selected_ids))

    measurements = query.all()
    for measurement in measurements:
        _record_simple_deletion(
            "glass_measurement_delete",
            "glass_measurement",
            measurement,
            f"Стеклопакет #{measurement.id}",
            f"Удалена позиция стеклопакета: {measurement.size_label() or 'без размера'}.",
            project_id=project.id,
            extra={
                "task_id": measurement.task_id,
                "apartment_id": measurement.apartment_id,
                "items": [_snapshot_model(item) for item in measurement.items],
            },
        )
        db.session.delete(measurement)
    db.session.commit()
    flash(f"Удалено позиций: {len(measurements)}", "success")
    return redirect(url_for("main.glass_measurements", tab=redirect_tab))


# Старые ссылки на отдельную страницу заказа оставлены как безопасный редирект.
@bp.route("/glass/order", methods=["GET", "POST"])
@login_required
def glass_order():
    return redirect(url_for("main.glass_measurements", tab="ordered"))


@bp.route("/glass/task/new", methods=["POST"])
@login_required
def glass_manual_task_new():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)

    apartment_id = request.form.get("apartment_id", type=int)
    text = (request.form.get("description") or "").strip()
    apartment = db.session.get(Apartment, apartment_id) if apartment_id else None
    default_point = WorkPoint.query.filter_by(point_number="22").order_by(WorkPoint.id.asc()).first()
    if default_point is None:
        default_point = WorkPoint(point_number="22", short_name="Прочее", original_column_name="Прочее", source_sheet_name="manual", is_active=True)
        db.session.add(default_point)
        db.session.flush()

    if not apartment or apartment.project_id != project.id:
        return jsonify({"ok": False, "message": "Выберите квартиру / коммерцию"}), 400
    if not text:
        return jsonify({"ok": False, "message": "Введите описание работы"}), 400

    source_uid = build_task_uid(
        project.name,
        apartment.construction_number or "",
        apartment.apartment_number or "",
        default_point.point_number,
        default_point.display_name,
        text,
    )
    if Task.query.filter_by(source_uid=source_uid).first():
        source_uid = stable_hash([source_uid, "glass-manual", datetime.utcnow().isoformat()])
    task = Task(
        source_uid=source_uid,
        project_id=project.id,
        apartment_id=apartment.id,
        work_point_id=default_point.id,
        title=default_point.display_name,
        description=text,
        source_cell_value=text,
        source_sheet_name="manual_glass",
        status=STATUS_NOT_STARTED,
        is_done=False,
        manually_edited=True,
        last_seen_at=datetime.utcnow(),
        source_hash=stable_hash([text]),
    )
    db.session.add(task)
    db.session.flush()
    measurement = _get_or_create_glass_measurement(task, status=GLASS_STATUS_NONE)
    measurement.status = GLASS_STATUS_NONE
    if not measurement.apartment_id:
        measurement.apartment_id = apartment.id
    db.session.commit()
    return jsonify({
        "ok": True,
        "message": "Замечание добавлено",
        "task_id": task.id,
        "apartment_label": apartment.label() if apartment else "—",
        "description": task.description or task.source_cell_value or "",
        "status_label": task.status_label(),
        "status_class": task.status_class(),
        "task_url": url_for("main.task_detail", task_id=task.id),
    })


@bp.route("/materials")
@login_required
def materials():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))

    active_tab = request.args.get("tab", "balance")
    allowed_tabs = {"balance", "requests", "writeoff", "history", "task"}
    if active_tab not in allowed_tabs:
        active_tab = "balance"

    material_requests = (
        MaterialRequest.query.options(selectinload(MaterialRequest.items), selectinload(MaterialRequest.author))
        .filter(MaterialRequest.project_id == project.id)
        .order_by(MaterialRequest.request_date.desc(), MaterialRequest.id.desc())
        .all()
    )
    balance_rows = _material_balance_rows(project.id)
    balance_options = _balance_options(project.id)
    writeoffs = (
        MaterialWriteOff.query.options(
            selectinload(MaterialWriteOff.items),
            selectinload(MaterialWriteOff.tasks).selectinload(Task.apartment),
            selectinload(MaterialWriteOff.tasks).selectinload(Task.work_point),
        )
        .filter(MaterialWriteOff.project_id == project.id)
        .order_by(MaterialWriteOff.writeoff_date.desc(), MaterialWriteOff.id.desc())
        .all()
    )
    writeoffs.sort(key=_writeoff_sort_value)
    if active_tab == "history":
        writeoffs = [_material_writeoff_history_view(writeoff) for writeoff in writeoffs]
    writeoff_tasks = _material_task_options(project.id, request.args) if active_tab == "writeoff" else []
    low_stock_count = sum(1 for row in balance_rows if 0 < float(row.get("balance") or 0) <= 5)
    return render_template(
        "materials.html",
        project=project,
        active_tab=active_tab,
        material_requests=material_requests,
        balance_rows=balance_rows,
        balance_options=balance_options,
        writeoffs=writeoffs,
        writeoff_tasks=writeoff_tasks,
        material_writeoff_premise_text=_material_writeoff_premise_text,
        material_writeoff_remark_lines=_material_writeoff_remark_lines,
        low_stock_count=low_stock_count,
        can_edit_materials=_can_edit_materials(),
        args=request.args,
        today=date.today(),
    )


@bp.route("/materials/balance/delete", methods=["POST"])
@login_required
def material_balance_delete():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)

    selected_keys = []
    if request.form.get("delete_all") == "1":
        selected_keys = [str(row.get("key") or "") for row in _material_balance_rows(project.id)]
    else:
        selected_keys = [str(value or "").strip() for value in request.form.getlist("material_keys") if str(value or "").strip()]

    selected_pairs = {_normalize_material_identity(*parsed) for parsed in (_split_material_key(value) for value in selected_keys) if parsed}
    if not selected_pairs:
        flash("Выберите хотя бы одну позицию баланса", "warning")
        return redirect(url_for("main.materials", tab="balance"))

    request_items = (
        MaterialRequestItem.query.join(MaterialRequest)
        .filter(MaterialRequest.project_id == project.id)
        .all()
    )
    writeoff_items = (
        MaterialWriteOffItem.query.join(MaterialWriteOff)
        .filter(MaterialWriteOff.project_id == project.id)
        .all()
    )

    removed_items = 0
    touched_requests: set[MaterialRequest] = set()
    touched_writeoffs: set[MaterialWriteOff] = set()
    deleted_request_item_snapshots: list[dict] = []
    deleted_writeoff_item_snapshots: list[dict] = []

    for item in request_items:
        if _normalize_material_identity(item.name, item.unit) not in selected_pairs:
            continue
        GlassMeasurement.query.filter(GlassMeasurement.material_request_item_id == item.id).update(
            {"material_request_item_id": None},
            synchronize_session=False,
        )
        touched_requests.add(item.request)
        deleted_request_item_snapshots.append(_snapshot_model(item))
        db.session.delete(item)
        removed_items += 1

    for item in writeoff_items:
        if _normalize_material_identity(item.name, item.unit) not in selected_pairs:
            continue
        touched_writeoffs.add(item.writeoff)
        deleted_writeoff_item_snapshots.append(_snapshot_model(item))
        db.session.delete(item)
        removed_items += 1

    db.session.flush()
    for material_request in list(touched_requests):
        if material_request is not None and not material_request.items:
            _record_simple_deletion(
                "material_request_delete_empty_after_balance",
                "material_request",
                material_request,
                material_request.title or f"Заявка #{material_request.id}",
                "Заявка удалена автоматически после удаления всех строк баланса.",
                project_id=project.id,
                extra={"items": [_snapshot_model(item) for item in material_request.items]},
            )
            db.session.delete(material_request)
    for writeoff in list(touched_writeoffs):
        if writeoff is not None and not writeoff.items:
            measurement_ids = _measurement_ids_for_writeoff(writeoff.id)
            _record_simple_deletion(
                "material_writeoff_delete_empty_after_balance",
                "material_writeoff",
                writeoff,
                f"Списание #{writeoff.id}",
                "Списание удалено автоматически после удаления всех строк баланса.",
                project_id=project.id,
                extra={
                    "items": [_snapshot_model(item) for item in writeoff.items],
                    "task_ids": [task.id for task in writeoff.tasks],
                    "measurement_ids": measurement_ids,
                },
            )
            _unlink_measurements_from_writeoff(writeoff.id)
            writeoff.tasks.clear()
            db.session.delete(writeoff)

    _record_deletion_action(
        "material_balance_delete",
        "material_balance",
        None,
        "Удаление материалов из баланса",
        f"Удалены материалы из баланса: {len(selected_pairs)} поз.; затронуто строк: {removed_items}.",
        {
            "selected_pairs": sorted([{"name": name, "unit": unit} for name, unit in selected_pairs], key=lambda row: (row["name"], row["unit"])),
            "removed_items": removed_items,
            "request_items": deleted_request_item_snapshots,
            "writeoff_items": deleted_writeoff_item_snapshots,
        },
        project_id=project.id,
    )
    db.session.commit()
    flash(f"Удалено позиций баланса: {len(selected_pairs)}. Затронуто строк материалов: {removed_items}", "success")
    return redirect(url_for("main.materials", tab="balance"))


@bp.route("/materials/request/<int:request_id>")
@login_required
def material_request_detail(request_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    material_request = (
        MaterialRequest.query.options(selectinload(MaterialRequest.items), selectinload(MaterialRequest.author))
        .filter(MaterialRequest.id == request_id, MaterialRequest.project_id == project.id)
        .first()
    ) or abort(404)
    max_rows = len(material_request.items) + 1
    return render_template(
        "material_request_detail.html",
        project=project,
        material_request=material_request,
        can_edit_materials=_can_edit_materials(),
        max_rows=max_rows,
        today=date.today(),
    )


@bp.route("/materials/request/<int:request_id>/rename", methods=["POST"])
@login_required
def material_request_rename(request_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)
    material_request = (
        MaterialRequest.query.filter(MaterialRequest.id == request_id, MaterialRequest.project_id == project.id).first()
        or abort(404)
    )
    title = (request.form.get("title") or "").strip()
    if not title:
        flash("Название заявки не может быть пустым", "warning")
        return redirect(request.referrer or url_for("main.materials", tab="requests"))
    material_request.title = title
    db.session.commit()
    flash("Название заявки обновлено", "success")
    return redirect(request.referrer or url_for("main.materials", tab="requests"))


@bp.route("/materials/request/<int:request_id>/update", methods=["POST"])
@login_required
def material_request_update(request_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)
    material_request = (
        MaterialRequest.query.filter(MaterialRequest.id == request_id, MaterialRequest.project_id == project.id).first()
        or abort(404)
    )
    title = (request.form.get("title") or "").strip()
    request_date = parse_date(request.form.get("request_date"))
    if not title:
        flash("Название заявки не может быть пустым", "warning")
        return redirect(url_for("main.material_request_detail", request_id=request_id))
    if request_date is None:
        flash("Введите корректную дату заявки", "warning")
        return redirect(url_for("main.material_request_detail", request_id=request_id))
    try:
        rows = _read_material_rows_from_form(limit=50)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("main.material_request_detail", request_id=request_id))
    if not rows:
        flash("Добавьте хотя бы одну позицию материала", "warning")
        return redirect(url_for("main.material_request_detail", request_id=request_id))
    old_items = list(material_request.items)
    old_item_ids_set = {item.id for item in old_items if item.id}
    measurement_request_groups = _measurement_request_groups(old_items) if _is_measurement_material_request(material_request) else []
    old_item_ids = [item.id for item in old_items if item.id]
    linked_measurements = (
        GlassMeasurement.query.filter(GlassMeasurement.material_request_item_id.in_(old_item_ids)).all()
        if old_item_ids
        else []
    )
    linked_by_old_item_id = {measurement.material_request_item_id: measurement for measurement in linked_measurements}
    for measurement in linked_measurements:
        measurement.material_request_item = None
    material_request.title = title
    material_request.request_date = request_date
    material_request.comment = (request.form.get("comment") or material_request.comment or "").strip() or None
    material_request.items.clear()
    new_items_by_old_id: dict[int, MaterialRequestItem] = {}
    for row in rows:
        new_item = MaterialRequestItem(name=str(row["name"]), quantity=float(row["quantity"]), unit=str(row["unit"]))
        material_request.items.append(new_item)
        source_item_id = row.get("source_item_id")
        if isinstance(source_item_id, int) and source_item_id in old_item_ids_set:
            new_items_by_old_id[source_item_id] = new_item
            linked_measurement = linked_by_old_item_id.get(source_item_id)
            if linked_measurement is not None:
                linked_measurement.material_request_item = new_item
    if _is_measurement_material_request(material_request):
        _sync_measurement_writeoffs_from_request_groups(material_request, measurement_request_groups, new_items_by_old_id)
    db.session.commit()
    flash("Заявка обновлена", "success")
    return redirect(url_for("main.material_request_detail", request_id=request_id))


@bp.route("/materials/request/<int:request_id>/delete", methods=["POST"])
@login_required
def material_request_delete(request_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)
    material_request = (
        MaterialRequest.query.options(selectinload(MaterialRequest.items))
        .filter(MaterialRequest.id == request_id, MaterialRequest.project_id == project.id)
        .first()
        or abort(404)
    )
    deleted_writeoff_ids: list[int] = []
    if _is_measurement_material_request(material_request):
        deleted_writeoff_ids = _delete_measurement_request_writeoffs(material_request)
    item_ids = [item.id for item in material_request.items]
    if item_ids:
        GlassMeasurement.query.filter(GlassMeasurement.material_request_item_id.in_(item_ids)).update(
            {"material_request_item_id": None},
            synchronize_session=False,
        )
    _record_simple_deletion(
        "material_request_delete",
        "material_request",
        material_request,
        material_request.title or f"Заявка #{material_request.id}",
        f"Удалена заявка на материалы: {material_request.title or material_request.id}.",
        project_id=project.id,
        extra={
            "items": [_snapshot_model(item) for item in material_request.items],
            "deleted_writeoff_ids": deleted_writeoff_ids,
        },
    )
    db.session.delete(material_request)
    db.session.commit()
    flash("Заявка удалена", "success")
    return redirect(url_for("main.materials", tab="requests"))


@bp.route("/materials/requests/delete", methods=["POST"])
@login_required
def material_requests_bulk_delete():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)

    selected_ids = []
    for raw_id in request.form.getlist("request_ids"):
        try:
            selected_ids.append(int(raw_id))
        except (TypeError, ValueError):
            pass

    query = MaterialRequest.query.options(selectinload(MaterialRequest.items)).filter(MaterialRequest.project_id == project.id)
    if request.form.get("delete_all") != "1":
        if not selected_ids:
            flash("Выберите хотя бы одну заявку", "warning")
            return redirect(url_for("main.materials", tab="requests"))
        query = query.filter(MaterialRequest.id.in_(selected_ids))

    requests_to_delete = query.all()
    deleted_writeoffs_by_request: dict[int, list[int]] = {}
    for material_request in requests_to_delete:
        if _is_measurement_material_request(material_request):
            deleted_writeoffs_by_request[material_request.id] = _delete_measurement_request_writeoffs(material_request)
    item_ids = [item.id for material_request in requests_to_delete for item in material_request.items if item.id]
    if item_ids:
        GlassMeasurement.query.filter(GlassMeasurement.material_request_item_id.in_(item_ids)).update(
            {"material_request_item_id": None},
            synchronize_session=False,
        )
    for material_request in requests_to_delete:
        _record_simple_deletion(
            "material_request_delete",
            "material_request",
            material_request,
            material_request.title or f"Заявка #{material_request.id}",
            f"Удалена заявка на материалы: {material_request.title or material_request.id}.",
            project_id=project.id,
            extra={
                "items": [_snapshot_model(item) for item in material_request.items],
                "deleted_writeoff_ids": deleted_writeoffs_by_request.get(material_request.id, []),
            },
        )
        db.session.delete(material_request)
    db.session.commit()
    flash(f"Удалено заявок: {len(requests_to_delete)}", "success")
    return redirect(url_for("main.materials", tab="requests"))


@bp.route("/materials/request/<int:request_id>/export")
@login_required
def material_request_export(request_id: int):
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    material_request = (
        MaterialRequest.query.options(selectinload(MaterialRequest.items))
        .filter(MaterialRequest.id == request_id, MaterialRequest.project_id == project.id)
        .first()
        or abort(404)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"
    ws.append(["Дата", "Название заявки", "Наименование", "Количество", "Ед. измерения"])
    request_title = material_request.title or material_request.comment or f"Заявка №{material_request.id}"
    for item in material_request.items:
        ws.append([
            material_request.request_date.strftime("%d.%m.%Y") if material_request.request_date else "",
            request_title,
            item.name,
            fmt_quantity(item.quantity),
            item.unit,
        ])
    if not material_request.items:
        ws.append([material_request.request_date.strftime("%d.%m.%Y") if material_request.request_date else "", request_title, "", "", ""])
    _style_excel_header(ws)
    filename = f"{project.name}_{request_title}_{date.today().strftime('%Y-%m-%d')}.xlsx".replace("/", "-")
    return _make_excel_response(wb, filename)


@bp.route("/materials/write-off/<int:writeoff_id>/edit", methods=["GET", "POST"])
@login_required
def material_writeoff_edit(writeoff_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)
    writeoff = (
        MaterialWriteOff.query.options(selectinload(MaterialWriteOff.items), selectinload(MaterialWriteOff.tasks).selectinload(Task.apartment))
        .filter(MaterialWriteOff.id == writeoff_id, MaterialWriteOff.project_id == project.id)
        .first()
        or abort(404)
    )
    if request.method == "POST":
        writeoff_date = parse_date(request.form.get("writeoff_date"))
        if writeoff_date is None:
            flash("Введите корректную дату списания", "warning")
            return redirect(url_for("main.material_writeoff_edit", writeoff_id=writeoff.id))
        try:
            rows = _read_material_rows_from_form(limit=20)
        except ValueError as exc:
            flash(str(exc), "danger")
            rows = []
        if not rows:
            flash("Добавьте хотя бы одну позицию материала", "warning")
            return redirect(url_for("main.material_writeoff_edit", writeoff_id=writeoff.id))
        writeoff.writeoff_date = writeoff_date
        writeoff.items.clear()
        for row in rows:
            writeoff.items.append(MaterialWriteOffItem(name=str(row["name"]), quantity=float(row["quantity"]), unit=str(row["unit"])))
        db.session.commit()
        flash("Списание обновлено", "success")
        return redirect(url_for("main.materials", tab="history"))
    return render_template("material_writeoff_edit.html", project=project, writeoff=writeoff, max_rows=max(10, len(writeoff.items) + 3))


@bp.route("/materials/write-off/<int:writeoff_id>/delete", methods=["POST"])
@login_required
def material_writeoff_delete(writeoff_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)
    writeoff = MaterialWriteOff.query.filter(MaterialWriteOff.id == writeoff_id, MaterialWriteOff.project_id == project.id).first() or abort(404)
    measurement_ids = _measurement_ids_for_writeoff(writeoff.id)
    _record_simple_deletion(
        "material_writeoff_delete",
        "material_writeoff",
        writeoff,
        f"Списание #{writeoff.id}",
        f"Удалено списание материалов от {writeoff.writeoff_date.strftime('%d.%m.%Y') if writeoff.writeoff_date else 'без даты'}.",
        project_id=project.id,
        extra={
            "items": [_snapshot_model(item) for item in writeoff.items],
            "task_ids": [task.id for task in writeoff.tasks],
            "measurement_ids": measurement_ids,
        },
    )
    _unlink_measurements_from_writeoff(writeoff.id)
    writeoff.tasks.clear()
    db.session.delete(writeoff)
    db.session.commit()
    flash("Списание удалено", "success")
    return redirect(url_for("main.materials", tab="history"))


@bp.route("/materials/write-offs/delete", methods=["POST"])
@login_required
def material_writeoffs_bulk_delete():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)

    selected_ids = []
    for raw_id in request.form.getlist("writeoff_ids"):
        try:
            selected_ids.append(int(raw_id))
        except (TypeError, ValueError):
            pass

    query = MaterialWriteOff.query.filter(MaterialWriteOff.project_id == project.id)
    if request.form.get("delete_all") != "1":
        if not selected_ids:
            flash("Выберите хотя бы одно списание", "warning")
            return redirect(url_for("main.materials", tab="history"))
        query = query.filter(MaterialWriteOff.id.in_(selected_ids))

    writeoffs_to_delete = query.all()
    for writeoff in writeoffs_to_delete:
        measurement_ids = _measurement_ids_for_writeoff(writeoff.id)
        _record_simple_deletion(
            "material_writeoff_delete",
            "material_writeoff",
            writeoff,
            f"Списание #{writeoff.id}",
            f"Удалено списание материалов от {writeoff.writeoff_date.strftime('%d.%m.%Y') if writeoff.writeoff_date else 'без даты'}.",
            project_id=project.id,
            extra={
                "items": [_snapshot_model(item) for item in writeoff.items],
                "task_ids": [task.id for task in writeoff.tasks],
                "measurement_ids": measurement_ids,
            },
        )
        _unlink_measurements_from_writeoff(writeoff.id)
        writeoff.tasks.clear()
        db.session.delete(writeoff)
    db.session.commit()
    flash(f"Удалено списаний: {len(writeoffs_to_delete)}", "success")
    return redirect(url_for("main.materials", tab="history"))


def _make_excel_response(workbook: Workbook, download_name: str):
    _style_excel_workbook_for_download(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _style_excel_workbook_for_download(workbook: Workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                if not alignment.vertical:
                    alignment.vertical = "top"
                cell.alignment = alignment
                cell.border = EXCEL_DOWNLOAD_BORDER
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions


def _style_excel_header(ws):
    fill = PatternFill(fill_type="solid", start_color=EXCEL_HEADER_FILL_COLOR, end_color=EXCEL_HEADER_FILL_COLOR)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = EXCEL_DOWNLOAD_BORDER
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 3, 14), 60)


@bp.route("/materials/expense/export")
@login_required
def material_expense_export():
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    writeoffs = (
        MaterialWriteOff.query.options(
            selectinload(MaterialWriteOff.items),
            selectinload(MaterialWriteOff.tasks).selectinload(Task.apartment),
            selectinload(MaterialWriteOff.tasks).selectinload(Task.work_point),
        )
        .filter(MaterialWriteOff.project_id == project.id)
        .order_by(MaterialWriteOff.writeoff_date.desc(), MaterialWriteOff.id.desc())
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Расход материала"
    ws.append(["Дата", "№", "Перечень замечаний", "Потраченный материал"])
    for writeoff in writeoffs:
        material_lines = [f"{display_material_name(item.name)} — {fmt_quantity(item.quantity)} {item.unit}" for item in writeoff.items]
        tasks = list(writeoff.tasks)
        manual_lines = _material_writeoff_remark_lines(writeoff) if not tasks else []
        if not tasks:
            tasks = [None]
        start_row = ws.max_row + 1
        for index, task in enumerate(tasks):
            premise = _excel_premise_label(task.apartment) if task and task.apartment else ""
            if not premise and not task:
                premise = _material_writeoff_premise_text(writeoff)
            remark = (task.description or task.source_cell_value or "").strip() if task else ""
            if not remark and not task and manual_lines:
                remark = manual_lines[index] if index < len(manual_lines) else manual_lines[0]
            ws.append([
                writeoff.writeoff_date.strftime("%d.%m.%Y") if writeoff.writeoff_date else "",
                premise or "—",
                remark,
                "\n".join(material_lines),
            ])
        end_row = ws.max_row
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
    _style_excel_header(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    filename = f"{project.name}_расход_материала_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return _make_excel_response(wb, filename)


@bp.route("/materials/request/new", methods=["GET", "POST"])
@login_required
def material_request_new():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)

    form_state = {
        "title": "",
        "request_date": date.today().isoformat(),
        "rows": [{"name": "", "quantity": "", "unit": ""}],
    }
    invalid_fields: dict[int, set[str]] = {}

    if request.method == "POST":
        posted_names = request.form.getlist("name[]")[:10]
        posted_quantities = request.form.getlist("quantity[]")[:10]
        posted_units = request.form.getlist("unit[]")[:10]
        posted_row_count = max(1, len(posted_names), len(posted_quantities), len(posted_units))
        form_state = {
            "title": request.form.get("title", ""),
            "request_date": request.form.get("request_date", "") or date.today().isoformat(),
            "rows": [
                {
                    "name": posted_names[index] if index < len(posted_names) else "",
                    "quantity": posted_quantities[index] if index < len(posted_quantities) else "",
                    "unit": posted_units[index] if index < len(posted_units) else "",
                }
                for index in range(posted_row_count)
            ],
        }
        active_row_indexes = [
            index
            for index, row in enumerate(form_state["rows"])
            if any(str(row[key]).strip() for key in ("name", "quantity", "unit"))
        ]
        indexes_to_validate = active_row_indexes or [0]
        for index in indexes_to_validate:
            row = form_state["rows"][index]
            missing = set()
            if not str(row["name"]).strip():
                missing.add("name")
            if _parse_quantity(row["quantity"]) is None:
                missing.add("quantity")
            if not str(row["unit"]).strip():
                missing.add("unit")
            if missing:
                invalid_fields[index] = missing

        try:
            rows = _read_material_rows_from_form(limit=10)
        except ValueError as exc:
            flash(str(exc), "danger")
            rows = []
        if invalid_fields:
            rows = []
        elif not rows:
            invalid_fields[0] = {"name", "quantity", "unit"}
            flash("Добавьте хотя бы одну позицию материала", "warning")
        else:
            request_date = parse_date(request.form.get("request_date")) or date.today()
            material_request = MaterialRequest(project_id=project.id, author_id=current_user.id, request_date=request_date, title=(request.form.get("title") or "").strip() or None)
            for row in rows:
                material_request.items.append(
                    MaterialRequestItem(
                        name=str(row["name"]),
                        quantity=float(row["quantity"]),
                        unit=str(row["unit"]),
                    )
                )
            db.session.add(material_request)
            db.session.commit()
            flash("Заявка на материал добавлена", "success")
            return redirect(url_for("main.materials", tab="requests"))

    return render_template(
        "material_request_form.html",
        project=project,
        max_rows=10,
        today=date.today(),
        form_state=form_state,
        invalid_fields=invalid_fields,
    )


@bp.route("/materials/write-off", methods=["GET", "POST"])
@login_required
def material_writeoff_new():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)
    if request.method == "GET":
        redirect_args = request.args.to_dict()
        redirect_args["tab"] = "writeoff"
        return redirect(url_for("main.materials", **redirect_args))

    if request.method == "POST":
        selected_task_ids = []
        for raw_id in request.form.getlist("task_ids"):
            try:
                selected_task_ids.append(int(raw_id))
            except (TypeError, ValueError):
                pass
        selected_tasks = (
            Task.query.options(selectinload(Task.apartment), selectinload(Task.work_point))
            .filter(Task.project_id == project.id, Task.id.in_(selected_task_ids or [-1]))
            .with_for_update()
            .all()
        )
        order_map = {task_id: index for index, task_id in enumerate(selected_task_ids)}
        selected_tasks.sort(key=lambda task: order_map.get(task.id, 10**9))
        already_written_off_ids = {
            task_id
            for (task_id,) in (
                db.session.query(material_writeoff_tasks.c.task_id)
                .filter(material_writeoff_tasks.c.task_id.in_([task.id for task in selected_tasks] or [-1]))
                .all()
            )
        }
        if already_written_off_ids:
            db.session.rollback()
            flash("Одно или несколько выбранных замечаний уже есть в истории списаний. Повторное списание отменено.", "warning")
            return redirect(url_for("main.materials", tab="writeoff"))
        try:
            rows = _read_balance_writeoff_rows(project.id)
        except ValueError as exc:
            flash(str(exc), "danger")
            rows = None
        if not selected_tasks:
            flash("Выберите одно или несколько замечаний", "warning")
            db.session.rollback()
            return redirect(url_for("main.materials", tab="writeoff"))
        elif rows is None:
            db.session.rollback()
            return redirect(url_for("main.materials", tab="writeoff"))
        else:
            writeoff_date = parse_date(request.form.get("writeoff_date")) or date.today()
            if request.form.get("action") == "distribute":
                for row in rows:
                    allocations = _distribute_material_quantity(selected_tasks, float(row["quantity"]))
                    for task in selected_tasks:
                        quantity = allocations.get(task.id)
                        if not quantity:
                            continue
                        writeoff = MaterialWriteOff(project_id=project.id, author_id=current_user.id, writeoff_date=writeoff_date, comment="auto_distributed")
                        writeoff.tasks = [task]
                        writeoff.items.append(
                            MaterialWriteOffItem(
                                name=str(row["name"]),
                                quantity=float(quantity),
                                unit=str(row["unit"]),
                            )
                        )
                        db.session.add(writeoff)
                db.session.commit()
                if False:
                    message = f"Отправлено в синхронизацию: {conflict_count}"
                    if created_count:
                        message += f". Добавлено новых замечаний: {created_count}"
                    if duplicate_count:
                        message += f". Уже были в базе: {duplicate_count}"
                    flash(message, "warning")
                    return redirect(url_for("main.sync_conflicts"))
                flash(f"Материал распределён между замечаниями: {len(selected_tasks)}", "success")
                return redirect(url_for("main.materials", tab="history"))
            writeoff = MaterialWriteOff(project_id=project.id, author_id=current_user.id, writeoff_date=writeoff_date, comment=None)
            writeoff.tasks = selected_tasks
            for row in rows:
                writeoff.items.append(
                    MaterialWriteOffItem(
                        name=str(row["name"]),
                        quantity=float(row["quantity"]),
                        unit=str(row["unit"]),
                    )
                )
            db.session.add(writeoff)
            db.session.commit()
            flash("Материал списан на выбранные замечания", "success")
            return redirect(url_for("main.materials", tab="history"))

    return redirect(url_for("main.materials", tab="writeoff"))


def _project_apartment_options(project_id: int) -> list[Apartment]:
    grouped: dict[str, list[Apartment]] = {}
    for apartment in Apartment.query.filter(Apartment.project_id == project_id).all():
        if _is_visible_apartment_row(apartment):
            grouped.setdefault(_apartment_group_key(apartment), []).append(apartment)
    apartments = [_pick_apartment_representative(group) for group in grouped.values()]
    return sorted(
        apartments,
        key=lambda apartment: (
            apartment.premise_type == "commercial",
            _apartment_number_sort_value(apartment.display_number(fallback_to_id=False)),
            apartment.building or "",
        ),
    )


def _project_work_point_options() -> list[WorkPoint]:
    return (
        WorkPoint.query.filter(WorkPoint.point_number.in_(CONTRACTOR_POINT_LABELS.keys()))
        .order_by(WorkPoint.point_number.asc())
        .all()
    )


def _remark_point_options() -> list[dict[str, str]]:
    return [{"number": number, "label": label} for number, label in CONTRACTOR_POINT_LABELS.items()]


def _get_or_create_manual_work_point(point_number: str | None = None) -> WorkPoint:
    number = str(point_number or "22").strip() or "22"
    label = CONTRACTOR_POINT_LABELS.get(number, "Прочее")
    point = WorkPoint.query.filter_by(point_number=number).order_by(WorkPoint.id.asc()).first()
    if point is None:
        point = WorkPoint(
            point_number=number,
            short_name=label,
            original_column_name=label,
            source_sheet_name="manual",
            is_active=True,
        )
        db.session.add(point)
        db.session.flush()
    return point


def _apartment_group_for_project(apartment: Apartment, project_id: int) -> list[Apartment]:
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project_id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    return group or [apartment]


def _apply_inspection_date_to_group(apartments: list[Apartment], inspection_date: date | None) -> None:
    if not inspection_date:
        return
    formatted = inspection_date.strftime("%d.%m.%Y")
    for item in apartments:
        existing_dates = {value for value in (item.inspection_date, item.first_inspection_date, item.reinspection_date) if value}
        if inspection_date in existing_dates:
            item.first_inspection_present = True
            continue
        if not item.inspection_date and not item.first_inspection_date:
            item.inspection_date = inspection_date
            item.first_inspection_date = inspection_date
            item.first_inspection_present = True
            continue
        if not item.first_inspection_date:
            item.first_inspection_date = inspection_date
            item.first_inspection_present = True
            continue
        if not item.reinspection_date:
            item.reinspection_date = inspection_date
            item.first_inspection_present = True
            continue
        note = str(item.inspection_note or "").strip()
        extra_line = f"Дополнительный осмотр: {formatted}"
        if extra_line not in note:
            item.inspection_note = f"{note}\n{extra_line}".strip()
        item.first_inspection_present = True


def _create_manual_remark_task(
    *,
    project: Project,
    apartment: Apartment,
    point_number: str | None,
    text: str,
    source_sheet_name: str = "manual",
    action: str = "manual_created",
) -> Task:
    work_point = _get_or_create_manual_work_point(point_number)
    source_uid = build_task_uid(
        project.name,
        apartment.construction_number or "",
        apartment.apartment_number or "",
        work_point.point_number,
        work_point.display_name,
        text,
    )
    if Task.query.filter_by(source_uid=source_uid).first():
        source_uid = stable_hash([source_uid, source_sheet_name, datetime.utcnow().isoformat()])
    task = Task(
        source_uid=source_uid,
        project_id=project.id,
        apartment_id=apartment.id,
        work_point_id=work_point.id,
        title=work_point.display_name,
        description=text,
        source_cell_value=text,
        source_sheet_name=source_sheet_name,
        status=STATUS_NOT_STARTED,
        is_done=False,
        completed_date=None,
        manually_edited=True,
        last_seen_at=datetime.utcnow(),
        source_hash=stable_hash([text]),
    )
    db.session.add(task)
    db.session.flush()
    log_change(task, action, None, None, text, user_id=current_user.id)
    return task


def _task_effective_remark_text(task: Task) -> str:
    return (task.description or task.source_cell_value or "").strip()


def _project_pending_conflicts_query(project_id: int):
    return (
        SyncConflict.query.outerjoin(Task, SyncConflict.task_id == Task.id)
        .outerjoin(Apartment, SyncConflict.apartment_id == Apartment.id)
        .filter(SyncConflict.status == "pending")
        .filter(or_(Task.project_id == project_id, Apartment.project_id == project_id))
    )


def _pending_task_sync_conflict(task_id: int) -> SyncConflict | None:
    return (
        SyncConflict.query.filter(
            SyncConflict.task_id == task_id,
            SyncConflict.status == "pending",
            SyncConflict.target_type == "task",
            or_(
                SyncConflict.field_name.is_(None),
                SyncConflict.field_name.in_(("source_cell_value", "description")),
            ),
        )
        .order_by(SyncConflict.id.desc())
        .first()
    )


def _remark_task_candidates(project_id: int, apartment_id: int, work_point_id: int) -> list[Task]:
    return (
        Task.query.filter(
            Task.project_id == project_id,
            Task.apartment_id == apartment_id,
            Task.work_point_id == work_point_id,
            Task.is_archived.is_(False),
        )
        .order_by(Task.is_done.asc(), Task.id.desc())
        .all()
    )


def _find_existing_remark_duplicate_or_conflict(
    *,
    project: Project,
    apartment: Apartment,
    point_number: str | None,
    text: str,
) -> tuple[WorkPoint, Task | None, Task | None]:
    work_point = _get_or_create_manual_work_point(point_number)
    target_text = normalize_text(text)
    duplicate_task = None
    conflict_task = None
    for task in _remark_task_candidates(project.id, apartment.id, work_point.id):
        task_text = normalize_text(_task_effective_remark_text(task))
        pending_conflict = _pending_task_sync_conflict(task.id)
        pending_text = normalize_text(pending_conflict.new_value or "") if pending_conflict else ""
        if target_text and (task_text == target_text or pending_text == target_text):
            duplicate_task = task
            break
        if conflict_task is None and (task_text or pending_conflict is not None):
            conflict_task = task
    return work_point, duplicate_task, conflict_task


def _queue_remark_sync_conflict(
    *,
    task: Task,
    new_text: str,
    source_type: str,
    source_name: str | None = None,
    sheet_name: str | None = None,
    field_label: str | None = None,
    cell_address: str | None = None,
) -> SyncConflict | None:
    old_text = _task_effective_remark_text(task)
    old_hash = task.source_hash or cell_hash(old_text)
    new_hash = cell_hash(new_text)
    if old_hash == new_hash and normalize_text(old_text) == normalize_text(new_text):
        return None
    existing = _pending_task_sync_conflict(task.id)
    if existing is None:
        existing = SyncConflict(
            task_id=task.id,
            target_type="task",
            field_name="description",
            field_label=field_label or "Замечание",
            source_type=source_type,
            source_name=source_name,
            sheet_name=sheet_name,
            cell_address=cell_address,
            old_value=old_text,
            new_value=new_text,
            old_hash=old_hash,
            new_hash=new_hash,
        )
        db.session.add(existing)
        return existing
    existing.field_name = "description"
    existing.field_label = field_label or existing.field_label or "Замечание"
    existing.source_type = source_type
    existing.source_name = source_name
    existing.sheet_name = sheet_name
    existing.cell_address = cell_address
    existing.old_value = old_text
    existing.new_value = new_text
    existing.old_hash = old_hash
    existing.new_hash = new_hash
    return existing


def _work_point_conflict_label(work_point: WorkPoint | None) -> str:
    if work_point is None:
        return "Замечание"
    point_number = str(work_point.point_number or "").strip()
    display_name = str(work_point.display_name or "").strip()
    if point_number and display_name:
        cleaned_name = re.sub(rf"^\s*{re.escape(point_number)}\s*[\.\-:)]?\s*", "", display_name).strip()
        if cleaned_name:
            return f"Пункт {point_number}: {cleaned_name}"
    if point_number:
        return f"Пункт {point_number}"
    return display_name or "Замечание"


def _save_remark_with_sync_fallback(
    *,
    project: Project,
    apartment: Apartment,
    point_number: str | None,
    text: str,
    created_source_sheet_name: str,
    created_action: str,
    conflict_source_type: str,
    conflict_source_name: str | None = None,
    conflict_sheet_name: str | None = None,
) -> str:
    work_point, duplicate_task, conflict_task = _find_existing_remark_duplicate_or_conflict(
        project=project,
        apartment=apartment,
        point_number=point_number,
        text=text,
    )
    if duplicate_task is not None:
        return "duplicate"
    if conflict_task is not None:
        _queue_remark_sync_conflict(
            task=conflict_task,
            new_text=text,
            source_type=conflict_source_type,
            source_name=conflict_source_name,
            sheet_name=conflict_sheet_name or conflict_source_name or created_source_sheet_name,
            field_label=_work_point_conflict_label(work_point),
            cell_address=None,
        )
        return "conflict"
    _create_manual_remark_task(
        project=project,
        apartment=apartment,
        point_number=work_point.point_number,
        text=text,
        source_sheet_name=created_source_sheet_name,
        action=created_action,
    )
    return "created"


def _start_snapshot_sync_log(*, project: Project | None, source_type: str, source_name: str | None = None) -> SyncLog:
    sync_log = SyncLog(
        source_type=(source_type or "").strip() or "manual",
        source_name=((source_name or "").strip()[:255] or None),
        started_at=datetime.utcnow(),
        status="running",
        project_id=project.id if project else None,
    )
    sync_log.rollback_data = build_project_rollback_data(project.id if project else None)
    db.session.add(sync_log)
    db.session.commit()
    return sync_log


def _finish_snapshot_sync_log(
    sync_log: SyncLog,
    *,
    created_count: int = 0,
    updated_count: int = 0,
    missing_count: int = 0,
    status: str = "success",
    error_message: str | None = None,
) -> SyncLog:
    sync_log.created_count = int(created_count or 0)
    sync_log.updated_count = int(updated_count or 0)
    sync_log.missing_count = int(missing_count or 0)
    sync_log.status = (status or "success").strip() or "success"
    sync_log.error_message = ((error_message or "").strip() or None)
    sync_log.finished_at = datetime.utcnow()
    db.session.add(sync_log)
    db.session.commit()
    return sync_log


def _pdf_preview_to_dict(preview, project: Project) -> dict:
    apartment = None
    if preview.apartment_number:
        apartment = (
            Apartment.query.filter(
                Apartment.project_id == project.id,
                Apartment.apartment_number == str(preview.apartment_number).strip(),
            )
            .order_by(Apartment.id.asc())
            .first()
        )
    warnings = list(preview.warnings or [])
    if preview.apartment_number and apartment is None:
        warnings.append(f"Квартира {preview.apartment_number} не найдена в выбранном объекте.")
    return {
        "filename": preview.filename,
        "template_ok": bool(getattr(preview, "template_ok", False)),
        "project_ok": preview.project_ok,
        "project_name": (getattr(preview, "project_name", "") or "").strip(),
        "project_prefix": preview.project_prefix,
        "used_ocr": bool(getattr(preview, "used_ocr", False)),
        "apartment_number": preview.apartment_number or "",
        "apartment_id": apartment.id if apartment else None,
        "inspection_date": preview.inspection_date.isoformat() if preview.inspection_date else "",
        "remarks": [
            {"point_number": remark.point_number, "description": remark.description, "active": remark.active}
            for remark in preview.remarks
        ],
        "warnings": warnings,
    }


def _pdf_previews_from_form(form) -> list[dict]:
    previews: list[dict] = []
    act_count = form.get("act_count", type=int) or 0
    for act_idx in range(act_count):
        row_count = form.get(f"act_{act_idx}_row_count", type=int) or 0
        preview = {
            "filename": (form.get(f"act_{act_idx}_filename") or f"Акт {act_idx + 1}").strip(),
            "template_ok": form.get(f"act_{act_idx}_template_ok") == "1",
            "project_ok": form.get(f"act_{act_idx}_project_ok") == "1",
            "project_name": (form.get(f"act_{act_idx}_project_name") or "").strip(),
            "project_prefix": (form.get(f"act_{act_idx}_project_prefix") or "").strip(),
            "used_ocr": form.get(f"act_{act_idx}_used_ocr") == "1",
            "apartment_number": (form.get(f"act_{act_idx}_apartment_number") or "").strip(),
            "apartment_id": form.get(f"act_{act_idx}_apartment_id", type=int),
            "inspection_date": (form.get(f"act_{act_idx}_inspection_date") or "").strip(),
            "remarks": [],
            "warnings": [],
        }
        warning_text = (form.get(f"act_{act_idx}_warnings") or "").strip()
        if warning_text:
            preview["warnings"] = [item.strip() for item in warning_text.split("\n") if item.strip()]
        for row_idx in range(row_count):
            preview["remarks"].append(
                {
                    "point_number": (form.get(f"act_{act_idx}_row_{row_idx}_point") or "22").strip() or "22",
                    "description": (form.get(f"act_{act_idx}_row_{row_idx}_description") or "").strip(),
                    "active": form.get(f"act_{act_idx}_row_{row_idx}_active") == "1",
                }
            )
        previews.append(preview)
    return previews


@bp.route("/tasks/new", methods=["GET", "POST"])
@login_required
def task_new():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)

    apartments = _project_apartment_options(project.id)
    points = _remark_point_options()
    add_mode = (request.form.get("add_mode") or request.args.get("mode") or "").strip()
    manual_kind = (request.form.get("manual_kind") or request.args.get("kind") or "").strip()

    if request.method == "POST":
        if add_mode != "manual":
            return redirect(url_for("main.task_recognition"))

        apartment_id = request.form.get("apartment_id", type=int)
        apartment = db.session.get(Apartment, apartment_id) if apartment_id else None
        if not apartment or apartment.project_id != project.id:
            flash("Выберите квартиру / коммерцию", "warning")
        elif manual_kind == "act":
            inspection_date = parse_date(request.form.get("inspection_date"))
            created_count = 0
            conflict_count = 0
            duplicate_count = 0
            prepared_entries: list[tuple[str, str]] = []
            for point in points:
                point_number = point["number"]
                text = (request.form.get(f"description_{point_number}") or "").strip()
                if not text or is_no_remark_text(text):
                    continue
                prepared_entries.append((point_number, text))
            if not prepared_entries:
                db.session.rollback()
                flash("Заполните хотя бы одно замечание по пункту", "warning")
            else:
                apartment_label = apartment.full_label() if apartment.premise_type == "commercial" else apartment.label()
                sync_log = _start_snapshot_sync_log(
                    project=project,
                    source_type="manual_act",
                    source_name=f"Ручной акт / {apartment_label}",
                )
                try:
                    all_entries_are_duplicates = True
                    for point_number, text in prepared_entries:
                        _, duplicate_task, _ = _find_existing_remark_duplicate_or_conflict(
                            project=project,
                            apartment=apartment,
                            point_number=point_number,
                            text=text,
                        )
                        if duplicate_task is None:
                            all_entries_are_duplicates = False
                            break
                    if all_entries_are_duplicates:
                        db.session.rollback()
                        _finish_snapshot_sync_log(sync_log)
                        flash("Точно такой же акт уже подгружен", "danger")
                        return render_template(
                            "task_form.html",
                            project=project,
                            apartments=apartments,
                            points=points,
                            add_mode=add_mode,
                            manual_kind=manual_kind,
                        )

                    target_group = _apartment_group_for_project(apartment, project.id)
                    _apply_inspection_date_to_group(target_group, inspection_date)
                    for point_number, text in prepared_entries:
                        outcome = _save_remark_with_sync_fallback(
                            project=project,
                            apartment=apartment,
                            point_number=point_number,
                            text=text,
                            created_source_sheet_name="manual_act",
                            created_action="manual_act_created",
                            conflict_source_type="manual_act",
                            conflict_source_name="Ручной акт",
                            conflict_sheet_name="Ручной акт",
                        )
                        if outcome == "created":
                            created_count += 1
                        elif outcome == "conflict":
                            conflict_count += 1
                        else:
                            duplicate_count += 1

                    if created_count or conflict_count:
                        db.session.commit()
                        _finish_snapshot_sync_log(
                            sync_log,
                            created_count=created_count,
                            missing_count=conflict_count,
                        )
                        if conflict_count:
                            message = f"Отправлено в синхронизацию: {conflict_count}"
                            if created_count:
                                message += f". Добавлено новых замечаний: {created_count}"
                            if duplicate_count:
                                message += f". Уже были в базе: {duplicate_count}"
                            flash(message, "warning")
                            return redirect(url_for("main.sync_conflicts"))
                        message = f"Добавлено замечаний из акта: {created_count}"
                        if duplicate_count:
                            message += f". Уже были в базе: {duplicate_count}"
                        flash(message, "success")
                        return redirect(url_for("main.task_list", status=STATUS_NOT_STARTED))

                    db.session.rollback()
                    _finish_snapshot_sync_log(sync_log)
                    flash("Ничего не сохранено: проверьте заполненные пункты акта", "warning")
                except Exception as exc:
                    db.session.rollback()
                    _finish_snapshot_sync_log(
                        sync_log,
                        created_count=created_count,
                        missing_count=conflict_count,
                        status="error",
                        error_message=str(exc),
                    )
                    raise
        else:
            manual_kind = "single"
            text = (request.form.get("description") or "").strip()
            if not text:
                flash("Введите описание работы", "warning")
            else:
                outcome = _save_remark_with_sync_fallback(
                    project=project,
                    apartment=apartment,
                    point_number="22",
                    text=text,
                    created_source_sheet_name="manual",
                    created_action="manual_created",
                    conflict_source_type="manual",
                    conflict_source_name="Ручное добавление",
                    conflict_sheet_name="Ручное добавление",
                )
                if outcome == "conflict":
                    db.session.commit()
                    if False and conflict_count:
                        message = f"Отправлено в синхронизацию: {conflict_count}"
                        if created_count:
                            message += f". Сохранено новых замечаний: {created_count}"
                        if blocked_count:
                            message += f". Актов пропущено: {blocked_count}"
                        if duplicate_act_names:
                            message += f". Уже подгружены: {', '.join(duplicate_act_names)}"
                        flash(message, "warning")
                        return redirect(url_for("main.sync_conflicts"))
                    flash("Замечание отправлено в синхронизацию: выберите, что оставить, а что заменить", "warning")
                    return redirect(url_for("main.sync_conflicts"))
                if outcome == "duplicate":
                    db.session.rollback()
                    flash("Точно такое же замечание уже добавлено", "warning")
                    return render_template(
                        "task_form.html",
                        project=project,
                        apartments=apartments,
                        points=points,
                        add_mode=add_mode,
                        manual_kind=manual_kind,
                    )
                db.session.commit()
                flash("Замечание добавлено со статусом не выполнено", "success")
                return redirect(url_for("main.task_list", status=STATUS_NOT_STARTED))

    return render_template(
        "task_form.html",
        project=project,
        apartments=apartments,
        points=points,
        add_mode=add_mode,
        manual_kind=manual_kind,
    )


@bp.route("/tasks/recognition", methods=["GET", "POST"])
@login_required
def task_recognition():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)

    apartments = _project_apartment_options(project.id)
    points = _remark_point_options()
    previews: list[dict] = []

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "preview":
            files = [file for file in request.files.getlist("pdf_files") if file and file.filename]
            if not files:
                flash("Загрузите PDF акт", "warning")
            else:
                if len(files) > 3:
                    flash("Можно загрузить максимум 3 акта за раз. Взял первые 3 файла.", "warning")
                for file in files[:3]:
                    try:
                        validate_upload(file, {"pdf"})
                    except ValueError as exc:
                        previews.append({
                            "filename": file.filename or "PDF",
                            "template_ok": False,
                            "project_ok": False,
                            "project_name": "",
                            "project_prefix": "",
                            "apartment_number": "",
                            "apartment_id": None,
                            "inspection_date": "",
                            "remarks": [],
                            "warnings": [str(exc)],
                        })
                        continue
                    preview = recognize_pdf_act(file.stream, file.filename or "PDF", project.name)
                    previews.append(_pdf_preview_to_dict(preview, project))
        elif action == "save":
            previews = _pdf_previews_from_form(request.form)
            if request.form.get("confirm_import") != "1":
                flash("Перед занесением нужно подтвердить сохранение распознанных замечаний", "warning")
            else:
                act_count = request.form.get("act_count", type=int) or 0
                created_count = 0
                conflict_count = 0
                blocked_count = 0
                duplicate_act_names: list[str] = []
                sync_log_source_names: list[str] = []
                for act_idx in range(act_count):
                    act_filename = (request.form.get(f"act_{act_idx}_filename") or f"PDF-акт {act_idx + 1}").strip()
                    if act_filename and act_filename not in sync_log_source_names:
                        sync_log_source_names.append(act_filename)
                sync_log = _start_snapshot_sync_log(
                    project=project,
                    source_type="pdf_recognition",
                    source_name=", ".join(sync_log_source_names) or "PDF-акты",
                )
                try:
                    for act_idx in range(act_count):
                        if request.form.get(f"act_{act_idx}_template_ok") != "1":
                            blocked_count += 1
                            continue
                        if request.form.get(f"act_{act_idx}_project_ok") != "1":
                            blocked_count += 1
                            continue
                        apartment_id = request.form.get(f"act_{act_idx}_apartment_id", type=int)
                        apartment = db.session.get(Apartment, apartment_id) if apartment_id else None
                        if not apartment or apartment.project_id != project.id:
                            blocked_count += 1
                            continue
                        act_filename = (request.form.get(f"act_{act_idx}_filename") or f"PDF-акт {act_idx + 1}").strip()
                        inspection_date = parse_date(request.form.get(f"act_{act_idx}_inspection_date"))
                        row_count = request.form.get(f"act_{act_idx}_row_count", type=int) or 0
                        prepared_rows: list[tuple[str, str]] = []
                        duplicate_row_count = 0
                        for row_idx in range(row_count):
                            if request.form.get(f"act_{act_idx}_row_{row_idx}_active") != "1":
                                continue
                            text = (request.form.get(f"act_{act_idx}_row_{row_idx}_description") or "").strip()
                            point_number = (request.form.get(f"act_{act_idx}_row_{row_idx}_point") or "22").strip()
                            if not text or is_no_remark_text(text):
                                continue
                            prepared_rows.append((point_number, text))
                            _, duplicate_task, _ = _find_existing_remark_duplicate_or_conflict(
                                project=project,
                                apartment=apartment,
                                point_number=point_number,
                                text=text,
                            )
                            if duplicate_task is not None:
                                duplicate_row_count += 1
                        if prepared_rows and duplicate_row_count == len(prepared_rows):
                            duplicate_act_names.append(act_filename)
                            continue
                        if prepared_rows:
                            _apply_inspection_date_to_group(_apartment_group_for_project(apartment, project.id), inspection_date)
                        for point_number, text in prepared_rows:
                            outcome = _save_remark_with_sync_fallback(
                                project=project,
                                apartment=apartment,
                                point_number=point_number,
                                text=text,
                                created_source_sheet_name="pdf_recognition",
                                created_action="pdf_recognition_created",
                                conflict_source_type="pdf_recognition",
                                conflict_source_name=act_filename,
                                conflict_sheet_name=act_filename,
                            )
                            if outcome == "created":
                                created_count += 1
                            elif outcome == "conflict":
                                conflict_count += 1
                    if conflict_count:
                        db.session.commit()
                        _finish_snapshot_sync_log(
                            sync_log,
                            created_count=created_count,
                            missing_count=conflict_count,
                        )
                        if duplicate_act_names:
                            flash("Точно такой же акт уже подгружен", "danger")
                        message = f"Отправлено в синхронизацию: {conflict_count}"
                        if created_count:
                            message += f". Сохранено новых замечаний: {created_count}"
                        if blocked_count:
                            message += f". Актов пропущено: {blocked_count}"
                        flash(message, "warning")
                        return redirect(url_for("main.sync_conflicts"))
                    if created_count or conflict_count:
                        db.session.commit()
                        _finish_snapshot_sync_log(
                            sync_log,
                            created_count=created_count,
                            missing_count=conflict_count,
                        )
                        if duplicate_act_names:
                            flash("Точно такой же акт уже подгружен", "danger")
                        message = f"Сохранено замечаний: {created_count}"
                        if blocked_count:
                            message += f". Актов пропущено: {blocked_count}"
                        flash(message, "success")
                        return redirect(url_for("main.task_recognition"))
                    if duplicate_act_names:
                        db.session.rollback()
                        _finish_snapshot_sync_log(sync_log)
                        flash("Точно такой же акт уже подгружен", "danger")
                        return render_template(
                            "task_recognition.html",
                            project=project,
                            apartments=apartments,
                            points=points,
                            previews=previews,
                        )
                    db.session.rollback()
                    _finish_snapshot_sync_log(sync_log)
                    flash("Ничего не сохранено: проверьте ЖК, квартиру и выбранные строки", "warning")
                except Exception as exc:
                    db.session.rollback()
                    _finish_snapshot_sync_log(
                        sync_log,
                        created_count=created_count,
                        missing_count=conflict_count,
                        status="error",
                        error_message=str(exc),
                    )
                    raise

    return render_template(
        "task_recognition.html",
        project=project,
        apartments=apartments,
        points=points,
        previews=previews,
    )


@bp.route("/materials/task/new", methods=["GET", "POST"])
@login_required
def material_manual_task_new():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if not _can_edit_materials():
        abort(403)

    balance_options = _balance_options(project.id)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"

    def fail(message: str, status_code: int = 400):
        if wants_json:
            return jsonify({"ok": False, "message": message}), status_code
        flash(message, "danger")
        return None

    if request.method == "POST":
        task_name = (request.form.get("task_name") or "").strip()
        premise_text = (request.form.get("premise_text") or "").strip()
        try:
            rows = _read_balance_writeoff_rows(project.id)
        except ValueError as exc:
            failed_response = fail(str(exc))
            if failed_response is not None:
                return failed_response
            rows = None
        if not task_name:
            failed_response = fail("Введите наименование")
            if failed_response is not None:
                return failed_response
        elif rows is None:
            pass
        else:
            writeoff = MaterialWriteOff(
                project_id=project.id,
                author_id=current_user.id,
                writeoff_date=parse_date(request.form.get("writeoff_date")) or date.today(),
                comment=_build_manual_writeoff_comment(task_name, premise_text),
            )
            for row in rows:
                writeoff.items.append(MaterialWriteOffItem(name=str(row["name"]), quantity=float(row["quantity"]), unit=str(row["unit"])))
            db.session.add(writeoff)
            db.session.commit()
            message = "Ручное списание добавлено"
            if wants_json:
                return jsonify({"ok": True, "message": message, "writeoff_id": writeoff.id})
            flash(message, "success")
            return redirect(url_for("main.materials", tab="task"))

    return render_template("material_task_form.html", project=project, balance_options=balance_options, today=date.today())


def _task_list_response(contractor_page: bool = False):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    ensure_default_categories()
    category_id = request.args.get("category_id", type=int)
    categories = WorkCategory.query.filter_by(is_active=True).order_by(WorkCategory.sort_order.asc()).all()
    all_cat = next((category for category in categories if (category.name or "").strip().lower() == "все"), None)
    if contractor_page:
        category_id = all_cat.id if all_cat else None
    elif not category_id and categories:
        category_id = all_cat.id if all_cat else categories[0].id
    section_id = request.args.get("section_id", type=int)
    if section_id and not contractor_page:
        category_id = section_id

    query_args = request.args.copy()
    contractor_options = []
    selected_contractor = None
    if contractor_page:
        # В разделе "Подрядчики" показываем ту же таблицу замечаний, но группируем/фильтруем по пунктам 10-22.
        # Не принудительно фильтруем по статусу "Подрядчик", иначе вкладка пустая до ручной разметки задач.
        query_args["sort"] = "apartment"
        contractor_options = (
            Contractor.query
            .options(selectinload(Contractor.work_points), selectinload(Contractor.apartments))
            .filter_by(project_id=project.id)
            .order_by(func.lower(Contractor.name).asc(), Contractor.id.asc())
            .all()
        )
        selected_contractor = next(
            (contractor for contractor in contractor_options if contractor.id == request.args.get("contractor_id", type=int)),
            None,
        )
    query = build_task_query(query_args, category_id=category_id, project_id=project.id)
    if contractor_page:
        query = _filter_tasks_for_contractor(query, selected_contractor)
    if current_user.role in WORKER_ROLES:
        query = query.filter(Task.responsible_id == current_user.id)
    # Список помещений для выбора Excel строим по всему отфильтрованному набору,
    # а не только по текущей странице пагинации.
    is_mobile_request = _is_mobile_phone_request()
    if is_mobile_request:
        premise_options = []
    else:
        premise_ids = [
            apartment_id
            for (apartment_id,) in (
                query.order_by(None)
                .with_entities(Task.apartment_id)
                .filter(Task.apartment_id.isnot(None))
                .distinct()
                .all()
            )
        ]
        premise_apartments = (
            Apartment.query.filter(Apartment.id.in_(premise_ids)).all()
            if premise_ids
            else []
        )
        premise_options = _premise_options_from_apartments(premise_apartments)

    page = request.args.get("page", 1, type=int)
    per_page = 10 if is_mobile_request else 20
    pagination = query.options(selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items)).paginate(page=page, per_page=per_page, error_out=False)
    active_category = next((category for category in categories if category.id == category_id), None)

    prev_args = request.args.to_dict(flat=False)
    next_args = request.args.to_dict(flat=False)
    if pagination.has_prev:
        prev_args["page"] = pagination.prev_num
    if pagination.has_next:
        next_args["page"] = pagination.next_num
    users = _active_users_for_project(project.id)
    points = WorkPoint.query.filter_by(is_active=True).order_by(WorkPoint.point_number.asc()).all()
    finishing_types = [
        x[0]
        for x in db.session.query(distinct(Apartment.finishing_type))
        .filter(Apartment.project_id == project.id, Apartment.finishing_type.isnot(None))
        .all()
    ]
    excel_export_args = {
        key: value
        for key, value in request.args.to_dict(flat=False).items()
        if key not in {"page", "task_ids", "premise_ids"}
    }
    if not contractor_page and category_id:
        excel_export_args["category_id"] = category_id
    premise_storage_key = _excel_premise_storage_key(
        "contractors-premises" if contractor_page else f"remarks-{category_id}-premises",
        project.id,
        excel_export_args,
    )
    return render_template(
        "task_list.html",
        tasks=pagination.items,
        pagination=pagination,
        categories=categories,
        active_category_id=category_id,
        users=users,
        points=points,
        finishing_types=finishing_types,
        args=request.args,
        active_category=active_category,
        contractor_page=contractor_page,
        list_endpoint="main.contractors_list" if contractor_page else "main.task_list",
        contractor_options=contractor_options,
        selected_contractor=selected_contractor,
        contractor_excel_selection_url=(url_for("main.contractors_excel_selection", **request.args.to_dict(flat=False)) if contractor_page else ""),
        remarks_excel_selection_url=(url_for("main.remarks_excel_selection", category_id=category_id, **{key: value for key, value in request.args.to_dict(flat=False).items() if key not in {"category_id", "page", "task_ids", "premise_ids"}}) if (not contractor_page and category_id) else ""),
        category_filter_args={key: value for key, value in request.args.to_dict(flat=False).items() if key not in {"category_id", "section_id", "page"}},
        premise_options=premise_options,
        premise_storage_key=premise_storage_key,
        page_title="Подрядчики" if contractor_page else "Замечания",
        page_subtitle="Раздел для работы с подрядчиками объекта" if contractor_page else "Работа с замечаниями по выбранному объекту.",
        today=date.today(),
        prev_args=prev_args,
        next_args=next_args,
    )


def _apartment_number_sort_value(value: str | None):
    text = str(value or "").strip()
    if not text:
        return (1, 0, "")
    if text.isdigit():
        return (0, int(text), text)
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        return (0, int(digits), text)
    return (1, 0, text.lower())


def _task_apartment_sort_value(task: Task):
    apartment = getattr(task, "apartment", None)
    done_rank = 1 if getattr(task, "is_done", False) else 0
    if not apartment:
        return (done_rank, 1, 1, (1, 0, ""), "", "", 0)
    return (
        done_rank,
        0,
        0 if (apartment.premise_type or "apartment") != "commercial" else 1,
        _apartment_number_sort_value(apartment.apartment_number or apartment.construction_number),
        str(apartment.building or "").strip().lower(),
        str(task.work_point.point_number if task.work_point else "").strip(),
        int(task.id or 0),
    )


def _task_apartment_sort_value_no_done(task: Task):
    apartment = getattr(task, "apartment", None)
    if not apartment:
        return (1, 1, (1, 0, ""), "", "", 0)
    return (
        0,
        0 if (apartment.premise_type or "apartment") != "commercial" else 1,
        _apartment_number_sort_value(apartment.apartment_number or apartment.construction_number),
        str(apartment.building or "").strip().lower(),
        str(task.work_point.point_number if task.work_point else "").strip(),
        int(task.id or 0),
    )


def _writeoff_sort_value(writeoff: MaterialWriteOff):
    tasks = list(writeoff.tasks or [])
    if not tasks:
        return (1, (1, 0, ""), 0, 0)
    best_task = min(tasks, key=_task_apartment_sort_value)
    return (
        0,
        _task_apartment_sort_value(best_task),
        -int(writeoff.writeoff_date.toordinal() if writeoff.writeoff_date else 0),
        -int(writeoff.id or 0),
    )


def _material_writeoff_manual_comment(writeoff: MaterialWriteOff) -> str:
    comment = (writeoff.comment or "").strip()
    if not comment:
        return ""
    if comment == "auto_distributed":
        return ""
    if comment.startswith(f"{MEASUREMENT_WRITEOFF_COMMENT_PREFIX}:"):
        return ""
    payload = _manual_writeoff_comment_payload(comment)
    if payload:
        return str(payload.get("text") or "").strip()
    return comment


def _manual_writeoff_comment_payload(comment: str | None) -> dict[str, str] | None:
    text = (comment or "").strip()
    if not text.startswith(MANUAL_WRITEOFF_COMMENT_PREFIX):
        return None
    try:
        payload = json.loads(text[len(MANUAL_WRITEOFF_COMMENT_PREFIX):])
    except (TypeError, ValueError):
        return None
    if not isinstance(payload, dict):
        return None
    return {str(key): str(value or "").strip() for key, value in payload.items()}


def _build_manual_writeoff_comment(task_name: str, premise: str) -> str:
    task_name = str(task_name or "").strip()
    premise = " ".join(str(premise or "").split())
    if not premise:
        return task_name
    return MANUAL_WRITEOFF_COMMENT_PREFIX + json.dumps(
        {"text": task_name, "premise": premise},
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _material_writeoff_manual_premise(writeoff: MaterialWriteOff) -> str:
    payload = _manual_writeoff_comment_payload(writeoff.comment)
    return str(payload.get("premise") or "").strip() if payload else ""


def _material_writeoff_premise_text(writeoff: MaterialWriteOff) -> str:
    labels: list[str] = []
    for task in list(writeoff.tasks or []):
        if task.apartment:
            label = task.apartment.label()
            if label and label not in labels:
                labels.append(label)
    if labels:
        return ", ".join(labels)
    manual_premise = _material_writeoff_manual_premise(writeoff)
    if manual_premise:
        return manual_premise
    return "Ручное списание" if _material_writeoff_manual_comment(writeoff) else "—"


def _material_writeoff_remark_lines(writeoff: MaterialWriteOff) -> list[str]:
    lines = [
        (task.description or task.source_cell_value or "").strip()
        for task in list(writeoff.tasks or [])
        if (task.description or task.source_cell_value or "").strip()
    ]
    if lines:
        return lines
    manual_comment = _material_writeoff_manual_comment(writeoff)
    return [manual_comment] if manual_comment else []


def _material_writeoff_history_view(writeoff: MaterialWriteOff):
    manual_comment = _material_writeoff_manual_comment(writeoff)
    if list(writeoff.tasks or []) or not manual_comment:
        return writeoff
    premise = _material_writeoff_manual_premise(writeoff) or "Ручное списание"
    pseudo_apartment = SimpleNamespace(label=lambda: premise)
    pseudo_task = SimpleNamespace(
        apartment=pseudo_apartment,
        description=manual_comment,
        source_cell_value=manual_comment,
    )
    return SimpleNamespace(
        id=writeoff.id,
        writeoff_date=writeoff.writeoff_date,
        items=writeoff.items,
        tasks=[pseudo_task],
        comment=writeoff.comment,
    )


def _apartment_identity_text(apartment: Apartment) -> str:
    return apartment.display_number(fallback_to_id=False)


def _clean_apartment_key(text: str) -> str:
    return " ".join(str(text or "").strip().lower().replace("ё", "е").split())


def _is_service_apartment_row(apartment: Apartment) -> bool:
    """Отсекает строки Excel вроде '1 корпус' и '1 подъезд', чтобы они не выглядели как квартиры."""
    text = _clean_apartment_key(_apartment_identity_text(apartment))
    if not text:
        return False
    service_words = ("корпус", "подъезд", "очеред", "секц", "итог", "дом")
    return any(word in text for word in service_words)


def _is_orphan_bogus_apartment_row(apartment: Apartment) -> bool:
    if (apartment.premise_type or "apartment") != "apartment":
        return False
    apartment_number = str(apartment.apartment_number or "").strip()
    construction_number = str(apartment.construction_number or "").strip()
    if not apartment_number or apartment_number != construction_number:
        return False
    if len(apartment_number) <= 24:
        return False
    return not list(apartment.tasks or [])


def _is_visible_apartment_row(apartment: Apartment) -> bool:
    text = _apartment_identity_text(apartment)
    return bool(text) and not _is_service_apartment_row(apartment) and not _is_orphan_bogus_apartment_row(apartment)


def _apartment_group_key(apartment: Apartment) -> str:
    text = _clean_apartment_key(_apartment_identity_text(apartment))
    premise_type = apartment.premise_type or "apartment"
    if premise_type == "commercial":
        building = str(apartment.building or "").strip()
        source = str(apartment.source_row_id or "").strip()
        number = text or str(apartment.id)
        return f"{premise_type}:building:{building}:num:{number}:src:{source}"
    if not text:
        return f"{premise_type}:id:{apartment.id}"
    if text.isdigit():
        return f"{premise_type}:num:{int(text)}"
    return f"{premise_type}:num:{text}"


def _pick_apartment_representative(apartments: list[Apartment]) -> Apartment:
    def score(apartment: Apartment):
        return (
            1 if apartment.finishing_type else 0,
            1 if apartment.app_deadline_date else 0,
            1 if apartment.deadline_date else 0,
            len(apartment.tasks or []),
            -(apartment.id or 0),
        )

    return sorted(apartments, key=score, reverse=True)[0]


def _apartment_group_mode(apartments: list[Apartment]) -> str:
    if apartments and any(_is_unsold_apartment(apartment) for apartment in apartments):
        return "не продана"
    if any(apartment.is_app_mode for apartment in apartments):
        return "АПП"
    return "не принята"


def _is_unsold_apartment(apartment: Apartment) -> bool:
    return is_apartment_unsold(apartment)


def _apartment_mode(apartment: Apartment) -> str:
    if _is_unsold_apartment(apartment):
        return "не продана"
    return "АПП" if apartment.is_app_mode else "не принята"


def _task_point_number(task: Task) -> str:
    return str(task.work_point.point_number if task.work_point else "").strip()


def _task_point_name(task: Task) -> str:
    if not task.work_point:
        return ""
    return (task.work_point.display_name or "").lower()


def _is_wall_task(task: Task) -> bool:
    return _task_point_number(task) in WALL_POINT_NUMBERS or "стен" in _task_point_name(task)


def _auto_po_status(tasks: list[Task]) -> str:
    main_tasks = [
        task
        for task in tasks
        if not task.is_archived
        and not task.is_missing_in_latest_sync
        and _task_point_number(task) in MAIN_WORK_POINT_NUMBERS
    ]
    wall_tasks = [task for task in main_tasks if _is_wall_task(task)]
    walls_done = bool(wall_tasks) and all(task.is_done for task in wall_tasks)
    more_than_half_done = bool(main_tasks) and sum(1 for task in main_tasks if task.is_done) > len(main_tasks) / 2
    return PO_STATUS_TO_THROW if walls_done or more_than_half_done else PO_STATUS_NOT_READY


def _po_status_for_group(apartments: list[Apartment], tasks: list[Task]) -> str | None:
    if _apartment_group_mode(apartments) == "АПП":
        return None
    manual = next((apartment.po_status for apartment in apartments if apartment.po_status_manual and apartment.po_status in PO_STATUS_LABELS), None)
    return manual or _auto_po_status(tasks)


def _apartment_inspection_comment(apartments: list[Apartment]) -> str | None:
    for apartment in apartments:
        note = str(apartment.inspection_note or "").strip()
        if note and not note.startswith("__inspection_schedule__:"):
            return note
    return None


def _apartment_manual_comment(apartments: list[Apartment]) -> str | None:
    for apartment in apartments:
        if apartment.comment:
            return apartment.comment
    return None


def _apartment_history_anchor_task(apartments: list[Apartment]) -> Task | None:
    tasks = [task for apartment in apartments for task in list(apartment.tasks or [])]
    if not tasks:
        apartment = _pick_apartment_representative(apartments) if apartments else None
        if apartment is None:
            return None
        work_point = WorkPoint.query.filter_by(point_number="history", source_sheet_name="apartment_history").first()
        if work_point is None:
            work_point = WorkPoint(
                point_number="history",
                short_name="История помещения",
                original_column_name="История помещения",
                source_sheet_name="apartment_history",
                is_active=False,
            )
            db.session.add(work_point)
            db.session.flush()
        source_uid = stable_hash(["apartment-history", apartment.project_id, apartment.id])
        task = Task(
            source_uid=source_uid,
            project_id=apartment.project_id,
            apartment_id=apartment.id,
            work_point_id=work_point.id,
            title="История помещения",
            description="Служебная запись истории помещения",
            source_cell_value="Служебная запись истории помещения",
            source_sheet_name="apartment_history",
            status=STATUS_NOT_STARTED,
            is_done=False,
            is_archived=True,
            manually_edited=True,
            last_seen_at=datetime.utcnow(),
            source_hash=stable_hash(["apartment-history-anchor"]),
        )
        db.session.add(task)
        db.session.flush()
        return task
    return sorted(
        tasks,
        key=lambda task: (
            0 if (task.source_sheet_name or "") == "apartment_history" else 1,
            1 if task.is_archived or task.is_missing_in_latest_sync else 0,
            task.created_at or datetime.min,
            task.id or 0,
        ),
    )[0]


def _log_apartment_field_change(apartments: list[Apartment], field_name: str, old_value: object, new_value: object) -> tuple[ChangeLog, Task] | None:
    old_text = str(old_value or "").strip()
    new_text = str(new_value or "").strip()
    if old_text == new_text:
        return None
    anchor_task = _apartment_history_anchor_task(apartments)
    if anchor_task is None:
        return None
    entry = log_change(anchor_task, "apartment_field_update", field_name, old_text, new_text, user_id=current_user.id)
    return entry, anchor_task


def _parse_inspection_schedule_marker(note: str | None) -> date | datetime | None:
    text = str(note or "").strip()
    prefix = "__inspection_schedule__:"
    if not text.startswith(prefix):
        return None
    payload = text[len(prefix):].strip()
    if not payload:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(payload, fmt)
            if fmt == "%Y-%m-%d":
                return parsed.date()
            return parsed
        except ValueError:
            continue
    return None


def _apartment_inspection_value(apartment: Apartment) -> date | datetime | None:
    scheduled = _parse_inspection_schedule_marker(apartment.inspection_note)
    if scheduled is not None:
        return scheduled
    return apartment.inspection_date or apartment.first_inspection_date


def _inspection_sort_datetime(value: date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.combine(value, datetime.max.time().replace(microsecond=0))


def _format_inspection_display_value(value: date | datetime | None) -> str:
    if value is None:
        return "—"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    return format_ru_date(value)


def _apartment_inspection_date(apartments: list[Apartment]) -> date | None:
    dates: list[date] = []
    for apartment in apartments:
        value = _apartment_inspection_value(apartment)
        if isinstance(value, datetime):
            dates.append(value.date())
        elif isinstance(value, date):
            dates.append(value)
    return min(dates) if dates else None


def _apartment_inspection_display(apartments: list[Apartment]) -> str:
    values = [_apartment_inspection_value(apartment) for apartment in apartments]
    values = [value for value in values if value is not None]
    if not values:
        return "—"
    chosen = min(values, key=lambda value: _inspection_sort_datetime(value) or datetime.max)
    return _format_inspection_display_value(chosen)


def _apartment_inspection_status_class(status: str | None) -> str:
    if status == "Будет":
        return "status-pill-warning"
    if status == "Был":
        return "status-pill-danger"
    return "status-pill-muted"


def _apartment_inspection_status(apartments: list[Apartment]) -> str | None:
    # Для непроданных квартир статус осмотра фиксируем как «Не был».
    # В интерфейсе для них нельзя переключить это состояние вручную.
    if apartments and any(_is_unsold_apartment(apartment) for apartment in apartments):
        return "Не был"
    if any(apartment.first_inspection_present for apartment in apartments):
        return "Был"
    return "Не был"


def _group_remark_deadline(apartments: list[Apartment]) -> date | None:
    deadlines = [apartment.effective_app_deadline_date() for apartment in apartments]
    deadlines = [deadline for deadline in deadlines if deadline]
    return min(deadlines) if deadlines else None


def _deadline_status(deadline: date | None, stored_status: str | None = None) -> dict | None:
    # «Без замечаний» и пустой срок показываем как «Нет срока», без тревожной плашки.
    if not deadline:
        return None
    days_left = (deadline - date.today()).days
    if days_left < 0:
        return {"label": "Срок истёк", "class": "expired", "days_left": days_left}
    if days_left <= 15:
        return {"label": "Истекает", "class": "expiring", "days_left": days_left}
    return None


def _remark_deadline_status_for_group(apartments: list[Apartment], deadline: date | None, stored_status: str | None = None) -> dict | None:
    status = _deadline_status(deadline, stored_status)
    if status and _group_avr_status(apartments) == AVR_STATUS_SIGNED:
        return {"label": "Устранили", "class": "resolved", "days_left": status.get("days_left")}
    return status


def _app_deadline_display(apartment: Apartment) -> str:
    return apartment.app_deadline_label()


def _group_app_deadline_display(apartments: list[Apartment]) -> str:
    deadline = _group_remark_deadline(apartments)
    if deadline:
        return deadline.strftime("%d.%m.%Y")
    for apartment in apartments:
        raw = str(getattr(apartment, "app_deadline_raw", None) or "").strip()
        if raw and not Apartment._is_no_deadline_text(raw):
            return raw
    return "Нет срока"


def _group_avr_status(apartments: list[Apartment]) -> str:
    if any(apartment.avr_status == AVR_STATUS_SIGNED for apartment in apartments):
        return AVR_STATUS_SIGNED
    return AVR_STATUS_NEEDED


def _group_avr_signed_date(apartments: list[Apartment]) -> date | None:
    dates = [apartment.avr_signed_date for apartment in apartments if apartment.avr_signed_date]
    return min(dates) if dates else None


def _group_app_deadline_status(apartments: list[Apartment]) -> str | None:
    deadline = _group_remark_deadline(apartments)
    if not deadline:
        return None
    days_left = (deadline - date.today()).days
    if days_left < 0:
        return APP_DEADLINE_EXPIRED
    if days_left <= 15:
        return APP_DEADLINE_EXPIRING
    return None


def _apartment_inspection_status(apartments: list[Apartment]) -> str | None:
    if apartments and any(_is_unsold_apartment(apartment) for apartment in apartments):
        return "Был"
    values = [_apartment_inspection_value(apartment) for apartment in apartments]
    values = [value for value in values if value is not None]
    if values:
        nearest = min(values, key=lambda value: _inspection_sort_datetime(value) or datetime.max)
        nearest_dt = _inspection_sort_datetime(nearest)
        if nearest_dt is not None:
            return "Будет" if nearest_dt >= datetime.now() else "Был"
    if any(apartment.first_inspection_present for apartment in apartments):
        return "Был"
    return "Не был"


def _is_app_inspection_locked(apartments: list[Apartment]) -> bool:
    return _apartment_group_mode(apartments) == "АПП" and _apartment_inspection_date(apartments) is not None


def _premise_label(apartment: Apartment) -> str:
    return "Комм" if (apartment.premise_type or "apartment") == "commercial" else "кв"


def _floor_from_construction_number(value: str | None) -> str:
    parts = [part.strip() for part in str(value or "").split("-")]
    return parts[1] if len(parts) >= 2 and parts[1] else ""


def _avr_floor(apartment: Apartment | None) -> str:
    if apartment is None:
        return ""
    return _floor_from_construction_number(apartment.construction_number) or (apartment.floor or "")


def _avr_owner_options(owner_name: str | None) -> list[str]:
    text = str(owner_name or "").strip()
    if not text:
        return []
    parts = [part.strip(" \t\r\n") for part in re.split(r"[\r\n;]+|\s+/\s+", text) if part.strip()]
    seen = set()
    options = []
    for part in parts:
        key = " ".join(part.lower().split())
        if key and key not in seen:
            seen.add(key)
            options.append(part)
    return options or [text]


def _contractor_point_options() -> list[dict[str, str]]:
    points = (
        WorkPoint.query.filter(WorkPoint.point_number.in_(CONTRACTOR_POINT_LABELS.keys()))
        .order_by(WorkPoint.point_number.asc())
        .all()
    )
    existing_numbers = {str(point.point_number).strip() for point in points}
    options = []
    for number, label in CONTRACTOR_POINT_LABELS.items():
        if number in existing_numbers:
            options.append({"number": number, "label": label})
    return options


def _contractor_apartment_options(project_id: int) -> list[dict]:
    apartments = Apartment.query.filter(Apartment.project_id == project_id).all()
    groups: dict[str, list[Apartment]] = {}
    for apartment in apartments:
        if not _is_visible_apartment_row(apartment):
            continue
        groups.setdefault(_apartment_group_key(apartment), []).append(apartment)

    options = []
    for group in groups.values():
        representative = _pick_apartment_representative(group)
        options.append({
            "id": representative.id,
            "apartment_ids": sorted(apartment.id for apartment in group),
            "label": representative.full_label(),
            "building": str(representative.building or "").strip(),
            "entrance": str(representative.entrance or "").strip(),
            "floor": str(representative.floor or "").strip(),
            "finishing_type": str(representative.finishing_type or "").strip(),
            "number": re.sub(r"\D+", "", str(representative.apartment_number or representative.construction_number or "").strip()),
            "_sort_key": (
                0 if (representative.premise_type or "apartment") == "apartment" else 1,
                _apartment_number_sort_value(representative.apartment_number or representative.construction_number),
                str(representative.building or "").strip().lower(),
            ),
        })
    options.sort(key=lambda option: option["_sort_key"])
    return options


def _group_project_apartments(project_id: int, include_activity: bool = True) -> list[list[Apartment]]:
    load_options = [
        selectinload(Apartment.tasks).selectinload(Task.work_point),
        selectinload(Apartment.tasks).selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items),
    ]
    if include_activity:
        load_options.extend([
            selectinload(Apartment.tasks).selectinload(Task.comments),
            selectinload(Apartment.tasks).selectinload(Task.changes),
        ])
    apartments = (
        Apartment.query.options(*load_options)
        .filter(Apartment.project_id == project_id)
        .all()
    )
    groups: dict[str, list[Apartment]] = {}
    for apartment in apartments:
        if not _is_visible_apartment_row(apartment):
            continue
        groups.setdefault(_apartment_group_key(apartment), []).append(apartment)
    return list(groups.values())


def _build_apartment_overview(apartment_or_group, include_activity: bool = True) -> dict:
    apartments = apartment_or_group if isinstance(apartment_or_group, list) else [apartment_or_group]
    apartment = _pick_apartment_representative(apartments)
    all_tasks = sorted(
        [task for item in apartments for task in list(item.tasks or [])],
        key=lambda task: (
            1 if task.is_done else 0,
            int(task.work_point.point_number) if task.work_point and str(task.work_point.point_number).isdigit() else 9999,
            task.created_at or datetime.min,
            task.id,
        ),
    )
    tasks = [task for task in all_tasks if (task.source_sheet_name or "") != "apartment_history"]
    active_tasks = [task for task in tasks if not task.is_archived and not task.is_missing_in_latest_sync]
    done_tasks = [task for task in active_tasks if task.is_done]
    left_tasks = [task for task in active_tasks if not task.is_done]
    problem_tasks = [task for task in active_tasks if task.status == "problem"]
    replaced_glass = [task.glass_measurement for task in tasks if task.glass_measurement and task.glass_measurement.status == GLASS_STATUS_REPLACED]
    ordered_glass = [task.glass_measurement for task in tasks if task.glass_measurement and task.glass_measurement.status == GLASS_STATUS_ORDERED]
    comments = []
    changes = []
    if include_activity:
        for task in tasks:
            for comment in task.comments:
                comments.append({"task": task, "comment": comment})
        for task in all_tasks:
            for change in task.changes:
                changes.append({"task": task, "change": change})
        comments.sort(key=lambda row: row["comment"].created_at, reverse=True)
        changes.sort(key=lambda row: row["change"].created_at, reverse=True)
    total = len(active_tasks)
    done = len(done_tasks)
    left = len(left_tasks)
    mode = _apartment_group_mode(apartments)
    percent = round((done / total * 100), 1) if total else (100 if mode == "АПП" else 0)
    remark_deadline = _group_remark_deadline(apartments)
    app_deadline_status = _group_app_deadline_status(apartments)
    return {
        "apartment": apartment,
        "premise_label": _premise_label(apartment),
        "apartments": apartments,
        "group_count": len(apartments),
        "mode": mode,
        "inspection_comment": _apartment_inspection_comment(apartments),
        "manual_comment": _apartment_manual_comment(apartments),
        "inspection_date": _apartment_inspection_date(apartments),
        "inspection_display": _apartment_inspection_display(apartments),
        "inspection_status": _apartment_inspection_status(apartments),
        "inspection_status_class": _apartment_inspection_status_class(_apartment_inspection_status(apartments)),
        "tasks": tasks,
        "active_tasks": active_tasks,
        "done_tasks": done_tasks,
        "left_tasks": left_tasks,
        "problem_tasks": problem_tasks,
        "comments": comments,
        "changes": changes,
        "total": total,
        "done": done,
        "left": left,
        "problem": len(problem_tasks),
        "percent": percent,
        "remark_deadline": remark_deadline,
        "remark_deadline_display": _group_app_deadline_display(apartments),
        "remark_deadline_status": _remark_deadline_status_for_group(apartments, remark_deadline, app_deadline_status),
        "avr_status": _group_avr_status(apartments),
        "avr_signed_date": _group_avr_signed_date(apartments),
        "show_avr": mode == "АПП" and (apartment.premise_type or "apartment") == "apartment",
        "po_status": _po_status_for_group(apartments, active_tasks),
        "has_ordered_glass": bool(ordered_glass),
        "has_replaced_glass": bool(replaced_glass),
        "glass_status_label": glass_measurement_action_label(replaced_glass[0] if replaced_glass else (ordered_glass[0] if ordered_glass else None)),
    }


def _apartment_live_stats_for_task(task: Task) -> dict[str, int | float] | None:
    if not task.apartment_id:
        return None
    apartment = db.session.get(Apartment, task.apartment_id)
    if not apartment:
        return None
    group_key = _apartment_group_key(apartment)
    apartments = (
        Apartment.query.options(
            selectinload(Apartment.tasks).selectinload(Task.work_point),
            selectinload(Apartment.tasks).selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items),
        )
        .filter(Apartment.project_id == task.project_id)
        .all()
    )
    group = [
        item
        for item in apartments
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ] or [apartment]
    overview = _build_apartment_overview(group, include_activity=False)
    return {
        "total": overview["total"],
        "done": overview["done"],
        "left": overview["left"],
        "problem": overview["problem"],
        "percent": overview["percent"],
    }


@bp.route("/apartments")
@login_required
def apartments():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    is_mobile_request = _is_mobile_phone_request()
    has_active_filters = any(
        str(value or "").strip()
        for key, value in request.args.items()
        if key != "page"
    )
    per_page = 10 if is_mobile_request else 20
    mobile_page = 1
    mobile_total_pages = 1
    mobile_pagination_args = request.args.to_dict(flat=True)
    mobile_pagination_args.pop("page", None)
    try:
        mobile_page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        mobile_page = 1

    if not has_active_filters:
        groups = _group_project_apartments(project.id, include_activity=False)
        groups.sort(key=lambda group: _apartment_number_sort_value(
            _pick_apartment_representative(group).display_number(fallback_to_id=False)
        ))
        total_count = len(groups)
        mobile_total_pages = max(1, (total_count + per_page - 1) // per_page)
        mobile_page = min(mobile_page, mobile_total_pages)
        start = (mobile_page - 1) * per_page
        visible_groups = groups[start:start + per_page]
        rows = [_build_apartment_overview(group, include_activity=False) for group in visible_groups]
        premise_selectors = []
        po_only = False
        po_alert_count = 0
        for group in groups:
            tasks = [
                task
                for apartment in group
                for task in list(apartment.tasks or [])
                if (task.source_sheet_name or "") != "apartment_history"
                and not task.is_archived
                and not task.is_missing_in_latest_sync
            ]
            if _po_status_for_group(group, tasks) == PO_STATUS_TO_THROW:
                po_alert_count += 1
    else:
        all_rows = [
            _build_apartment_overview(group, include_activity=False)
            for group in _group_project_apartments(project.id, include_activity=False)
        ]
        rows, premise_selectors, po_only = _filtered_apartment_overview_rows(
            project.id, request.args, source_rows=all_rows
        )
        total_count = len(rows)
        mobile_total_pages = max(1, (total_count + per_page - 1) // per_page)
        mobile_page = min(mobile_page, mobile_total_pages)
        start = (mobile_page - 1) * per_page
        rows = rows[start:start + per_page]
        po_alert_count = sum(1 for row in all_rows if row.get("po_status") == PO_STATUS_TO_THROW)
    finishing_types = [
        x[0]
        for x in db.session.query(distinct(Apartment.finishing_type))
        .filter(Apartment.project_id == project.id, Apartment.finishing_type.isnot(None), Apartment.finishing_type != "")
        .order_by(Apartment.finishing_type.asc())
        .all()
    ]
    return render_template(
        "apartments.html",
        rows=rows,
        args=request.args,
        export_args=request.args.to_dict(flat=True),
        finishing_types=finishing_types,
        total_count=total_count,
        mobile_page=mobile_page,
        mobile_total_pages=mobile_total_pages,
        mobile_pagination_args=mobile_pagination_args,
        po_only=po_only,
        po_alert_count=po_alert_count,
        po_status_labels=PO_STATUS_LABELS,
        po_status_classes=PO_STATUS_CLASSES,
        avr_status_needed=AVR_STATUS_NEEDED,
        avr_status_signed=AVR_STATUS_SIGNED,
    )


def _apartment_overview_search_haystack(row: dict) -> str:
    values: list[str] = []
    for apartment in row.get("apartments") or []:
        values.extend([
            apartment.label(),
            apartment.full_label(),
            apartment.detail_label(),
            apartment.apartment_number or "",
            apartment.construction_number or "",
            apartment.owner_name or "",
            apartment.phone or "",
            apartment.finishing_type or "",
            apartment.comment or "",
            apartment.inspection_note or "",
        ])
    values.append(str(row.get("glass_status_label") or ""))
    values.append(str(row.get("mode") or ""))
    values.append(str(PO_STATUS_LABELS.get(row.get("po_status"), row.get("po_status") or "")))
    return " ".join(values).lower().replace("ё", "е")


def _inspection_sort_rank(row: dict, order: str) -> int:
    status = str(row.get("inspection_status") or "").strip()
    was_statuses = {"Был"}
    not_was_statuses = {"Не был", "Будет"}
    if order == "was_first":
        if status in was_statuses:
            return 0
        if status in not_was_statuses:
            return 1
        return 2
    if order == "not_was_first":
        if status in not_was_statuses:
            return 0
        if status in was_statuses:
            return 1
        return 2
    return 0


def _filtered_apartment_overview_rows(
    project_id: int,
    args,
    source_rows: list[dict] | None = None,
) -> tuple[list[dict], list[dict], bool]:
    q = (args.get("q") or "").strip()
    premise_selectors, _tail_query = parse_multi_premise_search(q)
    po_only = args.get("po") == "1"
    inspection_filter = (args.get("inspection_status") or "").strip()
    inspection_order = (args.get("inspection_order") or "").strip()
    app_status_filter = (args.get("app_status") or "").strip()
    avr_status_filter = (args.get("avr_status") or "").strip()
    po_status_filter = (args.get("po_status") or "").strip()
    rows = []
    overview_rows = source_rows
    if overview_rows is None:
        overview_rows = [_build_apartment_overview(group) for group in _group_project_apartments(project_id)]
    for row in overview_rows:
        if q:
            premise_selectors, tail_query = parse_multi_premise_search(q)
            search_mode, search_value = detect_search_mode(q)
            if premise_selectors:
                if not any(
                    premise_matches_selector(item, selector)
                    for item in row["apartments"]
                    for selector in premise_selectors
                ):
                    continue
                if tail_query:
                    needle = tail_query.lower().replace("ё", "е")
                    haystack = _apartment_overview_search_haystack(row)
                    if needle not in haystack:
                        continue
            elif search_mode in {"premise_number", "premise_number_or_building", "commercial_pair", "construction_number"}:
                if not any(premise_matches_search(item, search_mode, search_value) for item in row["apartments"]):
                    continue
            else:
                needle = search_value.lower().replace("ё", "е")
                haystack = _apartment_overview_search_haystack(row)
                if needle not in haystack:
                    continue
        if po_only and row.get("po_status") != PO_STATUS_TO_THROW:
            continue
        if app_status_filter == "accepted" and row.get("mode") != "АПП":
            continue
        if app_status_filter == "not_accepted" and row.get("mode") != "не принята":
            continue
        if avr_status_filter in {AVR_STATUS_NEEDED, AVR_STATUS_SIGNED} and (not row.get("show_avr") or row.get("avr_status") != avr_status_filter):
            continue
        if po_status_filter and row.get("po_status") != po_status_filter:
            continue
        if inspection_filter == "was" and row.get("inspection_status") != "Был":
            continue
        if inspection_filter == "not_was" and row.get("inspection_status") not in {"Не был", "Будет"}:
            continue
        rows.append(row)

    if premise_selectors:
        rows.sort(
            key=lambda row: (
                next(
                    (
                        index
                        for index, selector in enumerate(premise_selectors)
                        if any(premise_matches_selector(item, selector) for item in row["apartments"])
                    ),
                    len(premise_selectors),
                ),
                _inspection_sort_rank(row, inspection_order),
                _apartment_number_sort_value(row["apartment"].display_number(fallback_to_id=False)),
            )
        )
    else:
        rows.sort(
            key=lambda row: (
                _inspection_sort_rank(row, inspection_order),
                _apartment_number_sort_value(row["apartment"].display_number(fallback_to_id=False)),
            )
        )
    return rows, premise_selectors, po_only


def _apartments_export_filename_stem(project_name: str | None) -> str:
    text = str(project_name or "").strip()
    match = re.search(r"\b(\d+)\s+Квартал\s+(\d+)\b", text, flags=re.IGNORECASE)
    if match:
        return f"Квартал {match.group(1)}-{match.group(2)}"
    return _safe_filename_part(text or "квартиры")


@bp.route("/apartments/export")
@login_required
def apartments_export():
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    rows, _premise_selectors, _po_only = _filtered_apartment_overview_rows(project.id, request.args)
    wb = Workbook()
    ws = wb.active
    ws.title = "Квартиры"
    ws.append([
        "Помещение",
        "Отделка",
        "Режим",
        "Осмотр",
        "Срок устранения",
        "Комментарий",
    ])
    for row in rows:
        apartment = row["apartment"]
        ws.append([
            apartment.full_label(),
            apartment.finishing_type or "",
            row.get("mode") or "",
            row.get("inspection_status") or "",
            row.get("remark_deadline_display") or "",
            row.get("manual_comment") or row.get("inspection_comment") or "",
        ])
    _style_excel_header(ws)
    ws.column_dimensions["F"].width = 48
    filename = f"{_apartments_export_filename_stem(project.name)}_квартиры.xlsx"
    return _make_excel_response(wb, filename)


@bp.route("/avr", methods=["GET", "POST"])
@login_required
def avr():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if request.method == "POST" and current_user.role == "viewer":
        abort(403)

    apartments = _project_apartment_options(project.id)
    apartment_rows = []
    for apartment in apartments:
        number = apartment.apartment_number or apartment.construction_number or ""
        floor = _avr_floor(apartment)
        owner_options = _avr_owner_options(apartment.owner_name)
        inspection_date = format_input_date(apartment.first_inspection_date)
        apartment_rows.append(
            {
                "id": apartment.id,
                "label": apartment.label(),
                "premise_type": apartment.premise_type or "apartment",
                "number": number,
                "floor": floor,
                "owner": owner_options[0] if owner_options else (apartment.owner_name or ""),
                "owner_options": owner_options,
                "owner_selected": owner_options[:1] if owner_options else ([apartment.owner_name] if apartment.owner_name else []),
                "inspection_date": inspection_date,
                "address": project.address or "",
                "phrase": default_avr_phrase(apartment.first_inspection_date),
            }
        )

    selected_apartment = None
    selected_id = request.form.get("apartment_id") or request.args.get("apartment_id")
    if selected_id:
        try:
            selected_apartment = next((item for item in apartments if item.id == int(selected_id)), None)
        except ValueError:
            selected_apartment = None
    today_input = date.today().strftime("%Y-%m-%d")
    selected_floor = _avr_floor(selected_apartment) if selected_apartment else ""
    selected_owner_options = _avr_owner_options(selected_apartment.owner_name if selected_apartment else "")
    defaults = {
        "apartment_id": str(selected_apartment.id) if selected_apartment else "",
        "apartment_number": (selected_apartment.apartment_number or selected_apartment.construction_number or "") if selected_apartment else "",
        "floor": selected_floor,
        "premise_type": (selected_apartment.premise_type or "apartment") if selected_apartment else "apartment",
        "owner_name": (selected_owner_options[0] if selected_owner_options else (selected_apartment.owner_name or "")) if selected_apartment else "",
        "owner_options": selected_owner_options,
        "owner_selected": selected_owner_options[:1] if selected_owner_options else ([selected_apartment.owner_name] if selected_apartment and selected_apartment.owner_name else []),
        "developer_representative": "Худовердиев В.С.",
        "act_date": today_input,
        "inspection_date": format_input_date(selected_apartment.first_inspection_date) if selected_apartment else "",
        "address": project.address or "",
        "completion_phrase": default_avr_phrase(selected_apartment.first_inspection_date) if selected_apartment else default_avr_phrase(None),
    }

    if request.method == "POST":
        if selected_apartment is None:
            flash("Выберите квартиру для формирования АВР.", "warning")
            return render_template("avr.html", apartments=apartment_rows, defaults=defaults, today=today_input)

        selected_owners = [owner.strip() for owner in request.form.getlist("owner_names") if owner.strip()]
        owner_name = "\n".join(selected_owners) if selected_owners else (request.form.get("owner_name") or "").strip()
        values = defaults | {
            "apartment_number": (request.form.get("apartment_number") or defaults["apartment_number"]).strip(),
            "floor": (request.form.get("floor") or defaults["floor"]).strip(),
            "premise_type": selected_apartment.premise_type or "apartment",
            "owner_name": owner_name,
            "owner_selected": selected_owners or ([owner_name] if owner_name else []),
            "developer_representative": (request.form.get("developer_representative") or "Худовердиев В.С.").strip(),
            "act_date": format_doc_date(request.form.get("act_date") or today_input),
            "inspection_date": format_doc_date(request.form.get("inspection_date") or defaults["inspection_date"]),
            "address": (request.form.get("address") or project.address or "").strip(),
            "completion_phrase": (request.form.get("completion_phrase") or default_avr_phrase(request.form.get("inspection_date") or defaults["inspection_date"])).strip(),
        }

        if (values["premise_type"] != "commercial" and not values["floor"]) or not values["owner_name"]:
            flash("Заполните этаж и ФИО собственника.", "warning")
            form_defaults = defaults | values
            return render_template("avr.html", apartments=apartment_rows, defaults=form_defaults, today=today_input)

        exports_dir = Path(current_app.config["EXPORT_FOLDER"]) / "avr"
        filename = safe_avr_filename(project.name, values["apartment_number"], selected_apartment.premise_type)
        output_path = exports_dir / filename
        try:
            build_avr_docx(output_path, values)
        except Exception as exc:
            current_app.logger.exception("AVR document generation failed")
            flash(f"Не удалось сформировать АВР: {exc}", "danger")
            return render_template("avr.html", apartments=apartment_rows, defaults=defaults, today=today_input)

        return send_file(output_path, as_attachment=True, download_name=filename)

    return render_template("avr.html", apartments=apartment_rows, defaults=defaults, today=today_input)


@bp.route("/documents")
@login_required
def documents():
    if _setting_bool("hide_documents_section"):
        return _documents_under_development_response()

    return render_template("documents.html", document_type=None)


@bp.route("/documents/addendum", methods=["GET", "POST"])
@login_required
def documents_addendum():
    if _setting_bool("hide_documents_section"):
        return _documents_under_development_response()
    if current_user.role == "viewer":
        abort(403)

    result = None
    fields = addendum_fields_for_template()
    field_keys = addendum_field_keys()
    options = addendum_options_for_template()

    if request.method == "POST":
        uploads_dir = Path(current_app.config["UPLOAD_FOLDER"]) / "documents"
        exports_dir = Path(current_app.config["EXPORT_FOLDER"]) / "documents"
        uploads_dir.mkdir(parents=True, exist_ok=True)
        exports_dir.mkdir(parents=True, exist_ok=True)

        upload = request.files.get("document")
        template_variant = "materials"
        source_path = None
        prepared_source_path = None
        source_label = f"vstroenny_shablon_{template_variant}.docx"

        if upload and upload.filename:
            filename_lower = upload.filename.lower()
            try:
                validate_upload(upload, ["docx", "doc"], max_size=current_app.config.get("MAX_UPLOAD_FILE_BYTES"))
            except ValueError as exc:
                flash(str(exc), "warning")
                return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)
            if not filename_lower.endswith((".docx", ".doc")):
                flash("Поддерживаются файлы Word .docx и .doc. .doc будет автоматически сконвертирован в .docx, если на сервере есть LibreOffice или Microsoft Word.", "warning")
                return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)

            source_name = safe_docx_filename(upload.filename, "source")
            if filename_lower.endswith(".doc"):
                source_name = source_name.replace(".docx", ".doc")
            source_path = uploads_dir / source_name
            upload.save(source_path)
            source_label = upload.filename

            try:
                prepared_source_path = prepare_uploaded_word_file(source_path, uploads_dir, template_variant)
                is_addendum, validation_message = validate_addendum_template(prepared_source_path)
                if not is_addendum:
                    flash(validation_message, "danger")
                    return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)
                if filename_lower.endswith(".doc"):
                    flash("Файл .doc загружен. Система автоматически подготовила его к генерации доп. соглашения.", "info")
            except Exception as exc:
                current_app.logger.exception("Document flow addendum source preparation failed")
                if filename_lower.endswith(".doc"):
                    flash(f"Не удалось подготовить .doc: {exc}", "danger")
                else:
                    flash(f"Не удалось подготовить шаблон Word: {exc}", "danger")
                return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)
        else:
            flash("Загрузите файл Word .docx или .doc — без файла доп. соглашение не формируется и скачать его нельзя.", "warning")
            return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)

        output_name = safe_docx_filename(source_label, "dop-soglashenie")
        output_path = exports_dir / output_name

        field_values = {key: (request.form.get(key) or "").strip() for key in field_keys}
        field_values["owner_count"] = (request.form.get("owner_count") or "1").strip()
        selected_options = request.form.getlist("options")
        try:
            changes = build_addendum_docx(prepared_source_path, output_path, field_values, selected_options)
        except Exception as exc:
            current_app.logger.exception("Document flow addendum failed")
            flash(f"Не удалось подготовить документ: {exc}", "danger")
            return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)

        result = {
            "filename": output_name,
            "changes": changes,
            "selected_count": len(selected_options),
            "replace_count": sum(1 for item in changes if item.kind == "replace"),
            "append_count": sum(1 for item in changes if item.kind == "append"),
        }
        flash("Доп. соглашение подготовлено.", "success")

    return render_template("documents.html", document_type="addendum", fields=fields, options=options, result=result)


@bp.route("/documents/download/<path:filename>")
@login_required
def documents_download(filename: str):
    if _setting_bool("hide_documents_section"):
        return redirect(url_for("main.documents"))
    exports_dir = (Path(current_app.config["EXPORT_FOLDER"]) / "documents").resolve()
    file_path = (exports_dir / filename).resolve()
    if exports_dir not in file_path.parents or not file_path.exists():
        abort(404)
    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@bp.route("/apartments/<int:apartment_id>/po-status", methods=["POST"])
@login_required
def update_apartment_po_status(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    status = (request.form.get("po_status") or "").strip()
    if status not in PO_STATUS_LABELS:
        abort(400)
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    target_group = group or [apartment]
    old_status = _build_apartment_overview(target_group).get("po_status")
    for item in group or [apartment]:
        item.po_status = status
        item.po_status_manual = True
    logged_change = _log_apartment_field_change(target_group, "po_status", old_status, status)
    history_entry = None
    if logged_change:
        history_entry = _build_change_history_entry(logged_change[0], task=logged_change[1], users_cache={})
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify({
            "ok": True,
            "message": "Внутренний статус обновлён",
            "po_status": status,
            "po_label": PO_STATUS_LABELS.get(status, status),
            "history_entry": history_entry,
        })
    flash("Внутренний статус обновлён", "success")
    return redirect(request.referrer or url_for("main.apartments"))


@bp.route("/apartments/<int:apartment_id>/inspection-status", methods=["POST"])
@login_required
def update_apartment_inspection_status(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    status = (request.form.get("inspection_status") or "").strip()
    if status not in {"was", "not_was"}:
        abort(400)
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    target_group = group or [apartment]
    if _is_app_inspection_locked(target_group) and status != "not_was":
        message = "В режиме АПП осмотр с датой изменить нельзя"
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": message}), 400
        flash(message, "warning")
        return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))
    if any(_is_unsold_apartment(item) for item in target_group):
        for item in target_group:
            item.first_inspection_present = False
            item.first_inspection_date = None
            item.inspection_date = None
        db.session.commit()
        flash("У непроданной квартиры осмотр фиксируется автоматически: Не был", "info")
        return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))

    was_present = status == "was"
    for item in target_group:
        item.first_inspection_present = was_present
        if was_present:
            restored_date = item.inspection_date or item.first_inspection_date or item.inspection_date_backup
            item.first_inspection_date = restored_date or date.today()
            item.inspection_date = restored_date or item.first_inspection_date
            item.inspection_date_backup = item.inspection_date
        else:
            previous_date = item.inspection_date or item.first_inspection_date
            if previous_date:
                item.inspection_date_backup = previous_date
            item.inspection_date = None
            item.first_inspection_date = None
            item.first_inspection_present = False
            if _parse_inspection_schedule_marker(item.inspection_note):
                item.inspection_note = None
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        overview = _build_apartment_overview(target_group)
        return jsonify({
            "ok": True,
            "message": "Статус осмотра обновлён",
            "inspection_date": overview.get("inspection_date").isoformat() if overview.get("inspection_date") else "",
            "inspection_date_label": overview.get("inspection_display") or "—",
            "inspection_status": overview.get("inspection_status") or "",
            "inspection_status_class": overview.get("inspection_status_class") or "status-pill-muted",
        })
    flash("Статус осмотра обновлён", "success")
    return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))


@bp.route("/apartments/<int:apartment_id>/inspection-date", methods=["POST"])
@login_required
def update_apartment_inspection_date(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)

    inspection_date = parse_date(request.form.get("inspection_date"))
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    target_group = group or [apartment]
    if _is_app_inspection_locked(target_group):
        message = "В режиме АПП осмотр с датой изменить нельзя"
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": message}), 400
        flash(message, "warning")
        return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))
    for item in target_group:
        item.inspection_date = inspection_date
        item.first_inspection_date = inspection_date
        item.first_inspection_present = inspection_date is not None
        if inspection_date is not None:
            item.inspection_date_backup = inspection_date
    db.session.commit()

    message = "Дата осмотра обновлена" if inspection_date else "Дата осмотра очищена"
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        overview = _build_apartment_overview(target_group)
        return jsonify({
            "ok": True,
            "message": message,
            "inspection_date": inspection_date.isoformat() if inspection_date else "",
            "inspection_date_label": overview.get("inspection_display") or "—",
            "inspection_status": overview.get("inspection_status") or "",
            "inspection_status_class": overview.get("inspection_status_class") or "status-pill-muted",
        })
    flash(message, "success")
    return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))


@bp.route("/apartments/<int:apartment_id>/inspection-note", methods=["POST"])
@login_required
def update_apartment_inspection_note(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    note = (request.form.get("inspection_note") or "").strip()
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    target_group = group or [apartment]
    old_note = _apartment_inspection_comment(target_group) or ""
    for item in group or [apartment]:
        item.inspection_note = note or None
    history_change = _log_apartment_field_change(target_group, "apartment_inspection_note", old_note, note)
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        history_entry = _build_change_history_entry(history_change[0], task=history_change[1]) if history_change else None
        return jsonify({"ok": True, "message": "Комментарий сохранен", "inspection_note": note, "history_entry": history_entry})
    flash("Комментарий сохранен", "success")
    return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))


@bp.route("/apartments/<int:apartment_id>/comment", methods=["POST"])
@login_required
def update_apartment_comment(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    comment = (request.form.get("apartment_comment") or "").strip()
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    target_group = group or [apartment]
    old_comment = _apartment_manual_comment(target_group) or ""
    for item in group or [apartment]:
        item.comment = comment or None
    history_change = _log_apartment_field_change(target_group, "apartment_comment", old_comment, comment)
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        history_entry = _build_change_history_entry(history_change[0], task=history_change[1]) if history_change else None
        return jsonify({"ok": True, "message": "Комментарий сохранен", "comment": comment, "history_entry": history_entry})
    flash("Комментарий сохранен", "success")
    return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))


@bp.route("/apartments/<int:apartment_id>/avr-status", methods=["POST"])
@login_required
def update_apartment_avr_status(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    status = (request.form.get("avr_status") or "").strip()
    if status not in {AVR_STATUS_NEEDED, AVR_STATUS_SIGNED}:
        abort(400)
    signed_date = parse_date(request.form.get("avr_signed_date"))
    if status == AVR_STATUS_SIGNED and signed_date is None:
        signed_date = date.today()

    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    target_group = group or [apartment]
    old_status = _group_avr_status(target_group)
    old_signed_date = next((item.avr_signed_date for item in target_group if item.avr_signed_date), None)
    for item in target_group:
        item.avr_status = status
        item.avr_signed_date = signed_date if status == AVR_STATUS_SIGNED else None
    status_history_change = _log_apartment_field_change(target_group, "avr_status", old_status, status)
    date_history_change = _log_apartment_field_change(
        target_group,
        "avr_signed_date",
        old_signed_date.isoformat() if old_signed_date else "",
        signed_date.isoformat() if signed_date and status == AVR_STATUS_SIGNED else "",
    )
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        history_entry = None
        for history_change in (date_history_change, status_history_change):
            if history_change:
                history_entry = _build_change_history_entry(history_change[0], task=history_change[1], users_cache={})
                break
        return jsonify({
            "ok": True,
            "message": "Статус АВР обновлен",
            "avr_status": status,
            "avr_signed_date": signed_date.isoformat() if signed_date and status == AVR_STATUS_SIGNED else "",
            "avr_signed_date_label": format_ru_date(signed_date) if signed_date and status == AVR_STATUS_SIGNED else "",
            "history_entry": history_entry,
        })
    flash("Статус АВР обновлен", "success")
    return redirect(request.referrer or url_for("main.apartment_detail", apartment_id=apartment.id))


@bp.route("/apartments/<int:apartment_id>")
@login_required
def apartment_detail(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    group_key = _apartment_group_key(apartment)
    group = [
        item
        for item in (
            Apartment.query.options(
                selectinload(Apartment.tasks).selectinload(Task.work_point),
                selectinload(Apartment.tasks).selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items),
                selectinload(Apartment.tasks).selectinload(Task.responsible),
                selectinload(Apartment.tasks).selectinload(Task.comments).selectinload(TaskComment.user),
                selectinload(Apartment.tasks).selectinload(Task.changes),
            )
            .filter(Apartment.project_id == project.id)
            .all()
        )
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    if not group:
        group = [apartment]
    overview = _build_apartment_overview(group)
    users_cache: dict[int, str] = {}
    history_entries_all = [
        _build_change_history_entry(item["change"], task=item["task"], users_cache=users_cache)
        for item in overview["changes"]
        if not _is_legacy_problem_comment_change(item["change"], item["task"])
    ]
    history_page = max(request.args.get("history_page", 1, type=int), 1)
    history_per_page = 12
    history_total = len(history_entries_all)
    history_pages = max((history_total + history_per_page - 1) // history_per_page, 1)
    if history_page > history_pages:
        history_page = history_pages
    history_start = (history_page - 1) * history_per_page
    history_entries = history_entries_all[history_start:history_start + history_per_page]
    requested_back_url = request.args.get("back")
    back_url = requested_back_url if _is_local_redirect(requested_back_url) else url_for("main.apartments")
    return render_template(
        "apartment_detail.html",
        row=overview,
        history_entries=history_entries,
        history_page=history_page,
        history_pages=history_pages,
        apartment=overview["apartment"],
        back_url=back_url,
        today=date.today(),
        po_status_labels=PO_STATUS_LABELS,
        po_status_classes=PO_STATUS_CLASSES,
        avr_status_needed=AVR_STATUS_NEEDED,
        avr_status_signed=AVR_STATUS_SIGNED,
    )


@bp.route("/apartments/<int:apartment_id>/remarks/export")
@login_required
def apartment_remarks_export(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    group_key = _apartment_group_key(apartment)
    apartment_ids = [
        item.id
        for item in Apartment.query.filter(Apartment.project_id == project.id).all()
        if _is_visible_apartment_row(item) and _apartment_group_key(item) == group_key
    ]
    if not apartment_ids:
        apartment_ids = [apartment.id]
    tasks = (
        Task.query.join(Apartment).join(WorkPoint)
        .options(selectinload(Task.apartment), selectinload(Task.work_point), selectinload(Task.responsible))
        .filter(
            Task.project_id == project.id,
            Task.apartment_id.in_(apartment_ids),
            Task.is_archived.is_(False),
        )
        .order_by(
            cast(WorkPoint.point_number, Integer).asc(),
            WorkPoint.point_number.asc(),
            Task.updated_at.desc(),
        )
        .all()
    )
    premise_label = apartment.full_label() if apartment.premise_type == "commercial" else apartment.label()
    filename_prefix = f"{project.name}_{premise_label}_помещение_замечания"
    path = export_remark_tasks_excel(tasks, filename_prefix, title=f"Карточка помещения: {premise_label}")
    return send_file(path, as_attachment=True, download_name=Path(path).name)


@bp.route("/report")
@login_required
def work_report():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    today = date.today()
    start = today - timedelta(days=55)
    # Include ordinary completed work and remarks closed with a concession.
    report_statuses = (STATUS_DONE, STATUS_CONCESSION)
    tasks = (
        Task.query.join(Apartment)
        .join(WorkPoint)
        .filter(
            Task.project_id == project.id,
            Task.is_done.is_(True),
            Task.status.in_(report_statuses),
            WorkPoint.point_number.notin_(DOP_AGREEMENT_POINT_NUMBERS),
            Task.completed_date.isnot(None),
            Task.completed_date >= start,
        )
        .order_by(Task.completed_date.desc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc())
        .all()
    )
    buckets = []
    bucket_index = 1
    cursor = today
    while cursor >= start:
        period_start = cursor - timedelta(days=6)
        period_tasks = [task for task in tasks if period_start <= task.completed_date.date() <= cursor]
        buckets.append({"index": bucket_index, "start": period_start, "end": cursor, "tasks": period_tasks})
        bucket_index += 1
        cursor = period_start - timedelta(days=1)
    return render_template("work_report.html", buckets=buckets, project=project)


@bp.route("/report/export")
@login_required
def work_report_export():
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER, ROLE_VERIFIER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    today = date.today()
    week = request.args.get("week", type=int)
    if week and week > 0:
        end = today - timedelta(days=(week - 1) * 7)
        start = end - timedelta(days=6)
        filename_prefix = f"{project.name}_отчет_{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}"
    else:
        start = today - timedelta(days=55)
        end = today
        filename_prefix = f"{project.name}_отчет_весь период"
    # Держим Excel-выгрузку синхронной с экраном отчёта.
    report_statuses = (STATUS_DONE, STATUS_CONCESSION)
    base_query = Task.query.filter(
        Task.project_id == project.id,
        Task.is_done.is_(True),
        Task.status.in_(report_statuses),
        Task.work_point.has(WorkPoint.point_number.notin_(DOP_AGREEMENT_POINT_NUMBERS)),
        Task.completed_date.isnot(None),
        Task.completed_date >= start,
        Task.completed_date <= end + timedelta(days=1),
    )
    task_count, last_updated = base_query.with_entities(func.count(Task.id), func.max(Task.updated_at)).one()
    cache_stamp = last_updated.strftime("%Y%m%d%H%M%S") if last_updated else "empty"
    cache_key = f"report-v3-concessions_count-{task_count}_updated-{cache_stamp}"
    path = build_export_path(filename_prefix, cache_key=cache_key)
    download_name = f"{_safe_filename_part(filename_prefix)}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    if path.exists():
        return send_file(path, as_attachment=True, download_name=download_name)
    tasks = (
        base_query.options(selectinload(Task.apartment))
        .join(Apartment)
        .order_by(Task.completed_date.desc(), cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc())
        .all()
    )
    path = export_report_tasks_excel(tasks, filename_prefix, cache_key=cache_key)
    return send_file(path, as_attachment=True, download_name=download_name)


def _apartment_task_groups(query):
    rows = []
    for apartment in query.order_by(cast(Apartment.apartment_number, Integer).asc(), Apartment.apartment_number.asc()).all():
        tasks = [task for task in apartment.tasks if not task.is_archived]
        visible = [task for task in tasks if task.work_point and task.work_point.point_number in VISIBLE_WORK_POINT_NUMBERS]
        done = [task for task in visible if task.is_done]
        left = [task for task in visible if not task.is_done]
        painter_tasks = [task for task in visible if task.work_point and task.work_point.point_number in {"10", "11", "12"}]
        painter_done = bool(painter_tasks) and all(task.is_done for task in painter_tasks)
        rows.append({"apartment": apartment, "done": done, "left": left, "total": len(visible), "painter_done": painter_done})
    return rows


@bp.route("/notifications")
@login_required
def notifications():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    tab = request.args.get("tab", "po")
    today = date.today()
    base = Apartment.query.filter(Apartment.project_id == project.id)
    if tab == "archive":
        apartments = base.filter(Apartment.avr_archived_at.isnot(None))
        rows = _apartment_task_groups(apartments)
    elif tab == "60":
        apartments = base.filter(
            Apartment.app_deadline_date.isnot(None),
            Apartment.app_deadline_date <= today + timedelta(days=15),
            Apartment.avr_archived_at.is_(None),
        )
        rows = _apartment_task_groups(apartments)
    else:
        apartments = base.filter(Apartment.deadline_date.is_(None), Apartment.finishing_type.ilike("%бел%"), Apartment.avr_archived_at.is_(None))
        rows = [
            row
            for row in _apartment_task_groups(apartments)
            if row["total"] > 0 and (len(row["done"]) > row["total"] / 2 or row["painter_done"])
        ]
    return render_template("notifications.html", rows=rows, tab=tab, today=today)


@bp.route("/notifications/<int:apartment_id>/avr", methods=["POST"])
@login_required
def archive_apartment_avr(apartment_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    apartment = db.session.get(Apartment, apartment_id) or abort(404)
    if apartment.project_id != project.id:
        abort(404)
    apartment.avr_archived_at = datetime.utcnow()
    for task in apartment.tasks:
        task.is_archived = True
    db.session.commit()
    flash("Квартира перенесена в архив АВР.", "success")
    return redirect(url_for("main.notifications", tab="60"))

def _delete_task_with_relations(task: Task, project_id: int) -> None:
    _record_simple_deletion(
        "task_delete",
        "task",
        task,
        f"Замечание #{task.id}",
        f"Удалено ручное замечание: {(task.description or task.source_cell_value or '')[:180]}",
        project_id=project_id,
        extra={
            "apartment_label": task.apartment.label() if task.apartment else None,
            "work_point": task.work_point.display_name if task.work_point else None,
            "comments": _snapshot_children(task.comments),
            "changes": _snapshot_children(task.changes),
            "glass_measurement": (
                _snapshot_model(
                    task.glass_measurement,
                    extra={"items": _snapshot_children(task.glass_measurement.items)},
                )
                if task.glass_measurement
                else None
            ),
            "material_writeoff_ids": [writeoff.id for writeoff in list(task.material_writeoffs)],
            "comment_count": len(task.comments),
            "change_count": len(task.changes),
        },
    )
    for writeoff in list(task.material_writeoffs):
        writeoff.tasks.remove(task)
    db.session.delete(task)


@bp.route("/tasks/delete-selected", methods=["POST"])
@login_required
def tasks_bulk_delete():
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)

    next_url = request.form.get("next")
    task_ids = _selected_task_ids_from_form()
    if not task_ids:
        flash("Выберите хотя бы одно замечание для удаления", "warning")
        return _safe_redirect(next_url, "main.task_list")

    order_map = {task_id: index for index, task_id in enumerate(task_ids)}
    tasks = (
        Task.query.options(
            selectinload(Task.apartment),
            selectinload(Task.work_point),
            selectinload(Task.comments),
            selectinload(Task.changes),
            selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items),
        )
        .filter(Task.project_id == project.id, Task.id.in_(task_ids))
        .all()
    )
    tasks.sort(key=lambda task: order_map.get(task.id, len(order_map)))

    if not tasks:
        flash("Выбранные замечания не найдены", "warning")
        return _safe_redirect(next_url, "main.task_list")

    deleted_count = 0
    for task in tasks:
        _delete_task_with_relations(task, project.id)
        deleted_count += 1

    if deleted_count:
        db.session.commit()

    if deleted_count == 1:
        flash("Замечание удалено", "success")
    elif deleted_count > 1:
        flash(f"Удалено замечаний: {deleted_count}", "success")

    return _safe_redirect(next_url, "main.task_list")


@bp.route("/tasks/<int:task_id>")
@login_required
def task_detail(task_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if current_user.role in WORKER_ROLES and task.responsible_id != current_user.id:
        abort(403)
    edit_form = TaskEditForm(obj=task)
    edit_form.responsible_id.choices = [("", "Не назначен")] + [
        (str(u.id), u.full_name or u.username)
        for u in _active_users_for_project(project.id)
    ]
    edit_form.responsible_id.data = str(task.responsible_id or "")
    edit_form.planned_date.data = task.planned_date.isoformat() if task.planned_date else ""
    comment_form = CommentForm()
    visible_changes = [change for change in task.changes if not _is_legacy_problem_comment_change(change, task)]
    users_cache: dict[int, str] = {}
    history_entries = [_build_change_history_entry(change, task=task, users_cache=users_cache) for change in visible_changes]
    other_tasks = (
        Task.query.join(WorkPoint)
        .options(selectinload(Task.work_point), selectinload(Task.glass_measurement).selectinload(GlassMeasurement.items))
        .filter(
            Task.project_id == project.id,
            Task.apartment_id == task.apartment_id,
            Task.id != task.id,
            Task.is_archived.is_(False),
        )
        .order_by(Task.is_done.asc(), WorkPoint.point_number.asc(), Task.updated_at.desc())
        .limit(30)
        .all()
    )
    requested_back_url = request.args.get("back")
    back_url = requested_back_url if _is_local_redirect(requested_back_url) else url_for("main.task_list")
    return render_template(
        "task_detail.html",
        task=task,
        edit_form=edit_form,
        comment_form=comment_form,
        visible_changes=visible_changes,
        history_entries=history_entries,
        other_open_tasks=other_tasks,
        back_url=back_url,
        can_delete_task=current_user.role != "viewer",
    )


@bp.route("/tasks/<int:task_id>/delete", methods=["POST"])
@login_required
def task_delete(task_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    if current_user.role == "viewer":
        abort(403)
    task = (
        Task.query.filter(Task.id == task_id, Task.project_id == project.id)
        .first()
        or abort(404)
    )
    _delete_task_with_relations(task, project.id)
    db.session.commit()
    flash("Замечание удалено", "success")
    return redirect(url_for("main.task_list"))


@bp.route("/tasks/<int:task_id>/update", methods=["POST"])
@login_required
def update_task(task_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if not can_change_task(current_user, task):
        abort(403)
    form = TaskEditForm()
    form.responsible_id.choices = [("", "Не назначен")] + [
        (str(u.id), u.full_name or u.username)
        for u in _active_users_for_project(project.id)
    ]
    if form.validate_on_submit():
        fields = {
            "status": form.status.data,
            "priority": form.priority.data,
            "responsible_id": int(form.responsible_id.data) if form.responsible_id.data else None,
            "planned_date": parse_date(form.planned_date.data),
            "comment": form.comment.data,
        }
        if fields["responsible_id"]:
            responsible = db.session.get(User, fields["responsible_id"])
            if not _user_can_work_in_project(responsible, project):
                flash("Выберите корректного исполнителя", "danger")
                return redirect(url_for("main.task_detail", task_id=task.id))
            if not _assignment_task_allowed_for_user(task, responsible):
                flash("Эта задача не относится к пунктам, отмеченным для выбранного исполнителя в распределении", "danger")
                return redirect(url_for("main.task_detail", task_id=task.id))
        for field, new_value in fields.items():
            old_value = getattr(task, field)
            if old_value != new_value:
                setattr(task, field, new_value)
                log_change(task, "field_update", field, old_value, new_value)
        if is_problem_details_required(task.status, task.comment):
            flash("Для статуса 'Проблема' нужно заполнить описание проблемы", "danger")
            return redirect(url_for("main.task_detail", task_id=task.id))
        task.is_done = task.status in DONE_STATUSES
        if task.is_done and not task.completed_date:
            task.completed_date = datetime.utcnow()
        if not task.is_done and task.completed_date:
            task.completed_date = None
        task.manually_edited = True
        db.session.commit()
        flash("Карточка задачи обновлена", "success")
    else:
        flash("Проверьте поля формы", "danger")
    return redirect(url_for("main.task_detail", task_id=task.id))


@bp.route("/tasks/<int:task_id>/comment", methods=["POST"])
@login_required
def add_task_comment(task_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if not can_change_task(current_user, task):
        abort(403)
    form = CommentForm()
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    if form.validate_on_submit():
        comment = TaskComment(task_id=task.id, user_id=current_user.id, body=form.body.data)
        db.session.add(comment)
        log_change(task, "comment_added", "comment", "", form.body.data)
        db.session.commit()
        if wants_json:
            history_change = next((change for change in reversed(task.changes) if change.action == "comment_added"), None)
            history_entry = _build_change_history_entry(history_change, task=task, users_cache={}) if history_change else None
            return jsonify({
                "ok": True,
                "message": "Комментарий добавлен",
                "comment": {
                    "author": current_user.full_name or current_user.username,
                    "timestamp": format_ru_datetime(comment.created_at),
                    "body": comment.body,
                },
                "history_entry": history_entry,
            })
        flash("Комментарий добавлен", "success")
    elif wants_json:
        return jsonify({"ok": False, "message": "Не удалось добавить комментарий"}), 400
    return redirect(url_for("main.task_detail", task_id=task.id))


@bp.route("/tasks/<int:task_id>/status/<status>", methods=["POST"])
@login_required
def quick_status(task_id: int, status: str):
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if not can_change_task(current_user, task):
        abort(403)
    if status not in TASK_STATUSES:
        abort(400)
    if is_problem_details_required(status, request.form.get("problem_comment") or task.comment):
        if wants_json:
            return jsonify({"ok": False, "message": "Для статуса 'Проблема' нужно заполнить описание проблемы"}), 400
        flash("Для статуса 'Проблема' нужно заполнить описание проблемы", "danger")
        return redirect(request.referrer or url_for("main.task_list"))
    if status == "problem":
        problem_comment = (request.form.get("problem_comment") or "").strip()
        if problem_comment:
            task.comment = problem_comment
    change_task_status(task, status, user_id=current_user.id)
    if request.form.get("sync_google_format") == "1":
        try:
            update_task_strike_in_google_sheet(task)
        except Exception as exc:
            if wants_json:
                return jsonify({"ok": False, "message": f"Статус сохранён, но зачёркивание в Google Sheets не применилось: {exc}"}), 500
            flash(f"Статус сохранён, но зачёркивание в Google Sheets не применилось: {exc}", "warning")
            return redirect(request.referrer or url_for("main.task_list"))
    if wants_json:
        history_change = (
            ChangeLog.query
            .filter_by(task_id=task.id, action="status_change", field_name="status")
            .order_by(ChangeLog.id.desc())
            .first()
        )
        history_entry = _build_change_history_entry(history_change, task=task, users_cache={}) if history_change else None
        apartment_stats = _apartment_live_stats_for_task(task)
        return jsonify(
            {
                "ok": True,
                "task_id": task.id,
                "status": task.status,
                "status_label": task.status_label(),
                "status_class": task.status_class(),
                "is_done": task.is_done,
                "message": f"Статус изменён: {TASK_STATUSES[status]['label']}",
                "history_entry": history_entry,
                "apartment_stats": apartment_stats,
            }
        )
    flash(f"Статус изменён: {TASK_STATUSES[status]['label']}", "success")
    return redirect(request.referrer or url_for("main.task_list"))


def _validate_uploaded_excel_kind(path: Path, expected_kind: str) -> None:
    remarks_info = inspect_remarks_workbook(path)
    transfer_info = inspect_transfer_workbook(path)
    if expected_kind == "remarks":
        if remarks_info.get("ok"):
            return
        if transfer_info.get("ok"):
            raise ValueError(
                "Похоже, вы загрузили таблицу статистики передач в блок замечаний. "
                "Загрузите этот файл в блок «Excel-файл передач»."
            )
        raise ValueError(
            "Не удалось распознать таблицу замечаний. "
            "Загрузите исходный Excel замечаний или файл, который подходит для этого раздела."
        )
    if expected_kind == "transfers":
        if transfer_info.get("ok"):
            return
        if remarks_info.get("ok"):
            raise ValueError(
                "Похоже, вы загрузили таблицу замечаний в блок статистики передач. "
                "Загрузите этот файл в блок «Excel-файл замечаний»."
            )
        raise ValueError(
            "Не удалось распознать таблицу статистики передач. "
            "Загрузите именно Excel-файл статистики передач, а не другую вкладку или таблицу."
        )
    raise ValueError("Неизвестный тип Excel-импорта.")


@bp.route("/upload-excel", methods=["GET", "POST"])
@login_required
def upload_excel():
    if not can_manage_sync(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    form = UploadExcelForm()
    transfer_form = UploadExcelForm(prefix="transfer")
    preview = None
    upload_kind = request.form.get("upload_kind")
    if upload_kind == "remarks" and form.validate_on_submit():
        try:
            validate_upload(form.file.data, ["xlsx"], max_size=current_app.config.get("MAX_UPLOAD_FILE_BYTES"))
            path = save_upload(form.file.data)
            _validate_uploaded_excel_kind(path, "remarks")
            result = sync_excel_file(path, project_name=project.name)
            set_setting(f"latest_excel_path_project_{project.id}", str(path))
            pending_conflicts = _project_pending_conflicts_query(project.id).count()
            latest_log = SyncLog.query.filter(SyncLog.project_id == project.id).order_by(SyncLog.started_at.desc()).first()
            if latest_log:
                latest_log.missing_count = pending_conflicts
                db.session.commit()
            if pending_conflicts:
                flash(
                    f"Добавлено новых - {result.get('created_count', 0)} , несостыковок - {pending_conflicts}",
                    "warning",
                )
                return redirect(url_for("main.sync_conflicts"), code=303)
            flash(
                f"Добавлено новых - {result.get('created_count', 0)} , несостыковок - {pending_conflicts}",
                "success",
            )
        except Exception as exc:
            current_app.logger.exception("Excel import failed")
            flash(f"Ошибка загрузки Excel: {exc}", "danger")
    elif upload_kind == "transfers" and transfer_form.validate_on_submit():
        try:
            validate_upload(transfer_form.file.data, ["xlsx"], max_size=current_app.config.get("MAX_UPLOAD_FILE_BYTES"))
            path = save_upload(transfer_form.file.data)
            _validate_uploaded_excel_kind(path, "transfers")
            result = sync_transfer_statistics(path, project_name=project.name)
            flash(
                "Статистика передач обновлена: "
                f"принято - {result.get('accepted_count', 0)}, "
                f"ждёт - {result.get('waiting_count', 0)}, "
                f"не продано - {result.get('unsold_count', 0)}",
                "success",
            )
        except Exception as exc:
            current_app.logger.exception("Transfer statistics import failed")
            flash(f"Не удалось загрузить статистику передач: {exc}", "danger")
    return render_template("upload_excel.html", form=form, transfer_form=transfer_form, preview=preview)


@bp.route("/export/tasks")
@login_required
def export_tasks():
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    category_id = request.args.get("category_id", type=int)
    tasks = _export_tasks_from_request(request.args.copy(), project.id, category_id=category_id).all()
    path = export_tasks_to_excel(tasks, filename_prefix=project.name)
    return send_file(path, as_attachment=True, download_name=Path(path).name)


@bp.route("/export/category/<int:category_id>")
@login_required
def export_category_tasks(category_id: int):
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    category = db.session.get(WorkCategory, category_id) or abort(404)
    tasks = _export_tasks_from_request(request.args.copy(), project.id, category_id=category_id).all()
    path = export_remark_tasks_excel(tasks, f"{project.name}_{category.name}", title=category.name)
    return send_file(path, as_attachment=True, download_name=Path(path).name)


@bp.route("/export/category/<int:category_id>/pdf")
@login_required
def export_category_tasks_pdf(category_id: int):
    if not can_export(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    category = db.session.get(WorkCategory, category_id) or abort(404)
    tasks = build_task_query({}, category_id=category_id, project_id=project.id).all()
    path = export_tasks_pdf(tasks, f"{project.name}_{category.name}", title=f"Замечания - {category.name}")
    return send_file(path, as_attachment=True, download_name=Path(path).name, mimetype="application/pdf")


@bp.route("/export/source-with-strikes")
@login_required
def export_source_with_strikes():
    if not can_export(current_user):
        abort(403)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    try:
        project = selected_project()
        if project is None:
            if wants_json:
                return jsonify({"ok": False, "message": "Сначала выберите объект."}), 400
            return redirect(url_for("main.objects"))
        source_path = get_setting(f"latest_excel_path_project_{project.id}")
        if not source_path:
            if wants_json:
                return jsonify({"ok": False, "message": "Для этого объекта ещё не загружали Excel с замечаниями."}), 400
            flash("Для этого объекта ещё не загружали Excel с замечаниями.", "warning")
            return redirect(url_for("main.upload_excel"))
        resolved_source_path = resolve_source_excel_with_strikes_path(source_path=source_path, project_id=project.id)
        if str(resolved_source_path) != str(Path(source_path)):
            set_setting(f"latest_excel_path_project_{project.id}", str(resolved_source_path))
            db.session.commit()
        path = export_source_excel_with_strikes(
            source_path=str(resolved_source_path),
            project_name=project.name,
            project_id=project.id,
        )
        return send_file(path, as_attachment=True, download_name=Path(path).name)
    except FileNotFoundError as exc:
        tasks = (
            Task.query.filter(Task.project_id == project.id)
            .order_by(Task.apartment_id.asc(), Task.id.asc())
            .all()
        ) if project is not None else []
        if tasks:
            path = export_source_excel_reconstructed(tasks, filename_prefix=f"{project.name}_замечания")
            return send_file(path, as_attachment=True, download_name=Path(path).name)
        message = str(exc) or "Не удалось найти исходный Excel с замечаниями. Загрузите его заново."
        if wants_json:
            return jsonify({"ok": False, "message": message}), 400
        flash(message, "warning")
        return redirect(url_for("main.upload_excel"))
    except Exception as exc:
        if wants_json:
            return jsonify({"ok": False, "message": str(exc) or "Не удалось подготовить Excel."}), 500
        flash(str(exc), "danger")
        return redirect(url_for("main.dashboard"))


@bp.route("/mappings", methods=["GET", "POST"])
@login_required
def mapping_settings():
    if not can_manage_mapping(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    ensure_default_categories()
    hidden_point_numbers = {"7", "9", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32", "33"}
    categories_to_show = [
        category
        for category in WorkCategory.query.filter(WorkCategory.is_active.is_(True)).order_by(WorkCategory.sort_order.asc()).all()
        if (category.name or "").strip().lower() not in {"все", "доп.соглашение"}
    ]
    point_ids_for_project = [
        point_id for (point_id,) in db.session.query(Task.work_point_id).filter(Task.project_id == project.id).distinct().all()
    ]
    points = (
        WorkPoint.query.filter(WorkPoint.id.in_(point_ids_for_project or [-1]), WorkPoint.is_active.is_(True))
        .order_by(WorkPoint.point_number.asc())
        .all()
    )
    if request.method == "POST":
        wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
        allowed_point_ids = {point.id for point in points}
        for category in categories_to_show:
            selected = request.form.getlist(f"category_{category.id}")
            point_ids = [int(x) for x in selected if x.isdigit() and int(x) in allowed_point_ids]
            update_category_points(category.id, point_ids, commit=False)
        db.session.commit()
        if wants_json:
            return jsonify({"ok": True, "message": f"Распределение сохранено для объекта: {project.name}"})
        flash(f"Распределение сохранено для объекта: {project.name}", "success")
        return redirect(url_for("main.mapping_settings"))
    return render_template(
        "mapping_settings.html",
        categories=categories_to_show,
        points=points,
        hidden_point_numbers=hidden_point_numbers,
        project=project,
    )




@bp.route("/settings", methods=["GET", "POST"])
@login_required
def site_settings():
    if current_user.role != ROLE_ADMIN:
        abort(403)
    if request.method == "POST":
        _set_setting_bool("hide_documents_section", request.form.get("hide_documents_section") == "1")
        _set_setting_bool("mobile_version_under_development", request.form.get("mobile_version_under_development") == "1")
        _set_setting_bool("site_maintenance_mode", request.form.get("site_maintenance_mode") == "1")
        _set_setting_bool("two_factor_every_login", request.form.get("two_factor_every_login") == "1")
        allowed_section_keys = {choice["key"] for choice in SECTION_LOCK_CHOICES}
        _set_setting_csv("blocked_site_sections", [key for key in request.form.getlist("blocked_site_sections") if key in allowed_section_keys])
        db.session.commit()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(ok=True)
        flash("Настройки сохранены", "success")
        return redirect(url_for("main.site_settings"))
    return render_template(
        "settings.html",
        hide_documents_section=_setting_bool("hide_documents_section"),
        mobile_version_under_development=_setting_bool("mobile_version_under_development"),
        site_maintenance_mode=_setting_bool("site_maintenance_mode"),
        two_factor_every_login=_setting_bool("two_factor_every_login"),
        blocked_site_sections=_setting_csv("blocked_site_sections"),
        section_lock_choices=SECTION_LOCK_CHOICES,
    )


@bp.route("/account", methods=["GET", "POST"])
@login_required
def account():
    user = db.session.get(User, current_user.id) or abort(404)
    pending_secret = session.get("account_2fa_pending_secret")
    provisioning = ""
    qr_data_uri = ""

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "start_2fa":
            pending_secret = generate_totp_secret()
            session["account_2fa_pending_secret"] = pending_secret
            flash("Отсканируйте QR-код или введите ключ вручную, затем подтвердите кодом.", "info")
        elif action == "confirm_2fa":
            secret = pending_secret or user.two_factor_secret
            code = request.form.get("two_factor_code")
            if verify_totp(secret, code):
                user.two_factor_secret = secret
                user.two_factor_enabled = True
                user.two_factor_confirmed_at = datetime.utcnow()
                db.session.commit()
                session.pop("account_2fa_pending_secret", None)
                pending_secret = None
                flash("Двухэтапная аутентификация подключена.", "success")
            else:
                flash("Код не подошёл. Проверьте время на телефоне и попробуйте ещё раз.", "danger")
        elif action == "disable_2fa":
            code = request.form.get("two_factor_code")
            if user.two_factor_enabled and not verify_totp(user.two_factor_secret, code):
                flash("Введите действующий код, чтобы отключить двухэтапную аутентификацию.", "danger")
            else:
                user.two_factor_enabled = False
                user.two_factor_secret = None
                user.two_factor_confirmed_at = None
                db.session.commit()
                session.pop("account_2fa_pending_secret", None)
                pending_secret = None
                flash("Двухэтапная аутентификация отключена.", "success")
        return redirect(url_for("main.account"))

    if pending_secret:
        provisioning = provisioning_uri(user.username, pending_secret)
        qr_data_uri = qr_svg_data_uri(provisioning)

    return render_template(
        "account.html",
        pending_secret=pending_secret,
        provisioning_uri=provisioning,
        qr_data_uri=qr_data_uri,
    )


@bp.route("/users", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def users():
    form = UserForm()
    all_projects = Project.query.order_by(Project.name.asc(), Project.id.asc()).all()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data.strip()).first()
        if user:
            flash("Пользователь с таким логином уже есть", "danger")
        else:
            project = selected_project()
            requested_project_ids = {
                value for value in request.form.getlist("project_ids") if str(value).isdigit()
            }
            valid_project_ids = {project.id for project in all_projects}
            project_ids = sorted({int(value) for value in requested_project_ids if int(value) in valid_project_ids})
            if not project_ids:
                flash("Выберите хотя бы один объект.", "danger")
                users = User.query.order_by(User.created_at.desc()).all()
                return render_template("users.html", users=users, form=form, project=project, all_projects=all_projects)
            user = User(
                username=form.username.data.strip(),
                full_name=form.full_name.data.strip() if form.full_name.data else None,
                role=form.role.data,
                is_active=True,
            )
            user.set_project_access(project_ids, all_projects=False)
            password = form.password.data
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash("Поздравляем, пользователь создан.", "success")
            return redirect(url_for("main.users"))
    project = selected_project()
    users = User.query.order_by(User.created_at.desc()).all()
    if project:
        users = [
            user for user in users
            if user.role in {ROLE_ADMIN, ROLE_MANAGER, ROLE_VERIFIER} or user.can_access_project(project)
        ]
    return render_template("users.html", users=users, form=form, project=project, all_projects=all_projects)


@bp.route("/users/<int:user_id>/projects", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def user_update_projects(user_id: int):
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    user = db.session.get(User, user_id) or abort(404)
    if user.role == ROLE_ADMIN:
        if wants_json:
            return jsonify(ok=False, message="Разработчику всегда доступны все объекты."), 400
        flash("Разработчику всегда доступны все объекты.", "info")
        return redirect(url_for("main.users"))
    all_projects = Project.query.with_entities(Project.id).all()
    valid_project_ids = {project_id for (project_id,) in all_projects}
    requested_ids = {
        int(value) for value in request.form.getlist("project_ids") if str(value).isdigit()
    }
    project_ids = sorted(requested_ids & valid_project_ids)
    if not project_ids:
        if wants_json:
            return jsonify(ok=False, message="Выберите хотя бы один объект."), 400
        flash("Выберите хотя бы один объект.", "danger")
        return redirect(url_for("main.users"))
    user.set_project_access(project_ids, all_projects=False)
    db.session.commit()
    if wants_json:
        count = len(project_ids)
        label = "1 объект" if count == 1 else (f"{count} объекта" if 2 <= count <= 4 else f"{count} объектов")
        return jsonify(ok=True, project_ids=project_ids, count=count, label=label, message="Доступ к объектам сохранён.")
    flash("Доступ к объектам обновлён.", "success")
    return redirect(url_for("main.users"))


@bp.route("/users/<int:user_id>/name", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def user_update_name(user_id: int):
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"
    user = db.session.get(User, user_id) or abort(404)
    _abort_if_user_outside_current_project(user)
    full_name = str(request.form.get("full_name") or "").strip()
    if len(full_name) > 160:
        if wants_json:
            return jsonify(ok=False, message="Имя не должно превышать 160 символов."), 400
        flash("Имя не должно превышать 160 символов.", "danger")
        return redirect(url_for("main.users"))
    user.full_name = full_name or None
    db.session.commit()
    if wants_json:
        return jsonify(ok=True, full_name=user.full_name or "", label=user.full_name or "—", message="Имя сохранено.")
    flash("Имя пользователя обновлено.", "success")
    return redirect(url_for("main.users"))


@bp.route("/users/<int:user_id>/captcha", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def user_toggle_captcha(user_id: int):
    user = db.session.get(User, user_id) or abort(404)
    _abort_if_user_outside_current_project(user)
    user.captcha_disabled = request.form.get("captcha_disabled") == "1"
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(
            ok=True,
            captcha_disabled=user.captcha_disabled,
            message="Настройка капчи сохранена.",
        )
    flash("Настройка капчи для пользователя обновлена.", "success")
    return redirect(url_for("main.users"))


@bp.route("/users/<int:user_id>/password", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def user_set_password(user_id: int):
    user = db.session.get(User, user_id) or abort(404)
    _abort_if_user_outside_current_project(user)
    if user.id == current_user.id:
        flash("Нельзя изменить пароль текущего пользователя здесь", "danger")
        return redirect(url_for("main.users"))
    password = (request.form.get("password") or "").strip()
    if len(password) < 8:
        flash("Пароль должен быть минимум 8 символов", "danger")
        return redirect(url_for("main.users"))
    user.set_password(password)
    db.session.commit()
    flash("Пароль обновлён", "success")
    return redirect(url_for("main.users"))


@bp.route("/users/<int:user_id>/password", methods=["GET"])
@login_required
@role_required(ROLE_ADMIN)
def user_set_password_page(user_id: int):
    user = db.session.get(User, user_id) or abort(404)
    _abort_if_user_outside_current_project(user)
    if user.id == current_user.id:
        flash("Нельзя изменить пароль текущего пользователя здесь", "danger")
        return redirect(url_for("main.users"))
    form = UserPasswordForm()
    return render_template("user_password.html", user=user, form=form)


@bp.route("/users/<int:user_id>/delete/confirm", methods=["GET"])
@login_required
@role_required(ROLE_ADMIN)
def user_delete_confirm(user_id: int):
    user = db.session.get(User, user_id) or abort(404)
    _abort_if_user_outside_current_project(user)
    if user.id == current_user.id:
        flash("Нельзя удалить текущего пользователя", "danger")
        return redirect(url_for("main.users"))
    return render_template("user_delete_confirm.html", user=user)


@bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def user_delete(user_id: int):
    user = db.session.get(User, user_id) or abort(404)
    _abort_if_user_outside_current_project(user)
    if user.id == current_user.id:
        flash("Нельзя удалить текущего пользователя", "danger")
        return redirect(url_for("main.users"))
    project = selected_project()
    _record_simple_deletion(
        "user_delete",
        "user",
        user,
        user.full_name or user.username,
        f"Удалён аккаунт пользователя: {user.full_name or user.username}.",
        project_id=project.id if project else None,
        extra={"project_ids": sorted(user.project_access_ids), "all_projects": user.can_access_all_projects},
    )
    db.session.delete(user)
    db.session.commit()
    flash("Аккаунт удалён", "success")
    return redirect(url_for("main.users"))


@bp.route("/sync-logs")
@login_required
def sync_logs():
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    logs = SyncLog.query.filter(SyncLog.project_id == project.id).order_by(SyncLog.started_at.desc()).limit(100).all()
    return render_template("sync_logs.html", logs=logs, project=project)


def _sync_log_snapshot_payload(log: SyncLog | None) -> dict:
    if log is None or not log.rollback_data:
        return {}
    try:
        payload = json.loads(log.rollback_data)
    except (TypeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _sync_log_after_payload(log: SyncLog, project_id: int) -> tuple[dict, bool]:
    next_log = (
        SyncLog.query.filter(
            SyncLog.project_id == project_id,
            SyncLog.started_at > log.started_at,
            SyncLog.rollback_data.isnot(None),
        )
        .order_by(SyncLog.started_at.asc(), SyncLog.id.asc())
        .first()
    )
    payload = _sync_log_snapshot_payload(next_log)
    if payload:
        return payload, True
    return {
        "apartments": [
            {
                "id": apartment.id,
                "apartment_number": apartment.apartment_number,
                "construction_number": apartment.construction_number,
                "premise_type": apartment.premise_type,
                "building": apartment.building,
                "owner_name": apartment.owner_name,
                "finishing_type": apartment.finishing_type,
                "is_unsold": apartment.is_unsold,
                "is_app_mode": apartment.is_app_mode,
            }
            for apartment in Apartment.query.filter_by(project_id=project_id).all()
        ],
        "tasks": [
            {
                "id": task.id,
                "apartment_id": task.apartment_id,
                "work_point_id": task.work_point_id,
                "description": task.description,
                "source_cell_value": task.source_cell_value,
                "source_sheet_name": task.source_sheet_name,
                "source_cell_address": task.source_cell_address,
                "is_missing_in_latest_sync": task.is_missing_in_latest_sync,
            }
            for task in Task.query.filter_by(project_id=project_id).all()
        ],
    }, False


def _sync_snapshot_premise_number(apartment: dict | None) -> str:
    apartment = apartment or {}
    premise_type = apartment.get("premise_type") or "apartment"
    for candidate in (apartment.get("apartment_number"), apartment.get("construction_number")):
        number = str(candidate or "").strip()
        if not number:
            continue
        if premise_type == "commercial" or looks_like_apartment_identifier(number):
            return number
    return ""


def _sync_snapshot_premise_label(apartment: dict | None) -> str:
    apartment = apartment or {}
    number = _sync_snapshot_premise_number(apartment) or "—"
    if (apartment.get("premise_type") or "apartment") == "commercial":
        building = str(apartment.get("building") or "").strip()
        return f"Комм. {number}" + (f", корпус {building}" if building else "")
    return f"кв {number}"


def _sync_log_conflicts(log: SyncLog, project_id: int) -> list[SyncConflict]:
    next_log = (
        SyncLog.query.filter(SyncLog.project_id == project_id, SyncLog.started_at > log.started_at)
        .order_by(SyncLog.started_at.asc(), SyncLog.id.asc())
        .first()
    )
    query = (
        SyncConflict.query.outerjoin(Task, SyncConflict.task_id == Task.id)
        .outerjoin(Apartment, SyncConflict.apartment_id == Apartment.id)
        .filter(SyncConflict.created_at >= log.started_at - timedelta(seconds=5))
        .filter(or_(Task.project_id == project_id, Apartment.project_id == project_id))
    )
    if next_log is not None:
        query = query.filter(SyncConflict.created_at < next_log.started_at)
    elif log.finished_at is not None:
        query = query.filter(SyncConflict.created_at <= log.finished_at + timedelta(seconds=5))
    return query.order_by(SyncConflict.created_at.asc(), SyncConflict.id.asc()).all()


@bp.route("/sync-logs/<int:log_id>/details")
@login_required
def sync_log_details(log_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    log = db.session.get(SyncLog, log_id) or abort(404)
    if not log.project_id or log.project_id != project.id:
        abort(404)

    after_payload, exact_snapshot = _sync_log_after_payload(log, project.id)
    apartments = {
        int(item["id"]): item
        for item in (after_payload.get("apartments") or [])
        if isinstance(item, dict) and item.get("id") is not None
    }
    details: list[dict] = []
    current_task_ids = {
        task_id for (task_id,) in db.session.query(Task.id).filter(Task.project_id == project.id).all()
    }

    if log.source_type in {"excel", "google_sheets"}:
        task_rows = [
            item for item in (after_payload.get("tasks") or [])
            if isinstance(item, dict) and item.get("is_missing_in_latest_sync")
        ]
        work_point_ids = {
            int(item["work_point_id"])
            for item in task_rows
            if item.get("work_point_id") is not None
        }
        work_points = {
            point.id: point
            for point in WorkPoint.query.filter(WorkPoint.id.in_(work_point_ids)).all()
        } if work_point_ids else {}
        for item in task_rows:
            point = work_points.get(int(item.get("work_point_id") or 0))
            if point is not None and str(point.point_number or "") not in VISIBLE_WORK_POINT_NUMBERS:
                continue
            apartment = apartments.get(int(item.get("apartment_id") or 0))
            if not _sync_snapshot_premise_number(apartment):
                continue
            details.append({
                "premise": _sync_snapshot_premise_label(apartment),
                "field": point.display_name if point else "Замечание",
                "value": str(item.get("description") or item.get("source_cell_value") or "").strip(),
                "location": " · ".join(filter(None, [
                    str(item.get("source_sheet_name") or "").strip(),
                    str(item.get("source_cell_address") or "").strip(),
                ])),
                "reason": "Не найдено в загруженной таблице",
                "task_id": item.get("id") if item.get("id") in current_task_ids else None,
            })
    elif log.source_type == "transfer_excel":
        for apartment in apartments.values():
            if (
                not _sync_snapshot_premise_number(apartment)
                or apartment.get("is_unsold")
                or apartment.get("is_app_mode")
            ):
                continue
            details.append({
                "premise": _sync_snapshot_premise_label(apartment),
                "field": "Статус передачи",
                "value": str(apartment.get("owner_name") or apartment.get("finishing_type") or "").strip(),
                "location": "",
                "reason": "Ожидает приёмки (не АПП)",
                "task_id": None,
            })

    conflicts = _sync_log_conflicts(log, project.id)
    details.sort(key=lambda item: (_apartment_number_sort_value(item["premise"]), item["field"]))
    return render_template(
        "sync_log_details.html",
        log=log,
        project=project,
        details=details,
        conflicts=conflicts,
        exact_snapshot=exact_snapshot,
    )


@bp.route("/sync-logs/<int:log_id>/delete", methods=["POST"])
@login_required
def delete_sync_log(log_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    log = db.session.get(SyncLog, log_id) or abort(404)
    if not log.project_id or log.project_id != project.id:
        abort(404)
    delete_from = log.started_at
    if not log.rolled_back_at:
        ok, message = apply_sync_rollback(log)
        if not ok:
            flash(message, "warning")
            return redirect(url_for("main.sync_logs"))
        affected_logs = SyncLog.query.filter(
            SyncLog.project_id == project.id,
            SyncLog.started_at >= delete_from,
        ).all()
    else:
        affected_logs = [log]
    deleted_logs = len(affected_logs)
    for affected_log in affected_logs:
        _record_simple_deletion(
            "sync_log_delete",
            "sync_log",
            affected_log,
            f"Синхронизация #{affected_log.id}",
            f"Удалена запись журнала синхронизации: {affected_log.source_type} {affected_log.source_name or ''}.",
            project_id=project.id,
        )
        db.session.delete(affected_log)
    _refresh_sync_dashboard_settings(project.id)
    db.session.commit()
    flash(f"Синхронизация откатана, данные загрузки удалены, записей журнала удалено: {deleted_logs}", "success")
    return redirect(url_for("main.sync_logs"))


@bp.route("/sync-logs/<int:log_id>/rollback", methods=["POST"])
@login_required
def rollback_sync_log(log_id: int):
    if current_user.role not in {ROLE_ADMIN, ROLE_MANAGER}:
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    log = db.session.get(SyncLog, log_id) or abort(404)
    if not log.project_id or log.project_id != project.id:
        abort(404)
    ok, message = apply_sync_rollback(log)
    if ok:
        _refresh_sync_dashboard_settings(project.id)
        db.session.commit()
    flash(message, "success" if ok else "warning")
    return redirect(url_for("main.sync_logs"))


def _parse_conflict_value_for_field(field_name: str | None, value: str | None):
    text = (value or "").strip()
    if field_name in {"inspection_date", "reinspection_date", "deadline_date", "remark_deadline_date", "app_deadline_date", "avr_signed_date"}:
        return parse_date(text)
    if field_name == "is_app_mode":
        return text.lower() in {"1", "true", "yes", "да", "апп"}
    if field_name in {"owner_name", "phone", "finishing_type", "entrance", "floor", "app_deadline_raw", "app_deadline_status", "avr_status", "comment", "inspection_note"}:
        return text or None
    return text or None


def _apply_sync_conflict_new_value(conflict: SyncConflict) -> None:
    if (conflict.target_type or "task") == "apartment":
        apartment = conflict.apartment or (conflict.task.apartment if conflict.task else None)
        if apartment is None or not conflict.field_name:
            return
        setattr(apartment, conflict.field_name, _parse_conflict_value_for_field(conflict.field_name, conflict.new_value))
        return

    task = conflict.task
    if not task:
        return
    if conflict.field_name == "status":
        old_status = task.status
        new_status = str(conflict.new_value or "").strip() or STATUS_NOT_STARTED
        task.status = new_status
        task.is_done = new_status in DONE_STATUSES
        if task.is_done:
            task.completed_date = task.completed_date or datetime.utcnow()
        elif old_status in DONE_STATUSES:
            task.completed_date = None
        task.manually_edited = True
        if old_status != new_status:
            log_change(task, "status_change", "status", old_status, new_status, user_id=current_user.id)
        return
    task.source_cell_value = conflict.new_value
    task.source_hash = conflict.new_hash
    # Пользователь нажал «Принять новое» — значит новый текст из Excel должен стать видимым текстом замечания.
    if conflict.field_name in {None, "source_cell_value", "description"}:
        task.description = conflict.new_value


def _create_task_from_sync_conflict(conflict: SyncConflict) -> Task | None:
    task = conflict.task
    if not task or not task.project or not task.apartment or not task.work_point:
        return None
    new_text = str(conflict.new_value or "").strip()
    if not new_text:
        return None
    if normalize_text(_task_effective_remark_text(task)) == normalize_text(new_text):
        return None
    for sibling in _remark_task_candidates(task.project_id, task.apartment_id, task.work_point_id):
        if sibling.id == task.id:
            continue
        if normalize_text(_task_effective_remark_text(sibling)) == normalize_text(new_text):
            return sibling
        pending_conflict = _pending_task_sync_conflict(sibling.id)
        if pending_conflict and normalize_text(pending_conflict.new_value or "") == normalize_text(new_text):
            return sibling
    source_sheet_name = (task.source_sheet_name or "").strip() or "manual"
    if (conflict.source_type or "").strip() == "pdf_recognition":
        source_sheet_name = "pdf_recognition"
    elif (conflict.source_type or "").strip() == "manual_act":
        source_sheet_name = "manual_act"
    elif (conflict.source_type or "").strip() == "manual":
        source_sheet_name = "manual"
    elif (conflict.source_type or "").strip() == "excel" and conflict.sheet_name:
        source_sheet_name = conflict.sheet_name
    return _create_manual_remark_task(
        project=task.project,
        apartment=task.apartment,
        point_number=task.work_point.point_number,
        text=new_text,
        source_sheet_name=source_sheet_name,
        action="sync_conflict_keep_both_created",
    )


@bp.route("/conflicts")
@login_required
def sync_conflicts():
    if not can_manage_sync(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    conflicts = (
        _project_pending_conflicts_query(project.id)
        .order_by(SyncConflict.created_at.desc())
        .limit(500)
        .all()
    )
    return render_template("sync_conflicts.html", conflicts=conflicts, status_labels=TASK_STATUSES)


@bp.route("/conflicts/<int:conflict_id>/<action>", methods=["POST"])
@login_required
def resolve_conflict(conflict_id: int, action: str):
    if not can_manage_sync(current_user):
        abort(403)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    conflict = db.session.get(SyncConflict, conflict_id) or abort(404)
    conflict_project_id = None
    if conflict.task:
        conflict_project_id = conflict.task.project_id
    elif conflict.apartment:
        conflict_project_id = conflict.apartment.project_id
    if conflict_project_id != project.id:
        abort(404)
    if conflict.status != "pending":
        return redirect(url_for("main.sync_conflicts"))
    if action not in {"keep_old", "apply_new", "keep_both"}:
        abort(400)
    if action == "apply_new":
        _apply_sync_conflict_new_value(conflict)
    elif action == "keep_both":
        if (conflict.target_type or "task") != "task" or conflict.field_name not in {None, "source_cell_value", "description"}:
            abort(400)
        _create_task_from_sync_conflict(conflict)
    conflict.status = action
    conflict.resolved_at = datetime.utcnow()
    conflict.resolved_by_user_id = current_user.id
    db.session.commit()
    flash("Конфликт синхронизации решён", "success")
    return redirect(url_for("main.sync_conflicts"))


@bp.route("/conflicts/bulk/<action>", methods=["POST"])
@login_required
def resolve_conflicts_bulk(action: str):
    if not can_manage_sync(current_user):
        abort(403)
    if action not in {"keep_old", "apply_new"}:
        abort(400)
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    conflicts = _project_pending_conflicts_query(project.id).all()
    changed = 0
    for conflict in conflicts:
        if action == "apply_new":
            _apply_sync_conflict_new_value(conflict)
        conflict.status = action
        conflict.resolved_at = datetime.utcnow()
        conflict.resolved_by_user_id = current_user.id
        changed += 1
    db.session.commit()
    flash(f"Готово: обработано несостыковок {changed}", "success")
    return redirect(url_for("main.sync_conflicts"))


@bp.route("/tasks/<int:task_id>/inline-text", methods=["POST"])
@login_required
def inline_update_text(task_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if not can_change_task(current_user, task):
        abort(403)
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text") or "").strip()
    old_text = task.description
    task.description = text or None
    if old_text != task.description:
        log_change(task, "field_update", "description", old_text, task.description, user_id=current_user.id)
    task.manually_edited = True
    db.session.commit()
    history_change = next((change for change in reversed(task.changes) if change.action == "field_update" and change.field_name == "description"), None)
    history_entry = _build_change_history_entry(history_change, task=task, users_cache={}) if history_change else None
    return jsonify({"ok": True, "text": task.description or "", "history_entry": history_entry})


@bp.route("/tasks/<int:task_id>/split", methods=["POST"])
@login_required
def split_task_remark(task_id: int):
    project = selected_project()
    if project is None:
        return redirect(url_for("main.objects"))
    task = db.session.get(Task, task_id) or abort(404)
    if task.project_id != project.id:
        abort(404)
    if not can_change_task(current_user, task):
        abort(403)
    payload = request.get_json(silent=True) or {}
    current_text = " ".join(str(payload.get("current_text") or "").split()).strip()
    new_text = " ".join(str(payload.get("new_text") or "").split()).strip()
    current_status = str(payload.get("current_status") or task.status or STATUS_NOT_STARTED).strip()
    new_status = str(payload.get("new_status") or STATUS_NOT_STARTED).strip()

    if not current_text or not new_text:
        return jsonify({"ok": False, "message": "Заполните обе части замечания"}), 400
    if current_text == new_text:
        return jsonify({"ok": False, "message": "Тексты частей должны отличаться"}), 400
    if current_status not in TASK_STATUSES or new_status not in TASK_STATUSES:
        return jsonify({"ok": False, "message": "Выбран некорректный статус"}), 400
    if current_status == "problem" or new_status == "problem":
        return jsonify({"ok": False, "message": "Статус «Проблема» задайте после разделения через обычную кнопку проблемы"}), 400

    previous_text = task.description or task.source_cell_value or ""
    task.description = current_text
    task.manually_edited = True
    if previous_text != current_text:
        split_payload = json.dumps({"first": current_text, "second": new_text}, ensure_ascii=False)
        log_change(task, "manual_split", "description", previous_text, split_payload, user_id=current_user.id)
    change_task_status(task, current_status, user_id=current_user.id, commit=False)

    point_number = task.work_point.point_number if task.work_point and task.work_point.point_number else "22"
    new_task = _create_manual_remark_task(
        project=project,
        apartment=task.apartment,
        point_number=point_number,
        text=new_text,
        source_sheet_name="manual_split",
        action="manual_split_created",
    )
    change_task_status(new_task, new_status, user_id=current_user.id, commit=False)
    db.session.commit()
    return jsonify(
        {
            "ok": True,
            "message": "Замечание разделено на две части",
            "current_task_id": task.id,
            "new_task_id": new_task.id,
            "current_text": task.description or "",
            "new_text": new_task.description or new_task.source_cell_value or "",
        }
    )

