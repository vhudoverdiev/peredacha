from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from flask import current_app
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from app import db
from app.models import Project, SyncLog, Task, WorkPoint
from app.services.task_service import VISIBLE_WORK_POINT_NUMBERS, set_setting, sync_rows
from app.services.changelog_service import log_change
from app.services.sync_rollback import build_project_rollback_data


def _cell_rgb(cell: Any) -> tuple[int, int, int] | None:
    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "fill_type", None):
        return None
    color = getattr(fill, "fgColor", None)
    if not color:
        return None
    rgb = None
    if getattr(color, "type", None) == "rgb":
        rgb = getattr(color, "rgb", None)
    elif getattr(color, "type", None) == "indexed":
        # openpyxl indexed palette: index 52 is a common orange/yellow fill.
        indexed = getattr(color, "indexed", None)
        indexed_map = {52: "F4B183", 53: "FCE4D6", 44: "FFC000", 45: "FFFF00"}
        rgb = indexed_map.get(indexed)
    if not rgb or not isinstance(rgb, str):
        return None
    rgb = rgb[-6:]
    try:
        return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return None


def _is_orange_unsold_fill(cell: Any) -> bool:
    rgb = _cell_rgb(cell)
    if not rgb:
        return False
    r, g, b = rgb
    # Orange/yellow fills used in Excel/Google Sheets for unsold premises.
    return r >= 210 and 95 <= g <= 215 and b <= 120


def save_upload(file: FileStorage) -> Path:
    folder = Path(current_app.config["UPLOAD_FOLDER"])
    folder.mkdir(parents=True, exist_ok=True)
    filename = secure_filename(file.filename or "table.xlsx")
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    target = folder / f"{timestamp}_{filename}"
    file.save(target)
    return target


def worksheet_to_rows(path: Path, sheet_name: str | None = None) -> tuple[str, list[list[Any]]]:
    # For imports we need computed values, not formulas like "=A113+1".
    # If the file was saved from Excel/Sheets, cached results are typically present.
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    return ws.title, rows


def worksheet_to_rows_with_strikes(path: Path, sheet_name: str | None = None) -> tuple[str, list[list[Any]], set[tuple[int, int]], set[tuple[int, int]]]:
    """
    Return (sheet_title, rows, struck_cells, orange_cells).
    Coordinates are (row_index_1_based, col_index_1_based).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows: list[list[Any]] = []
    struck: set[tuple[int, int]] = set()
    orange: set[tuple[int, int]] = set()
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        out_row: list[Any] = []
        for c_idx, cell in enumerate(row, start=1):
            out_row.append(cell.value)
            if getattr(getattr(cell, "font", None), "strike", False):
                struck.add((r_idx, c_idx))
            if _is_orange_unsold_fill(cell):
                orange.add((r_idx, c_idx))
        rows.append(out_row)
    return ws.title, rows, struck, orange


def workbook_sheets_to_rows_with_strikes(path: Path) -> list[tuple[str, list[list[Any]], set[tuple[int, int]], set[tuple[int, int]]]]:
    """Читает все листы Excel, чтобы квартиры и коммерция попадали в CRM одним импортом."""
    wb = load_workbook(path, data_only=True)
    result: list[tuple[str, list[list[Any]], set[tuple[int, int]], set[tuple[int, int]]]] = []
    active_title = wb.active.title
    for ws in wb.worksheets:
        # Основной лист замечаний + отдельный лист коммерции.
        # Служебные листы по категориям не импортируем, чтобы не плодить дубли задач.
        if ws.title != active_title and "коммер" not in (ws.title or "").strip().lower():
            continue
        rows: list[list[Any]] = []
        struck: set[tuple[int, int]] = set()
        orange: set[tuple[int, int]] = set()
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            out_row: list[Any] = []
            for c_idx, cell in enumerate(row, start=1):
                out_row.append(cell.value)
                if getattr(getattr(cell, "font", None), "strike", False):
                    struck.add((r_idx, c_idx))
                if _is_orange_unsold_fill(cell):
                    orange.add((r_idx, c_idx))
            rows.append(out_row)
        if any(any(str(value or "").strip() for value in row) for row in rows):
            result.append((ws.title, rows, struck, orange))
    return result


def mark_missing_tasks(project_id: int, seen_uids: set[str]) -> int:
    query = Task.query.filter(Task.project_id == project_id)
    query = query.filter(Task.work_point.has(WorkPoint.point_number.in_(VISIBLE_WORK_POINT_NUMBERS)))
    if seen_uids:
        query = query.filter(~Task.source_uid.in_(seen_uids))
    missing_count = 0
    for task in query.all():
        if not task.is_missing_in_latest_sync:
            task.is_missing_in_latest_sync = True
            log_change(task, "missing_in_latest_sync", "is_missing_in_latest_sync", False, True)
        missing_count += 1
    return missing_count


def preview_excel(path: Path, limit: int = 20) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"sheets": [], "active_sheet": wb.active.title, "rows": []}
    for ws in wb.worksheets:
        result["sheets"].append({"title": ws.title, "max_row": ws.max_row, "max_column": ws.max_column})
    for idx, row in enumerate(wb.active.iter_rows(values_only=True), start=1):
        result["rows"].append(["" if value is None else value for value in row])
        if idx >= limit:
            break
    return result


def sync_excel_file(path: Path, sheet_name: str | None = None, project_name: str = "100 Квартал 7 очередь") -> dict:
    project = Project.query.filter_by(name=project_name).first()
    sync_log = SyncLog(source_type="excel", source_name=str(path), started_at=datetime.utcnow(), status="running", project_id=project.id if project else None)
    sync_log.rollback_data = build_project_rollback_data(project.id if project else None)
    db.session.add(sync_log)
    db.session.commit()
    try:
        if sheet_name:
            sheets = [worksheet_to_rows_with_strikes(path, sheet_name=sheet_name)]
        else:
            sheets = workbook_sheets_to_rows_with_strikes(path)

        total_result = {"created_count": 0, "updated_count": 0, "missing_count": 0, "seen_uids": set()}
        for actual_sheet_name, rows, struck_cells, orange_cells in sheets:
            result = sync_rows(
                rows,
                sheet_name=actual_sheet_name,
                project_name=project_name,
                source_name=f"{path.name} / {actual_sheet_name}",
                struck_cells=struck_cells,
                unsold_color_cells=orange_cells,
                mark_missing=False,
            )
            total_result["created_count"] += int(result.get("created_count", 0))
            total_result["updated_count"] += int(result.get("updated_count", 0))
            total_result["seen_uids"].update(result.get("seen_uids", set()) or set())

        project = Project.query.filter_by(name=project_name).first()
        sync_log.project_id = project.id if project else sync_log.project_id
        if project:
            total_result["missing_count"] = mark_missing_tasks(project.id, total_result["seen_uids"])

        set_setting("latest_excel_path", str(path))
        sync_log.created_count = int(total_result.get("created_count", 0))
        sync_log.updated_count = int(total_result.get("updated_count", 0))
        sync_log.missing_count = int(total_result.get("missing_count", 0))
        sync_log.status = "success"
        sync_log.finished_at = datetime.utcnow()
        db.session.commit()
        return total_result
    except Exception as exc:
        db.session.rollback()
        sync_log.status = "error"
        sync_log.error_message = str(exc)
        sync_log.finished_at = datetime.utcnow()
        db.session.add(sync_log)
        db.session.commit()
        raise
