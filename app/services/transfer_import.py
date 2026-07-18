from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app import db
from app.models import Apartment, Project, SyncLog
from app.services.task_service import (
    apply_app_deadline_logic,
    get_or_create_project,
    is_service_premise_text,
    is_unsold_owner_name,
    looks_like_apartment_identifier,
    normalize_apartment_number_cell,
    normalize_building_marker,
    normalize_finishing_type,
    normalize_number_cell,
    parse_date,
)
from app.services.sync_rollback import build_project_rollback_data
from app.services.uid_service import normalize_text


APP_DATE_RE = re.compile(r"\bапп\b[^\d]*(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?", re.IGNORECASE)
APP_MODE_RE = re.compile(r"\bапп\b", re.IGNORECASE)


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _find_header_map(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    exact_headers = {
        "number": ("№ кв", "№ квартиры"),
        "owner_name": ("ф.и.о. дольщиков", "фио дольщиков", "ф.и.о дольщиков"),
        "phone": ("телефон",),
        "finishing_type": ("вид отделки",),
        "inspection_note": ("дата осмотра",),
        "first_inspection_date": ("дата первичного осмотра",),
        "reinspection_date": ("дата повторного осмотра",),
        "app_signed_date": ("дата подписания апп",),
        "remark_deadline_date": ("срок устранения замечаний по апп",),
    }
    aliases = {
        "number": ("№ кв", "№ ком", "помещения"),
        "owner_name": ("ф.и.о", "фио", "дольщик", "дольщиков"),
        "phone": ("телефон",),
        "finishing_type": ("вид отделки",),
        "comment": ("комментар",),
    }
    for row_idx, row in enumerate(rows):
        normalized = [_normalize_header(value) for value in row]
        mapping: dict[str, int] = {}
        for col_idx, header in enumerate(normalized):
            if not header:
                continue
            for field, candidates in exact_headers.items():
                if field not in mapping and header in candidates:
                    mapping[field] = col_idx
        for col_idx, header in enumerate(normalized):
            if not header:
                continue
            for field, needles in aliases.items():
                if field not in mapping and any(needle in header for needle in needles):
                    mapping[field] = col_idx
        if "number" in mapping and "inspection_note" in mapping:
            return row_idx, mapping
    raise ValueError("Не найдены заголовки таблицы статистики передач.")


def _is_transfer_header_map(mapping: dict[str, int]) -> bool:
    if "number" not in mapping or "inspection_note" not in mapping:
        return False
    extra_fields = {
        "owner_name",
        "phone",
        "finishing_type",
        "first_inspection_date",
        "reinspection_date",
        "remark_deadline_date",
        "app_signed_date",
    }
    extra_hits = sum(1 for field in extra_fields if field in mapping)
    return extra_hits >= 2


def _value_at(row: list[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _cell_at(row: list[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _cell_rgb(cell: Any) -> tuple[int, int, int] | None:
    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "fill_type", None):
        return None
    color = getattr(fill, "fgColor", None)
    if not color:
        return None
    rgb = None
    if getattr(color, "type", None) == "rgb":
        rgb = getattr(color, "rgb", None)
    elif getattr(color, "type", None) == "indexed":
        indexed = getattr(color, "indexed", None)
        indexed_map = {52: "F4B183", 53: "FCE4D6", 44: "FFC000", 45: "FFFF00"}
        rgb = indexed_map.get(indexed)
    if not rgb or not isinstance(rgb, str):
        return None
    rgb = rgb[-6:]
    try:
        return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return None


def _is_green_fill(cell: Any) -> bool:
    rgb = _cell_rgb(cell)
    if not rgb:
        return False
    r, g, b = rgb
    return g >= 120 and g > r + 25 and g > b + 25


def _is_orange_unsold_fill(cell: Any) -> bool:
    rgb = _cell_rgb(cell)
    if not rgb:
        return False
    r, g, b = rgb
    # Google Sheets / Excel orange used for "не продано" is close to #FBBC04.
    # Yellow section headers are skipped before this check, so they are not counted.
    return r >= 210 and 95 <= g <= 215 and b <= 120


def _is_section_row(value: Any) -> bool:
    return is_service_premise_text(normalize_number_cell(value))


def _parse_app_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return None
    if isinstance(value, date):
        return None
    text = str(value).strip()
    match = APP_DATE_RE.search(text)
    if not match:
        return None
    day = int(match.group(1))
    month = int(match.group(2))
    year_text = match.group(3)
    if year_text:
        year = int(year_text)
        if year < 100:
            year += 2000
    else:
        year = datetime.today().year
    return date(year, month, day)


def _parse_inspection_schedule(value: Any) -> datetime | date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text or _is_app_mode(text):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            if "H" not in fmt:
                return parsed.date()
            return parsed
        except ValueError:
            continue
    parsed_date = parse_date(text)
    return parsed_date


def _inspection_schedule_marker(value: datetime | date | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        value = value.replace(microsecond=0)
        payload = value.isoformat(timespec="seconds")
    else:
        payload = value.isoformat()
    return f"__inspection_schedule__:{payload}"


def _is_app_mode(value: Any) -> bool:
    return bool(APP_MODE_RE.search(str(value or "").strip()))


def _is_unsold(owner_name: str | None) -> bool:
    return is_unsold_owner_name(owner_name)


def inspect_transfer_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    matched_sheets: list[str] = []
    for ws in wb.worksheets:
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        if not rows or not any(any(str(value or "").strip() for value in row) for row in rows):
            continue
        try:
            _header_index, mapping = _find_header_map(rows)
        except ValueError:
            continue
        if _is_transfer_header_map(mapping):
            matched_sheets.append(ws.title)
    return {
        "ok": bool(matched_sheets),
        "matched_sheets": matched_sheets,
        "sheet_count": len(wb.worksheets),
    }


def sync_transfer_statistics(path: Path, project_name: str) -> dict[str, int]:
    project = get_or_create_project(project_name)
    sync_log = SyncLog(source_type="transfer_excel", source_name=str(path), started_at=datetime.utcnow(), status="running", project_id=project.id)
    sync_log.rollback_data = build_project_rollback_data(project.id)
    db.session.add(sync_log)
    db.session.commit()
    try:
        wb = load_workbook(path, data_only=True)
        result = {"created_count": 0, "updated_count": 0, "accepted_count": 0, "waiting_count": 0, "unsold_count": 0}

        for ws in wb.worksheets:
            is_commercial_sheet = "коммер" in (ws.title or "").strip().lower()
            current_building: str | None = None
            commercial_numbers_in_building: set[str] = set()
            if is_commercial_sheet and not getattr(project, "has_commercial", True):
                continue
            if not is_commercial_sheet and not getattr(project, "has_apartments", True):
                continue
            cell_rows = [list(row) for row in ws.iter_rows()]
            rows = [[cell.value for cell in row] for row in cell_rows]
            if not rows:
                continue
            header_index, mapping = _find_header_map(rows)
            if not _is_transfer_header_map(mapping):
                continue
            for row_idx, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                cell_row = cell_rows[row_idx - 1] if row_idx - 1 < len(cell_rows) else []
                number_cell = _cell_at(cell_row, mapping.get("number"))
                owner_cell = _cell_at(cell_row, mapping.get("owner_name"))
                raw_number = _value_at(row, mapping.get("number"))
                building_marker = normalize_building_marker(raw_number)
                if is_commercial_sheet and building_marker:
                    current_building = building_marker
                    commercial_numbers_in_building = set()
                    continue
                if raw_number is None or _is_section_row(raw_number):
                    continue
                apartment_number = normalize_apartment_number_cell(raw_number)
                if not apartment_number:
                    continue
                if not is_commercial_sheet and not looks_like_apartment_identifier(apartment_number):
                    continue

                if is_commercial_sheet:
                    if current_building is None:
                        current_building = "1"
                    if apartment_number in commercial_numbers_in_building:
                        try:
                            current_building = str(int(current_building or "1") + 1)
                        except ValueError:
                            current_building = "2"
                        commercial_numbers_in_building = set()
                    commercial_numbers_in_building.add(apartment_number)

                owner_name_raw = str(_value_at(row, mapping.get("owner_name")) or "").strip() or None
                unsold_by_text = is_unsold_owner_name(owner_name_raw)
                unsold_by_color = _is_orange_unsold_fill(owner_cell) or _is_orange_unsold_fill(number_cell)
                owner_has_real_name = bool(owner_name_raw and not unsold_by_text)
                # Оранжевая отметка означает «не продано» только если в ФИО нет
                # реального собственника. Если собственник указан (например, кв. 180),
                # квартира не должна попадать в непроданные.
                is_unsold = bool(unsold_by_text or (unsold_by_color and not owner_has_real_name))
                owner_name = "не продано" if is_unsold else owner_name_raw
                phone = None if is_unsold else (str(_value_at(row, mapping.get("phone")) or "").strip() or None)
                finishing_type = normalize_finishing_type(_value_at(row, mapping.get("finishing_type")))
                inspection_note = _value_at(row, mapping.get("inspection_note"))
                first_inspection_value = _value_at(row, mapping.get("first_inspection_date"))
                first_inspection_present = bool(str(first_inspection_value or "").strip())
                inspection_text = str(inspection_note or "").strip()
                # Принято / АПП считаем по цвету ячейки номера помещения:
                # зелёная ячейка в колонке «№ кв» / «№ ком.» = принято.
                accepted_date = _parse_app_date(inspection_note)
                scheduled_inspection = _parse_inspection_schedule(inspection_note)
                is_app_mode = bool((_is_green_fill(number_cell) and not is_unsold) or accepted_date)

                premise_type = "commercial" if is_commercial_sheet else "apartment"
                if is_commercial_sheet:
                    source_row_id = normalize_text(f"{project.id}|commercial|building-{current_building or 'unknown'}|{apartment_number}")
                    apartment = Apartment.query.filter_by(project_id=project.id, source_row_id=source_row_id).first()
                else:
                    source_row_id = None
                    apartment = Apartment.query.filter_by(project_id=project.id, apartment_number=apartment_number, premise_type="apartment").first()
                created = apartment is None
                if apartment is None:
                    apartment = Apartment(project_id=project.id, apartment_number=apartment_number, premise_type=premise_type)
                    db.session.add(apartment)

                # Если основная таблица «Статистика» уже отметила квартиру оранжевым цветом,
                # таблица передач не должна сбрасывать этот признак простым отсутствием текста
                # «не продано». Сброс произойдёт при следующей загрузке основной статистики,
                # если оранжевая отметка там будет снята.
                if (
                    not is_commercial_sheet
                    and not is_unsold
                    and bool(getattr(apartment, "is_unsold", False))
                    and not owner_has_real_name
                ):
                    is_unsold = True
                    owner_name = "не продано" if is_unsold_owner_name(apartment.owner_name) else owner_name_raw

                apartment.premise_type = premise_type
                if is_commercial_sheet:
                    apartment.building = current_building
                if source_row_id:
                    apartment.source_row_id = source_row_id
                apartment.owner_name = owner_name
                apartment.is_unsold = is_unsold
                apartment.phone = phone
                apartment.finishing_type = finishing_type
                apartment.inspection_date = accepted_date or (scheduled_inspection.date() if isinstance(scheduled_inspection, datetime) else scheduled_inspection)
                apartment.first_inspection_date = parse_date(first_inspection_value)
                apartment.first_inspection_present = bool(first_inspection_present or scheduled_inspection)
                apartment.reinspection_date = None
                apartment.deadline_date = accepted_date
                # Dates use the internal schedule marker; free-form text in the
                # same source column is the apartment inspection comment.
                apartment.inspection_note = _inspection_schedule_marker(scheduled_inspection) or inspection_text or None
                apartment.is_app_mode = is_app_mode
                remark_deadline_value = _value_at(row, mapping.get("remark_deadline_date"))
                if remark_deadline_value is not None and str(remark_deadline_value).strip():
                    apply_app_deadline_logic(apartment, remark_deadline_value)
                elif accepted_date:
                    apply_app_deadline_logic(apartment, accepted_date + timedelta(days=60))
                if not is_app_mode and not apartment.po_status_manual:
                    apartment.po_status = "not_ready"

                if created:
                    result["created_count"] += 1
                else:
                    result["updated_count"] += 1
                if is_unsold:
                    result["unsold_count"] += 1
                elif is_app_mode:
                    result["accepted_count"] += 1
                else:
                    result["waiting_count"] += 1

        if not any(result[key] for key in ("created_count", "updated_count", "accepted_count", "waiting_count", "unsold_count")):
            raise ValueError("Не удалось распознать таблицу статистики передач: в файле не найдены ожидаемые колонки.")

        sync_log.created_count = result["created_count"]
        sync_log.updated_count = result["updated_count"]
        sync_log.missing_count = result["waiting_count"]
        sync_log.status = "success"
        sync_log.finished_at = datetime.utcnow()
        db.session.commit()
        return result
    except Exception as exc:
        db.session.rollback()
        sync_log.status = "error"
        sync_log.error_message = str(exc)
        sync_log.finished_at = datetime.utcnow()
        db.session.add(sync_log)
        db.session.commit()
        raise
