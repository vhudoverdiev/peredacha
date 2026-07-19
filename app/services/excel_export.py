from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
import re
from typing import Iterable

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from app.models import Apartment, SyncLog, Task, STATUS_DONE, STATUS_FINISHERS, STATUS_CONTRACTOR, STATUS_GUARANTEE, STATUS_CONCESSION
from app.services.excel_import import inspect_remarks_workbook
from app.services.task_service import get_setting


EXCEL_HEADER_FILL_COLOR = "FFE2F0D9"
HEADER_FILL = PatternFill(fill_type="solid", start_color=EXCEL_HEADER_FILL_COLOR, end_color=EXCEL_HEADER_FILL_COLOR)
REPORT_HEADER_FILL = PatternFill(fill_type="solid", start_color=EXCEL_HEADER_FILL_COLOR, end_color=EXCEL_HEADER_FILL_COLOR)
THIN_BORDER = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))


def style_header_row(ws) -> None:
    ws.row_dimensions[1].height = 32
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_report_header_row(ws) -> None:
    ws.row_dimensions[1].height = 32
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = REPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_borders(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def enable_wrap_text(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            if not alignment.vertical:
                alignment.vertical = "top"
            cell.alignment = alignment


def set_column_widths(ws, widths) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def auto_adjust_row_heights(ws) -> None:
    for row_idx in range(2, ws.max_row + 1):
        max_len = 0
        for cell in ws[row_idx]:
            value = str(cell.value or "")
            if value:
                max_len = max(max_len, len(value))
        if max_len > 240:
            ws.row_dimensions[row_idx].height = 90
        elif max_len > 120:
            ws.row_dimensions[row_idx].height = 62
        elif max_len > 60:
            ws.row_dimensions[row_idx].height = 45
        else:
            ws.row_dimensions[row_idx].height = 30


def apply_worksheet_style(ws, widths=None, report_header: bool = False) -> None:
    if report_header:
        style_report_header_row(ws)
    else:
        style_header_row(ws)
    apply_borders(ws)
    if widths:
        set_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    auto_adjust_row_heights(ws)


def build_export_path(filename_prefix: str, cache_key: str | None = None) -> Path:
    folder = Path(current_app.config["EXPORT_FOLDER"])
    folder.mkdir(parents=True, exist_ok=True)
    stem = f"{_safe_filename_part(filename_prefix)}_{datetime.now().strftime('%Y-%m-%d')}"
    if cache_key:
        stem = f"{stem}_{_safe_filename_part(cache_key)}"
    return folder / f"{stem}.xlsx"


def _write_only_header_row(ws, headers: list[str], fill: PatternFill) -> None:
    row = []
    for value in headers:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = Font(bold=True, color="111827")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        row.append(cell)
    ws.append(row)


EXPORT_HEADERS = [
    "Помещение",
    "Строительный номер",
    "Собственник",
    "Телефон",
    "Вид отделки",
    "Пункт",
    "Название пункта",
    "Текст замечания",
    "Ответственный",
    "Статус",
    "Приоритет",
    "Плановая дата",
    "Фактическая дата",
    "Примечание",
    "Нет в новой таблице",
]
SOURCE_EXPORT_HEADERS = EXPORT_HEADERS[:9]

EXPORT_STATUS_SUFFIXES = {
    STATUS_DONE: "(лб)",
    STATUS_FINISHERS: "(чистовики)",
    STATUS_CONTRACTOR: "(подрядчик)",
}


def export_tasks_to_excel(tasks: Iterable[Task], filename_prefix: str = "tasks_export") -> Path:
    path = build_export_path(filename_prefix)

    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи CRM"
    ws.append(EXPORT_HEADERS)

    for task in tasks:
        ws.append(
            [
                (_excel_premise_label(task.apartment) if task.apartment else ""),
                task.apartment.construction_number,
                task.apartment.owner_name,
                task.apartment.phone,
                task.apartment.finishing_type,
                task.work_point.point_number,
                task.work_point.display_name,
                task.description or task.source_cell_value,
                task.responsible.full_name if task.responsible else "",
                task.status_label(),
                task.priority,
                task.planned_date.isoformat() if task.planned_date else "",
                task.completed_date.isoformat(sep=" ", timespec="minutes") if task.completed_date else "",
                task.comment,
                "Да" if task.is_missing_in_latest_sync else "Нет",
            ]
        )
        if task.is_done:
            for cell in ws[ws.max_row]:
                new_font = copy(cell.font)
                new_font.strike = True
                cell.font = new_font

    apply_worksheet_style(ws, [18, 18, 28, 18, 18, 10, 30, 100, 28, 20, 14, 18, 20, 42, 18])
    wb.save(path)
    return path


def export_source_excel_reconstructed(tasks: Iterable[Task], filename_prefix: str) -> Path:
    tasks = list(tasks)
    path = build_export_path(filename_prefix)

    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи CRM"
    ws.append(SOURCE_EXPORT_HEADERS)

    for task in tasks:
        ws.append(
            [
                (_excel_premise_label(task.apartment) if task.apartment else ""),
                task.apartment.construction_number if task.apartment else "",
                task.apartment.owner_name if task.apartment else "",
                task.apartment.phone if task.apartment else "",
                task.apartment.finishing_type if task.apartment else "",
                task.work_point.point_number if task.work_point else "",
                task.work_point.display_name if task.work_point else "",
                _task_export_value(task, task.source_cell_value or task.description or ""),
                task.responsible.full_name if task.responsible else "",
            ]
        )
        if task.is_done:
            for cell in ws[ws.max_row]:
                new_font = copy(cell.font)
                new_font.strike = True
                cell.font = new_font

    apply_worksheet_style(ws, [18, 18, 28, 18, 18, 10, 30, 100, 28])
    wb.save(path)
    return path


def _excel_premise_label(apartment: Apartment | None) -> str:
    if apartment is None:
        return ""
    number = str(apartment.apartment_number or apartment.construction_number or f"ID {apartment.id}").strip()
    if apartment.premise_type == "commercial":
        return _normalize_report_export_text(_excel_commercial_label(number, apartment.building))
    return _normalize_report_export_text(apartment.label())


def _excel_commercial_label(number: str | None, building: str | None = None) -> str:
    text = str(number or "").strip()
    text = re.sub(r"^коммерци[яи]\s*", "", text, flags=re.IGNORECASE).strip()
    building_text = str(building or "").strip()

    pair_match = re.match(r"^к?\s*(\d+)\s*/\s*(?:к|корпус)?\s*(\d+)\s*$", text, flags=re.IGNORECASE)
    if pair_match:
        commercial_number, parsed_building = pair_match.groups()
        return f"коммерция {commercial_number}/корпус {building_text or parsed_building}"

    simple_match = re.match(r"^к?\s*(\d+)\s*$", text, flags=re.IGNORECASE)
    if simple_match:
        commercial_number = simple_match.group(1)
        if building_text:
            return f"коммерция {commercial_number}/корпус {building_text}"
        return f"коммерция {commercial_number}"

    if building_text and "корпус" not in text.lower() and "/" not in text:
        return f"коммерция {text}/корпус {building_text}".strip()
    return f"коммерция {text}".strip()


def _excel_premise_finish_label(apartment: Apartment | None) -> str:
    premise = _excel_premise_label(apartment) or "—"
    finish = (apartment.finishing_type or "").strip() if apartment else ""
    return f"{premise}\n{finish}" if finish else premise


def _excel_finish_label(apartment: Apartment | None) -> str:
    return (apartment.finishing_type or "").strip() if apartment else ""


def _task_group_key(task: Task) -> tuple[int, str]:
    apartment = task.apartment
    if apartment:
        return (apartment.id, _excel_premise_label(apartment))
    return (0, "")


def _group_tasks_by_apartment(tasks: Iterable[Task]) -> list[tuple[Apartment | None, list[Task]]]:
    groups: dict[tuple[int, str], dict[str, object]] = {}
    for task in tasks:
        key = _task_group_key(task)
        if key not in groups:
            groups[key] = {"apartment": task.apartment, "tasks": []}
        groups[key]["tasks"].append(task)
    return [(group["apartment"], group["tasks"]) for group in groups.values()]


def _task_remark_text(task: Task) -> str:
    text = str(task.description or task.source_cell_value or "").strip()
    if not text:
        return ""

    point = task.work_point
    if not point:
        return text

    prefix_candidates = []
    if point.display_name:
        prefix_candidates.append(str(point.display_name).strip())
    if point.original_column_name:
        prefix_candidates.append(str(point.original_column_name).strip())
    if point.short_name:
        prefix_candidates.append(str(point.short_name).strip())

    def strip_point_name_prefix(value: str) -> str:
        for prefix in sorted({item for item in prefix_candidates if item}, key=len, reverse=True):
            if value.lower().startswith(prefix.lower()):
                cleaned = value[len(prefix):].lstrip(" .:-–—\t")
                if cleaned:
                    return cleaned.strip()
        return value

    text = strip_point_name_prefix(text)

    point_number = re.escape(str(point.point_number).strip()) if point.point_number else r"\d{1,3}"
    text = re.sub(
        rf"^\s*(?:пункт|п\.?)?\s*№?\s*{point_number}\s*(?:[.)]|[:\-–—])?\s*",
        "",
        text,
        count=1,
        flags=re.IGNORECASE,
    )
    return strip_point_name_prefix(text.strip())


def _combined_task_lines(tasks: list[Task], include_status: bool = True, include_point: bool = False) -> str:
    lines = []
    for index, task in enumerate(tasks, start=1):
        point = task.work_point.display_name if include_point and task.work_point else ""
        text = _task_remark_text(task)
        status = f" ({task.status_label()})" if include_status else ""
        prefix = f"{index}. " if len(tasks) > 1 else ""
        if point:
            lines.append(f"{prefix}{point}: {text}{status}".strip())
        else:
            lines.append(f"{prefix}{text}{status}".strip())
    return "\n".join(lines)


def _task_completed_by_label(task: Task) -> str:
    if task.status in {STATUS_DONE, STATUS_FINISHERS}:
        return "Личная бригада"
    if task.status == STATUS_CONCESSION:
        return "Отступные"
    if task.status == STATUS_GUARANTEE:
        contractor_names = str(task.status_label() or "").strip()
        if contractor_names and contractor_names.lower() != "гарантия":
            return contractor_names
        return "Подрядчик не указан"
    if task.status == STATUS_CONTRACTOR:
        return "Подрядчик"
    return ""


def _combined_completed_by(tasks: list[Task]) -> str:
    labels = []
    seen = set()
    for task in tasks:
        label = _task_completed_by_label(task)
        if label and label not in seen:
            seen.add(label)
            labels.append(label)
    return "\n".join(labels)


def export_simple_tasks_excel(tasks: Iterable[Task], filename_prefix: str, title: str = "Задачи", *, report_header: bool = False, include_point_in_remarks: bool = False) -> Path:
    tasks = list(tasks)
    path = build_export_path(filename_prefix)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(["Помещение", "Отделка", "Замечания"])
    for apartment, group_tasks in _group_tasks_by_apartment(tasks):
        ws.append(
            [
                _excel_premise_label(apartment) or "—",
                _excel_finish_label(apartment),
                _combined_task_lines(group_tasks, include_point=include_point_in_remarks),
            ]
        )
    apply_worksheet_style(ws, [24, 24, 120], report_header=report_header)
    wb.save(path)
    return path


def _style_remark_export_title(ws, column_count: int) -> None:
    if column_count <= 0:
        return
    if column_count > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    ws.row_dimensions[1].height = 32
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14, color="111827")
    title_cell.fill = REPORT_HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, column_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = THIN_BORDER
        if col_idx > 1:
            cell.fill = REPORT_HEADER_FILL

    ws.row_dimensions[2].height = 32
    for cell in ws[2]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A3"
    if ws.max_row >= 2 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"


def export_remark_tasks_excel(tasks: Iterable[Task], filename_prefix: str, title: str = "Замечания") -> Path:
    tasks = list(tasks)
    path = build_export_path(filename_prefix)

    wb = Workbook()
    ws_open = wb.active
    ws_open.title = "Не выполненные"
    ws_done = wb.create_sheet("Выполненные")
    open_headers = ["Помещение", "Отделка", "Замечания"]
    done_headers = [*open_headers, "Кем выполнено"]
    if title:
        ws_open.append([title])
        ws_done.append([title])
    ws_open.append(open_headers)
    ws_done.append(done_headers)

    sheet_groups = (
        (ws_open, [task for task in tasks if not task.is_done], False),
        (ws_done, [task for task in tasks if task.is_done], True),
    )
    for ws, grouped_tasks, include_completed_by in sheet_groups:
        for apartment, group_tasks in _group_tasks_by_apartment(grouped_tasks):
            remark_text = _combined_task_lines(group_tasks, include_status=False, include_point=False)
            row = [
                _excel_premise_label(apartment) or "—",
                _excel_finish_label(apartment),
                remark_text,
            ]
            if include_completed_by:
                row.append(_combined_completed_by(group_tasks))
            ws.append(row)
            _apply_remark_strike_style(ws.cell(row=ws.max_row, column=3), remark_text)

    for ws, headers, widths in (
        (ws_open, open_headers, [24, 24, 120]),
        (ws_done, done_headers, [24, 24, 100, 28]),
    ):
        apply_worksheet_style(ws, widths)
        if title:
            _style_remark_export_title(ws, len(headers))
    wb.save(path)
    return path


def _apply_remark_strike_style(cell, value: object) -> None:
    text = "" if value is None else str(value)
    if not text:
        return
    if text.lstrip().startswith("-"):
        new_font = copy(cell.font)
        new_font.strike = True
        cell.font = new_font
        return
    rich_value = _quoted_strike_rich_text(text)
    if rich_value is not None:
        cell.value = rich_value


def _quoted_strike_rich_text(text: str) -> CellRichText | None:
    quote_pairs = {'"': '"', "«": "»"}
    parts = []
    plain_buffer = []
    index = 0
    found = False

    def flush_plain() -> None:
        if plain_buffer:
            parts.append("".join(plain_buffer))
            plain_buffer.clear()

    while index < len(text):
        opener = text[index]
        closer = quote_pairs.get(opener)
        if not closer:
            plain_buffer.append(text[index])
            index += 1
            continue
        end = text.find(closer, index + 1)
        if end == -1:
            plain_buffer.append(text[index])
            index += 1
            continue
        flush_plain()
        parts.append(opener)
        inner = text[index + 1:end]
        if inner:
            parts.append(TextBlock(InlineFont(strike=True), inner))
            found = True
        parts.append(closer)
        index = end + 1
    flush_plain()
    return CellRichText(parts) if found else None


def _clean_report_remark(value: object) -> str:
    return _normalize_report_export_text(value)


def _normalize_report_export_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"^\s*-\s*", "", text)
    text = re.sub(r"\(\s*лб\s*\)", "", text, flags=re.IGNORECASE)
    text = text.replace('"', "")
    text = re.sub(r"\s{2,}", " ", text).strip(" -\u2013\u2014")
    if not text:
        return ""
    return text[:1].upper() + text[1:]


def _report_task_remark(task: Task) -> str:
    remark = _clean_report_remark(task.description or task.source_cell_value)
    if task.status == STATUS_CONCESSION:
        return f"{remark} (Выданы отступные)" if remark else "Выданы отступные"
    return remark


REPORT_STATUS_SHEETS = (
    ("Выполненные", STATUS_DONE),
    ("Не выполненные", None),
    ("Отступные", STATUS_CONCESSION),
    ("Подрядчики", STATUS_CONTRACTOR),
    ("Гарантия", STATUS_GUARANTEE),
    ("Чистовики", STATUS_FINISHERS),
)


def _report_status_sheet(task: Task) -> str:
    for sheet_name, status in REPORT_STATUS_SHEETS:
        if status is not None and task.status == status:
            return sheet_name
    return "Не выполненные"


def _report_status_label(task: Task) -> str:
    label = str(task.status_label() or task.status or "").strip()
    if task.status == STATUS_GUARANTEE and label and label.lower() != "гарантия":
        return f"Гарантия: {label}"
    return label


def _report_task_executor(task: Task) -> str:
    if task.status != STATUS_DONE or not task.responsible:
        return ""
    return str(task.responsible.full_name or task.responsible.username or "").strip()


def _report_task_row(task: Task, *, include_executor: bool = False) -> list[str]:
    row = [
        (_excel_premise_label(task.apartment) if task.apartment else ""),
        _report_task_remark(task),
        _report_status_label(task),
    ]
    if include_executor:
        row.append(_report_task_executor(task))
    row.append(task.completed_date.strftime("%d.%m.%Y") if task.completed_date else "")
    return row


def export_report_tasks_excel(
    tasks: Iterable[Task],
    filename_prefix: str,
    cache_key: str | None = None,
    *,
    split_by_status: bool = False,
    include_executor: bool = False,
) -> Path:
    path = build_export_path(filename_prefix, cache_key=cache_key)
    tasks = list(tasks)

    wb = Workbook(write_only=True)
    headers = ["Помещение", "Замечание", "Статус"]
    widths = [18, 110, 28]
    if include_executor:
        headers.append("Исполнитель")
        widths.append(28)
    headers.append("Дата выполнения")
    widths.append(20)
    if split_by_status:
        grouped_tasks = {sheet_name: [] for sheet_name, _ in REPORT_STATUS_SHEETS}
        for task in tasks:
            grouped_tasks[_report_status_sheet(task)].append(task)
        for sheet_name, _ in REPORT_STATUS_SHEETS:
            ws = wb.create_sheet(title=sheet_name)
            set_column_widths(ws, widths)
            _write_only_header_row(ws, headers, REPORT_HEADER_FILL)
            for task in grouped_tasks[sheet_name]:
                ws.append(_report_task_row(task, include_executor=include_executor))
    else:
        ws = wb.create_sheet(title="Отчет")
        simple_headers = ["Помещение", "Замечание"]
        simple_widths = [18, 110]
        if include_executor:
            simple_headers.append("Исполнитель")
            simple_widths.append(28)
        simple_headers.append("Дата выполнения")
        simple_widths.append(20)
        set_column_widths(ws, simple_widths)
        _write_only_header_row(ws, simple_headers, REPORT_HEADER_FILL)
        for task in tasks:
            row = [
                (_excel_premise_label(task.apartment) if task.apartment else ""),
                _report_task_remark(task),
            ]
            if include_executor:
                row.append(_report_task_executor(task))
            row.append(task.completed_date.strftime("%d.%m.%Y") if task.completed_date else "")
            ws.append(row)
    wb.save(path)
    return path


def _append_status_suffix(value: object, suffix: str) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return suffix
    # Если отметка уже была в исходнике/предыдущем экспорте — заменяем её на актуальную,
    # а не добавляем второй хвост.
    known_suffixes = sorted((s for s in EXPORT_STATUS_SUFFIXES.values()), key=len, reverse=True)
    changed = True
    while changed:
        changed = False
        lower = text.lower().rstrip()
        for existing_suffix in known_suffixes:
            if lower.endswith(existing_suffix.lower()):
                text = text[: -len(existing_suffix)].rstrip()
                changed = True
                break
    return f"{text} {suffix}"


def _prefix_dash_for_struck_cell(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return "-"
    return text if text.startswith("-") else f"- {text}"


def _task_export_value(task: Task, current_cell_value: object) -> str:
    # Если текст правили в CRM — выгружаем именно актуальный текст из CRM.
    base = task.description or task.source_cell_value or current_cell_value or ""
    suffix = EXPORT_STATUS_SUFFIXES.get(task.status)
    value = _append_status_suffix(base, suffix) if suffix else str(base)
    return _prefix_dash_for_struck_cell(value)


def _is_tmc_column(ws, column_index: int) -> bool:
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        value = str(ws.cell(row=row_idx, column=column_index).value or "").lower()
        normalized = value.replace("ё", "е")
        if "отступ" in normalized and "тмц" in normalized:
            return True
    return False


def _safe_filename_part(value: str | None) -> str:
    text = re.sub(r"[\\/:*?\"<>|]+", " ", str(value or "").strip())
    text = re.sub(r"\s+", " ", text).strip()
    return text or "object"


def _remarks_source_quality(path: Path) -> tuple[int, int] | None:
    try:
        if not path.exists():
            return None
        info = inspect_remarks_workbook(path)
        if not info.get("ok"):
            return None
        matched_sheets = info.get("matched_sheets") or []
        return (len(matched_sheets), int(info.get("sheet_count") or 0))
    except Exception:
        return None


def resolve_source_excel_with_strikes_path(source_path: str | None = None, project_id: int | None = None) -> Path:
    if source_path:
        direct_source = Path(source_path)
        if direct_source.exists():
            return direct_source

    candidates: list[Path] = []
    seen: set[str] = set()

    def add_candidate(candidate: str | Path | None) -> None:
        if not candidate:
            return
        path = Path(candidate)
        key = str(path)
        if key in seen:
            return
        seen.add(key)
        candidates.append(path)

    add_candidate(source_path)
    if project_id is not None:
        add_candidate(get_setting(f"latest_excel_path_project_{project_id}"))
        recent_logs = (
            SyncLog.query.filter(
                SyncLog.project_id == project_id,
                SyncLog.source_type == "excel",
                SyncLog.status == "success",
                SyncLog.source_name.isnot(None),
            )
            .order_by(SyncLog.started_at.desc(), SyncLog.id.desc())
            .limit(25)
            .all()
        )
        for log in recent_logs:
            add_candidate(log.source_name)
    add_candidate(get_setting("latest_excel_path"))

    best_candidate: Path | None = None
    best_quality: tuple[int, int] = (-1, -1)
    for candidate in candidates:
        quality = _remarks_source_quality(candidate)
        if quality and quality > best_quality:
            best_candidate = candidate
            best_quality = quality

    if best_candidate is not None:
        return best_candidate

    if source_path:
        source_file = Path(source_path)
        if not source_file.exists():
            raise FileNotFoundError(f"Файл не найден: {source_file}")
    raise FileNotFoundError("Не найден корректный исходный Excel замечаний для полного экспорта. Загрузите исходный файл замечаний заново.")


def export_source_excel_with_strikes(source_path: str | None = None, project_name: str | None = None, project_id: int | None = None) -> Path:
    """Return a copy of the latest source Excel with CRM statuses applied to source cells."""
    source_file = resolve_source_excel_with_strikes_path(source_path=source_path, project_id=project_id)

    folder = Path(current_app.config["EXPORT_FOLDER"])
    folder.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    target = folder / f"{_safe_filename_part(project_name)}_{today}.xlsx"

    wb = load_workbook(source_file)
    tasks = (
        Task.query.filter(Task.is_done.is_(True))
        .filter(Task.source_sheet_name.isnot(None))
        .filter(Task.source_row_index.isnot(None))
        .filter(Task.source_column_index.isnot(None))
    )
    if project_name:
        tasks = tasks.filter(Task.project.has(name=project_name))
    tasks = tasks.order_by(Task.updated_at.asc(), Task.id.asc()).all()
    for task in tasks:
        if not task.source_sheet_name or not task.source_row_index or not task.source_column_index:
            continue
        if task.source_sheet_name not in wb.sheetnames:
            continue
        ws = wb[task.source_sheet_name]
        cell = ws.cell(row=task.source_row_index, column=task.source_column_index)
        cell.value = _task_export_value(task, cell.value)
        if not _is_tmc_column(ws, task.source_column_index):
            new_font = copy(cell.font)
            new_font.strike = True
            cell.font = new_font
    manual_query = Task.query.filter(Task.is_done.is_(True), Task.manually_edited.is_(True))
    if project_name:
        manual_query = manual_query.filter(Task.project.has(name=project_name))
    manual_tasks = [
        task for task in manual_query.order_by(Task.updated_at.asc(), Task.id.asc()).all()
        if not task.source_sheet_name or not task.source_row_index or not task.source_column_index
    ]
    if manual_tasks:
        ws_manual = wb.create_sheet("Добавлено вручную")
        ws_manual.append(["Помещение", "Пункт", "Замечание", "Статус", "Дата выполнения"])
        for task in manual_tasks:
            ws_manual.append([
                _excel_premise_label(task.apartment) if task.apartment else "",
                task.work_point.display_name if task.work_point else "",
                task.description or task.source_cell_value or "",
                task.status_label(),
                task.completed_date.strftime("%d.%m.%Y") if task.completed_date else "",
            ])
        apply_worksheet_style(ws_manual, [20, 36, 100, 20, 18])
    wb.save(target)
    return target


def export_glass_measurements_excel(rows: Iterable[dict], filename_prefix: str = "glass_measurements") -> Path:
    """Export rows from the glass measurement screen.

    rows is a list of dictionaries with keys: task, measurement, status/status_label.
    """
    folder = Path(current_app.config["EXPORT_FOLDER"])
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{_safe_filename_part(filename_prefix)}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Сделать замер"
    ws.append([
        "№ квартиры/помещения",
        "Корпус",
        "Тип помещения",
        "Текст замечания",
        "Подрядчик / категория",
        "Статус по стеклам",
        "Ширина",
        "Высота",
        "Количество",
        "Тип стеклопакета/стекла",
        "Комментарий",
    ])
    for row in rows:
        task = row.get("task")
        measurement = row.get("measurement")
        apartment = task.apartment if task else None
        ws.append([
            _excel_premise_label(apartment) if apartment else "",
            getattr(apartment, "building", "") or "",
            "Коммерция" if apartment and apartment.premise_type == "commercial" else "Квартира",
            (task.description or task.source_cell_value or "") if task else "",
            task.work_point.display_name if task and task.work_point else "",
            row.get("status_label") or (measurement.status_label() if measurement else "Без замера"),
            getattr(measurement, "width", None) or "",
            getattr(measurement, "height", None) or "",
            getattr(measurement, "quantity", None) or "",
            getattr(measurement, "glass_type", None) or "",
            getattr(measurement, "comment", None) or "",
        ])
    apply_worksheet_style(ws, [20, 14, 16, 100, 30, 22, 14, 14, 14, 28, 42])
    wb.save(path)
    return path
