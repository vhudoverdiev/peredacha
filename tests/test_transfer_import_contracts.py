from datetime import date, datetime
import unittest

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.services.transfer_import import (
    _apply_tint,
    _find_header_map,
    _inspection_schedule_marker,
    _is_app_mode,
    _is_green_fill,
    _is_not_inspected_fill,
    _is_orange_unsold_fill,
    _is_transfer_header_map,
    _parse_app_date,
    _parse_inspection_schedule,
    _theme_color_rgb,
)


class TransferImportContractsTests(unittest.TestCase):
    def test_header_detection_requires_number_and_inspection_note_plus_domain_fields(self):
        rows = [
            ["metadata", "", ""],
            ["№ кв", "Ф.И.О. дольщиков", "Телефон", "Дата осмотра", "Вид отделки"],
            ["1", "Owner", "+7999", "2026-07-29 10:00", "White box"],
        ]

        header_index, mapping = _find_header_map(rows)

        self.assertEqual(header_index, 1)
        self.assertEqual(mapping["number"], 0)
        self.assertEqual(mapping["inspection_note"], 3)
        self.assertTrue(_is_transfer_header_map(mapping))
        self.assertFalse(_is_transfer_header_map({"number": 0, "inspection_note": 1}))

    def test_app_date_parser_extracts_explicit_app_date_without_treating_plain_dates_as_app(self):
        self.assertEqual(_parse_app_date("АПП 05.08.2026"), date(2026, 8, 5))
        self.assertEqual(_parse_app_date("апп: 05/08/26"), date(2026, 8, 5))
        self.assertIsNone(_parse_app_date("05.08.2026"))
        self.assertIsNone(_parse_app_date(date(2026, 8, 5)))

    def test_inspection_schedule_parser_ignores_app_mode_and_preserves_datetime_precision(self):
        self.assertEqual(_parse_inspection_schedule("2026-07-29 10:30"), datetime(2026, 7, 29, 10, 30))
        self.assertEqual(_parse_inspection_schedule("29.07.2026"), date(2026, 7, 29))
        self.assertIsNone(_parse_inspection_schedule("АПП 05.08.2026"))
        self.assertTrue(_is_app_mode("записан АПП"))

    def test_inspection_schedule_marker_serializes_date_and_datetime_distinctly(self):
        self.assertEqual(
            _inspection_schedule_marker(datetime(2026, 7, 29, 10, 30, 15, 999)),
            "__inspection_schedule__:2026-07-29T10:30:15",
        )
        self.assertEqual(_inspection_schedule_marker(date(2026, 7, 29)), "__inspection_schedule__:2026-07-29")
        self.assertIsNone(_inspection_schedule_marker(None))

    def test_fill_detection_distinguishes_warm_not_inspected_green_and_orange_unsold_colors(self):
        workbook = Workbook()
        sheet = workbook.active
        red = sheet["A1"]
        red.fill = PatternFill(fill_type="solid", fgColor="FF0000")
        green = sheet["A2"]
        green.fill = PatternFill(fill_type="solid", fgColor="00B050")
        orange = sheet["A3"]
        orange.fill = PatternFill(fill_type="solid", fgColor="FFFBBC04")

        self.assertTrue(_is_not_inspected_fill(red))
        self.assertFalse(_is_green_fill(red))
        self.assertTrue(_is_green_fill(green))
        self.assertFalse(_is_not_inspected_fill(green))
        self.assertTrue(_is_orange_unsold_fill(orange))

    def test_theme_color_and_tint_helpers_handle_invalid_xml_and_lighten_or_darken_rgb(self):
        theme_xml = """
        <a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
          <a:themeElements>
            <a:clrScheme name="Office">
              <a:accent1><a:srgbClr val="336699"/></a:accent1>
            </a:clrScheme>
          </a:themeElements>
        </a:theme>
        """

        self.assertEqual(_theme_color_rgb(theme_xml, 4), (51, 102, 153))
        self.assertIsNone(_theme_color_rgb("<bad", 4))
        self.assertGreater(_apply_tint((100, 100, 100), 0.5)[0], 100)
        self.assertLess(_apply_tint((100, 100, 100), -0.5)[0], 100)
        self.assertEqual(_apply_tint((100, 100, 100), "bad"), (100, 100, 100))


if __name__ == "__main__":
    unittest.main()
