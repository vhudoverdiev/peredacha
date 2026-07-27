import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from config import Config
from app import create_app, db, login_manager
from app.models import (
    Apartment,
    GlassMeasurement,
    Project,
    ROLE_ADMIN,
    STATUS_NOT_STARTED,
    Task,
    User,
    WorkPoint,
)


class TestConfig(Config):
    TESTING = True
    SECRET_KEY = "glass-all-work-points-split-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    WTF_CSRF_ENABLED = False
    SESSION_COOKIE_SECURE = False


class GlassAllWorkPointsSplitTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app(TestConfig)
        self.previous_session_protection = login_manager.session_protection
        login_manager.session_protection = None
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()

        self.project = Project(name="Glass all work points QA")
        self.user = User(
            username="glass-all-points-admin",
            password_hash="unused",
            role=ROLE_ADMIN,
            all_projects_access=True,
        )
        self.apartment = Apartment(project=self.project, apartment_number="104")
        self.point_one = WorkPoint(
            point_number="1",
            original_column_name="Пункт 1",
            source_sheet_name="qa-point-1",
        )
        self.point_sixteen = WorkPoint(
            point_number="16",
            original_column_name="Пункт 16",
            source_sheet_name="qa-point-16",
        )
        self.point_twenty_two = WorkPoint(
            point_number="22",
            original_column_name="Прочее",
            source_sheet_name="manual",
        )
        db.session.add_all(
            [
                self.project,
                self.user,
                self.apartment,
                self.point_one,
                self.point_sixteen,
                self.point_twenty_two,
            ]
        )
        db.session.flush()

        self.non_glass_point_task = Task(
            source_uid="glass-all-point-one",
            project=self.project,
            apartment=self.apartment,
            work_point=self.point_one,
            description="Замечание из пункта 1",
            source_sheet_name="qa",
        )
        self.window_point_task = Task(
            source_uid="glass-all-point-sixteen",
            project=self.project,
            apartment=self.apartment,
            work_point=self.point_sixteen,
            description="Замечание из пункта 16",
            source_sheet_name="qa",
        )
        self.non_glass_ordered_task = Task(
            source_uid="glass-ordered-point-one",
            project=self.project,
            apartment=self.apartment,
            work_point=self.point_one,
            description="Заказанный замер из пункта 1",
            source_sheet_name="qa",
        )
        self.non_glass_ordered_measurement = GlassMeasurement(
            project=self.project,
            apartment=self.apartment,
            task=self.non_glass_ordered_task,
            status="ordered",
        )
        self.manual_glass_task = Task(
            source_uid="glass-manual-before-split",
            project=self.project,
            apartment=self.apartment,
            work_point=self.point_twenty_two,
            description="Большое замечание для разделения",
            source_cell_value="Большое замечание для разделения",
            source_sheet_name="manual_glass",
            manually_edited=True,
        )
        db.session.add_all(
            [
                self.non_glass_point_task,
                self.window_point_task,
                self.non_glass_ordered_task,
                self.non_glass_ordered_measurement,
                self.manual_glass_task,
            ]
        )
        db.session.commit()

        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = str(self.user.id)
            session["_fresh"] = True
            session["session_version"] = int(self.user.session_version or 0)
            session["current_project_id"] = self.project.id

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()
        login_manager.session_protection = self.previous_session_protection

    def _desktop_get(self, query: str):
        return self.client.get(
            f"/glass-measurements{query}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        )

    def test_all_tab_contains_tasks_from_every_work_point(self):
        response = self._desktop_get("?tab=all")
        page = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Замечание из пункта 1", page)
        self.assertIn("Замечание из пункта 16", page)
        self.assertIn(
            f'action="/glass/{self.non_glass_point_task.id}/need-measure"',
            page,
        )

    def test_desktop_ordered_uses_all_points_without_changing_mobile_scope(self):
        db.session.add_all(
            [
                GlassMeasurement(
                    project=self.project,
                    apartment=self.apartment,
                    task=self.manual_glass_task,
                    status="ordered",
                ),
                GlassMeasurement(
                    project=self.project,
                    apartment=self.apartment,
                    task=self.window_point_task,
                    status="measure_needed",
                ),
            ]
        )
        db.session.commit()

        desktop_response = self._desktop_get("?tab=ordered")
        mobile_response = self.client.get(
            "/glass-measurements?tab=ordered",
            headers={"User-Agent": "Mozilla/5.0 (Linux; Android 15; Mobile)"},
        )

        self.assertIn(
            "Заказанный замер из пункта 1",
            desktop_response.get_data(as_text=True),
        )
        self.assertNotIn(
            "Заказанный замер из пункта 1",
            mobile_response.get_data(as_text=True),
        )

        desktop_order_response = self._desktop_get("?tab=order")
        mobile_order_response = self.client.get(
            "/glass-measurements?tab=order",
            headers={"User-Agent": "Mozilla/5.0 (Linux; Android 15; Mobile)"},
        )
        self.assertIn(
            "Замечание из пункта 16",
            desktop_order_response.get_data(as_text=True),
        )
        self.assertNotIn(
            "Замечание из пункта 16",
            mobile_order_response.get_data(as_text=True),
        )

        desktop_export = self.client.get(
            "/glass/order/export?scope=order",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        )
        mobile_export = self.client.get(
            "/glass/order/export?scope=order",
            headers={"User-Agent": "Mozilla/5.0 (Linux; Android 15; Mobile)"},
        )
        desktop_sheet = load_workbook(
            BytesIO(desktop_export.data),
            read_only=True,
        ).active
        mobile_sheet = load_workbook(
            BytesIO(mobile_export.data),
            read_only=True,
        ).active
        desktop_cells = "\n".join(
            str(value or "")
            for row in desktop_sheet.iter_rows(values_only=True)
            for value in row
        )
        mobile_cells = "\n".join(
            str(value or "")
            for row in mobile_sheet.iter_rows(values_only=True)
            for value in row
        )
        self.assertIn("Замечание из пункта 16", desktop_cells)
        self.assertNotIn("Замечание из пункта 16", mobile_cells)

    def test_split_manual_glass_task_second_part_can_be_measured(self):
        db.session.add(
            GlassMeasurement(
                project=self.project,
                apartment=self.apartment,
                task=self.manual_glass_task,
                status="ordered",
            )
        )
        db.session.commit()
        split_response = self.client.post(
            f"/tasks/{self.manual_glass_task.id}/split",
            json={
                "current_text": "Первая часть большого замечания",
                "new_text": "Вторая часть — отдельный новый замер",
                "current_status": STATUS_NOT_STARTED,
                "new_status": STATUS_NOT_STARTED,
            },
        )

        self.assertEqual(split_response.status_code, 200)
        split_payload = split_response.get_json()
        self.assertTrue(split_payload["ok"])
        split_task = db.session.get(Task, split_payload["new_task_id"])
        self.assertEqual(split_task.source_sheet_name, "manual_split")
        self.assertEqual(split_task.work_point.point_number, "22")

        all_response = self._desktop_get("?tab=all")
        all_page = all_response.get_data(as_text=True)
        self.assertIn("Вторая часть — отдельный новый замер", all_page)
        self.assertIn(f'action="/glass/{split_task.id}/need-measure"', all_page)

        measure_response = self.client.post(
            f"/glass/{split_task.id}/need-measure",
            data={"return_tab": "all"},
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "Accept": "application/json",
            },
        )

        self.assertEqual(measure_response.status_code, 200)
        self.assertTrue(measure_response.get_json()["ok"])
        measurement = GlassMeasurement.query.filter_by(task_id=split_task.id).one()
        self.assertEqual(measurement.status, "measure_needed")

        order_response = self._desktop_get(
            "?tab=order&q=%D0%92%D1%82%D0%BE%D1%80%D0%B0%D1%8F"
        )
        self.assertIn(
            "Вторая часть — отдельный новый замер",
            order_response.get_data(as_text=True),
        )

    def test_need_measure_ajax_does_not_force_reload_when_filtered_row_disappears(self):
        response = self.client.post(
            f"/glass/{self.non_glass_point_task.id}/need-measure",
            data={"return_tab": "all"},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "X-Requested-With": "XMLHttpRequest",
                "Accept": "application/json",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content_type, "application/json")
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["task_id"], self.non_glass_point_task.id)

        script = Path("app/static/script.js").read_text(encoding="utf-8")
        start = script.index("const submitGlassNeedMeasureForm")
        end = script.index("const bindGlassReturnForm", start)
        need_measure_block = script[start:end]
        self.assertIn("replaceGlassNeedMeasureFormLocally(form, data)", need_measure_block)
        self.assertIn("document.addEventListener('submit', event =>", need_measure_block)
        self.assertNotIn("location.reload", need_measure_block)

    def test_order_delete_uses_ajax_without_redirecting_the_page(self):
        measurement = GlassMeasurement(
            project=self.project,
            apartment=self.apartment,
            task=self.window_point_task,
            status="measure_needed",
        )
        db.session.add(measurement)
        db.session.commit()
        measurement_id = measurement.id

        page_response = self._desktop_get("?tab=order")
        page = page_response.get_data(as_text=True)
        self.assertIn('class="js-glass-delete-form"', page)
        self.assertIn(f'id="glass-order-delete-{measurement_id}"', page)

        response = self.client.post(
            "/glass/delete",
            data={
                "scope": "order",
                "measurement_ids": str(measurement_id),
            },
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "X-Requested-With": "XMLHttpRequest",
                "Accept": "application/json",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content_type, "application/json")
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["scope"], "order")
        self.assertEqual(payload["deleted_count"], 1)
        self.assertIsNone(db.session.get(GlassMeasurement, measurement_id))

        script = Path("app/static/script.js").read_text(encoding="utf-8")
        start = script.index("const bindGlassDeleteForm")
        end = script.index("const submitGlassNeedMeasureForm", start)
        delete_block = script[start:end]
        self.assertIn("event.preventDefault()", delete_block)
        self.assertIn("fetch(form.action", delete_block)
        self.assertIn("orderCard.remove()", delete_block)
        self.assertNotIn("location.reload", delete_block)
        init_start = script.index("const initGlassPageScope")
        init_end = script.index("initGlassPageScope(document)", init_start)
        init_block = script[init_start:init_end]
        self.assertIn("if (!isMobileViewport())", init_block)
        self.assertIn(".js-glass-delete-form", init_block)


if __name__ == "__main__":
    unittest.main()
