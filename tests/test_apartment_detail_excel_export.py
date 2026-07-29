import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app import create_app, db, login_manager
from app.models import Apartment, Project, ROLE_ADMIN, STATUS_DONE, Task, User, WorkPoint
from config import Config


class TestConfig(Config):
    TESTING = True
    SECRET_KEY = "apartment-detail-export-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    WTF_CSRF_ENABLED = False
    SESSION_COOKIE_SECURE = False


class ApartmentDetailExcelExportTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()

        class RuntimeConfig(TestConfig):
            EXPORT_FOLDER = str(Path(self.tempdir.name) / "exports")

        self.app = create_app(RuntimeConfig)
        self.previous_session_protection = login_manager.session_protection
        login_manager.session_protection = None
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()

        self.project = Project(name="Apartment Export QA")
        self.user = User(
            username="apartment-export-admin",
            password_hash="unused",
            role=ROLE_ADMIN,
            all_projects_access=True,
        )
        self.apartment = Apartment(
            project=self.project,
            apartment_number="42",
            finishing_type="Белая",
            owner_name="Иванов Иван Иванович",
            phone="+7 900 123-45-67",
        )
        open_point = WorkPoint(point_number="1", short_name="Пункт 1", source_sheet_name="qa")
        done_point = WorkPoint(point_number="2", short_name="Пункт 2", source_sheet_name="qa")
        self.open_task = Task(
            source_uid="apartment-export-open",
            project=self.project,
            apartment=self.apartment,
            work_point=open_point,
            description="Открытое замечание",
        )
        self.done_task = Task(
            source_uid="apartment-export-done",
            project=self.project,
            apartment=self.apartment,
            work_point=done_point,
            description="Выполненное замечание",
            status=STATUS_DONE,
            is_done=True,
        )
        db.session.add_all(
            [
                self.project,
                self.user,
                self.apartment,
                open_point,
                done_point,
                self.open_task,
                self.done_task,
            ]
        )
        db.session.commit()

        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = str(self.user.id)
            session["_fresh"] = True
            session["session_version"] = int(self.user.session_version or 0)
            session["current_project_id"] = self.project.id

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()
        login_manager.session_protection = self.previous_session_protection
        self.tempdir.cleanup()

    def test_desktop_apartment_card_has_excel_download_link(self):
        response = self.client.get(
            f"/apartments/{self.apartment.id}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        )

        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn(f"/apartments/{self.apartment.id}/remarks/export", html)
        self.assertIn("apartment-detail-export-btn", html)
        self.assertIn("d-none d-md-inline-flex", html)
        self.assertIn("<span>Скачать Excel</span>", html)

    def test_apartment_export_download_contains_all_apartment_remarks(self):
        response = self.client.get(f"/apartments/{self.apartment.id}/remarks/export")

        try:
            self.assertEqual(response.status_code, 200)
            self.assertIn("attachment", response.headers.get("Content-Disposition", ""))
            self.assertIn(".xlsx", response.headers.get("Content-Disposition", ""))

            workbook = load_workbook(BytesIO(response.data))
            try:
                self.assertIn("Не выполненные", workbook.sheetnames)
                self.assertIn("Выполненные", workbook.sheetnames)

                open_rows = list(workbook["Не выполненные"].iter_rows(values_only=True))
                done_rows = list(workbook["Выполненные"].iter_rows(values_only=True))

                self.assertEqual(open_rows[0][0], "Карточка помещения: кв 42")
                self.assertEqual(open_rows[1][:3], ("Помещение", "Отделка", "Замечания"))
                self.assertEqual(open_rows[2][:3], ("Кв 42", "Белая", "Открытое замечание"))

                self.assertEqual(done_rows[0][0], "Карточка помещения: кв 42")
                self.assertEqual(done_rows[1][:4], ("Помещение", "Отделка", "Замечания", "Кем выполнено"))
                self.assertEqual(done_rows[2][:4], ("Кв 42", "Белая", "Выполненное замечание", "Личная бригада"))
            finally:
                workbook.close()
        finally:
            response.close()

    def test_apartment_detail_renders_owner_and_phone_on_desktop_and_mobile(self):
        for user_agent in (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Mozilla/5.0 (Linux; Android 15; Mobile)",
        ):
            with self.subTest(user_agent=user_agent):
                response = self.client.get(
                    f"/apartments/{self.apartment.id}",
                    headers={"User-Agent": user_agent},
                )

                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
                self.assertIn("ФИО собственника", html)
                self.assertIn("Иванов Иван Иванович", html)
                self.assertIn("Номер телефона", html)
                self.assertIn("+7 900 123-45-67", html)


if __name__ == "__main__":
    unittest.main()
