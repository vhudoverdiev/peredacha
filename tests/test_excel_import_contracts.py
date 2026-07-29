from datetime import datetime, timedelta
from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import Config
from app import create_app, db
from app.models import Apartment, Project, SyncLog, Task, WorkPoint
from app.services.excel_import import (
    inspect_remarks_workbook,
    mark_missing_tasks,
    mark_stale_running_sync_logs,
    preview_excel,
    save_upload,
    workbook_sheets_to_rows_with_strikes,
    worksheet_to_rows,
    worksheet_to_rows_with_strikes,
)
from werkzeug.datastructures import FileStorage


class TestConfig(Config):
    TESTING = True
    SECRET_KEY = "excel-import-contracts-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    WTF_CSRF_ENABLED = False
    SESSION_COOKIE_SECURE = False


class ExcelImportContractsTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app(TestConfig)
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()

    def _workbook_path(self, name="excel-import-contract.xlsx"):
        return self.app.config["UPLOAD_FOLDER"] / name if hasattr(self.app.config["UPLOAD_FOLDER"], "__truediv__") else None

    def _save_workbook(self, filename: str, workbook: Workbook):
        from pathlib import Path

        path = Path(self.app.config["UPLOAD_FOLDER"]) / filename
        workbook.save(path)
        return path

    def test_mark_stale_running_sync_logs_only_marks_expired_logs_for_selected_project(self):
        project = Project(name="Stale sync project")
        other_project = Project(name="Other project")
        now = datetime(2026, 7, 29, 12, 0)
        stale = SyncLog(project=project, source_type="excel", status="running", started_at=now - timedelta(minutes=6))
        fresh = SyncLog(project=project, source_type="excel", status="running", started_at=now - timedelta(minutes=4))
        other = SyncLog(project=other_project, source_type="excel", status="running", started_at=now - timedelta(minutes=10))
        db.session.add_all([project, other_project, stale, fresh, other])
        db.session.commit()

        changed = mark_stale_running_sync_logs(project.id, now=now)

        self.assertEqual(changed, 1)
        self.assertEqual(stale.status, "error")
        self.assertEqual(stale.finished_at, now)
        self.assertIn("не завершилась", stale.error_message)
        self.assertEqual(fresh.status, "running")
        self.assertEqual(other.status, "running")

    def test_mark_missing_tasks_flags_visible_active_tasks_and_ignores_archived_hidden_and_seen(self):
        project = Project(name="Missing tasks project")
        apartment = Apartment(project=project, apartment_number="1")
        visible_point = WorkPoint(point_number="10")
        hidden_point = WorkPoint(point_number="5")
        seen = Task(source_uid="seen", project=project, apartment=apartment, work_point=visible_point)
        missing = Task(source_uid="missing", project=project, apartment=apartment, work_point=visible_point)
        hidden = Task(source_uid="hidden", project=project, apartment=apartment, work_point=hidden_point)
        archived = Task(source_uid="archived", project=project, apartment=apartment, work_point=visible_point, is_archived=True)
        db.session.add_all([project, apartment, visible_point, hidden_point, seen, missing, hidden, archived])
        db.session.commit()

        changed = mark_missing_tasks(project.id, {"seen"})
        db.session.commit()

        self.assertEqual(changed, 1)
        self.assertFalse(seen.is_missing_in_latest_sync)
        self.assertTrue(missing.is_missing_in_latest_sync)
        self.assertFalse(hidden.is_missing_in_latest_sync)
        self.assertFalse(archived.is_missing_in_latest_sync)

    def test_save_upload_uses_secure_filename_and_timestamped_upload_folder(self):
        file = FileStorage(stream=BytesIO(b"content"), filename="../../bad name.xlsx")
        fixed_now = datetime(2026, 7, 29, 12, 34, 56)

        with patch("app.services.excel_import.datetime") as datetime_mock:
            datetime_mock.utcnow.return_value = fixed_now
            path = save_upload(file)

        self.assertTrue(path.exists())
        self.assertEqual(path.name, "20260729_123456_bad_name.xlsx")
        self.assertEqual(path.read_bytes(), b"content")

    def test_worksheet_readers_return_values_struck_cells_and_orange_unsold_cells(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Remarks"
        sheet["A1"] = "Header"
        sheet["A2"] = "Done remark"
        sheet["A2"].font = Font(strike=True)
        sheet["B2"] = "Unsold"
        sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFBBC04")
        path = self._save_workbook("strikes-orange.xlsx", workbook)

        title, rows = worksheet_to_rows(path)
        struck_title, struck_rows, struck, orange = worksheet_to_rows_with_strikes(path)

        self.assertEqual(title, "Remarks")
        self.assertEqual(rows[1][0], "Done remark")
        self.assertEqual(struck_title, "Remarks")
        self.assertEqual(struck_rows[1][1], "Unsold")
        self.assertIn((2, 1), struck)
        self.assertIn((2, 2), orange)

    def test_workbook_sheet_detection_keeps_only_remark_like_sheets_and_preview_is_bounded(self):
        workbook = Workbook()
        remark = workbook.active
        remark.title = "Remarks"
        remark.append(["Квартира", "Пункт 10"])
        remark.append(["1", "Replace handle"])
        empty = workbook.create_sheet("Empty")
        empty.append([""])
        unrelated = workbook.create_sheet("Notes")
        unrelated.append(["Name", "Comment"])
        unrelated.append(["Only", "notes"])
        path = self._save_workbook("remark-detection.xlsx", workbook)

        sheets = workbook_sheets_to_rows_with_strikes(path)
        info = inspect_remarks_workbook(path)
        preview = preview_excel(path, limit=1)

        self.assertEqual([sheet[0] for sheet in sheets], ["Remarks"])
        self.assertEqual(info["matched_sheets"], ["Remarks"])
        self.assertEqual(info["sheet_count"], 3)
        self.assertEqual(preview["active_sheet"], "Remarks")
        self.assertEqual(len(preview["rows"]), 1)


if __name__ == "__main__":
    unittest.main()
