import re
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from config import Config
from app import create_app, db, login_manager
from app.models import Apartment, Project, ROLE_ADMIN, User
from app.services.task_service import dashboard_stats
from app.services.transfer_import import inspect_transfer_workbook, sync_transfer_statistics


class TestConfig(Config):
    TESTING = True
    SECRET_KEY = "dashboard-inspection-fill-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    WTF_CSRF_ENABLED = False
    SESSION_COOKIE_SECURE = False


GOOGLE_THEME = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Google">
  <a:themeElements>
    <a:clrScheme name="Google">
      <a:dk1><a:srgbClr val="000000"/></a:dk1>
      <a:lt1><a:srgbClr val="FFFFFF"/></a:lt1>
      <a:dk2><a:srgbClr val="000000"/></a:dk2>
      <a:lt2><a:srgbClr val="FFFFFF"/></a:lt2>
      <a:accent1><a:srgbClr val="4285F4"/></a:accent1>
      <a:accent2><a:srgbClr val="EA4335"/></a:accent2>
      <a:accent3><a:srgbClr val="FBBC04"/></a:accent3>
      <a:accent4><a:srgbClr val="34A853"/></a:accent4>
      <a:accent5><a:srgbClr val="FF6D01"/></a:accent5>
      <a:accent6><a:srgbClr val="46BDC6"/></a:accent6>
      <a:hlink><a:srgbClr val="1155CC"/></a:hlink>
      <a:folHlink><a:srgbClr val="1155CC"/></a:folHlink>
    </a:clrScheme>
  </a:themeElements>
</a:theme>
"""


class DashboardInspectionFillStatsTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app(TestConfig)
        self.previous_session_protection = login_manager.session_protection
        login_manager.session_protection = None
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()
        self.tempdir = tempfile.TemporaryDirectory()

        self.project = Project(name="Inspection fill QA")
        self.user = User(username="inspection-fill-admin", password_hash="unused", role=ROLE_ADMIN)
        db.session.add_all([self.project, self.user])
        db.session.commit()

        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = str(self.user.id)
            session["_fresh"] = True
            session["session_version"] = int(self.user.session_version or 0)
            session["current_project_id"] = self.project.id

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()
        self.tempdir.cleanup()
        login_manager.session_protection = self.previous_session_protection

    def _statistics_workbook(self) -> Path:
        workbook = Workbook()
        workbook.loaded_theme = GOOGLE_THEME
        sheet = workbook.active
        sheet.title = "Статистика"
        sheet.append([
            "№ кв",
            "Ф.И.О. дольщиков",
            "Телефон",
            "Вид отделки",
            "Дата осмотра",
            "Дата первичного осмотра",
        ])
        rows = [
            ["1", "Owner 1", "+1", "Белая", datetime(2026, 7, 1, 10, 0), datetime(2026, 7, 1)],
            ["2", "Owner 2", "+2", "Белая", datetime(2026, 7, 2, 10, 0), datetime(2026, 7, 2)],
            ["3", "Owner 3", "+3", "Белая", datetime(2099, 7, 3, 10, 0), None],
            ["4", "Owner 4", "+4", "Белая", None, None],
            ["5", "Owner 5", "+5", "Белая", None, None],
        ]
        for row in rows:
            sheet.append(row)

        # Theme accent2 is red, accent3 is yellow and accent4 is green.
        sheet["E2"].fill = PatternFill(patternType="solid", fgColor=Color(theme=5))
        sheet["E3"].fill = PatternFill(patternType="solid", fgColor=Color(theme=6))
        sheet["E4"].fill = PatternFill(patternType="solid", fgColor=Color(theme=7))
        sheet["E5"].fill = PatternFill(patternType="solid", fgColor="FFFFFFFF")

        path = Path(self.tempdir.name) / "transfer-statistics.xlsx"
        workbook.save(path)
        return path

    def test_red_and_yellow_cells_are_the_only_not_inspected_rows(self):
        path = self._statistics_workbook()
        self.assertTrue(inspect_transfer_workbook(path)["ok"])

        result = sync_transfer_statistics(path, project_name=self.project.name)

        self.assertEqual(result["created_count"], 5)
        flags = {
            apartment.apartment_number: apartment.first_inspection_present
            for apartment in Apartment.query.order_by(Apartment.id.asc()).all()
        }
        self.assertEqual(
            flags,
            {"1": False, "2": False, "3": True, "4": True, "5": True},
        )
        stats = dashboard_stats(self.project.id)
        self.assertEqual(stats["inspected"], 3)
        self.assertEqual(stats["not_inspected"], 2)

    def test_desktop_and_mobile_dashboard_render_the_same_fill_based_counts(self):
        sync_transfer_statistics(self._statistics_workbook(), project_name=self.project.name)

        for user_agent in (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Mozilla/5.0 (Linux; Android 15; Mobile)",
        ):
            with self.subTest(user_agent=user_agent):
                response = self.client.get("/", headers={"User-Agent": user_agent})
                self.assertEqual(response.status_code, 200)
                html = response.get_data(as_text=True)
                inspection_card = re.search(
                    r'<article class="dashboard-info-card">\s*'
                    r'<div class="dashboard-info-title"><i class="bi bi-eye"></i>.*?</article>',
                    html,
                    re.DOTALL,
                )
                self.assertIsNotNone(inspection_card)
                self.assertIn("<b>3</b>", inspection_card.group(0))
                self.assertIn("<b>2</b>", inspection_card.group(0))


if __name__ == "__main__":
    unittest.main()
