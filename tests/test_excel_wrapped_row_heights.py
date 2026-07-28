import unittest
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services.excel_export import (
    _write_only_data_row,
    auto_adjust_row_heights,
    estimate_excel_row_height,
)


class ExcelWrappedRowHeightTests(unittest.TestCase):
    def test_explicit_remark_lines_expand_row_beyond_old_fixed_limit(self):
        remarks = "\n".join(f"{index}. Замечание по помещению" for index in range(1, 14))

        height = estimate_excel_row_height(["кв 1", "Белая", remarks], [24, 24, 120])

        self.assertGreater(height, 190)

    def test_long_wrapped_text_uses_column_width_and_preserves_existing_height(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 30
        sheet.append(["Помещение", "Замечания"])
        sheet.append(["кв 1", "Очень длинное замечание " * 20])
        sheet.append(["кв 2", "Короткое"])
        sheet.row_dimensions[3].height = 48

        auto_adjust_row_heights(sheet)

        self.assertGreater(sheet.row_dimensions[2].height, 120)
        self.assertEqual(sheet.row_dimensions[3].height, 48)

    def test_streaming_report_rows_keep_calculated_height_in_saved_xlsx(self):
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Отчёт")
        sheet.append(["Помещение", "Замечания"])
        remarks = "\n".join(f"{index}. Замечание" for index in range(1, 14))
        _write_only_data_row(sheet, ["кв 1", remarks], [18, 110], 2)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        saved_sheet = load_workbook(output)["Отчёт"]

        self.assertTrue(saved_sheet["B2"].alignment.wrap_text)
        self.assertGreater(saved_sheet.row_dimensions[2].height, 190)


if __name__ == "__main__":
    unittest.main()
