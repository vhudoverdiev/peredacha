import unittest
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from config import Config
from app import create_app, db, login_manager
from app.models import (
    Apartment,
    GlassMeasurement,
    MaterialRequest,
    MaterialRequestItem,
    Project,
    ROLE_ADMIN,
    Task,
    User,
    WorkPoint,
)
from app.routes import _material_request_display_rows


class TestConfig(Config):
    TESTING = True
    SECRET_KEY = "material-request-apartment-column-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    WTF_CSRF_ENABLED = False
    SESSION_COOKIE_SECURE = False


class MaterialRequestApartmentColumnTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app(TestConfig)
        self.previous_session_protection = login_manager.session_protection
        login_manager.session_protection = None
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()

        self.project = Project(name="Material request apartment QA")
        self.user = User(username="material-request-admin", password_hash="unused", role=ROLE_ADMIN)
        self.apartment = Apartment(project=self.project, apartment_number="8")
        self.work_point = WorkPoint(point_number="QA-MATERIAL", source_sheet_name="qa")
        self.task = Task(
            source_uid="material-request-apartment-column",
            project=self.project,
            apartment=self.apartment,
            work_point=self.work_point,
            description="Apartment column QA",
        )
        self.material_request = MaterialRequest(
            project=self.project,
            author=self.user,
            request_date=date(2026, 7, 20),
            title="Заявка из замеров №5",
            comment="Автоматическая заявка из замеров стеклопакетов",
        )
        self.first_item = MaterialRequestItem(
            name="Стеклопакет 635×2085 — 40мм кв 8",
            quantity=2,
            unit="шт",
        )
        self.second_item = MaterialRequestItem(
            name="Стеклопакет 412×1445 — 40мм кв 8",
            quantity=1,
            unit="шт",
        )
        self.material_request.items.extend([self.first_item, self.second_item])
        self.measurement = GlassMeasurement(
            project=self.project,
            task=self.task,
            apartment=self.apartment,
            status="ordered",
            ordered_at=date(2026, 7, 20),
            material_request_item=self.first_item,
        )
        db.session.add_all([self.project, self.user, self.work_point, self.material_request, self.measurement])
        db.session.commit()

        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = str(self.user.id)
            session["_fresh"] = True
            session["session_version"] = int(self.user.session_version or 0)
            session["current_project_id"] = self.project.id

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()
        login_manager.session_protection = self.previous_session_protection

    def test_desktop_detail_has_separate_apartment_column_and_clean_names(self):
        response = self.client.get(
            f"/materials/request/{self.material_request.id}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        )

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("№ квартиры", html)
        self.assertIn(">8</td>", html)
        self.assertIn("Стеклопакет 635×2085 — 40мм", html)
        self.assertNotIn("Стеклопакет 635×2085 — 40мм кв 8", html)
        self.assertNotIn("Стеклопакет 412×1445 — 40мм кв 8", html)

    def test_mobile_detail_layout_is_unchanged(self):
        response = self.client.get(
            f"/materials/request/{self.material_request.id}",
            headers={"User-Agent": "Mozilla/5.0 (Linux; Android 15; Mobile)"},
        )

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertNotIn("№ квартиры", html)
        self.assertIn("Стеклопакет 635×2085 — 40мм кв 8", html)

    def test_excel_export_has_apartment_column_and_clean_names(self):
        response = self.client.get(f"/materials/request/{self.material_request.id}/export")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
        sheet = workbook["Заявка"]
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            ("Дата", "Название заявки", "№ квартиры", "Наименование", "Количество", "Ед. измерения"),
        )
        self.assertEqual(rows[1][2], "8")
        self.assertEqual(rows[1][3], "Стеклопакет 635×2085 — 40мм")
        self.assertEqual(rows[2][2], "8")
        self.assertEqual(rows[2][3], "Стеклопакет 412×1445 — 40мм")
        workbook.close()

    def test_manual_request_name_is_not_modified_without_measurement_link(self):
        manual_request = MaterialRequest(
            project=self.project,
            author=self.user,
            request_date=date(2026, 7, 20),
            title="Ручная заявка",
        )
        manual_request.items.append(MaterialRequestItem(name="Профиль кв 8", quantity=1, unit="шт"))
        db.session.add(manual_request)
        db.session.commit()

        rows = _material_request_display_rows(manual_request)

        self.assertEqual(rows[0]["apartment_number"], "")
        self.assertEqual(rows[0]["display_name"], "Профиль кв 8")


if __name__ == "__main__":
    unittest.main()
