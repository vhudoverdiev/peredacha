from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import load_workbook

from app import create_app, db, login_manager
from app.models import Apartment, Project, ROLE_MANAGER, User
from config import Config


class TestConfig(Config):
    TESTING = True
    SECRET_KEY = "apartments-filtered-export-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    WTF_CSRF_ENABLED = False
    SESSION_COOKIE_SECURE = False


class ApartmentsFilteredExportTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app(TestConfig)
        self.previous_session_protection = login_manager.session_protection
        login_manager.session_protection = None
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()

        self.project = Project(name="Filtered export project")
        self.user = User(
            username="filtered-export-manager",
            role=ROLE_MANAGER,
            project_id=None,
            all_projects_access=True,
        )
        self.user.set_password("Strong-test-password-2026!")
        self.matching = Apartment(
            project=self.project,
            apartment_number="125",
            owner_name="Иванов",
            finishing_type="Белая",
        )
        self.other = Apartment(
            project=self.project,
            apartment_number="126",
            owner_name="Петров",
            finishing_type="Черновая",
        )
        db.session.add_all([self.project, self.user, self.matching, self.other])
        db.session.commit()

        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = str(self.user.id)
            session["_fresh"] = True
            session["session_version"] = int(self.user.session_version or 0)
            session["current_project_id"] = self.project.id

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()
        login_manager.session_protection = self.previous_session_protection

    def test_filtered_page_exposes_export_link_with_the_same_query(self):
        response = self.client.get("/apartments", query_string={"q": "125"})
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('data-ajax-pagination-sync="apartments-export"', html)
        self.assertIn('/apartments/export?q=125', html)

    def test_filtered_excel_contains_only_apartments_from_current_filter(self):
        response = self.client.get("/apartments/export", query_string={"q": "125"})
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        values = [
            str(cell or "")
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
        ]
        content = "\n".join(values)
        self.assertIn("125", content)
        self.assertNotIn("126", content)

    def test_ajax_refresh_updates_export_href_outside_the_cards_container(self):
        script = Path(self.app.root_path) / "static" / "script.js"
        source = script.read_text(encoding="utf-8")
        sync = source.index("currentRoot.querySelectorAll('[data-ajax-pagination-sync]')")
        history = source.index("window.history.pushState", sync)
        self.assertLess(sync, history)
        self.assertIn("document.documentElement.classList.contains('desktop-like-pointer')", source[:sync])
        self.assertIn("nextRoot.querySelector", source[sync:history])
        self.assertIn("morphPaginationNode(currentNode, nextNode)", source[sync:history])


if __name__ == "__main__":
    unittest.main()
